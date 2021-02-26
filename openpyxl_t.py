import pandas as pd
import os
import zipfile
from sqlalchemy import create_engine
from settings import Settings
from emailControl import EmailControl
import datetime
import xlwings as xw
import openpyxl
from openpyxl import Workbook, load_workbook  #可以向不同的sheet写入数据
from openpyxl.styles import Font, Border, Side, PatternFill, colors, Alignment # 设置字体风格为Times New Roman，大小为16，粗体、斜体，颜色蓝色

# wb1 = Workbook()				# 创建新工作簿时需要使用save保存路径生效
# wb1.active.title = "New Shit"
# ws1 = wb1.create_sheet('111')
# wb1.save(r'D:\Users\Administrator\Desktop\\输出文件\日本花费明细14表.xlsx')
# wb1.close


filePath = r'D:\Users\Administrator\Desktop\\输出文件\2020.12.18 日本本月产品花费55表.xlsx'

writer = pd.ExcelWriter(filePath, engine='openpyxl')         # 初始化写入对象
wb = load_workbook(filePath)

print(wb.sheetnames)
print(wb.active)
# wb.active.title = "New Shit"
# print(wb.active)

sheet = wb.get_sheet_by_name("直发成本2")
# sheet['C3'] = 'Hello world!00000099999'
print(sheet.max_row)    # 10     <-最大行数
print(sheet.max_column)    # 5     <-最大列数

# sheet.cell('A2:A{}'.format(sheet.max_row)).delete
# sheet['A2'].alignment = Alignment(horizontal='center',vertical='center')
for i in range(3,5):
	i = i + 1
	# print(i+1)
	# print(sheet.cell(2,i+1).value) 
	clVal = sheet.cell(2,i).value
	# AA ='A2'.format(AA,j)
	AB = 2
	# # rlVal = sheet.cell(2,i).value
	for j in range(sheet.max_row):
		j = j + 2
		# print(j+1) 
		# print(sheet.cell(j,i).value) 
		# print(sheet.cell(j,i).value == clVal)
		# print(sheet.cell(j,i).value != '')
		if sheet.cell(j,i).value == clVal and sheet.cell(j,i).value != '':
			sheet.merge_cells(start_row=AB, start_column=i, end_row=j, end_column=i)
			# sheet.cell(AB, i).value = clVal
			# sheet.cell(AB, i).alignment = Alignment(horizontal='center',vertical='center')
			sheet.cell(AB, i).alignment  = Alignment(horizontal='center',vertical='center')
			# sheet.cell(AB, i).alignment  = Alignment(horizontal='center',vertical='center')
			# sheet.merge_cells('{}:A{}'.format(AA,j))
		else:
			clVal = sheet.cell(j,i).value
			# AA = 'A{}'.format(j)
			AB = j
	# 	if sheet.cell(j,i).value == clVal and sheet.cell(j,i).value != '':
	# 		sheet.merge_cells('A{}:A{}'.format(2,j))





wb.save(filePath)

# sheet = wb.get_sheet_by_name("总表成本")
# # print(sheet["C"])
# # print(sheet["4"])
# print(sheet["C4"].value)    # c4     <-第C4格的值
# print(sheet.max_row)    # 10     <-最大行数
# print(sheet.max_column)    # 5     <-最大列数



# wb.save
# wb1.close


# writer.book = book                                           # 将数据写入excel中的sheet2表,sheet_name改变后即是新增一个sheet
# listTValue[0].to_excel(excel_writer=writer, sheet_name=list_value[0], index=False)
# writer.save()
# writer.close()




# import xlsxwriter
# from execl import My_DataFrame


# DF=My_DataFrame({'A':[1,2,2,2,3,3],'B':[1,1,1,1,1,1],'C':[1,1,1,1,1,1],'D':[1,1,1,1,1,1]})
# DF.my_mergewr_excel('000_2.xlsx',['A'],['B','C'])