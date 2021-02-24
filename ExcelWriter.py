import pandas as pd
import numpy as np
demo_excel = pd.ExcelFile(r'D:\demo.xlsx')


writer = pd.ExcelWriter(r'D:\demo.xlsx')
print(type(writer))



'''创建数据框1'''
df1 = pd.DataFrame({'V1':np.random.rand(100),
          'V2 ':np.random.rand(100),
          'V3':np.random.rand(100)})
df1.to_excel(writer,sheet_name='sheet1',index=False)
 
'''创建数据框2'''
df2 = pd.DataFrame({'V1':np.random.rand(100),
          'V2 ':np.random.rand(100),
          'V3':np.random.rand(100)})
df2.to_excel(writer,sheet_name='sheet2',index=False)
 
'''创建数据框3'''
df3 = pd.DataFrame({'V1':np.random.rand(100),
          'V2 ':np.random.rand(100),
          'V3':np.random.rand(100)})
df3.to_excel(writer,sheet_name='sheet3',index=False)
 
'''数据写出到excel文件中'''
writer.save()


import pandas as pd
from openpyxl import load_workbook
result2=[('a','2','ss'),('b','2','33'),('c','4','bbb')]#列表数据
writer = pd.ExcelWriter(r'D:\demo99.xlsx',engine='openpyxl')#可以向不同的sheet写入数据
book=load_workbook(r'D:\demo99.xlsx')
writer.book = book
df = pd.DataFrame(result2,columns=['xuhao','id','name'])#列表数据转为数据框
df.to_excel(writer, sheet_name='sheet2')#将数据写入excel中的sheet2表,sheet_name改变后即是新增一个sheet
writer.save()#保存


import pandas as pd
from openpyxl import load_workbook
result2=[('a','2','ss'),('b','2','33'),('c','4','bbb')]#需要新写入的数据
df = pd.DataFrame(result2,columns=['xuhao','id','name'])#列表数据转为数据框
df1 = pd.DataFrame(pd.read_excel(r'D:\demo99.xlsx',sheet_name='aa')) #读取原数据文件和表
writer = pd.ExcelWriter(r'D:\demo99.xlsx',engine='openpyxl')
book=load_workbook(r'D:\demo99.xlsx')
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
df_rows = df1.shape[0] #获取原数据的行数
df.to_excel(writer, sheet_name='aa',startrow=df_rows+1, index=False, header=False)#将数据写入excel中的aa表,从第一个空行开始写
writer.save()#保存
