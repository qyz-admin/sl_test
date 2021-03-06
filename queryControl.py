import pandas as pd
import os
import xlwings as xl
import pandas.io.formats.excel
from sqlalchemy import create_engine
from settings import Settings
from emailControl import EmailControl
from openpyxl import load_workbook  # 可以向不同的sheet写入数据
from openpyxl.styles import Font, Border, Side, PatternFill, colors, \
    Alignment  # 设置字体风格为Times New Roman，大小为16，粗体、斜体，颜色蓝色

from tkinter import messagebox

import datetime
import xlwings as xw


# -*- coding:utf-8 -*-
class QueryControl(Settings):
    def __init__(self):
        Settings.__init__(self)
        self.engine1 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql1['user'],
                                                                                    self.mysql1['password'],
                                                                                    self.mysql1['host'],
                                                                                    self.mysql1['port'],
                                                                                    self.mysql1['datebase']))
        self.engine2 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql2['user'],
                                                                                    self.mysql2['password'],
                                                                                    self.mysql2['host'],
                                                                                    self.mysql2['port'],
                                                                                    self.mysql2['datebase']))
        self.engine3 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql3['user'],
                                                                                    self.mysql3['password'],
                                                                                    self.mysql3['host'],
                                                                                    self.mysql3['port'],
                                                                                    self.mysql3['datebase']))
        self.e = EmailControl()

    def reSetEngine(self):
        self.engine1 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql1['user'],
                                                                                    self.mysql1['password'],
                                                                                    self.mysql1['host'],
                                                                                    self.mysql1['port'],
                                                                                    self.mysql1['datebase']))
        self.engine2 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql2['user'],
                                                                                    self.mysql2['password'],
                                                                                    self.mysql2['host'],
                                                                                    self.mysql2['port'],
                                                                                    self.mysql2['datebase']))

    def writeSqlReplace(self, dataFrame):
        dataFrame.to_sql('tem', con=self.engine1, index=False, if_exists='replace')

    def replaceInto(self, team, dfColumns):
        columns = list(dfColumns)
        columns = ', '.join(columns)
        if team == 'slrb':
            print(team + '---9')
            sql = 'REPLACE INTO {}({}, 添加时间) SELECT *, NOW() 添加时间 FROM tem; '.format(team, columns)
        else:
            print(team)
            sql = 'INSERT IGNORE INTO {}({}, 添加时间) SELECT *, NOW() 添加时间 FROM tem; '.format(team, columns)
        try:
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=100)
        except Exception as e:
            print('插入失败：', str(Exception) + str(e))

    def readSql(self, sql):
        db = pd.read_sql_query(sql=sql, con=self.engine1)
        # db = pd.read_sql(sql=sql, con=self.engine1) or team == 'slgat'
        return db
    # 团队花费明细查询（公用）（现用）
    def sl_tem_cost(self, team, tem):
        match = {'slgat_zqsb': '"台湾", "香港"',
                 'sltg_zqsb': '泰国',
                 'slxmt_zqsb': '"新加坡", "马来西亚", "菲律宾"',
                 'slrb_zqsb_rb': '日本'}
        match1 = {'slgat_zqsb': 'slgat_month',
                  'sltg_zqsb': 'sltg_month',
                  'slxmt_zqsb': 'slxmt_month',
                  'slrb_zqsb_rb': 'slrb_month'}
        match3 = {'新加坡': 'SG',
                  '马来西亚': 'MY',
                  '菲律宾': 'PH',
                  '日本': 'JP',
                  '香港': 'HK',
                  '台湾': 'TW',
                  '泰国': 'TH'}
        emailAdd = {'slgat_zqsb': 'giikinliujun@163.com',
                    'sltg_zqsb': '1845389861@qq.com',
                    'slxmt_zqsb': 'zhangjing@giikin.com',
                    'slzb': '直播团队',
                    'slyn': '越南',
                    'slrb_zqsb_rb': 'sunyaru@giikin.com'}
        emailAdd2 = {'sltg_zqsb': 'zhangjing@giikin.com'}
        start = datetime.datetime.now()
        endDate = (datetime.datetime.now()).strftime('%Y%m')
        print(endDate)
        startDate = (datetime.datetime.now().replace(day=1) - datetime.timedelta(days=1)).strftime('%Y%m')
        print(startDate)
        if datetime.datetime.now().day >= 9:
            end_Date = [startDate, endDate]
            start_Date = [startDate, endDate]
        else:
            end_Date = [startDate]
            start_Date = [startDate]
        listT = []  # 查询sql 存放池
        show_name = []  # 打印进度需要
        for i in range(len(end_Date)):
            # # 总花费明细表---查询
            # sql20 = '''SELECT *
            #         FROM (
            #             SELECT sl_zong.币种,
            #                     IFNULL(sl_zong.年月,'合计') 年月,
            #                     IFNULL(sl_zong.父级分类,'合计') 父级分类,
            #                     IFNULL(sl_zong.二级分类,'合计') 二级分类,
            #                     IFNULL(sl_zong.三级分类,'合计') 三级分类,
            #                     IFNULL(sl_zong.产品id,'合计') 产品id,
            #                     IFNULL(sl_zong.产品名称,'合计') 产品名称,
            #                     IFNULL(sl_zong.物流方式,'合计') 物流方式,
            #                     IFNULL(sl_zong.旬,'合计') 旬,
            #                     SUM(sl_zong.订单量) 订单量,
            #                     IFNULL(SUM(sl_zong_zf.`直发订单量`),0) 直发订单量,
            #                     (SUM(sl_zong.订单量) - IFNULL(SUM(sl_zong_zf.`直发订单量`),0)) AS 改派订单量,
            #                     IFNULL(SUM(sl_zong_zhifa.`已签收订单量`),0) 签收订单量,
            #                     IFNULL(SUM(sl_zong_jushou.`拒收订单量`),0) 拒收订单量,
            #                     SUM(sl_zong.总成本) / SUM(sl_zong.销售额)  AS '采购/销售额',
            #                     IFNULL(SUM(sl_zong_zf.`直发成本`),0) / SUM(sl_zong.销售额)  AS '直发采购/销售额',
            #                     SUM(sl_zong.物流运费) / SUM(sl_zong.销售额)  AS '运费占比',
            #                     SUM(sl_zong.手续费) / SUM(sl_zong.销售额)  AS '手续费占比',
            #                     IFNULL(SUM(sl_zong_zhifa.`已签收销售额`),0) / (IFNULL(SUM(sl_zong_zhifa.`已签收销售额`),0) + IFNULL(SUM(sl_zong_jushou.`拒收销售额`),0)) AS '金额签收/完成',
            #                     IFNULL(SUM(sl_zong_zhifa.`已签收销售额`),0) / SUM(sl_zong.销售额) AS '金额签收/总计',
            #                     (IFNULL(SUM(sl_zong_zhifa.`已签收销售额`),0) + IFNULL(SUM(sl_zong_jushou.`拒收销售额`),0)) / SUM(sl_zong.销售额) AS '金额完成占比',
            #                     IFNULL(SUM(sl_zong_zhifa.`已签收订单量`),0) / (IFNULL(SUM(sl_zong_zhifa.`已签收订单量`),0) + IFNULL(SUM(sl_zong_jushou.`拒收订单量`),0)) AS '数量签收/完成',
            #                     (IFNULL(SUM(sl_zong_zhifa.`已签收订单量`),0) + IFNULL(SUM(sl_zong_jushou.`拒收订单量`),0)) / SUM(sl_zong.订单量) AS '数量完成占比'
            #             FROM (
            #                     SELECT 币种,
            #                             年月,
            #                             父级分类,
            #                             二级分类,
            #                             三级分类,
            #                             产品id,
            #                             CONCAT(产品id, '#' ,产品名称) 产品名称,
            #                             物流方式,
            #                             旬,
            #                             COUNT(`订单编号`) 订单量,
            #                             SUM(`价格RMB`) 销售额,
            #                             SUM(`成本价`) 总成本,
            #                             SUM(`物流花费`) 物流运费,
            #                             SUM(`打包花费`) 打包花费,
            #                             SUM(`其它花费`) 手续费
            #                     FROM  {0} sl_cx
            #                     WHERE sl_cx.`币种` = '{1}'
            #                         AND sl_cx.`年月` >= '{start_Date}'
            #                         AND sl_cx.`年月` <= '{end_Date}'
            #                         AND sl_cx.`系统订单状态`!="已删除"
            #                     GROUP BY 币种,年月,父级分类,三级分类,产品名称,物流方式,旬
            #                     ORDER BY 币种,年月
            #                 ) sl_zong
            #         LEFT JOIN
            #                 (   SELECT 币种,
            #                             年月,
            #                             父级分类,
            #                             二级分类,
            #                             三级分类,
            #                             产品id,
            #                             CONCAT(产品id, '#' ,产品名称) 产品名称,
            #                             物流方式,
            #                             旬,
            #                             COUNT(`订单编号`) 直发订单量,
            #                             SUM(`价格RMB`) 销售额,
            #                             SUM(`成本价`) 直发成本,
            #                             SUM(`物流花费`) 物流运费,
            #                             SUM(`打包花费`) 打包花费,
            #                             SUM(`其它花费`) 手续费
            #                     FROM  {0} sl_cx_zf
            #                     WHERE sl_cx_zf.`币种` = '{1}'
            #                         AND sl_cx_zf.`年月` >= '{start_Date}'
            #                         AND sl_cx_zf.`年月` <= '{end_Date}'
            #                         AND sl_cx_zf.`系统订单状态`!="已删除"
            #                         AND sl_cx_zf.`是否改派` = "直发"
            #                     GROUP BY 币种,年月,父级分类,三级分类,产品名称,物流方式,旬
            #                     ORDER BY 币种,年月
            #             ) sl_zong_zf
            #                 ON sl_zong_zf.`币种` = sl_zong.`币种`
            #                     AND sl_zong_zf.`年月` = sl_zong.`年月`
            #                     AND sl_zong_zf.`父级分类` = sl_zong.`父级分类`
            #                     AND sl_zong_zf.`二级分类` = sl_zong.`二级分类`
            #                     AND sl_zong_zf.`三级分类` = sl_zong.`三级分类`
            #                     AND sl_zong_zf.`产品id` = sl_zong.`产品id`
            #                     AND sl_zong_zf.`产品名称` = sl_zong.`产品名称`
            #                     AND sl_zong_zf.`物流方式` = sl_zong.`物流方式`
            #                     AND sl_zong_zf.`旬` = sl_zong.`旬`
            #         LEFT JOIN
            #                 (   SELECT 币种,
            #                             年月,
            #                             父级分类,
            #                             二级分类,
            #                             三级分类,
            #                             产品id,
            #                             CONCAT(产品id, '#' ,产品名称) 产品名称,
            #                             物流方式,
            #                             旬,
            #                             COUNT(`订单编号`) 已签收订单量,
            #                             SUM(`价格RMB`) 已签收销售额,
            #                             SUM(`成本价`) 已签收成本,
            #                             SUM(`物流花费`) 已签收物流运费,
            #                             SUM(`打包花费`) 已签收打包花费,
            #                             SUM(`其它花费`) 已签收手续费
            #                     FROM  {0} sl_cx_zhifa
            #                     WHERE sl_cx_zhifa.`币种` = '{1}'
            #                         AND sl_cx_zhifa.`年月` >= '{start_Date}'
            #                         AND sl_cx_zhifa.`年月` <= '{end_Date}'
            #                         AND sl_cx_zhifa.`系统订单状态`!="已删除"
            #                         AND sl_cx_zhifa.`最终状态` = "已签收"
            #                     GROUP BY 币种,年月,父级分类,三级分类,产品名称,物流方式,旬
            #                     ORDER BY 币种,年月
            #             ) sl_zong_zhifa
            #             ON sl_zong_zhifa.`币种` = sl_zong.`币种`
            #                     AND sl_zong_zhifa.`年月` = sl_zong.`年月`
            #                     AND sl_zong_zhifa.`父级分类` = sl_zong.`父级分类`
            #                     AND sl_zong_zhifa.`二级分类` = sl_zong.`二级分类`
            #                     AND sl_zong_zhifa.`三级分类` = sl_zong.`三级分类`
            #                     AND sl_zong_zhifa.`产品id` = sl_zong.`产品id`
            #                     AND sl_zong_zhifa.`产品名称` = sl_zong.`产品名称`
            #                     AND sl_zong_zhifa.`物流方式` = sl_zong.`物流方式`
            #                     AND sl_zong_zhifa.`旬` = sl_zong.`旬`
            #         LEFT JOIN
            #                 (   SELECT 币种,
            #                             年月,
            #                             父级分类,
            #                             二级分类,
            #                             三级分类,
            #                             产品id,
            #                             CONCAT(产品id, '#' ,产品名称) 产品名称,
            #                             物流方式,
            #                             旬,
            #                             COUNT(`订单编号`) 拒收订单量,
            #                             SUM(`价格RMB`) 拒收销售额,
            #                             SUM(`成本价`) 拒收成本,
            #                             SUM(`物流花费`) 拒收物流运费,
            #                             SUM(`打包花费`) 拒收打包花费,
            #                             SUM(`其它花费`) 拒收手续费
            #                     FROM  {0} sl_cx_jushou
            #                     WHERE sl_cx_jushou.`币种` = '{1}'
            #                         AND sl_cx_jushou.`年月` >= '{start_Date}'
            #                         AND sl_cx_jushou.`年月` <= '{end_Date}'
            #                         AND sl_cx_jushou.`系统订单状态`!="已删除"
            #                         AND sl_cx_jushou.`最终状态` = "拒收"
            #                     GROUP BY 币种,年月,父级分类,三级分类,产品名称,物流方式,旬
            #                     ORDER BY 币种,年月
            #             ) sl_zong_jushou
            #             ON sl_zong_jushou.`币种` = sl_zong.`币种`
            #                     AND sl_zong_jushou.`年月` = sl_zong.`年月`
            #                     AND sl_zong_jushou.`父级分类` = sl_zong.`父级分类`
            #                     AND sl_zong_jushou.`二级分类` = sl_zong.`二级分类`
            #                     AND sl_zong_jushou.`三级分类` = sl_zong.`三级分类`
            #                     AND sl_zong_jushou.`产品id` = sl_zong.`产品id`
            #                     AND sl_zong_jushou.`产品名称` = sl_zong.`产品名称`
            #                     AND sl_zong_jushou.`物流方式` = sl_zong.`物流方式`
            #                     AND sl_zong_jushou.`旬` = sl_zong.`旬`
            #         GROUP BY sl_zong.父级分类,sl_zong.三级分类,sl_zong.产品名称,sl_zong.物流方式,sl_zong.旬
            #         with rollup
            #         ) sl_zong_wl
            #         WHERE sl_zong_wl.`旬` = '合计';'''.format(team, tem, start_Date=start_Date[i], end_Date=end_Date[i])
            # listT.append(sql20)
            # show_name.append(start_Date[i] + '月（总）详细花费数据…………')
            # 直发花费明细表---查询
            # sql30 = '''SELECT sl.`币种`,
            #                     sl.`年月`,
            #                     sl.`父级分类`,
            #                     sl.`三级分类`,
            #                     sl.`产品名称`,
            #                     sl.`物流方式`,
            #                     sl.`总订单量`,
            #                     sl.`直发订单量`,
            #                     sl.`直发退货量`,
            #                     sl.`已签收订单量`,
            #                     sl.`拒收订单量`,
            #                     sl.`总销售额` / sl.`总订单量` AS '总客单价',
            #                     sl.`直发成本` / sl.`总销售额` AS '直发采购/销售额',
            #                     sl.`直发物流运费` / sl.`总销售额` AS '运费占比',
            #                     sl.`直发手续费` / sl.`总销售额` AS '手续费占比',
            #                     sl.`已签收销售额` / (sl.`已签收销售额` + sl.`拒收销售额`) AS '金额签收/完成',
            #                     sl.`已签收销售额` / sl.`总销售额` AS '金额签收/总计',
            #                     (sl.`已签收销售额` + sl.`拒收销售额`) /  sl.`总销售额`  AS '金额完成占比',
            #                     sl.`已签收订单量` /  (sl.`已签收订单量` + sl.`拒收订单量`) AS '数量签收/完成',
            #                     (sl.`已签收订单量` + sl.`拒收订单量`) / sl.`总订单量` AS '数量完成占比'
            #         FROM {0} sl
            #          WHERE sl.`币种` = '{1}'
            #             AND sl.`旬` = CONVERT('合计' USING utf8) COLLATE utf8_general_ci
            #             AND sl.`年月` = CONVERT('{end_Date}' USING utf8) COLLATE utf8_general_ci;'''.format(match1[team], tem, end_Date=end_Date[i])
            sql30 = '''SELECT *
                    FROM (
                        SELECT sl_zong.币种,
                            IFNULL(sl_zong.年月,'合计') 年月,
                            IFNULL(sl_zong.父级分类,'合计') 父级分类,
                            IFNULL(sl_zong.二级分类,'合计') 二级分类,
                            IFNULL(sl_zong.三级分类,'合计') 三级分类,
                            IFNULL(sl_zong.产品id,'合计') 产品id,
                            IFNULL(sl_zong.产品名称,'合计') 产品名称,
                            IFNULL(sl_zong.物流方式,'合计') 物流方式,
                            IFNULL(sl_zong.旬,'合计') 旬,
                            SUM(sl_zong.直发订单量) 直发订单量,
                            IFNULL(SUM(sl_zong_zhifa.`已签收订单量`),0) 签收订单量,
                            IFNULL(SUM(sl_zong_jushou.`拒收订单量`),0) 拒收订单量,
                            SUM(sl_zong.直发成本) / SUM(sl_zong.销售额) AS '直发采购/销售额',
                            SUM(sl_zong.物流运费) / SUM(sl_zong.销售额) AS '运费占比',
                            SUM(sl_zong.手续费) / SUM(sl_zong.销售额) AS '手续费占比',
                            IFNULL(SUM(sl_zong_zhifa.`已签收销售额`),0) / (IFNULL(SUM(sl_zong_zhifa.`已签收销售额`),0) + IFNULL(SUM(sl_zong_jushou.`拒收销售额`),0)) AS '金额签收/完成',
                            IFNULL(SUM(sl_zong_zhifa.`已签收销售额`),0) / SUM(sl_zong.销售额) AS '金额签收/总计',
                            (IFNULL(SUM(sl_zong_zhifa.`已签收销售额`),0) + IFNULL(SUM(sl_zong_jushou.`拒收销售额`),0)) / SUM(sl_zong.销售额) AS '金额完成占比',
                            IFNULL(SUM(sl_zong_zhifa.`已签收订单量`),0) / (IFNULL(SUM(sl_zong_zhifa.`已签收订单量`),0) + IFNULL(SUM(sl_zong_jushou.`拒收订单量`),0)) AS '数量签收/完成',
                            (IFNULL(SUM(sl_zong_zhifa.`已签收订单量`),0) + IFNULL(SUM(sl_zong_jushou.`拒收订单量`),0)) / SUM(sl_zong.直发订单量) AS '数量完成占比'
                        FROM (
                            SELECT 币种,
                                    年月,
                                    父级分类,
                                    二级分类,
                                    三级分类,
                                    产品id,
                                    CONCAT(产品id, '#' ,产品名称) 产品名称,
                                    物流方式,
                                    旬,
                                    COUNT(`订单编号`) 直发订单量,
                                    SUM(`价格RMB`) 销售额,
                                    SUM(`成本价`) 直发成本,
                                    SUM(`物流花费`) 物流运费,
                                    SUM(`打包花费`) 打包花费,
                                    SUM(`其它花费`) 手续费
                            FROM  {0} sl_cx
                            WHERE sl_cx.`币种` = '{1}'
                                AND sl_cx.`年月` >= '{start_Date}'
                                AND sl_cx.`年月` <= '{end_Date}'
                                AND sl_cx.`系统订单状态` IN ('已审核', '待发货', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)', '待发货转审核') 
                                AND sl_cx.`是否改派` = "直发"
                            GROUP BY 币种,年月,父级分类,二级分类,三级分类,产品名称,物流方式,旬
                            ORDER BY 币种,年月
                            ) sl_zong
                        LEFT JOIN
                            (SELECT 币种,
                                    年月,
                                    父级分类,
                                    二级分类,
                                    三级分类,
                                    产品id,
                                    CONCAT(产品id, '#' ,产品名称) 产品名称,
                                    物流方式,
                                    旬,
                                    COUNT(`订单编号`) 已签收订单量,
                                    SUM(`价格RMB`) 已签收销售额,
                                    SUM(`成本价`) 已签收成本,
                                    SUM(`物流花费`) 已签收物流运费,
                                    SUM(`打包花费`) 已签收打包花费,
                                    SUM(`其它花费`) 已签收手续费
                            FROM  {0}	sl_cx_zhifa
                            WHERE sl_cx_zhifa.`币种` = '{1}'
                                AND sl_cx_zhifa.`年月` >= '{start_Date}'
                                AND sl_cx_zhifa.`年月` <= '{end_Date}'
                                AND sl_cx_zhifa.`系统订单状态` IN ('已审核', '待发货', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)', '待发货转审核') 
                                AND sl_cx_zhifa.`是否改派` = "直发"
                                AND sl_cx_zhifa.`最终状态` = "已签收"
                            GROUP BY 币种,年月,父级分类,二级分类,三级分类,产品名称,物流方式,旬
                            ORDER BY 币种,年月
                            ) sl_zong_zhifa
                            ON sl_zong_zhifa.`币种` = sl_zong.`币种`
                                AND sl_zong_zhifa.`年月` = sl_zong.`年月`
                                AND sl_zong_zhifa.`父级分类` = sl_zong.`父级分类`
                                AND sl_zong_zhifa.`二级分类` = sl_zong.`二级分类`
                                AND sl_zong_zhifa.`三级分类` = sl_zong.`三级分类`
                                AND sl_zong_zhifa.`产品id` = sl_zong.`产品id`
                                AND sl_zong_zhifa.`产品名称` = sl_zong.`产品名称`
                                AND sl_zong_zhifa.`物流方式` = sl_zong.`物流方式`
                                AND sl_zong_zhifa.`旬` = sl_zong.`旬`
                        LEFT JOIN
                            (SELECT 币种,
                                    年月,
                                    父级分类,
                                    二级分类,
                                    三级分类,
                                    产品id,
                                    CONCAT(产品id, '#' ,产品名称) 产品名称,
                                    物流方式,
                                    旬,
                                    COUNT(`订单编号`) 拒收订单量,
                                    SUM(`价格RMB`) 拒收销售额,
                                    SUM(`成本价`) 拒收成本,
                                    SUM(`物流花费`) 拒收物流运费,
                                    SUM(`打包花费`) 拒收打包花费,
                                    SUM(`其它花费`) 拒收手续费
                            FROM  {0} sl_cx_jushou
                            WHERE sl_cx_jushou.`币种` = '{1}'
                                AND sl_cx_jushou.`年月` >= '{start_Date}'
                                AND sl_cx_jushou.`年月` <= '{end_Date}'
                                AND sl_cx_jushou.`系统订单状态` IN ('已审核', '待发货', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)', '待发货转审核') 
                                AND sl_cx_jushou.`是否改派` = "直发"
                                AND sl_cx_jushou.`最终状态` = "拒收"
                            GROUP BY 币种,年月,父级分类,二级分类,三级分类,产品名称,物流方式,旬
                            ORDER BY 币种,年月
                            ) sl_zong_jushou
                        ON sl_zong_jushou.`币种` = sl_zong.`币种`
                            AND sl_zong_jushou.`年月` = sl_zong.`年月`
                            AND sl_zong_jushou.`父级分类` = sl_zong.`父级分类`
                            AND sl_zong_jushou.`二级分类` = sl_zong.`二级分类`
                            AND sl_zong_jushou.`三级分类` = sl_zong.`三级分类`
                            AND sl_zong_jushou.`产品id` = sl_zong.`产品id`
                            AND sl_zong_jushou.`产品名称` = sl_zong.`产品名称`
                            AND sl_zong_jushou.`物流方式` = sl_zong.`物流方式`
                            AND sl_zong_jushou.`旬` = sl_zong.`旬`
                    GROUP BY sl_zong.父级分类,sl_zong.二级分类,sl_zong.三级分类,sl_zong.产品名称,sl_zong.物流方式,sl_zong.旬
                    with rollup
                    ) sl_zong_wl
                    WHERE sl_zong_wl.`旬` = '合计';'''.format(team, tem, start_Date=start_Date[i], end_Date=end_Date[i])
            listT.append(sql30)
            show_name.append(start_Date[i] + '月（直发）详细花费数据…………')
            # 总成本父级
            sql40 = '''SELECT *
                        FROM(
                            (SELECT s1.团队,
                                    s1.年月,
                                    s1.品类,
                                    s1.销售额,
                                    s1.订单量,
                                    (s1.订单量 - s1.改派订单量) AS '直发订单量',
                                    (s1.订单量 - s1.改派订单量) / s1.订单量 AS '直发占比',
                                    s1.改派订单量,
                                    s1.改派订单量 / s1.订单量 AS '改派占比',
                                    s1.销售额 / s1.订单量 AS '客单价',
                                    s1.销售额 / s1.广告成本 AS 'ROI',
                                    s1.活跃产品数,
                                    s1.订单量 / s1.活跃产品数 AS 产能,
                                    s1.直发采购额 / s1.销售额 AS '直发采购/总销售额',
                                    s1.广告成本 / s1.销售额 AS '广告占比',
                                    s1.物流成本 / s1.销售额 AS '运费占比',
                                    s1.手续费 / s1.销售额 AS '手续费占比',
                                    (s1.直发采购额 + s1.广告成本 + s1.物流成本 + s1.手续费 ) AS '总成本',
                                    (s1.直发采购额 + s1.广告成本 + s1.物流成本 + s1.手续费 ) / s1.销售额 AS '总成本占比',
                                    s1.签收额 / (s1.签收额 + s1.拒收额) AS '金额签收/完成',
                                    s1.签收额 / s1.销售额 AS '金额签收/总计',
                                    (s1.签收额 + s1.拒收额) / s1.销售额 AS '金额完成占比',
                                    s1.签收量 / (s1.签收量 + s1.拒收量) AS '数量签收/完成',
                                    s1.签收量 / s1.订单量 AS '数量签收/总计',
                                    (s1.签收量 + s1.拒收量) / s1.订单量 AS '数量完成占比',
                                    s1.签收额 / (s1.签收额 + s1.拒收额) -( s1.直发采购额 + s1.广告成本 + s1.物流成本 + s1.手续费 ) AS 利润率,
                                    (s1.签收额 / (s1.签收额 + s1.拒收额) -( s1.直发采购额 + s1.广告成本 + s1.物流成本 + s1.手续费 ) /s1.销售额) * (s1.销售额 / s1.订单量) AS 利润值
                                FROM (SELECT EXTRACT(YEAR_MONTH FROM a.rq) AS 年月,
                                        b.pname AS 团队,
                                        c.uname AS leader,
                                        d.ppname AS 品类,
                                        SUM(a.orders) AS 订单量,
                                        COUNT(DISTINCT a.product_id) AS 活跃产品数,
                                        SUM(a.yqs) AS 签收量,
                                        SUM(a.yjs) AS 拒收量,
                                        SUM(a.salesRMB) AS 销售额,
                                        SUM(a.salesRMB_yqs) AS 签收额,
                                        SUM(a.salesRMB_yjs) AS 拒收额,
                                        SUM(a.gps) AS 改派订单量,
                            --          SUM(a.cgcost) AS 总采购额,
                            --          SUM(a.cgcost_zf) AS 直发采购额,
                            --          SUM(a.adcost) AS 广告成本,
                                        null 总采购额,
                                        null 直发采购额,
                                        null 广告成本,
                                        SUM(a.wlcost) AS 物流成本,
                                        SUM(a.qtcost) AS 手续费 
                                    FROM gk_order_day_kf a
                                        LEFT JOIN dim_currency_lang b ON a.currency_lang_id = b.id
                                        LEFT JOIN dim_area c on c.id = a.area_id
                                        LEFT JOIN dim_cate d on d.id = a.third_cate_id
                                        LEFT JOIN (SELECT * FROM gk_sale WHERE id IN (SELECT MAX(id) FROM gk_sale GROUP BY product_id ) ORDER BY id) e on e.id = a.product_id
                                    WHERE b.pcode = '{0}'
                                        AND EXTRACT(YEAR_MONTH FROM a.rq) >= '{start_Date}'
                                        AND EXTRACT(YEAR_MONTH FROM a.rq) <= '{end_Date}'
                                        AND c.uname = '王冰'
                                        AND a.beform <> 'mf'
                                        AND c.uid <> 10099  -- 过滤翼虎
                                     GROUP BY b.pname, c.uname, a.cate_id
                                    ORDER BY a.product_id
                                ) s1 WHERE s1.订单量 > 0 ORDER BY s1.订单量)
                            UNION ALL
                                (SELECT s3.团队,
						                s3.年月,
						                s3.品类,
						                s3.销售额,
						                s3.订单量,
						                (s3.订单量 - s3.改派订单量) AS 直发订单量,
						                (s3.订单量 - s3.改派订单量) / s3.订单量 AS 直发占比,
						                s3.改派订单量,
						                s3.改派订单量 / s3.订单量 AS 改派占比,
						                s3.销售额 / s3.订单量 客单价,
						                s3.销售额 / s3.广告成本 ROI,
						                S3.活跃产品数,
						                s3.订单量 / S3.活跃产品数 AS 产能,
						                s3.直发采购额 / s3.销售额 AS '直发采购/总销售额',
						                s3.广告成本 / s3.销售额 AS '广告占比',
						                s3.物流成本 / s3.销售额 AS '运费占比',
						                s3.手续费 / s3.销售额 AS '手续费占比',
						                (s3.广告成本 + s3.物流成本 + s3.手续费 + s3.直发采购额 ) AS '总成本',
						                (s3.广告成本 + s3.物流成本 + s3.手续费 + s3.直发采购额 ) / s3.销售额 AS '总成本占比',
						                s3.签收额 / (s3.拒收额 + s3.签收额) '金额签收/完成',
						                s3.签收额 / s3.销售额 '金额签收/总计',
						                (s3.签收额 + s3.拒收额) / s3.销售额 '金额完成占比',
						                s3.签收量 / (s3.拒收量 + s3.签收量) '数量签收/完成',
						                s3.签收量 / s3.订单量 '数量签收/总计',
						                (s3.拒收量 + s3.签收量) / s3.订单量 '数量完成占比',
						                s3.签收额 / (s3.签收额 + s3.拒收额) -( s3.直发采购额 + s3.广告成本 + s3.物流成本 + s3.手续费 ) AS 利润率,
						                (s3.签收额 / (s3.签收额 + s3.拒收额) -( s3.直发采购额 + s3.广告成本 + s3.物流成本 + s3.手续费 ) / s3.销售额) * (s3.销售额 / s3.订单量) AS 利润值
                                FROM (SELECT EXTRACT(YEAR_MONTH FROM a.rq) AS 年月,
						                b.pname AS 团队,
						                '合计' AS 品类,
						                SUM(a.salesRMB) 销售额,
						                SUM(a.orders) AS 订单量,
						                COUNT(DISTINCT a.product_id) AS 活跃产品数,
						                '' 产能,
						                SUM(a.yqs) AS 签收量,
						                SUM(a.yjs) AS 拒收量,
						                SUM(a.salesRMB_yqs) AS 签收额,
						                SUM(a.salesRMB_yjs) AS 拒收额,
						                SUM(a.gps) AS 改派订单量,
                            --          SUM(a.cgcost) AS 总采购额,
                            --          SUM(a.cgcost_zf) AS 直发采购额,
                            --          SUM(a.adcost) AS 广告成本,
                                        null 总采购额,
                                        null 直发采购额,
                                        null 广告成本,
						                SUM(a.wlcost) AS 物流成本,
						                SUM(a.qtcost) AS 手续费 
			                        FROM gk_order_day_kf a
				                        LEFT JOIN dim_currency_lang b ON a.currency_lang_id = b.id
				                        LEFT JOIN dim_area c on c.id = a.area_id
				                        LEFT JOIN dim_cate d on d.id = a.third_cate_id
                                        LEFT JOIN (SELECT * FROM gk_sale WHERE id IN (SELECT MAX(id) FROM gk_sale GROUP BY product_id ) ORDER BY id) e on e.id = a.product_id
                                    WHERE b.pcode = '{0}'
                                        AND EXTRACT(YEAR_MONTH FROM a.rq) >= '{start_Date}'
                                        AND EXTRACT(YEAR_MONTH FROM a.rq) <= '{end_Date}'
						                AND c.uname = '王冰'
						                AND a.beform <> 'mf'
						                AND c.uid <> 10099  -- 过滤翼虎
			                        GROUP BY b.pname, c.uname
                                ) s3)
                            ) s ORDER BY s.订单量'''.format(match3[tem], start_Date=start_Date[i], end_Date=end_Date[i])
            listT.append(sql40)
            show_name.append(start_Date[i] + '月（父级）成本数据…………')
            # 总成本二级
            sql41 = '''SELECT s1.团队,
                                s1.年月,
                                s1.二级分类,
                                s1.销售额,
                                s1.订单量,
                                (s1.订单量 - s1.改派订单量) AS '直发订单量',
                                (s1.订单量 - s1.改派订单量) / s1.订单量 AS '直发占比',
                                s1.改派订单量,
                                s1.改派订单量 / s1.订单量 AS '改派占比',
                                s1.销售额 / s1.订单量 AS '客单价',
                                s1.销售额 / s1.广告成本 AS 'ROI',
                                s1.活跃产品数,
                                s1.订单量 / s1.活跃产品数 AS 产能,
                                s1.直发采购额 / s1.销售额 AS '直发采购/总销售额',
                                s1.广告成本 / s1.销售额 AS '广告占比',
                                s1.物流成本 / s1.销售额 AS '运费占比',
                                s1.手续费 / s1.销售额 AS '手续费占比',
                                ( s1.直发采购额 + s1.广告成本 + s1.物流成本 + s1.手续费 ) AS '总成本',
                                ( s1.直发采购额 + s1.广告成本 + s1.物流成本 + s1.手续费 ) / s1.销售额 AS '总成本占比',
                                s1.签收额 / (s1.签收额 + s1.拒收额) AS '金额签收/完成',
                                s1.签收额 / s1.销售额 AS '金额签收/总计',
                                (s1.签收额 + s1.拒收额) / s1.销售额 AS '金额完成占比',
                                s1.签收量 / (s1.签收量 + s1.拒收量) AS '数量签收/完成',
                                s1.签收量 / s1.订单量 AS '数量签收/总计',
                                (s1.签收量 + s1.拒收量) / s1.订单量 AS '数量完成占比',
                                s1.签收额 / (s1.签收额 + s1.拒收额) -( s1.直发采购额 + s1.广告成本 + s1.物流成本 + s1.手续费 ) AS 利润率,
                                (s1.签收额 / (s1.签收额 + s1.拒收额) -( s1.直发采购额 + s1.广告成本 + s1.物流成本 + s1.手续费 ) /s1.销售额) * (s1.销售额 / s1.订单量) AS 利润值
                        FROM (  SELECT EXTRACT(YEAR_MONTH FROM a.rq) AS 年月,
                                        b.pname AS 团队,
                                        c.uname AS leader,
                                        d.pname AS 二级分类,
                                        SUM(a.orders) AS 订单量,
                                        COUNT(DISTINCT a.product_id) AS 活跃产品数,
                                        SUM(a.yqs) AS 签收量,
                                        SUM(a.yjs) AS 拒收量,
                                        SUM(a.salesRMB) AS 销售额,
                                        SUM(a.salesRMB_yqs) AS 签收额,
                                        SUM(a.salesRMB_yjs) AS 拒收额,
                                        SUM(a.gps) AS 改派订单量,
                            --          SUM(a.cgcost) AS 总采购额,
                            --          SUM(a.cgcost_zf) AS 直发采购额,
                            --          SUM(a.adcost) AS 广告成本,
                                        null 总采购额,
                                        null 直发采购额,
                                        null 广告成本,
                                        SUM(a.wlcost) AS 物流成本,
                                        SUM(a.qtcost) AS 手续费 
                                FROM gk_order_day_kf a
                                    LEFT JOIN dim_currency_lang b ON a.currency_lang_id = b.id
                                    LEFT JOIN dim_area c on c.id = a.area_id
                                    LEFT JOIN dim_cate d on d.id = a.third_cate_id
                                    LEFT JOIN (SELECT * FROM gk_sale WHERE id IN (SELECT MAX(id) FROM gk_sale GROUP BY product_id ) ORDER BY id) e on e.id = a.product_id
                                WHERE b.pcode = '{0}'
                                    AND EXTRACT(YEAR_MONTH FROM a.rq) >= '{start_Date}'
                                    AND EXTRACT(YEAR_MONTH FROM a.rq) <= '{end_Date}'
                                    AND c.uname = '王冰'
                                    AND a.beform <> 'mf'
                                    AND c.uid <> 10099  -- 过滤翼虎
                                GROUP BY b.pname, c.uname, a.second_cate_id
                                ORDER BY a.product_id
                        ) s1
                        WHERE s1.订单量 > 0
                        ORDER BY s1.订单量;'''.format(match3[tem], start_Date=start_Date[i], end_Date=end_Date[i])
            listT.append(sql41)
            show_name.append(start_Date[i] + '月（二级）成本数据…………')
            # 总成本三级
            sql42 = '''SELECT s1.团队,
                                s1.年月,
                                s1.三级分类,
                                s1.销售额,
                                s1.订单量,
                                (s1.订单量 - s1.改派订单量) AS '直发订单量',
                                (s1.订单量 - s1.改派订单量) / s1.订单量 AS '直发占比',
                                s1.改派订单量,
                                s1.改派订单量 / s1.订单量 AS '改派占比',
                                s1.销售额 / s1.订单量 AS '客单价',
                                s1.销售额 / s1.广告成本 AS 'ROI',
                                s1.活跃产品数,
                                s1.订单量 / s1.活跃产品数 AS 产能,
                                s1.直发采购额 / s1.销售额 AS '直发采购/总销售额',
                                s1.广告成本 / s1.销售额 AS '广告占比',
                                s1.物流成本 / s1.销售额 AS '运费占比',
                                s1.手续费 / s1.销售额 AS '手续费占比',
                                ( s1.直发采购额 + s1.广告成本 + s1.物流成本 + s1.手续费 ) AS '总成本',
                                ( s1.直发采购额 + s1.广告成本 + s1.物流成本 + s1.手续费 ) / s1.销售额 AS '总成本占比',
                                s1.签收额 / (s1.签收额 + s1.拒收额) AS '金额签收/完成',
                                s1.签收额 / s1.销售额 AS '金额签收/总计',
                                (s1.签收额 + s1.拒收额) / s1.销售额 AS '金额完成占比',
                                s1.签收量 / (s1.签收量 + s1.拒收量) AS '数量签收/完成',
                                s1.签收量 / s1.订单量 AS '数量签收/总计',
                                (s1.签收量 + s1.拒收量) / s1.订单量 AS '数量完成占比',
                                s1.签收额 / (s1.签收额 + s1.拒收额) -( s1.直发采购额 + s1.广告成本 + s1.物流成本 + s1.手续费 ) AS 利润率,
                                (s1.签收额 / (s1.签收额 + s1.拒收额) -( s1.直发采购额 + s1.广告成本 + s1.物流成本 + s1.手续费 ) /s1.销售额) * (s1.销售额 / s1.订单量) AS 利润值
                        FROM (  SELECT EXTRACT(YEAR_MONTH FROM a.rq) AS 年月,
                                        b.pname AS 团队,
                                        c.uname AS leader,
                                        d.`name` AS 三级分类,
                                        SUM(a.orders) AS 订单量,
                                        COUNT(DISTINCT a.product_id) AS 活跃产品数,
                                        SUM(a.yqs) AS 签收量,
                                        SUM(a.yjs) AS 拒收量,
                                        SUM(a.salesRMB) AS 销售额,
                                        SUM(a.salesRMB_yqs) AS 签收额,
                                        SUM(a.salesRMB_yjs) AS 拒收额,
                                        SUM(a.gps) AS 改派订单量,
                            --          SUM(a.cgcost) AS 总采购额,
                            --          SUM(a.cgcost_zf) AS 直发采购额,
                            --          SUM(a.adcost) AS 广告成本,
                                        null 总采购额,
                                        null 直发采购额,
                                        null 广告成本,
                                        SUM(a.wlcost) AS 物流成本,
                                        SUM(a.qtcost) AS 手续费 
                                FROM gk_order_day_kf a
                                    LEFT JOIN dim_currency_lang b ON a.currency_lang_id = b.id
                                    LEFT JOIN dim_area c on c.id = a.area_id
                                    LEFT JOIN dim_cate d on d.id = a.third_cate_id
                                    LEFT JOIN (SELECT * FROM gk_sale WHERE id IN (SELECT MAX(id) FROM gk_sale GROUP BY product_id ) ORDER BY id) e on e.id = a.product_id
                                WHERE b.pcode = '{0}'
                                    AND EXTRACT(YEAR_MONTH FROM a.rq) >= '{start_Date}'
                                    AND EXTRACT(YEAR_MONTH FROM a.rq) <= '{end_Date}'
                                    AND c.uname = '王冰'
                                    AND a.beform <> 'mf'
                                    AND c.uid <> 10099  -- 过滤翼虎
                                GROUP BY b.pname, c.uname, a.third_cate_id
                                ORDER BY a.product_id
                        ) s1
                        WHERE s1.订单量 > 0
                        ORDER BY s1.订单量;'''.format(match3[tem], start_Date=start_Date[i], end_Date=end_Date[i])
            listT.append(sql42)
            show_name.append(start_Date[i] + '月（三级）成本数据…………')
        listTValue = []  # 查询sql的结果 存放池
        for i, sql in enumerate(listT):
            print(i)
            print('正在获取 ' + tem + show_name[i])
            if i == 0 or i == 4:
                df = pd.read_sql_query(sql=sql, con=self.engine1)
            else:
                df = pd.read_sql_query(sql=sql, con=self.engine2)
            # print(df)
            columns = list(df.columns)  # 获取数据的标题名，转为列表
            columns_value = ['直发占比', '改派占比', '直发采购/总销售额', '广告占比', '总成本占比', '数量签收/总计',  '利润率', '采购/销售额', '直发采购/销售额', '运费占比', '手续费占比', '金额签收/完成', '金额签收/总计', '金额完成占比', '数量签收/完成', '数量完成占比']
            if '旬' in columns:
                df.drop(labels=['旬'], axis=1, inplace=True)  # 去掉多余的旬列表
            for column_val in columns_value:
                if column_val in columns:
                    df[column_val] = df[column_val].fillna(value=0)
                    df[column_val] = df[column_val].apply(lambda x: format(x, '.2%'))
            listTValue.append(df)
        print('查询耗时：', datetime.datetime.now() - start)
        today = datetime.date.today().strftime('%Y.%m.%d')
        sheet_name = ['直发成本', '父级成本', '二级成本', '三级成本']  # 生成的工作表的表名
        if len(listTValue) == 4:
            file_Path = []  # 发送邮箱文件使用
            filePath = 'D:\\Users\\Administrator\\Desktop\\输出文件\\{} {}上月产品花费表.xlsx'.format(today, tem)
            if os.path.exists(filePath):  # 判断是否有需要的表格
                print("正在使用(上月-单月)文件......")
                filePath = filePath
            else:  # 判断是否无需要的表格，进行初始化创建
                print("正在创建文件......")
                df0 = pd.DataFrame([])  # 创建空的dataframe数据框
                df0.to_excel(filePath, sheet_name='缓存使用', index=False)  # 备用：可以向不同的sheet写入数据（创建新的工作表并进行写入）
                filePath = filePath
            print('正在写入excel…………')
            writer = pd.ExcelWriter(filePath, engine='openpyxl')  # 初始化写入对象
            book = load_workbook(filePath)  # 可以向不同的sheet写入数据（对现有工作表的追加）
            writer.book = book  # 将数据写入excel中的sheet2表,sheet_name改变后即是新增一个sheet
            for i in range(len(listTValue)):
                listTValue[i].to_excel(excel_writer=writer, sheet_name=sheet_name[i], index=False)
            if '缓存使用' in book.sheetnames:  # 删除新建文档时的第一个工作表
                del book['缓存使用']
            writer.save()
            writer.close()
            print('输出文件成功…………')
            print('正在运行宏…………')
            app = xl.App(visible=False, add_book=False)           # 运行宏调整
            app.display_alerts = False
            wbsht = app.books.open('D:/Users/Administrator/Desktop/新版-格式转换(工具表).xlsm')
            wbsht1 = app.books.open(filePath)
            wbsht.macro('花费运行')()
            wbsht1.save()
            wbsht1.close()
            wbsht.close()
            app.quit()
            # self.xiugaiyangshi(filePath, sheet_name[1])
            file_Path.append(filePath)
            self.e.send(tem + '产品花费表', file_Path,
                        emailAdd[team])
            if tem == '泰国':
                self.e.send(tem + '产品花费表', file_Path,
                            emailAdd2[team])
        else:
            file_Path = []  # 发送邮箱文件使用
            filePath = 'D:\\Users\\Administrator\\Desktop\\输出文件\\{} {}上月产品花费表.xlsx'.format(today, tem)
            if os.path.exists(filePath):  # 判断是否有需要的表格
                print("正在使用(上月)文件......")
                filePath = filePath
            else:  # 判断是否无需要的表格，进行初始化创建
                print("正在创建文件......")
                df0 = pd.DataFrame([])  # 创建空的dataframe数据框
                df0.to_excel(filePath, index=False)  # 备用：可以向不同的sheet写入数据（创建新的工作表并进行写入）
                filePath = filePath
            print('正在写入excel…………')
            writer = pd.ExcelWriter(filePath, engine='openpyxl')  # 初始化写入对象
            book = load_workbook(filePath)  # 可以向不同的sheet写入数据（对现有工作表的追加）
            writer.book = book  # 将数据写入excel中的sheet2表,sheet_name改变后即是新增一个sheet
            listTValue[0].to_excel(excel_writer=writer, sheet_name=sheet_name[0], index=False)
            listTValue[1].to_excel(excel_writer=writer, sheet_name=sheet_name[1], index=False)
            listTValue[2].to_excel(excel_writer=writer, sheet_name=sheet_name[2], index=False)
            listTValue[3].to_excel(excel_writer=writer, sheet_name=sheet_name[3], index=False)
            # listTValue[1].to_excel(excel_writer=writer, sheet_name=sheet_name[1], index=False)
            writer.save()
            writer.close()
            print('输出(上月)文件成功…………')
            print('正在运行宏…………')
            app = xl.App(visible=False, add_book=False)           # 运行宏调整
            app.display_alerts = False
            wbsht = app.books.open('D:/Users/Administrator/Desktop/新版-格式转换(工具表).xlsm')
            wbsht1 = app.books.open(filePath)
            wbsht.macro('花费运行')()
            wbsht1.save()
            wbsht1.close()
            wbsht.close()
            app.quit()
            # self.xiugaiyangshi(filePath, sheet_name[1])
            file_Path.append(filePath)
            print('------分割线------')
            filePathT = 'D:\\Users\\Administrator\\Desktop\\输出文件\\{} {}本月产品花费表.xlsx'.format(today, tem)
            if os.path.exists(filePathT):  # 判断是否有需要的表格
                print("正在使用(本月)文件......")
                filePathT = filePathT
            else:  # 判断是否无需要的表格，进行初始化创建
                print("正在创建文件......")
                df0T = pd.DataFrame([])  # 创建空的dataframe数据框-2
                df0T.to_excel(filePathT, index=False)  # 备用：可以向不同的sheet写入数据（创建新的工作表并进行写入）-2
                filePathT = filePathT
            print('正在写入excel…………')
            writerT = pd.ExcelWriter(filePathT, engine='openpyxl')  # 初始化写入对象-2
            bookT = load_workbook(filePathT)  # 可以向不同的sheet写入数据（对现有工作表的追加）-2
            writerT.book = bookT  # 将数据写入excel中的sheet2表,sheet_name改变后即是新增一个sheet-2
            listTValue[4].to_excel(excel_writer=writerT, sheet_name=sheet_name[0], index=False)
            listTValue[5].to_excel(excel_writer=writerT, sheet_name=sheet_name[1], index=False)
            listTValue[6].to_excel(excel_writer=writerT, sheet_name=sheet_name[2], index=False)
            listTValue[7].to_excel(excel_writer=writerT, sheet_name=sheet_name[3], index=False)
            writerT.save()
            writerT.close()
            print('输出(本月)文件成功…………')
            print('正在运行宏…………')
            app = xl.App(visible=False, add_book=False)           # 运行宏调整
            app.display_alerts = False
            wbsht = app.books.open('D:/Users/Administrator/Desktop/新版-格式转换(工具表).xlsm')
            wbsht1 = app.books.open(filePathT)
            wbsht.macro('花费运行')()
            wbsht1.save()
            wbsht1.close()
            wbsht.close()
            app.quit()
            # self.xiugaiyangshi(filePathT, sheet_name[1])
            file_Path.append(filePathT)
            self.e.send(tem + '产品花费表', file_Path,
                        emailAdd[team])
            if tem == '泰国':
                self.e.send(tem + '产品花费表', file_Path,
                            emailAdd2[team])
            print('处理耗时：', datetime.datetime.now() - start)

    # 修改样式（备用）
    def xiugaiyangshi(self, filePath, sheetname):
        print('正在修改样式…………')
        wb = load_workbook(filePath)
        print(wb.sheetnames)
        sheet = wb[sheetname]
        for i in range(4, 5):
            for j in range(2, sheet.max_row):
                if sheet.cell(j, i).value == '合计' and sheet.cell(j, i + 1).value == '合计' and sheet.cell(j,
                                                                                                        i + 2).value == '合计':
                    for c in range(1, sheet.max_column + 1):
                        sheet.cell(j, c).fill = PatternFill(patternType='solid', fgColor='1874CD')
        for i in range(5, 6):
            for j in range(2, sheet.max_row):
                if sheet.cell(j, i - 1).value != '合计' and sheet.cell(j, i).value == '合计':
                    for c in range(1, sheet.max_column + 1):
                        sheet.cell(j, c).fill = PatternFill(patternType='solid', start_color='FFFF00',
                                                            end_color='FFFF00')
        for i in range(6, 7):
            for j in range(2, sheet.max_row):
                if sheet.cell(j, i).value == '合计' and sheet.cell(j, i + 1).value != '合计' and sheet.cell(j,
                                                                                                        i - 1).value != '合计':
                    for c in range(1, sheet.max_column + 1):
                        sheet.cell(j, c).font = Font(color='00FF0000')
        print('----已完成样式修改----')
    # 团队品类签收率（停用）
    def OrderQuan(self, team, tem):
        match1 = {'slgat': '港台',
                  'sltg': '泰国',
                  'slxmt': '新马',
                  'slzb': '直播团队',
                  'slyn': '越南',
                  'slrb': '日本'}
        match = {'slgat': '"神龙家族-港澳台"',
                 'sltg': '"神龙家族-泰国"',
                 'slxmt': '"神龙家族-新加坡", "神龙家族-马来西亚"',
                 'slzb': '"神龙家族-直播团队"',
                 'slyn': '"神龙家族-越南"',
                 'slrb': '"神龙家族-日本团队"'}
        # yesterday = (datetime.datetime.now()).strftime('%Y-%m-%d') + ' 23:59:59'
        # yesterday = (datetime.datetime.now().replace(month=5, day=31)).strftime('%Y-%m-%d')
        # yesterday = '2020-08-25'
        # print(yesterday)
        # last_month = (datetime.datetime.now().replace(day=1)).strftime('%Y-%m-%d')
        # last_month = (datetime.datetime.now().replace(month=5, day=27)).strftime('%Y-%m-%d')
        # last_month = '2020-08-15'
        # print(last_month)
        # -*- coding:utf-8 -*-
        sql = ''' SELECT IFNULL(ql.币种,'合计') 币种,IFNULL(ql.年月,'合计') 年月,IFNULL(ql.是否改派,'合计') 是否改派,IFNULL(ql.父级分类,'合计') 父级分类,IFNULL(ql.产品名称,'合计') 产品名称,IFNULL(ql.物流方式,'合计') 物流方式,IFNULL(ql.旬,'合计') 旬,签收,拒收,在途,未发货,未上线,已退货,理赔,自发头程丢件,已发货,已完成,全部,ql.签收 / ql.已完成 AS 完成签收, ql.签收 / ql.全部 AS 总计签收, ql.已完成 / ql.全部 AS 完成占比 ,ql.已发货 / ql.已完成 AS '已完成/已发货' , ql.已退货 / ql.全部 AS 退货率,'' 已发货占比,'' 已完成占比,'' 全部占比 FROM
        (SELECT qq.币种,qq.年月,qq.是否改派,qq.父级分类,qq.产品名称,qq.物流方式, qq.旬,sum(签收) 签收,sum(拒收) 拒收,sum(在途) 在途,sum(未发货) 未发货,sum(未上线) 未上线,sum(已退货) 已退货,sum(理赔) 理赔,sum(自发头程丢件) 自发头程丢件,sum(已发货) 已发货,sum(已完成) 已完成,sum(全部) 全部 FROM
        (SELECT q.币种,q.年月,q.是否改派,q.父级分类,q.产品名称,q.物流方式, q.旬,已签收 签收,拒收,在途,未发货,未上线,已退货,理赔,自发头程丢件,已发货,已完成 FROM
        (SELECT 币种,年月,是否改派,父级分类,CONCAT(产品id, 产品名称) 产品名称,物流方式, 旬,COUNT(最终状态) 已签收 FROM sl_tem sl				
        WHERE sl.最终状态 IN ('已签收')  
        GROUP BY 币种,年月,是否改派,父级分类,产品名称,物流方式,旬
        ORDER BY 币种,年月) q
        LEFT JOIN
        (SELECT 币种,年月,是否改派,父级分类,CONCAT(产品id, 产品名称) 产品名称,物流方式, 旬,COUNT(最终状态) 拒收 FROM sl_tem sl				
        WHERE sl.最终状态 IN ('拒收')  
        GROUP BY 币种,年月,是否改派,父级分类,产品名称,物流方式,旬
        ORDER BY 币种,年月) j
        ON q.`币种` = j.`币种` AND q.`年月` = j.`年月` AND q.`产品名称` = j.`产品名称` AND q.`物流方式` = j.`物流方式` AND q.`旬` = j.`旬` AND q.`父级分类` = j.`父级分类` AND q.`是否改派` = j.`是否改派`
        LEFT JOIN
        (SELECT 币种,年月,是否改派,父级分类,CONCAT(产品id, 产品名称) 产品名称,物流方式, 旬,COUNT(最终状态) 在途 FROM sl_tem sl 				
        WHERE sl.最终状态 IN ('在途')  
        GROUP BY 币种,年月,是否改派,父级分类,产品名称,物流方式,旬
        ORDER BY 币种,年月) zz
        ON  q.`币种` = zz.`币种` AND q.`年月` = zz.`年月`AND q.`产品名称` = zz.`产品名称`AND q.`物流方式` = zz.`物流方式`AND q.`旬` = zz.`旬` AND q.`父级分类` = zz.`父级分类` AND q.`是否改派` = zz.`是否改派`
        LEFT JOIN
        (SELECT 币种,年月,是否改派,父级分类,CONCAT(产品id, 产品名称) 产品名称,物流方式, 旬,COUNT(最终状态) 未发货 FROM sl_tem sl 				
        WHERE sl.最终状态 IN ('未发货')  
        GROUP BY 币种,年月,是否改派,父级分类,产品名称,物流方式,旬
        ORDER BY 币种,年月) wf
        ON  wf.`币种` = q.`币种` AND wf.`年月` = q.`年月`AND wf.`产品名称` = q.`产品名称`AND wf.`物流方式` = q.`物流方式`AND wf.`旬` = q.`旬` AND q.`父级分类` = wf.`父级分类` AND q.`是否改派` = wf.`是否改派`
        LEFT JOIN
        (SELECT 币种,年月,是否改派,父级分类,CONCAT(产品id, 产品名称) 产品名称,物流方式, 旬,COUNT(最终状态) 未上线 FROM sl_tem sl 				
        WHERE sl.最终状态 IN ('未上线') 
        GROUP BY 币种,年月,是否改派,父级分类,产品名称,物流方式,旬
        ORDER BY 币种,年月) ws
        ON  ws.`币种` = q.`币种` AND ws.`年月` = q.`年月`AND ws.`产品名称` = q.`产品名称`AND ws.`物流方式` = q.`物流方式`AND ws.`旬` = q.`旬`AND q.`父级分类` = ws.`父级分类` AND q.`是否改派` = ws.`是否改派`
        LEFT JOIN
        (SELECT 币种,年月,是否改派,父级分类,CONCAT(产品id, 产品名称) 产品名称,物流方式, 旬,COUNT(最终状态) 已退货 FROM sl_tem sl 				
        WHERE sl.最终状态 IN ('已退货')  
        GROUP BY 币种,年月,是否改派,父级分类,产品名称,物流方式,旬
        ORDER BY 币种,年月) th
        ON  q.`币种` = th.`币种` AND q.`年月` = th.`年月`AND q.`产品名称` = th.`产品名称`AND q.`物流方式` = th.`物流方式`AND q.`旬` = th.`旬`AND q.`父级分类` = th.`父级分类`AND q.`是否改派` = th.`是否改派`
        LEFT JOIN
        (SELECT 币种,年月,是否改派,父级分类,CONCAT(产品id, 产品名称) 产品名称,物流方式, 旬,COUNT(最终状态) 理赔 FROM sl_tem sl				
        WHERE sl.最终状态 IN ('理赔')  
        GROUP BY 币种,年月,是否改派,父级分类,产品名称,物流方式,旬
        ORDER BY 币种,年月) lp
        ON  lp.`币种` = q.`币种` AND lp.`年月` = q.`年月`AND lp.`产品名称` = q.`产品名称`AND lp.`物流方式` = q.`物流方式`AND lp.`旬` = q.`旬`AND q.`父级分类` = lp.`父级分类`AND q.`是否改派` = lp.`是否改派`
        LEFT JOIN
        (SELECT 币种,年月,是否改派,父级分类,CONCAT(产品id, 产品名称) 产品名称,物流方式, 旬,COUNT(最终状态) 自发头程丢件 FROM sl_tem sl				
        WHERE sl.最终状态 IN ('自发头程丢件') 
        GROUP BY 币种,年月,是否改派,父级分类,产品名称,物流方式,旬
        ORDER BY 币种,年月) zf
        ON  zf.`币种` = q.`币种` AND zf.`年月` = q.`年月`AND zf.`产品名称` = q.`产品名称`AND zf.`物流方式` = q.`物流方式`AND zf.`旬` = q.`旬`AND q.`父级分类` = zf.`父级分类`AND q.`是否改派` = zf.`是否改派`
        LEFT JOIN
        (SELECT 币种,年月,是否改派,父级分类,CONCAT(产品id, 产品名称) 产品名称,物流方式, 旬,COUNT(最终状态) 已发货 FROM sl_tem sl				
        WHERE sl.最终状态 IN ('已签收','拒收','理赔','已退货','在途','未上线')  
        GROUP BY 币种,年月,是否改派,父级分类,产品名称,物流方式,旬
        ORDER BY 币种,年月) fh
        ON  fh.`币种` = q.`币种` AND fh.`年月` = q.`年月`AND fh.`产品名称` = q.`产品名称`AND fh.`物流方式` = q.`物流方式`AND fh.`旬` = q.`旬`AND q.`父级分类` = fh.`父级分类`AND q.`是否改派` = fh.`是否改派`
        LEFT JOIN
        (SELECT 币种,年月,是否改派,父级分类,CONCAT(产品id, 产品名称) 产品名称,物流方式, 旬,COUNT(最终状态) 已完成 FROM sl_tem sl				
        WHERE sl.最终状态 IN ('已签收','拒收','理赔','已退货')  
        GROUP BY 币种,年月,是否改派,父级分类,产品名称,物流方式,旬
        ORDER BY 币种,年月) wc
        ON  wc.`币种` = q.`币种` AND wc.`年月` = q.`年月`AND wc.`产品名称` = q.`产品名称`AND wc.`物流方式` = q.`物流方式`AND wc.`旬` = q.`旬`AND q.`父级分类` = wc.`父级分类`AND q.`是否改派` = wc.`是否改派`
        ORDER BY 币种,年月) qq
        LEFT JOIN
        (SELECT 币种,年月,是否改派,父级分类,CONCAT(产品id, 产品名称) 产品名称,物流方式, 旬,COUNT(最终状态) 全部 FROM sl_tem sl				
        WHERE sl.最终状态 IN ('已签收','拒收','理赔','已退货','在途','未上线','未发货')  
        GROUP BY 币种,年月,是否改派,父级分类,产品名称,物流方式,旬
        ORDER BY 币种,年月) qb
        ON  qb.`币种` = qq.`币种` AND qb.`年月` = qq.`年月`AND qb.`产品名称` = qq.`产品名称`AND qb.`物流方式` = qq.`物流方式`AND qb.`旬` = qq.`旬`AND qq.`父级分类` = qb.`父级分类`AND qq.`是否改派` = qb.`是否改派`
        GROUP BY 年月,是否改派,父级分类,产品名称,物流方式,旬
        with rollup) ql;'''.format(team, tem)
        print('正在获取-' + match1[team] + '-品类签收率…………')
        df = pd.read_sql_query(sql=sql, con=self.engine1)
        print('----已获' + match1[team] + '-品类签收率')
        columns = list(df.columns)  # 获取数据的标题名，转为列表
        columns_value = ['退货率', '完成签收', '总计签收', '完成占比', '已完成/已发货']
        for column_val in columns_value:
            if column_val in columns:
                df[column_val] = df[column_val].fillna(value=0)
                df[column_val] = df[column_val].apply(lambda x: format(x, '.2%'))
        df.loc['全部'] = df.apply(lambda x: x.sum())
        df.drop(df.index[len(df) - 1], inplace=True)
        print(df)
        print('正在写入EXECL中…………')
        today = datetime.date.today().strftime('%Y.%m.%d')
        df.to_excel('D:\\Users\\Administrator\\Desktop\\输出文件\\{} 神龙{}---2200签收率.xlsx'.format(today, match1[team]),
                    sheet_name=match[team], index=False)
        print('----已写入excel')
        #  https://www.cnblogs.com/liming19680104/p/11648048.html 修改表格样式
        filePath = r'D:\\Users\\Administrator\\Desktop\\输出文件\\{} 神龙{}---2200签收率.xlsx'.format(today, match1[team])
        app = xw.App(visible=False, add_book=False)
        wb = app.books.open(filePath, update_links=False, read_only=True)
        print(match[team])
        sht = wb.sheets[match[team]]

        rng = sht.range('a1')
        rng.color = (233, 233, 235)
        sht.range("b2:g4").columns.autofit()
        print(wb.sheets[match[team]].range('d4').value)
        wb.close()
        app.quit()
    # 各团队全部订单表-函数（停用）
    def tgOrderQuan(self, team):
        match1 = {'slgat': '港台',
                  'sltg': '泰国',
                  'slxmt': '新马',
                  'slzb': '直播团队',
                  'slyn': '越南',
                  'slrb': '日本'}
        match = {'slgat': '"神龙家族-港澳台"',
                 'sltg': '"神龙家族-泰国"',
                 'slxmt': '"神龙家族-新加坡", "神龙家族-马来西亚"',
                 'slzb': '"神龙家族-直播团队"',
                 'slyn': '"神龙家族-越南"',
                 'slrb': '"神龙家族-日本团队"'}
        yesterday = (datetime.datetime.now() + datetime.timedelta(days=1)).strftime('%Y-%m-%d')
        # yesterday = '2020-08-25'
        print(yesterday)
        last_month = (datetime.datetime.now() - datetime.timedelta(days=2)).strftime('%Y-%m-%d')
        # last_month = '2020-08-15'
        print(last_month)
        sql = '''SELECT a.id,
                        a.order_number order_number,
                        dim_area.name area_id,
                        a.sale_id main_id,
                        a.ship_phone ship_phone,
                        a.ship_zip ship_zip,
                        a.amount amount,
                        tg_order_status.name order_status,
                        UPPER(a.waybill_number) waybill_number,
                        dim_payment.pay_name pay_type,
                        a.addtime addtime,
                        a.uptime update_time,
                        a.product_id goods_id, 
                        a.qty quantity,
                        dim_trans_way.all_name logistics_id,
                        '' op_id,
                        CONCAT(gk_sale.product_id, gk_sale.product_name) goods_name, 
                        IF(a.second=0,'直发','改派') secondsend_status,
                        IF(a.low_price=0,'否','是') low_price
                FROM gk_order a 
                    left join dim_area ON dim_area.id = a.area_id 
                    left join dim_payment on dim_payment.id = a.payment_id
                    left join gk_sale on gk_sale.product_id = a.product_id 
                    left join dim_trans_way on dim_trans_way.id = a.logistics_id
                    left join tg_order_status on tg_order_status.id = a.order_status
                WHERE a.rq >= '{}' AND a.rq <= '{}'
                    AND dim_area.name IN ({});'''.format(last_month, yesterday, match[team])
        print('正在获取最近 3 天订单…………')
        try:
            df = pd.read_sql_query(sql=sql, con=self.engine2)
            print('----已获取近 3 天订单')
            # print(df)
            print('正在写入缓存表中…………')
            df.to_sql('tem_tg', con=self.engine3, index=False, if_exists='replace')
        except Exception as e:
            print('更新缓存失败：', str(Exception) + str(e))
        print('++++更新缓存完成++++')
        print('正在写入全部订单表中…………')
        sql = 'REPLACE INTO 全部订单_{} SELECT *, NOW() 添加时间 FROM tem_tg;'.format(team)
        pd.read_sql_query(sql=sql, con=self.engine3, chunksize=100)
        print('----已写入全部订单表中')

if __name__ == '__main__':
    m = QueryControl()
    match1 = {'slgat': '港台',
              'sltg': '泰国',
              'slxmt': '新马',
              'slzb': '直播团队',
              'slyn': '越南',
              'slrb': '日本'}
    # messagebox.showinfo("提示！！！", "当前查询已完成--->>> 请前往（ 输出文件 ）查看")
    #  各团队全部订单表-函数
    # m.tgOrderQuan('sltg')

    # team = 'slgat'
    # for tem in ['台湾', '香港']:
    #     m.OrderQuan(team, tem)

    #  订单花费明细查询
    # match9 = {'slgat_zqsb': '港台',
    #           'sltg_zqsb': '泰国',
    #           'slxmt_zqsb': '新马',
    #           'slrb_zqsb_rb': '日本'}
    # team = 'sltg_zqsb'
    # m.sl_tem_cost(team, match9[team])

