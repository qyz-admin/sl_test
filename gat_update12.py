import pandas as pd
import os
import datetime
import xlwings

import requests
import json
import sys
from queue import Queue
from dateutil.relativedelta import relativedelta
from threading import Thread #  使用 threading 模块创建线程
import pandas.io.formats.excel

from sqlalchemy import create_engine
from settings import Settings
from emailControl import EmailControl
from openpyxl import load_workbook  # 可以向不同的sheet写入数据
from openpyxl.styles import Font, Border, Side, PatternFill, colors, \
    Alignment  # 设置字体风格为Times New Roman，大小为16，粗体、斜体，颜色蓝色

from mysqlControl import MysqlControl
# -*- coding:utf-8 -*-
class QueryUpdate(Settings):
    def __init__(self):
        Settings.__init__(self)
        self.session = requests.session()  # 实例化session，维持会话,可以让我们在跨请求时保存某些参数
        self.q = Queue()  # 多线程调用的函数不能用return返回值，用来保存返回值
        self.engine1 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql1['user'],
                                                                                    self.mysql1['password'],
                                                                                    self.mysql1['host'],
                                                                                    self.mysql1['port'],
                                                                                    self.mysql1['datebase']))
        self.engine2 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql2['user'],
                                                                                    self.mysql2['password'],
                                                                                    self.mysql2['host'],
                                                                                    self.mysql2['port'],
                                                                                    self.mysql2['datebase']))
        self.engine20 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql20['user'],
                                                                                    self.mysql20['password'],
                                                                                    self.mysql20['host'],
                                                                                    self.mysql20['port'],
                                                                                    self.mysql20['datebase']))
        self.engine3 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql3['user'],
                                                                                    self.mysql3['password'],
                                                                                    self.mysql3['host'],
                                                                                    self.mysql3['port'],
                                                                                    self.mysql3['datebase']))
        self.e = EmailControl()
        self.m = MysqlControl()
    def reSetEngine(self):
        self.engine1 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql1['user'],
                                                                                    self.mysql1['password'],
                                                                                    self.mysql1['host'],
                                                                                    self.mysql1['port'],
                                                                                    self.mysql1['datebase']))
        self.engine2 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql2['user'],
                                                                                    self.mysql2['password'],
                                                                                    self.mysql2['host'],
                                                                                    self.mysql2['port'],
                                                                                    self.mysql2['datebase']))

    # 获取签收表内容---港澳台更新签收总表(一)
    def readFormHost(self, team, write, last_time):
        start = datetime.datetime.now()
        path = r'D:\Users\Administrator\Desktop\需要用到的文件\数据库'
        dirs = os.listdir(path=path)
        # ---读取execl文件---
        for dir in dirs:
            filePath = os.path.join(path, dir)
            if dir[:2] != '~$':
                print(filePath)
                self.wbsheetHost(filePath, team, write, last_time)
        print('处理耗时：', datetime.datetime.now() - start)
    # 工作表的订单信息
    def wbsheetHost(self, filePath, team, write, last_time):
        match2 = {'slgat': '神龙港台',
                  'slgat_hfh': '火凤凰港台',
                  'slgat_hs': '红杉港台',
                  'slsc': '品牌',
                  'gat': '港台'}
        print('---正在获取 ' + match2[team] + ' 签收表的详情++++++')
        fileType = os.path.splitext(filePath)[1]
        app = xlwings.App(visible=False, add_book=False)
        app.display_alerts = False
        if 'xls' in fileType:
            wb = app.books.open(filePath, update_links=False, read_only=True)
            for sht in wb.sheets:
                try:
                    db = None
                    db = sht.used_range.options(pd.DataFrame, header=1, numbers=int, index=False).value
                    # print(db.columns)
                except Exception as e:
                    print('xxxx查看失败：' + sht.name, str(Exception) + str(e))
                if db is not None and len(db) > 0:
                    print('++++正在导入更新：' + sht.name + ' 共：' + str(len(db)) + '行',
                          'sheet共：' + str(sht.used_range.last_cell.row) + '行')
                    # 将返回的dateFrame导入数据库的临时表
                    self.writeCacheHost(db)
                    print('++++正在更新：' + sht.name + '--->>>到总订单')
                    # 将数据库的临时表替换进指定的总表
                    if write == '本期':
                        self.replaceSqlHost(team)
                    elif write == '上期':
                        self.replaceSqlHostTWO(team, last_time)
                    print('++++----->>>' + sht.name + '：订单更新完成++++')
                else:
                    print('----------数据为空导入失败：' + sht.name)
            wb.close()
        app.quit()      # 工作表的订单信息    # 工作表的订单信息

    def delete_order(self, dataFrame):    # 写入总表删除订单
        dataFrame.to_sql('delete_order', con=self.engine1, index=False, if_exists='replace')

    def writeCacheHost(self, dataFrame):    # 写入更新缓存表
        dataFrame.to_sql('gat_update', con=self.engine1, index=False, if_exists='replace')
    def replaceSqlHost(self, team):    # 更新-总表
        try:
            print('正在更新单表中......')
            sql = '''update {0}_order_list a, gat_update b
                                set a.`运单编号`= IF(b.`运单编号` = '', NULL, b.`运单编号`),
        		                    a.`是否改派`= IF(b.`是否改派` = '', NULL, b.`是否改派`),
        		                    a.`物流方式`= IF(b.`物流方式` = '', NULL, b.`物流方式`),
        		                    a.`物流名称`= IF(b.`物流名称` = '', NULL, b.`物流名称`),
        		                    a.`付款方式`= IF(b.`付款方式` = '', NULL, b.`付款方式`),
        		                    a.`产品id`= IF(b.`产品id` = '', NULL, b.`产品id`),
        		                    a.`产品名称`= IF(b.`产品名称` = '', NULL, b.`产品名称`),
        		                    a.`父级分类`= IF(b.`父级分类` = '', NULL, b.`父级分类`),
        		                    a.`二级分类`= IF(b.`二级分类` = '', NULL, b.`二级分类`),
        		                    a.`三级分类`= IF(b.`三级分类` = '', NULL, b.`三级分类`)
        		                where a.`订单编号`= b.`订单编号`;'''.format(team)
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=1000)
            print('正在更新总表中......')
            sql = '''update {0}_zqsb a, gat_update b
                                            set a.`运单编号`= IF(b.`运单编号` = '', NULL, b.`运单编号`),
                    		                    a.`是否改派`= IF(b.`是否改派` = '', NULL, b.`是否改派`),
                    		                    a.`物流方式`= IF(b.`物流方式` = '', NULL, b.`物流方式`),
                    		                    a.`物流名称`= IF(b.`物流名称` = '', NULL, b.`物流名称`),
                    		                    a.`付款方式`= IF(b.`付款方式` = '', NULL, b.`付款方式`),
                    		                    a.`产品id`= IF(b.`产品id` = '', NULL, b.`产品id`),
                    		                    a.`产品名称`= IF(b.`产品名称` = '', NULL, b.`产品名称`),
                    		                    a.`父级分类`= IF(b.`父级分类` = '', NULL, b.`父级分类`),
                    		                    a.`二级分类`= IF(b.`二级分类` = '', NULL, b.`二级分类`),
                    		                    a.`三级分类`= IF(b.`三级分类` = '', NULL, b.`三级分类`)
                    		                where a.`订单编号`= b.`订单编号`;'''.format(team)
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=1000)
        except Exception as e:
            print('更新失败：', str(Exception) + str(e))
        print('更新成功…………')
    def replaceSqlHostTWO(self, team, last_time):    # 更新上期-总表 DATE_SUB(CURDATE(), INTERVAL 1 month)
        try:
            print('正在获取往昔数据中......')
            sql = '''SELECT 年月, 旬, 日期, IF(团队 LIKE "红杉%","红杉",IF(团队 LIKE "金狮%","金狮", IF(团队 LIKE "火凤凰%","火凤凰", IF(团队 LIKE "神龙%","神龙",团队)))) 团队,
                            币种, 订单编号, 出货时间, 状态时间, 上线时间, 系统订单状态, 系统物流状态, 退货登记, 最终状态,是否改派,物流方式,
                            是否低价,产品id,产品名称,父级分类,二级分类,下单时间, 审核时间,仓储扫描时间,完结状态时间,if(价格RMB = '',null,价格RMB) 价格RMB, '{0}' 记录时间
                    FROM gat_update;'''.format(last_time)
            df = pd.read_sql_query(sql=sql, con=self.engine1)
            print('正在添加缓存中......')
            df.to_sql('gat_update_cp', con=self.engine1, index=False, if_exists='replace')
            print('正在数据添加中......')
            sql = '''REPLACE INTO qsb_{0} SELECT * FROM gat_update_cp; '''.format(team)
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=1000)
        except Exception as e:
            print('更新失败：', str(Exception) + str(e))
        print('更新成功…………')


    # 获取签收表内容---港澳台更新签收总表(一.附表)转存总表
    def makeSql(self, team):
        month_last = (datetime.datetime.now().replace(day=1) - datetime.timedelta(days=1)).strftime('%Y-%m') + '-01'
        month_yesterday = datetime.datetime.now().strftime('%Y-%m-%d')
        try:
            print('正在查询中'+ month_yesterday + '最近两个月的订单......')
            sql = '''SELECT 年月, 旬, 日期, IF(团队 LIKE "红杉%","红杉",IF(团队 LIKE "金狮%","金狮", IF(团队 LIKE "火凤凰%","火凤凰", IF(团队 LIKE "神龙%","神龙",团队)))) 团队,
                            币种, 订单编号, 出货时间, 状态时间, 上线时间, 系统订单状态, 系统物流状态, 退货登记, 最终状态,是否改派,物流方式,
                            是否低价,产品id,产品名称,父级分类,二级分类,下单时间, 审核时间,仓储扫描时间,完结状态时间,价格RMB, curdate() 记录时间
                    FROM {0}_zqsb a WHERE a.日期 >= '{1}' and a.日期 <= '{2}';'''.format(team, month_last, month_yesterday)
            df = pd.read_sql_query(sql=sql, con=self.engine1)
            print('正在添加缓存中......')
            df.to_sql('gat_update_cp', con=self.engine1, index=False, if_exists='replace')
            print('正在转存数据中......')
            sql = '''REPLACE INTO qsb_{0} SELECT * FROM gat_update_cp; '''.format(team)
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=1000)
        except Exception as e:
            print('转存失败：', str(Exception) + str(e))
        print('转存成功…………')




    # 导出需要更新的签收表---港澳台(二)
    def EportOrder(self, team):
        today = datetime.date.today().strftime('%Y.%m.%d')
        match = {'slgat': '神龙-港台',
                 'slgat_hfh': '火凤凰-港台',
                 'slgat_hs': '红杉-港台',
                 'slgat_js': '金狮-港台',
                 'gat': '港台',
                 'slsc': '品牌'}
        if team in ('gat'):
            month_last = (datetime.datetime.now().replace(day=1) - datetime.timedelta(days=1)).strftime('%Y-%m') + '-01'
            month_yesterday = datetime.datetime.now().strftime('%Y-%m-%d')
            month_begin = (datetime.datetime.now() - relativedelta(months=3)).strftime('%Y-%m-%d')
            print(month_begin)
        else:
            month_last = '2021-05-01'
            month_yesterday = '2021-06-010'
            month_begin = '2021-02-01'
        print('正在检查父级分类为空的信息---')
        sql = '''SELECT 订单编号,商品id,
        				        dp.product_id, dp.`name` product_name, dp.third_cate_id,
                                dc.ppname cate, dc.pname second_cate, dc.`name` third_cate
                        FROM (SELECT id,日期,`订单编号`,`商品id`,sl.`产品id`
                            FROM {0}_order_list sl
                            WHERE sl.`日期`> '{1}' AND (sl.`父级分类` IS NULL or sl.`父级分类`= '') AND ( NOT sl.`系统订单状态` IN ('已删除', '问题订单', '支付失败', '未支付'))
        			        ) s
                        LEFT JOIN (SELECT MAX(id),product_id,`name`,third_cate_id  FROM dim_product GROUP BY product_id ) dp ON  dp.product_id = s.`产品id`
                        LEFT JOIN dim_cate dc ON  dc.id = dp.third_cate_id;'''.format(team, month_begin)
        df = pd.read_sql_query(sql=sql, con=self.engine1)
        df.to_sql('tem_product_id', con=self.engine1, index=False, if_exists='replace')
        print('正在更新父级分类的详情…………')
        sql = '''update {0}_order_list a, tem_product_id b
                    		    set a.`父级分类`= IF(b.`cate` = '', a.`父级分类`, b.`cate`),
                    				a.`二级分类`= IF(b.`second_cate` = '', a.`二级分类`, b.`second_cate`),
                    				a.`三级分类`= IF(b.`third_cate` = '', a.`三级分类`, b.`third_cate`)
                    			where a.`订单编号`= b.`订单编号`;'''.format(team)
        pd.read_sql_query(sql=sql, con=self.engine1, chunksize=1000)
        print('更新完成+++')

        print('正在检查产品id为空的信息---')
        sql = '''SELECT 订单编号,商品id,
        				        dp.product_id, dp.`name` product_name, dp.third_cate_id
                        FROM (SELECT id,日期,`订单编号`,`商品id`,sl.`产品id`
                            FROM {0}_order_list sl
                            WHERE sl.`日期`> '{1}' AND (sl.`产品名称` IS NULL or sl.`产品名称`= '') AND ( NOT sl.`系统订单状态` IN ('已删除', '问题订单', '支付失败', '未支付'))
        			        ) s
                        LEFT JOIN (SELECT MAX(id),product_id,`name`,third_cate_id  FROM dim_product GROUP BY product_id ) dp ON dp.product_id = s.`产品id`;'''.format(
            team, month_begin)
        df = pd.read_sql_query(sql=sql, con=self.engine1)
        df.to_sql('tem_product_id', con=self.engine1, index=False, if_exists='replace')
        print('正在更新产品详情…………')
        sql = '''update {0}_order_list a, tem_product_id b
                    		    set a.`产品id`= IF(b.`product_id` = '',a.`产品id`, b.`product_id`),
                    		        a.`产品名称`= IF(b.`product_name` = '',a.`产品名称`, b.`product_name`)
                    			where a.`订单编号`= b.`订单编号`;'''.format(team)
        pd.read_sql_query(sql=sql, con=self.engine1, chunksize=1000)
        print('更新完成+++')

        print('正在获取---' + match[team] + ' ---更新数据内容…………')
        sql = '''SELECT 日期, 团队, a.订单编号 订单编号, a.运单编号 运单编号,IF(ISNULL(c.标准物流状态), b.物流状态, c.标准物流状态) 物流状态, c.`物流状态代码` 物流状态代码,系统订单状态, 
                        IF(ISNULL(d.订单编号), 系统物流状态, '已退货') 系统物流状态,
                        IF(ISNULL(d.订单编号), IF(ISNULL(系统物流状态), IF(ISNULL(c.标准物流状态) OR c.标准物流状态 = '未上线', IF(系统订单状态 IN ('已转采购', '待发货'), '未发货', '未上线') , c.标准物流状态), 系统物流状态), '已退货') 最终状态,
                        IF(是否改派='二次改派', '改派', 是否改派) 是否改派, 物流方式,物流名称,付款方式,产品id,产品名称,父级分类,二级分类,三级分类
                FROM 
                    (SELECT * FROM {0}_order_list 
                    WHERE a.日期 >= '{2}' AND a.日期 <= '{3}' AND a.系统订单状态 IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)')
                    ) a
                LEFT JOIN (SELECT * FROM {0} WHERE id IN (SELECT MAX(id) FROM {0} WHERE {0}.添加时间 > '{1}' GROUP BY 运单编号) ORDER BY id) b ON a.`运单编号` = b.`运单编号`
                LEFT JOIN {0}_logisitis_match c ON b.物流状态 = c.签收表物流状态
                LEFT JOIN {0}_return d ON a.订单编号 = d.订单编号
                ORDER BY a.`下单时间`;'''.format(team, month_begin, month_last, month_yesterday) # 港台查询函数导出
        df = pd.read_sql_query(sql=sql, con=self.engine1)
        print('正在写入excel…………')
        df.to_excel('D:\\Users\\Administrator\\Desktop\\输出文件\\{} {} 更新-签收表.xlsx'.format(today, match[team]),
                    sheet_name=match[team], index=False)
        print('----已写入excel')



    # 导出总的签收表---港澳台(三)
    def EportOrderBook(self, team):
        today = datetime.date.today().strftime('%Y.%m.%d')
        match = {'slgat': '神龙-港台',
                 'slgat_hfh': '火凤凰-港台',
                 'slgat_hs': '红杉-港台',
                 'slgat_js': '金狮-港台',
                 'gat': '港台',
                 'slsc': '品牌'}
        if team in ('gat'):
            month_last = (datetime.datetime.now().replace(day=1) - datetime.timedelta(days=1)).strftime('%Y-%m') + '-01'
            month_yesterday = datetime.datetime.now().strftime('%Y-%m-%d')
        else:
            month_last = '2021-05-01'
            month_yesterday = '2021-06-010'
        print('正在获取---' + match[team] + ' ---全部数据内容…………')
        sql = '''SELECT * FROM {0}_zqsb a WHERE a.日期 >= '{1}' AND a.日期 <= '{2}' ORDER BY a.`下单时间`;'''.format(team, month_last, month_yesterday)     # 港台查询函数导出
        df = pd.read_sql_query(sql=sql, con=self.engine1)
        print('正在写入---' + match[team] + ' ---临时缓存…………')             # 备用临时缓存表
        df.to_sql('d1_{0}'.format(team), con=self.engine1, index=False, if_exists='replace')

        for tem in ('"神龙家族-港澳台"|slgat', '"红杉家族-港澳台", "红杉家族-港澳台2"|slgat_hs', '"火凤凰-港澳台"|slgat_hfh', '"金狮-港澳台"|slgat_js'):
            tem1 = tem.split('|')[0]
            tem2 = tem.split('|')[1]
            sql = '''SELECT * FROM d1_{0} sl WHERE sl.`团队`in ({1});'''.format(team, tem1)
            df = pd.read_sql_query(sql=sql, con=self.engine1)
            df.to_sql('d1_{0}'.format(tem2), con=self.engine1, index=False, if_exists='replace')
            df.to_excel('D:\\Users\\Administrator\\Desktop\\输出文件\\{} {}签收表.xlsx'.format(today, match[tem2]),
                        sheet_name=match[tem2], index=False)
            print(tem2 + '----已写入excel')
            print('正在打印' + match[tem2] + ' 物流时效…………')
            self.m.data_wl(tem2)

    # 新版签收率-报表(刘姐看的)
    def qsb_new(self, team):  # 报表各团队近两个月的物流数据
        month_last = (datetime.datetime.now().replace(day=1) - datetime.timedelta(days=1)).strftime('%Y-%m') + '-01'
        month_last = '2021-06-01'
        month_now = datetime.datetime.now().strftime('%Y-%m-%d')
        match = {'gat': '港台-每日'}
        emailAdd = {'台湾': 'giikinliujun@163.com',
                    '香港': 'giikinliujun@163.com',
                    '品牌': 'sunyaru@giikin.com'}
        sql = '''DELETE FROM gat_zqsb
                WHERE gat_zqsb.`订单编号` IN (SELECT 订单编号
            								FROM gat_order_list 
            								WHERE gat_order_list.`系统订单状态` NOT IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)')
            								);'''
        print('正在清除总表的可能删除了的订单…………')
        pd.read_sql_query(sql=sql, con=self.engine1, chunksize=100)

        sql = '''DELETE FROM gat_zqsb gz 
                WHERE gz.`系统订单状态` = '已转采购' and gz.`是否改派` = '改派' and gz.`审核时间` >= '{0} 00:00:00' AND gz.`日期` >= '{1}';'''.format(month_now, month_last)
        print('正在清除不参与计算的今日改派订单…………')
        pd.read_sql_query(sql=sql, con=self.engine1, chunksize=100)


        sql = '''UPDATE gat_zqsb d
                SET d.`物流方式`= IF(d.`物流方式` LIKE '香港-易速配-顺丰%','香港-易速配-顺丰', IF(d.`物流方式` LIKE '台湾-天马-711%','台湾-天马-新竹', d.`物流方式`) )
                WHERE d.`是否改派` ='直发';'''
        print('正在修改-直发的物流渠道…………')
        pd.read_sql_query(sql=sql, con=self.engine1, chunksize=100)
        sql = '''UPDATE gat_zqsb d
                SET d.`物流方式`= IF(d.`物流方式` LIKE '香港-森鸿%','香港-森鸿-改派',
                                IF(d.`物流方式` LIKE '香港-立邦%','香港-立邦-改派',
    							IF(d.`物流方式` LIKE '香港-易速配%','香港-易速配-改派',
    							IF(d.`物流方式` LIKE '台湾-立邦普货头程-森鸿尾程%' OR d.`物流方式` LIKE '台湾-大黄蜂普货头程-森鸿尾程%' OR d.`物流方式` LIKE '台湾-森鸿-新竹%','森鸿',
    							IF(d.`物流方式` LIKE '台湾-天马-顺丰%','天马顺丰',
    							IF(d.`物流方式` LIKE '台湾-天马-新竹%' OR d.`物流方式` LIKE '台湾-天马-711%','天马新竹',
    							IF(d.`物流方式` LIKE '台湾-天马-黑猫%','天马黑猫',
    							IF(d.`物流方式` LIKE '台湾-易速配-龟山%' OR d.`物流方式` LIKE '台湾-易速配-新竹%' OR d.`物流方式` = '易速配','龟山',
    							IF(d.`物流方式` LIKE '台湾-速派-新竹%' OR d.`物流方式` LIKE '台湾-速派-711超商%','速派', 
    							IF(d.`物流方式` LIKE '台湾-大黄蜂普货头程-易速配尾程%' OR d.`物流方式` LIKE '台湾-立邦普货头程-易速配尾程%','龟山', d.`物流方式`)))  )  )  )  )  )  )  )
                WHERE d.`是否改派` ='改派';'''
        print('正在修改-改派的物流渠道…………')
        pd.read_sql_query(sql=sql, con=self.engine1, chunksize=100)

        filePath = []
        listT = []  # 查询sql的结果 存放池
        print('正在获取---' + match[team] + '---签收率…………')
        # 每日各线路
        print('正在获取---每日各线路…………')
        sql0 = '''SELECT 月份,地区, 家族,
                            SUM(s.昨日订单量) as 昨日订单量,
                            SUM(s.直发签收) as 直发签收,
                            SUM(s.直发拒收) as 直发拒收,
                            SUM(s.直发完成) as 直发完成,
                            SUM(s.直发总订单) as 直发总订单,
                            concat(ROUND(IFNULL(SUM(s.直发签收) / SUM(s.直发完成), 0) * 100,2),'%') as 直发完成签收,
                            concat(ROUND(IFNULL(SUM(s.直发签收) / SUM(s.直发总订单), 0) * 100,2),'%') as 直发总计签收,
                            concat(ROUND(IFNULL(SUM(s.直发完成) / SUM(s.直发总订单), 0) * 100,2),'%')as 直发完成占比,
                            SUM(s.改派签收) as 改派签收,
                            SUM(s.改派拒收) as 改派拒收,
                            SUM(s.改派完成) as 改派完成,
                            SUM(s.改派总订单) as 改派总订单,
                            concat(ROUND(IFNULL(SUM(s.改派签收) / SUM(s.改派完成), 0) * 100,2),'%') as 改派完成签收,
                            concat(ROUND(IFNULL(SUM(s.改派签收) / SUM(s.改派总订单), 0) * 100,2),'%') as 改派总计签收,
                            concat(ROUND(IFNULL(SUM(s.改派完成) / SUM(s.改派总订单), 0) * 100,2),'%') as 改派完成占比
                    FROM( SELECT IFNULL(cx.`年月`, '总计') 月份,
                                IFNULL(cx.币种, '总计') 地区,
                                IFNULL(cx.家族, '总计') 家族,  
                                SUM(IF(cx.`日期` = DATE_SUB(CURDATE(), INTERVAL 1 DAY),1,0)) as 昨日订单量,
                                SUM(IF(`是否改派` = '直发' AND 最终状态 = "已签收",1,0)) as 直发签收,
                                SUM(IF(`是否改派` = '直发' AND 最终状态 = "拒收",1,0)) as 直发拒收,
                                SUM(IF(`是否改派` = '直发' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 直发完成,
                                SUM(IF(`是否改派` = '直发',1,0)) as 直发总订单,
                                SUM(IF(`是否改派` = '改派' AND 最终状态 = "已签收",1,0)) as 改派签收,
                                SUM(IF(`是否改派` = '改派' AND 最终状态 = "拒收",1,0)) as 改派拒收,
                                SUM(IF(`是否改派` = '改派' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 改派完成,
                                SUM(IF(`是否改派` = '改派',1,0)) as 改派总订单
                            FROM (SELECT *,
                                   IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "%火凤凰%","火凤凰",IF(cc.团队 LIKE "%神龙%","神龙",IF(cc.团队 LIKE "%金狮%","金狮",IF(cc.团队 LIKE "%金鹏%","金鹏",cc.团队))))) as 家族 
                                    FROM gat_zqsb cc
                                    where cc.日期 >= '{0}' and cc.`运单编号` is not null 
                                  ) cx
                            GROUP BY cx.年月,cx.币种,cx.家族
                            WITH ROLLUP 
                        ) s
                        GROUP BY 月份,地区,家族
                        ORDER BY 月份 DESC,
                                FIELD( 地区, '台湾', '香港', '总计' ),
                                FIELD( 家族, '神龙', '火凤凰', '金狮', '金鹏', '红杉', '总计' );'''.format(month_last, team)
        df0 = pd.read_sql_query(sql=sql0, con=self.engine1)
        listT.append(df0)

        # 各月各线路
        print('正在获取---各月各线路…………')
        sql10 = '''SELECT *
                            FROM(SELECT IFNULL(cx.`年月`, '总计') 月份,
                                        IFNULL(cx.`币种`, '总计') 地区,
                                        IFNULL(cx.家族, '总计') 家族,
                                        COUNT(cx.`订单编号`) as 总单量,
            			                concat(ROUND(SUM(IF( 最终状态 = "已签收",1,0)) / SUM(IF( 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) * 100,2),'%') as 完成签收,
            			                concat(ROUND(SUM(IF( 最终状态 = "已签收",1,0)) / COUNT(cx.`订单编号`) * 100,2),'%') as 总计签收,
            			                concat(ROUND(SUM(IF( 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) / COUNT(cx.`订单编号`) * 100,2),'%') as 完成占比,
            			                concat(ROUND(SUM(IF( 最终状态 = "已退货",1,0)) / COUNT(cx.`订单编号`) * 100,2),'%') as 退款率,
            			                concat(ROUND(SUM(IF( 最终状态 = "已签收",价格RMB,0)) / SUM(价格RMB) * 100,2),'%') as '总计签收(金额)',
            			                ROUND(SUM(价格RMB) / COUNT(cx.`订单编号`),2) as 平均客单价,

                                        SUM(IF(`是否改派` = '直发',1,0))  as 直发单量,
            			                concat(ROUND(SUM(IF(`是否改派` = '直发' AND 最终状态 = "已签收",1,0)) / SUM(IF(`是否改派` = '直发' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) * 100,2),'%') as 直发完成签收,
            			                concat(ROUND(SUM(IF(`是否改派` = '直发' AND 最终状态 = "已签收",1,0)) / SUM(IF(`是否改派` = '直发',1,0)) * 100,2),'%') as 直发总计签收,
            			                concat(ROUND(SUM(IF(`是否改派` = '直发' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) / SUM(IF(`是否改派` = '直发',1,0)) * 100,2),'%') as 直发完成占比,
                                        concat(ROUND(SUM(IF(`是否改派` = '改派',1,0)) / COUNT(cx.`订单编号`) * 100,2),'%')as 改派占比,
            			                concat(ROUND(SUM(IF(`是否改派` = '改派' AND 最终状态 = "已签收",1,0)) / SUM(IF(`是否改派` = '直发' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) * 100,2),'%') as 改派完成签收,
            			                concat(ROUND(SUM(IF(`是否改派` = '改派' AND 最终状态 = "已签收",1,0)) / SUM(IF(`是否改派` = '改派',1,0)) * 100,2),'%') as 改派总计签收,
            			                concat(ROUND(SUM(IF(`是否改派` = '改派' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) / SUM(IF(`是否改派` = '改派',1,0)) * 100,2),'%') as 改派完成占比
                                FROM (SELECT *,
                                         IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "%火凤凰%","火凤凰",IF(cc.团队 LIKE "%神龙%","神龙",IF(cc.团队 LIKE "%金狮%","金狮",IF(cc.团队 LIKE "%金鹏%","金鹏",cc.团队))))) as 家族 
                                     FROM gat_zqsb cc where cc.`运单编号` is not null 
                                      ) cx									
                                GROUP BY cx.年月,cx.币种,cx.家族
                                WITH ROLLUP 
            	            ) s
                            ORDER BY 月份 DESC,
                                    FIELD( 地区, '台湾', '香港', '总计' ),
                                    FIELD( s.家族, '神龙', '火凤凰','金狮', '金鹏', '红杉', '总计' ),
                                    s.总单量 DESC;'''.format(team)
        df10 = pd.read_sql_query(sql=sql10, con=self.engine1)
        listT.append(df10)

        # 各月各线路---分旬
        print('正在获取---各月各线路---分旬…………')
        sql11 = '''SELECT *
                            FROM(SELECT IFNULL(cx.`年月`, '总计') 月份,
                                        IFNULL(cx.`旬`, '总计') 旬,
                                        IFNULL(cx.`币种`, '总计') 地区,
                                        IFNULL(cx.家族, '总计') 家族,
                                        COUNT(cx.`订单编号`) as 总单量,
            			                concat(ROUND(SUM(IF( 最终状态 = "已签收",1,0)) / SUM(IF( 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) * 100,2),'%') as 完成签收,
            			                concat(ROUND(SUM(IF( 最终状态 = "已签收",1,0)) / COUNT(cx.`订单编号`) * 100,2),'%') as 总计签收,
            			                concat(ROUND(SUM(IF( 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) / COUNT(cx.`订单编号`) * 100,2),'%') as 完成占比,
            			                concat(ROUND(SUM(IF( 最终状态 = "已退货",1,0)) / COUNT(cx.`订单编号`) * 100,2),'%') as 退款率,
            			                concat(ROUND(SUM(IF( 最终状态 = "已签收",价格RMB,0)) / SUM(价格RMB) * 100,2),'%') as '总计签收(金额)',
            			                ROUND(SUM(价格RMB) / COUNT(cx.`订单编号`),2) as 平均客单价,

                                        SUM(IF(`是否改派` = '直发',1,0))  as 直发单量,
            			                concat(ROUND(SUM(IF(`是否改派` = '直发' AND 最终状态 = "已签收",1,0)) / SUM(IF(`是否改派` = '直发' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) * 100,2),'%') as 直发完成签收,
            			                concat(ROUND(SUM(IF(`是否改派` = '直发' AND 最终状态 = "已签收",1,0)) / SUM(IF(`是否改派` = '直发',1,0)) * 100,2),'%') as 直发总计签收,
            			                concat(ROUND(SUM(IF(`是否改派` = '直发' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) / SUM(IF(`是否改派` = '直发',1,0)) * 100,2),'%') as 直发完成占比,
                                        concat(ROUND(SUM(IF(`是否改派` = '改派',1,0)) / COUNT(cx.`订单编号`) * 100,2),'%')as 改派占比,
            			                concat(ROUND(SUM(IF(`是否改派` = '改派' AND 最终状态 = "已签收",1,0)) / SUM(IF(`是否改派` = '直发' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) * 100,2),'%') as 改派完成签收,
            			                concat(ROUND(SUM(IF(`是否改派` = '改派' AND 最终状态 = "已签收",1,0)) / SUM(IF(`是否改派` = '改派',1,0)) * 100,2),'%') as 改派总计签收,
            			                concat(ROUND(SUM(IF(`是否改派` = '改派' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) / SUM(IF(`是否改派` = '改派',1,0)) * 100,2),'%') as 改派完成占比
                                FROM (SELECT *,
                                          IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "%火凤凰%","火凤凰",IF(cc.团队 LIKE "%神龙%","神龙",IF(cc.团队 LIKE "%金狮%","金狮",IF(cc.团队 LIKE "%金鹏%","金鹏",cc.团队))))) as 家族 
                                     FROM gat_zqsb cc where cc.`运单编号` is not null 
                                      )  cx									
                                GROUP BY cx.年月,cx.旬,cx.币种, cx.家族
                                WITH ROLLUP 
            	            ) s
                            ORDER BY 月份 DESC,旬,
                                    FIELD( 地区, '台湾', '香港', '总计' ),
                                    FIELD( s.家族, '神龙', '火凤凰','金狮', '金鹏', '红杉', '总计' ),
                                    s.总单量 DESC;'''.format(team)
        df11 = pd.read_sql_query(sql=sql11, con=self.engine1)
        listT.append(df11)

        # 各品类各线路
        print('正在获取---各品类各线路…………')
        sql12 = '''SELECT *
                            FROM(SELECT IFNULL(cx.`年月`, '总计') 月份,
                                        IFNULL(cx.`币种`, '总计') 地区,
                                        IFNULL(cx.`父级分类`, '总计') 父级分类,
                                        IFNULL(cx.家族, '总计') 家族,
                                        COUNT(cx.`订单编号`) as 总单量,
                                        concat(ROUND(SUM(IF( 最终状态 = "已签收",1,0)) / SUM(IF( 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) * 100,2),'%') as 完成签收,
                                        concat(ROUND(SUM(IF( 最终状态 = "已签收",1,0)) / COUNT(cx.`订单编号`) * 100,2),'%') as 总计签收,
                                        concat(ROUND(SUM(IF( 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) / COUNT(cx.`订单编号`) * 100,2),'%') as 完成占比,
                                        concat(ROUND(SUM(IF( 最终状态 = "已退货",1,0)) / COUNT(cx.`订单编号`) * 100,2),'%') as 退款率,
                                        concat(ROUND(SUM(IF( 最终状态 = "已签收",价格RMB,0)) / SUM(价格RMB) * 100,2),'%') as '总计签收(金额)',
                                        ROUND(SUM(价格RMB) / COUNT(cx.`订单编号`),2) as 平均客单价,

                                        SUM(IF(`是否改派` = '直发',1,0))  as 直发单量,
                                        concat(ROUND(SUM(IF(`是否改派` = '直发' AND 最终状态 = "已签收",1,0)) / SUM(IF(`是否改派` = '直发' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) * 100,2),'%') as 直发完成签收,
                                        concat(ROUND(SUM(IF(`是否改派` = '直发' AND 最终状态 = "已签收",1,0)) / SUM(IF(`是否改派` = '直发',1,0)) * 100,2),'%') as 直发总计签收,
                                        concat(ROUND(SUM(IF(`是否改派` = '直发' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) / SUM(IF(`是否改派` = '直发',1,0)) * 100,2),'%') as 直发完成占比,
                                        concat(ROUND(SUM(IF(`是否改派` = '改派',1,0)) / COUNT(cx.`订单编号`) * 100,2),'%')as 改派占比,
                                        concat(ROUND(SUM(IF(`是否改派` = '改派' AND 最终状态 = "已签收",1,0)) / SUM(IF(`是否改派` = '直发' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) * 100,2),'%') as 改派完成签收,
                                        concat(ROUND(SUM(IF(`是否改派` = '改派' AND 最终状态 = "已签收",1,0)) / SUM(IF(`是否改派` = '改派',1,0)) * 100,2),'%') as 改派总计签收,
                                        concat(ROUND(SUM(IF(`是否改派` = '改派' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) / SUM(IF(`是否改派` = '改派',1,0)) * 100,2),'%') as 改派完成占比
                                FROM (SELECT *,
                                          IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "%火凤凰%","火凤凰",IF(cc.团队 LIKE "%神龙%","神龙",IF(cc.团队 LIKE "%金狮%","金狮",IF(cc.团队 LIKE "%金鹏%","金鹏",cc.团队))))) as 家族 
                                     FROM gat_zqsb cc
                                      where cc.日期 >= '{0}' and cc.`运单编号` is not null 
                                    ) cx                                  
                                GROUP BY cx.年月,cx.币种,cx.父级分类,cx.家族
                                WITH ROLLUP 
                            ) s
                            ORDER BY 月份 DESC,
                                    FIELD( 地区, '台湾', '香港', '总计' ),
                                    FIELD( 父级分类, '居家百货', '电子电器', '服饰', '医药保健',  '鞋类', '美容个护', '包类','钟表珠宝','母婴玩具','总计' ),
                                    FIELD( s.家族, '神龙', '火凤凰','金狮', '金鹏', '红杉', '总计' ),
                                    s.总单量 DESC;'''.format(month_last, team)
        df12 = pd.read_sql_query(sql=sql12, con=self.engine1)
        listT.append(df12)

        # 各物流各线路
        print('正在获取---各物流各线路…………')
        sql13 = '''SELECT *
                        FROM(SELECT IFNULL(cx.`年月`, '总计') 月份,
                                    IFNULL(cx.`币种`, '总计') 地区,
                                    IFNULL(cx.`是否改派`, '总计') 是否改派,
                                    IFNULL(cx.`物流方式`, '总计') 物流方式,
                                    IFNULL(cx.家族, '总计') 家族,
                                    COUNT(cx.`订单编号`) as 总单量,
                                    concat(ROUND(SUM(IF( 最终状态 = "已签收",1,0)) / COUNT(cx.`订单编号`) * 100,2),'%') as 总计签收,
                                    concat(ROUND(SUM(IF( 最终状态 = "已签收",1,0)) / SUM(IF( 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) * 100,2),'%') as 完成签收,
                                    concat(ROUND(SUM(IF( 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) / COUNT(cx.`订单编号`) * 100,2),'%') as 完成占比,
                                    concat(ROUND(SUM(IF( 最终状态 = "已退货",1,0)) / COUNT(cx.`订单编号`) * 100,2),'%') as 退款率,
                                    concat(ROUND(SUM(IF( 最终状态 = "已签收",价格RMB,0)) / SUM(价格RMB) * 100,2),'%') as '总计签收(金额)',
                                    ROUND(SUM(价格RMB) / COUNT(cx.`订单编号`),2) as 平均客单价,

                                    SUM(IF(`是否改派` = '直发',1,0))  as 直发单量,
                                    concat(ROUND(SUM(IF(`是否改派` = '直发' AND 最终状态 = "已签收",1,0)) / SUM(IF(`是否改派` = '直发',1,0)) * 100,2),'%') as 直发总计签收,
                                    concat(ROUND(SUM(IF(`是否改派` = '直发' AND 最终状态 = "已签收",1,0)) / SUM(IF(`是否改派` = '直发' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) * 100,2),'%') as 直发完成签收,
                                    concat(ROUND(SUM(IF(`是否改派` = '直发' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) / SUM(IF(`是否改派` = '直发',1,0)) * 100,2),'%') as 直发完成占比,
                                    concat(ROUND(SUM(IF(`是否改派` = '改派',1,0)) / COUNT(cx.`订单编号`) * 100,2),'%')as 改派占比,
                                    concat(ROUND(SUM(IF(`是否改派` = '改派' AND 最终状态 = "已签收",1,0)) / SUM(IF(`是否改派` = '改派',1,0)) * 100,2),'%') as 改派总计签收,
                                    concat(ROUND(SUM(IF(`是否改派` = '改派' AND 最终状态 = "已签收",1,0)) / SUM(IF(`是否改派` = '直发' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) * 100,2),'%') as 改派完成签收,
                                    concat(ROUND(SUM(IF(`是否改派` = '改派' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) / SUM(IF(`是否改派` = '改派',1,0)) * 100,2),'%') as 改派完成占比
                            FROM (SELECT *, IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "%火凤凰%","火凤凰",IF(cc.团队 LIKE "%神龙%","神龙",IF(cc.团队 LIKE "%金狮%","金狮",IF(cc.团队 LIKE "%金鹏%","金鹏",cc.团队))))) as 家族
                                    FROM gat_zqsb cc
                                    where cc.日期 >= '{0}' and cc.`运单编号` is not null 
                                ) cx                                  
                            GROUP BY cx.年月,cx.币种,cx.是否改派,cx.物流方式,cx.家族
                            WITH ROLLUP
                        ) s
                        ORDER BY FIELD(月份, '202108', '202107', '202106', '202105', '202104', '总计' ),
                                FIELD(地区, '台湾', '香港', '总计' ),
                                FIELD(是否改派, '直发', '改派', '总计' ),
                                FIELD(物流方式, '台湾-大黄蜂普货头程-森鸿尾程','台湾-大黄蜂普货头程-易速配尾程', '台湾-立邦普货头程-森鸿尾程','台湾-立邦普货头程-易速配尾程', '台湾-森鸿-新竹-自发头程', '台湾-速派-711超商', '台湾-速派-新竹','台湾-天马-新竹','台湾-天马-顺丰','台湾-天马-黑猫','台湾-易速配-新竹',
                                    '香港-立邦-顺丰','香港-森鸿-SH渠道','香港-森鸿-顺丰渠道','香港-易速配-顺丰', '龟山','森鸿','速派','天马顺丰','天马新竹','香港-立邦-改派','香港-森鸿-改派','香港-易速配-改派','总计' ),
                                FIELD( s.家族, '神龙', '火凤凰','金狮','金鹏', '红杉', '总计' ),
                                s.总单量 DESC;'''.format(month_last, team)
        df13 = pd.read_sql_query(sql=sql13, con=self.engine1)
        listT.append(df13)

        # 同产品各团队的对比
        print('正在获取---同产品各团队的对比…………')
        sql14 = '''SELECT *,
    			            IF(神龙完成签收 = '0.00%' OR 神龙完成签收 IS NULL, 神龙完成签收, concat(ROUND(神龙完成签收-完成签收,2),'%')) as 神龙对比,
    			            IF(火凤凰完成签收 = '0.00%' OR 火凤凰完成签收 IS NULL, 火凤凰完成签收, concat(ROUND(火凤凰完成签收-完成签收,2),'%')) as 火凤凰对比,
    			            IF(金狮完成签收 = '0.00%' OR 金狮完成签收 IS NULL, 金狮完成签收, concat(ROUND(金狮完成签收-完成签收,2),'%')) as 金狮对比,
    			            IF(金鹏完成签收 = '0.00%' OR 金鹏完成签收 IS NULL, 金鹏完成签收, concat(ROUND(金鹏完成签收-完成签收,2),'%')) as 金鹏对比,
    			            IF(红杉完成签收 = '0.00%' OR 红杉完成签收 IS NULL,红杉完成签收, concat(ROUND(红杉完成签收-完成签收,2),'%')) as 红杉对比
                    FROM(SELECT IFNULL(cx.`年月`, '总计') 月份,
                                IFNULL(cx.币种, '总计') 地区,
                                IFNULL(cx.产品id, '总计') 产品id,
                                IFNULL(cx.产品名称, '总计') 产品名称,
                                IFNULL(cx.父级分类, '总计') 父级分类,
                                COUNT(cx.`订单编号`) as 总单量,
                                SUM(IF( 最终状态 = "已签收",1,0)) as 签收,
                                SUM(IF( 最终状态 = "拒收",1,0)) as 拒收,
                                concat(ROUND(SUM(IF(`是否改派` = '改派',1,0)) / COUNT(cx.`订单编号`) * 100,2),'%') as 改派占比,
                                concat(ROUND(SUM(IF( 最终状态 = "已签收",1,0)) / COUNT(cx.`订单编号`) * 100,2),'%') as 总计签收,
                            concat(ROUND(SUM(IF( 最终状态 = "已签收",1,0)) / SUM(IF( 最终状态 IN ("已签收","拒收","已退货","理赔"),1,0)) * 100,2),'%') as 完成签收,
                                concat(ROUND(SUM(IF( 最终状态 IN ("已签收","拒收","已退货","理赔"),1,0)) / COUNT(cx.`订单编号`) * 100,2),'%') as 完成占比,
                            SUM(IF(cx.家族 LIKE '神龙%',1,0)) as 神龙单量,
                                SUM(IF( cx.家族 LIKE '神龙%' AND 最终状态 = "已签收",1,0)) as 神龙签收,
                                SUM(IF( cx.家族 LIKE '神龙%' AND 最终状态 = "拒收",1,0)) as 神龙拒收,
                                concat(ROUND(SUM(IF(cx.家族 LIKE '神龙%' AND `是否改派` = '改派',1,0)) / SUM(IF(cx.家族 LIKE '神龙%',1,0)) * 100,2),'%') as 神龙改派占比,
                                concat(ROUND(SUM(IF(cx.家族 LIKE '神龙%' AND  最终状态 = "已签收",1,0)) / SUM(IF(cx.家族 LIKE '神龙%',1,0)) * 100,2),'%') as 神龙签收率,
                                concat(ROUND(SUM(IF(cx.家族 LIKE '神龙%' AND  最终状态 = "已签收",1,0)) / SUM(IF(cx.家族 LIKE '神龙%' AND 最终状态 IN ("已签收","拒收","已退货","理赔"),1,0)) * 100,2),'%') as 神龙完成签收,
                                concat(ROUND(SUM(IF(cx.家族 LIKE '神龙%' AND  最终状态 IN ("已签收","拒收","已退货","理赔"),1,0)) / SUM(IF(cx.家族 LIKE '神龙%',1,0)) * 100,2),'%') as 神龙完成占比,
                            SUM(IF(cx.家族 LIKE '火凤凰%',1,0)) as 火凤凰单量,
                                SUM(IF( cx.家族 LIKE '火凤凰%' AND 最终状态 = "已签收",1,0)) as 火凤凰签收,
                                SUM(IF( cx.家族 LIKE '火凤凰%' AND 最终状态 = "拒收",1,0)) as 火凤凰拒收,
                                concat(ROUND(SUM(IF(cx.家族 LIKE '火凤凰%' AND `是否改派` = '改派',1,0)) / SUM(IF(cx.家族 LIKE '火凤凰%',1,0)) * 100,2),'%') as 火凤凰改派占比,
                                concat(ROUND(SUM(IF(cx.家族 LIKE '火凤凰%' AND  最终状态 = "已签收",1,0)) / SUM(IF(cx.家族 LIKE '火凤凰%',1,0)) * 100,2),'%') as 火凤凰签收率,
                                 concat(ROUND(SUM(IF(cx.家族 LIKE '火凤凰%' AND  最终状态 = "已签收",1,0)) / SUM(IF(cx.家族 LIKE '火凤凰%' AND 最终状态 IN ("已签收","拒收","已退货","理赔"),1,0)) * 100,2),'%') as 火凤凰完成签收,
                                concat(ROUND(SUM(IF(cx.家族 LIKE '火凤凰%' AND  最终状态 IN ("已签收","拒收","已退货","理赔"),1,0)) / SUM(IF(cx.家族 LIKE '火凤凰%',1,0)) * 100,2),'%') as 火凤凰完成占比,
                            SUM(IF(cx.家族 LIKE '金狮%',1,0)) as 金狮单量,
                                SUM(IF( cx.家族 LIKE '金狮%' AND 最终状态 = "已签收",1,0)) as 金狮签收,
                                SUM(IF( cx.家族 LIKE '金狮%' AND 最终状态 = "拒收",1,0)) as 金狮拒收,
                                concat(ROUND(SUM(IF(cx.家族 LIKE '金狮%' AND `是否改派` = '改派',1,0)) / SUM(IF(cx.家族 LIKE '金狮%',1,0)) * 100,2),'%') as 金狮改派占比,
                                concat(ROUND(SUM(IF(cx.家族 LIKE '金狮%' AND  最终状态 = "已签收",1,0)) / SUM(IF(cx.家族 LIKE '金狮%',1,0)) * 100,2),'%') as 金狮签收率,
                                concat(ROUND(SUM(IF(cx.家族 LIKE '金狮%' AND  最终状态 = "已签收",1,0)) / SUM(IF(cx.家族 LIKE '金狮%' AND 最终状态 IN ("已签收","拒收","已退货","理赔"),1,0)) * 100,2),'%') as 金狮完成签收,
                                concat(ROUND(SUM(IF(cx.家族 LIKE '金狮%' AND  最终状态 IN ("已签收","拒收","已退货","理赔"),1,0)) / SUM(IF(cx.家族 LIKE '金狮%',1,0)) * 100,2),'%') as 金狮完成占比,
                            SUM(IF(cx.家族 LIKE '金鹏%',1,0)) as 金鹏单量,
                                SUM(IF( cx.家族 LIKE '金鹏%' AND 最终状态 = "已签收",1,0)) as 金鹏签收,
                                SUM(IF( cx.家族 LIKE '金鹏%' AND 最终状态 = "拒收",1,0)) as 金鹏拒收,
                                concat(ROUND(SUM(IF(cx.家族 LIKE '金鹏%' AND `是否改派` = '改派',1,0)) / SUM(IF(cx.家族 LIKE '金鹏%',1,0)) * 100,2),'%') as 金鹏改派占比,
                                concat(ROUND(SUM(IF(cx.家族 LIKE '金鹏%' AND  最终状态 = "已签收",1,0)) / SUM(IF(cx.家族 LIKE '金鹏%',1,0)) * 100,2),'%') as 金鹏签收率,
                                concat(ROUND(SUM(IF(cx.家族 LIKE '金鹏%' AND  最终状态 = "已签收",1,0)) / SUM(IF(cx.家族 LIKE '金鹏%' AND 最终状态 IN ("已签收","拒收","已退货","理赔"),1,0)) * 100,2),'%') as 金鹏完成签收,
                                concat(ROUND(SUM(IF(cx.家族 LIKE '金鹏%' AND  最终状态 IN ("已签收","拒收","已退货","理赔"),1,0)) / SUM(IF(cx.家族 LIKE '金鹏%',1,0)) * 100,2),'%') as 金鹏完成占比,
                            SUM(IF(cx.家族 LIKE '红杉%',1,0)) as 红杉单量,
                                SUM(IF( cx.家族 LIKE '红杉%' AND 最终状态 = "已签收",1,0)) as 红杉签收,
                                SUM(IF( cx.家族 LIKE '红杉%' AND 最终状态 = "拒收",1,0)) as 红杉拒收,
                                concat(ROUND(SUM(IF(cx.家族 LIKE '红杉%' AND `是否改派` = '改派',1,0)) / SUM(IF(cx.家族 LIKE '红杉%',1,0)) * 100,2),'%') as 红杉改派占比,
                                concat(ROUND(SUM(IF(cx.家族 LIKE '红杉%' AND  最终状态 = "已签收",1,0)) / SUM(IF(cx.家族 LIKE '红杉%',1,0)) * 100,2),'%') as 红杉签收率,
                                concat(ROUND(SUM(IF(cx.家族 LIKE '红杉%' AND  最终状态 = "已签收",1,0)) / SUM(IF(cx.家族 LIKE '红杉%' AND 最终状态 IN ("已签收","拒收","已退货","理赔"),1,0)) * 100,2),'%') as 红杉完成签收,
                                concat(ROUND(SUM(IF(cx.家族 LIKE '红杉%' AND  最终状态 IN ("已签收","拒收","已退货","理赔"),1,0)) / SUM(IF(cx.家族 LIKE '红杉%',1,0)) * 100,2),'%') as 红杉完成占比
                        FROM (SELECT *,IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "%火凤凰%","火凤凰",IF(cc.团队 LIKE "%神龙%","神龙",IF(cc.团队 LIKE "%金狮%","金狮",IF(cc.团队 LIKE "%金鹏%","金鹏",cc.团队))))) as 家族 
                            FROM gat_zqsb cc
                            where cc.日期 >= '2021-04-01' and cc.`运单编号` is not null 
                            ) cx
                        GROUP BY cx.年月,cx.币种,cx.产品id
                    WITH ROLLUP ) s
                    ORDER BY FIELD(月份,DATE_FORMAT(CURDATE(),'%Y%m'), DATE_FORMAT(DATE_SUB(CURDATE(),INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(CURDATE(),INTERVAL 2 MONTH),'%Y%m'),DATE_FORMAT(DATE_SUB(CURDATE(),INTERVAL 3 MONTH),'%Y%m'),'总计'),
                            FIELD(地区,'台湾','香港','总计'),
                            总单量 DESC;'''.format(month_last, team)
        df14 = pd.read_sql_query(sql=sql14, con=self.engine1)
        listT.append(df14)

        # 同产品各月的对比
        print('正在获取---同产品各月的对比…………')
        sql15 = '''SELECT *
                    FROM(SELECT IFNULL(cx.`家族`, '总计') 家族,
                                IFNULL(cx.币种, '总计') 地区,
                                IFNULL(cx.产品id, '总计') 产品id,
                                IFNULL(cx.产品名称, '总计') 产品名称,
                                IFNULL(cx.父级分类, '总计') 父级分类,
                                COUNT(cx.`订单编号`) as 总单量,
                            SUM(IF(cx.年月 = '202104',1,0)) as 04总单量,
                                concat(ROUND(SUM(IF(cx.年月 = '202104' AND 最终状态 = "已签收",1,0)) / SUM(IF(cx.年月 = '202104',1,0)) * 100,2),'%') as 04总计签收,
                                concat(ROUND(SUM(IF(cx.年月 = '202104' AND 最终状态 = "已签收",1,0)) / SUM(IF(cx.年月 = '202104' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) * 100,2),'%') as 04完成签收,
                                concat(ROUND(SUM(IF(cx.年月 = '202104' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) / SUM(IF(cx.年月 = '202104',1,0)) * 100,2),'%') as 04完成占比,
                            SUM(IF(cx.年月 = '202105',1,0)) as 05总单量,
                                concat(ROUND(SUM(IF(cx.年月 = '202105' AND 最终状态 = "已签收",1,0)) / SUM(IF(cx.年月 = '202105',1,0)) * 100,2),'%') as 05总计签收,
                                concat(ROUND(SUM(IF(cx.年月 = '202105' AND 最终状态 = "已签收",1,0)) / SUM(IF(cx.年月 = '202105' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) * 100,2),'%') as 05完成签收,
                                concat(ROUND(SUM(IF(cx.年月 = '202105' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) / SUM(IF(cx.年月 = '202105',1,0)) * 100,2),'%') as 05完成占比,
                            SUM(IF(cx.年月 = '202106',1,0)) as 06总单量,
                                concat(ROUND(SUM(IF(cx.年月 = '202106' AND 最终状态 = "已签收",1,0)) / SUM(IF(cx.年月 = '202106',1,0)) * 100,2),'%') as 06总计签收,
                                concat(ROUND(SUM(IF(cx.年月 = '202106' AND 最终状态 = "已签收",1,0)) / SUM(IF(cx.年月 = '202106' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) * 100,2),'%') as 06完成签收,
                                concat(ROUND(SUM(IF(cx.年月 = '202106' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) / SUM(IF(cx.年月 = '202106',1,0)) * 100,2),'%') as 06完成占比,        
                            SUM(IF(cx.年月 = '202107',1,0)) as 07总单量,
                                concat(ROUND(SUM(IF(cx.年月 = '202107' AND 最终状态 = "已签收",1,0)) / SUM(IF(cx.年月 = '202107',1,0)) * 100,2),'%') as 07总计签收,
                                concat(ROUND(SUM(IF(cx.年月 = '202107' AND 最终状态 = "已签收",1,0)) / SUM(IF(cx.年月 = '202107' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) * 100,2),'%') as 07完成签收,
                                concat(ROUND(SUM(IF(cx.年月 = '202107' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) / SUM(IF(cx.年月 = '202107',1,0)) * 100,2),'%') as 07完成占比,
                            SUM(IF(cx.年月 = '202108',1,0)) as 08总单量,
                                concat(ROUND(SUM(IF(cx.年月 = '202108' AND 最终状态 = "已签收",1,0)) / SUM(IF(cx.年月 = '202108',1,0)) * 100,2),'%') as 08总计签收,
                                concat(ROUND(SUM(IF(cx.年月 = '202108' AND 最终状态 = "已签收",1,0)) / SUM(IF(cx.年月 = '202108' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) * 100,2),'%') as 08完成签收,
                                concat(ROUND(SUM(IF(cx.年月 = '202108' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) / SUM(IF(cx.年月 = '202108',1,0)) * 100,2),'%') as 08完成占比
                        FROM (SELECT *,
                                    IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "%火凤凰%","火凤凰",IF(cc.团队 LIKE "%神龙%","神龙",IF(cc.团队 LIKE "%金狮%","金狮",IF(cc.团队 LIKE "%金鹏%","金鹏",cc.团队))))) as 家族 
                             FROM gat_zqsb cc where cc.`运单编号` is not null 
                             )  cx
                        GROUP BY cx.家族,cx.币种,cx.产品id
                        WITH ROLLUP 
                    ) s
                    ORDER BY FIELD(s.`家族`,'神龙','火凤凰','金狮','金鹏','红杉','总计'),
                            FIELD( 地区, '台湾', '香港', '总计' ),
                            s.总单量 DESC;'''
        df15 = pd.read_sql_query(sql=sql15, con=self.engine1)
        listT.append(df15)

        print('正在写入excel…………')
        today = datetime.date.today().strftime('%Y.%m.%d')
        file_path = 'D:\\Users\\Administrator\\Desktop\\输出文件\\{} {}-签收率.xlsx'.format(today, match[team])
        sheet_name = ['每日各线路', '各月各线路', '各月各线路分旬', '各品类各线路', '各物流各线路', '同产品各团队', '同产品各月']
        df0 = pd.DataFrame([])  # 创建空的dataframe数据框
        df0.to_excel(file_path, index=False)  # 备用：可以向不同的sheet写入数据（创建新的工作表并进行写入）
        writer = pd.ExcelWriter(file_path, engine='openpyxl')  # 初始化写入对象
        book = load_workbook(file_path)  # 可以向不同的sheet写入数据（对现有工作表的追加）
        writer.book = book  # 将数据写入excel中的sheet2表,sheet_name改变后即是新增一个sheet
        for i in range(len(listT)):
            listT[i].to_excel(excel_writer=writer, sheet_name=sheet_name[i], index=False)
        if 'Sheet1' in book.sheetnames:  # 删除新建文档时的第一个工作表
            del book['Sheet1']
        writer.save()
        writer.close()
        print('正在运行' + match[team] + '表宏…………')
        app = xlwings.App(visible=False, add_book=False)  # 运行宏调整
        app.display_alerts = False
        wbsht = app.books.open('D:/Users/Administrator/Desktop/新版-格式转换(工具表).xlsm')
        wbsht1 = app.books.open(file_path)
        wbsht.macro('zl_report_day')()
        wbsht1.save()
        wbsht1.close()
        wbsht.close()
        app.quit()
        print('----已写入excel ')
        # filePath.append(file_path)
        # self.e.send('{} 神龙-{}物流时效.xlsx'.format(today, tem1), filePath,
        #                 emailAdd[tem1])


    # report报表
    def qsb_report(self, team, day_yesterday, day_last):  # 获取各团队近两个月的物流数据
        match = {'gat': '港台'}
        filePath = []
        listT = []  # 查询sql的结果 存放池
        print('正在获取---' + match[team] + '---签收率…………')
        # 每日
        sql0 = '''SELECT 月份,地区, 家族,
                        SUM(s.昨日订单量) as 昨日订单量,
                        SUM(s.直发签收) as 直发签收,
                        SUM(s.直发完成) as 直发完成,
                        SUM(s.直发总订单) as 直发总订单,
                        IFNULL(SUM(s.直发签收) / SUM(s.直发完成), 0) as 直发完成签收,
                        IFNULL(SUM(s.直发签收) / SUM(s.直发总订单), 0) as 直发总计签收,
                        IFNULL(SUM(s.直发完成) / SUM(s.直发总订单), 0) as 直发完成占比,
                        SUM(s.改派签收) as 改派签收,
                        SUM(s.改派完成) as 改派完成,
                        SUM(s.改派总订单) as 改派总订单,
                        IFNULL(SUM(s.改派签收) / SUM(s.改派完成), 0) as 改派完成签收,
                        IFNULL(SUM(s.改派签收) / SUM(s.改派总订单), 0) as 改派总计签收,
                        IFNULL(SUM(s.改派完成) / SUM(s.改派总订单), 0) as 改派完成占比
                FROM( SELECT IFNULL(cx.`年月`, '总计') 月份,
                            IFNULL(cx.币种, '总计') 地区,
                            IFNULL(cx.团队, '总计') 家族,
                            SUM(IF(cx.`日期` = DATE_SUB(CURDATE(), INTERVAL 1 DAY),1,0)) as 昨日订单量,
                            SUM(IF(`是否改派` = '直发' AND 最终状态 = "已签收",1,0)) as 直发签收,
                            SUM(IF(`是否改派` = '直发' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 直发完成,
                            SUM(IF(`是否改派` = '直发',1,0)) as 直发总订单,
                            SUM(IF(`是否改派` = '改派' AND 最终状态 = "已签收",1,0)) as 改派签收,
                            SUM(IF(`是否改派` = '改派' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 改派完成,
                            SUM(IF(`是否改派` = '改派',1,0)) as 改派总订单
                        FROM  qsb_gat cx
                        WHERE cx.`记录时间` = '{1}'
                        GROUP BY cx.年月,cx.币种,cx.团队
                        WITH ROLLUP 
                    ) s
                    GROUP BY 月份,地区,家族
                    ORDER BY 月份 DESC,
                            FIELD( 地区, '台湾', '香港', '总计' ),
                            FIELD( 家族, '神龙', '火凤凰', '红杉', '金狮', '总计' );'''.format(team, day_yesterday)
        df0 = pd.read_sql_query(sql=sql0, con=self.engine1)
        listT.append(df0)

        # 总表
        sql = '''SELECT cx.币种 线路,
			                cx.团队 家族,
			                cx.年月 月份,
			                count(订单编号) as 总订单,
			                concat(ROUND(SUM(IF(最终状态 = "已签收",1,0)) / SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) * 100,2),'%') as 完成签收,
			                concat(ROUND(SUM(IF(最终状态 = "已签收",1,0)) /  count(订单编号) * 100,2),'%') as 总计签收,
			                concat(ROUND(SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) / count(订单编号) * 100,2),'%') as 完成占比,
			                null 序号
                    FROM qsb_gat cx
                    WHERE cx.`记录时间` = '{1}'
                    GROUP BY cx.币种,cx.团队,cx.年月
                    ORDER BY cx.币种,cx.团队,cx.年月;'''.format(team, day_yesterday)
        df = pd.read_sql_query(sql=sql, con=self.engine1)
        listT.append(df)
        # 总表-上月
        sql2 = '''SELECT 线路,家族,月份,总订单,完成签收,总计签收,完成占比,@rownum:=@rownum+1 AS 序号
	            FROM (SELECT cx.币种 线路,
        			        cx.团队 家族,
        			        cx.年月 月份,
        			        count(订单编号) as 总订单,
        			        concat(ROUND(SUM(IF(最终状态 = "已签收",1,0)) / SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) * 100,2),'%') as 完成签收,
        			        concat(ROUND(SUM(IF(最终状态 = "已签收",1,0)) /  count(订单编号) * 100,2),'%') as 总计签收,
        			        concat(ROUND(SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) / count(订单编号) * 100,2),'%') as 完成占比,
        			        @rownum:=0 
                        FROM qsb_gat cx
                        WHERE cx.`记录时间` = '{1}'
                        GROUP BY cx.币种,cx.团队,cx.年月
                    ) s
                ORDER BY s.线路,s.家族,s.月份;'''.format(team, day_last)
        df2 = pd.read_sql_query(sql=sql2, con=self.engine1)
        listT.append(df2)

        # 物流
        sql3 = '''SELECT s2.币种,s2.团队 家族,s2.年月,s2.是否改派,s2.物流方式,
						s2.总订单,
						concat(ROUND(s2.签收 / s2.完成 * 100,2),'%') 完成签收,
						concat(ROUND(s2.签收 / s2.总订单 * 100,2),'%') 总计签收,
						concat(ROUND(s2.完成 / s2.总订单 * 100,2),'%') 完成占比,
						concat(ROUND(s2.总订单 / s2.总订单量 * 100,2),'%') 单量占比,
						concat(ROUND(s2.签收金额 / s2.总计金额 * 100,2),'%') '总计签收(金额)',
						concat(ROUND(IF(s2.物流方式 like "总计", IF(@i > 1 ,@i - SUBSTRING_INDEX(@i,'.',1), @i), IF((@i:=@i + 总订单 / 总订单量) >1, @i - SUBSTRING_INDEX(@i,'.',1),  @i)) * 100,2),'%')  as 累计占比,
			            null 序号
				 FROM ( SELECT  IFNULL(s1.币种,'总计') as 币种,
                                IFNULL(s1.团队,'总计') as 团队,
                                IFNULL(s1.年月,'总计') as 年月,
                                IFNULL(s1.是否改派,'总计') as 是否改派,
                                IFNULL(s1.物流方式,'总计') as 物流方式,
								SUM(s1.签收) 签收,
								SUM(s1.完成) 完成,
								SUM(s1.总订单) 总订单,
								SUM(s1.总订单量) 总订单量,
								SUM(s1.签收金额) 签收金额,
								SUM(s1.总计金额) 总计金额
                        FROM (	SELECT cx.币种 as 币种,
								        cx.团队 as 团队,
								        cx.年月 as 年月,
								        cx.是否改派 as 是否改派,
								        cx.物流方式 as 物流方式,
                                        SUM(IF(最终状态 = "已签收",1,0)) as 签收,
                                        SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 完成,
                                        count(订单编号) as 总订单,
                                        总订单量,
                                        @i:=0,
										SUM(IF(最终状态 = "已签收",`价格RMB`,0)) as 签收金额,
										SUM(`价格RMB`) as 总计金额
                                FROM (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') cx
                                LEFT JOIN (SELECT 币种,团队,年月,物流方式,count(订单编号) as 总订单量
                                            FROM (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') dg
                                            GROUP BY dg.币种,dg.团队,dg.年月
                                          ) cx2 ON cx.币种 = cx2.币种 AND  cx.团队 = cx2.团队 AND  cx.年月 = cx2.年月
                                GROUP BY cx.`币种`,cx.`团队`, cx.`年月`, cx.`是否改派`, cx.`物流方式`
                                ORDER BY cx.`币种`,cx.`团队`, cx.`年月`, cx.`是否改派` DESC,总订单 DESC
						    ) s1
						    GROUP BY s1.`币种`,s1.`团队`, s1.`年月`, s1.`是否改派`, s1.`物流方式`
					   	    with rollup
					    ) s2
                ORDER BY    FIELD(s2.`币种`,'台湾','香港','总计'),
							FIELD(s2.`团队`,'神龙家族-港澳台','火凤凰-港澳台','红杉家族-港澳台','红杉家族-港澳台2','金狮-港澳台','总计'),
							FIELD(s2.`年月`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'),'总计'),
							FIELD(s2.`是否改派`,'直发','改派','总计'),
							FIELD(s2.`物流方式`,'总计'),
							s2.总订单 DESC;'''.format(team, day_yesterday)
        df3 = pd.read_sql_query(sql=sql3, con=self.engine1)
        listT.append(df3)
        # 物流-上月
        sql4 = '''SELECT 币种,团队 家族,年月,是否改派,物流方式,总订单,完成签收,总计签收,完成占比,单量占比,总计签收金额 as '总计签收(金额)',累计占比, @rownum:=@rownum+1 AS 序号
		        FROM ( SELECT s2.币种,
        							s2.团队,
        							s2.年月,
        							s2.是否改派,
        							s2.物流方式,
        							s2.总订单,
        							concat(ROUND(s2.签收 / s2.完成 * 100,2),'%') 完成签收,
        							concat(ROUND(s2.签收 / s2.总订单 * 100,2),'%') 总计签收,
        							concat(ROUND(s2.完成 / s2.总订单 * 100,2),'%') 完成占比,
        							concat(ROUND(s2.总订单 / s2.总订单量 * 100,2),'%') 单量占比,
        							concat(ROUND(s2.签收金额 / s2.总计金额 * 100,2),'%') 总计签收金额,
        							concat(ROUND(IF(s2.物流方式 like "总计", IF(@i > 1 ,@i - SUBSTRING_INDEX(@i,'.',1), @i), IF((@i:=@i + 总订单 / 总订单量) >1, @i - SUBSTRING_INDEX(@i,'.',1),  @i)) * 100,2),'%')  as 累计占比
        				 FROM ( SELECT  IFNULL(s1.币种,'总计') as 币种,
                                        IFNULL(s1.团队,'总计') as 团队,
                                        IFNULL(s1.年月,'总计') as 年月,
                                        IFNULL(s1.是否改派,'总计') as 是否改派,
                                        IFNULL(s1.物流方式,'总计') as 物流方式,
        								SUM(s1.签收) 签收,
        								SUM(s1.完成) 完成,
        								SUM(s1.总订单) 总订单,
        								SUM(s1.总订单量) 总订单量,
        								SUM(s1.签收金额) 签收金额,
        								SUM(s1.总计金额) 总计金额
                                FROM (	SELECT cx.币种 as 币种,
        								        cx.团队 as 团队,
        								        cx.年月 as 年月,
        								        cx.是否改派 as 是否改派,
        								        cx.物流方式 as 物流方式,
                                                SUM(IF(最终状态 = "已签收",1,0)) as 签收,
                                                SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 完成,
                                                count(订单编号) as 总订单,
                                                总订单量,
                                                @i:=0,
        										SUM(IF(最终状态 = "已签收",`价格RMB`,0)) as 签收金额,
        										SUM(`价格RMB`) as 总计金额
                                        FROM (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') cx
                                        LEFT JOIN (SELECT 币种,团队,年月,物流方式,count(订单编号) as 总订单量
                                                    FROM (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') dg
                                                    GROUP BY dg.币种,dg.团队,dg.年月
                                                ) cx2 ON cx.币种 = cx2.币种 AND  cx.团队 = cx2.团队 AND  cx.年月 = cx2.年月
                                        GROUP BY cx.`币种`,cx.`团队`, cx.`年月`, cx.`是否改派`, cx.`物流方式`
                                        ORDER BY cx.`币种`,cx.`团队`, cx.`年月`, cx.`是否改派` DESC,总订单 DESC
        						    ) s1
        						    GROUP BY s1.`币种`,s1.`团队`, s1.`年月`, s1.`是否改派`, s1.`物流方式`
        					   	    with rollup
        					    ) s2
                        ) s
                        ORDER BY FIELD(s.`币种`,'台湾','香港','总计'),
        					    FIELD(s.`团队`,'神龙家族-港澳台','火凤凰-港澳台','红杉家族-港澳台','红杉家族-港澳台2','金狮-港澳台','总计'),
        					    FIELD(s.`年月`, DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 3 MONTH),'%Y%m'),'总计'),
        					    FIELD(s.`是否改派`,'直发','改派','总计'),
        					    FIELD(s.`物流方式`,'总计'),
        					    s.总订单 DESC;'''.format(team, day_last)
        df4 = pd.read_sql_query(sql=sql4, con=self.engine1)
        listT.append(df4)

        # 品类
        sql5 = '''SELECT s2.币种,
								s2.团队 家族,
								s2.年月,
								s2.是否改派,
								s2.父级分类,
								s2.总订单,
								concat(ROUND(s2.签收 / s2.完成 * 100,2),'%') 完成签收,
								concat(ROUND(s2.签收 / s2.总订单 * 100,2),'%') 总计签收,
								concat(ROUND(s2.完成 / s2.总订单 * 100,2),'%') 完成占比,
								concat(ROUND(s2.总订单 / s2.总订单量 * 100,2),'%') 单量占比,
								concat(ROUND(s2.签收金额 / s2.总计金额 * 100,2),'%') '总计签收(金额)',
                                concat(ROUND(IF(s2.父级分类 like "总计", IF(@i > 1 ,@i - SUBSTRING_INDEX(@i,'.',1), @i), IF((@i:=@i + 总订单 / 总订单量) >1, @i - SUBSTRING_INDEX(@i,'.',1),  @i)) * 100,2),'%')  as 累计占比,
			                    null 序号
				 FROM (
                        SELECT  IFNULL(s1.币种,'总计') as 币种,
                                IFNULL(s1.团队,'总计') as 团队,
                                IFNULL(s1.年月,'总计') as 年月,
                                IFNULL(s1.是否改派,'总计') as 是否改派,
                                IFNULL(s1.父级分类,'总计') as 父级分类,
								SUM(s1.签收) 签收,
								SUM(s1.完成) 完成,
								SUM(s1.总订单) 总订单,
								SUM(s1.总订单量) 总订单量,
								SUM(s1.签收金额) 签收金额,
								SUM(s1.总计金额) 总计金额
                        FROM (	 SELECT cx.币种 as 币种,
								        cx.团队 as 团队,
								        cx.年月 as 年月,
								        cx.是否改派 as 是否改派,
								        cx.父级分类 as 父级分类,
                                        SUM(IF(最终状态 = "已签收",1,0)) as 签收,
                                        SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 完成,
                                        count(订单编号) as 总订单,
                                        总订单量,
                                        @i:=0,
										SUM(IF(最终状态 = "已签收",`价格RMB`,0)) as 签收金额,
										SUM(`价格RMB`) as 总计金额
                                FROM (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') cx
                                LEFT JOIN  (SELECT 币种,团队,年月,父级分类,count(订单编号) as 总订单量
                                            FROM (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') dg
                                            GROUP BY dg.币种,dg.团队,dg.年月
                                            ) cx2  ON cx.币种 = cx2.币种 AND  cx.团队 = cx2.团队 AND  cx.年月 = cx2.年月
                                GROUP BY cx.`币种`,cx.`团队`, cx.`年月`, cx.`是否改派`, cx.`父级分类`
                                ORDER BY cx.`币种`,cx.`团队`, cx.`年月`, cx.`是否改派` DESC,总订单 DESC
							) s1
						GROUP BY s1.`币种`,s1.`团队`, s1.`年月`, s1.`是否改派`, s1.`父级分类`
					   	with rollup
				 ) s2
				 ORDER BY	FIELD(s2.`币种`,'台湾','香港','总计'),
							FIELD(s2.`团队`,'神龙家族-港澳台','火凤凰-港澳台','红杉家族-港澳台','红杉家族-港澳台2','金狮-港澳台','总计'),
							FIELD(s2.`年月`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'),'总计'),
							FIELD(s2.`是否改派`,'直发','改派','总计'),
							FIELD(s2.`父级分类`,'总计'),
							s2.总订单 DESC;'''.format(team, day_yesterday)
        df5 = pd.read_sql_query(sql=sql5, con=self.engine1)
        listT.append(df5)
        # 品类-上月
        sql5 = '''SELECT 币种,团队 家族,年月,是否改派,父级分类,总订单,完成签收,总计签收,完成占比,单量占比,总计签收金额 as `总计签收(金额)`,累计占比, @rownum:=@rownum+1 AS 序号
		        FROM (SELECT s2.币种,
        								s2.团队,
        								s2.年月,
        								s2.是否改派,
        								s2.父级分类,
        								s2.总订单,
        								concat(ROUND(s2.签收 / s2.完成 * 100,2),'%') 完成签收,
        								concat(ROUND(s2.签收 / s2.总订单 * 100,2),'%') 总计签收,
        								concat(ROUND(s2.完成 / s2.总订单 * 100,2),'%') 完成占比,
        								concat(ROUND(s2.总订单 / s2.总订单量 * 100,2),'%') 单量占比,
        								concat(ROUND(s2.签收金额 / s2.总计金额 * 100,2),'%') 总计签收金额,
                                        concat(ROUND(IF(s2.父级分类 like "总计", IF(@i > 1 ,@i - SUBSTRING_INDEX(@i,'.',1), @i), IF((@i:=@i + 总订单 / 总订单量) >1, @i - SUBSTRING_INDEX(@i,'.',1),  @i)) * 100,2),'%')  as 累计占比
        				 FROM (
                                SELECT  IFNULL(s1.币种,'总计') as 币种,
                                        IFNULL(s1.团队,'总计') as 团队,
                                        IFNULL(s1.年月,'总计') as 年月,
                                        IFNULL(s1.是否改派,'总计') as 是否改派,
                                        IFNULL(s1.父级分类,'总计') as 父级分类,
        								SUM(s1.签收) 签收,
        								SUM(s1.完成) 完成,
        								SUM(s1.总订单) 总订单,
        								SUM(s1.总订单量) 总订单量,
        								SUM(s1.签收金额) 签收金额,
        								SUM(s1.总计金额) 总计金额
                                FROM (	 SELECT cx.币种 as 币种,
        								        cx.团队 as 团队,
        								        cx.年月 as 年月,
        								        cx.是否改派 as 是否改派,
        								        cx.父级分类 as 父级分类,
                                                SUM(IF(最终状态 = "已签收",1,0)) as 签收,
                                                SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 完成,
                                                count(订单编号) as 总订单,
                                                总订单量,
                                                @i:=0,
        										SUM(IF(最终状态 = "已签收",`价格RMB`,0)) as 签收金额,
        										SUM(`价格RMB`) as 总计金额
                                        FROM (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') cx
                                        LEFT JOIN  (SELECT 币种,团队,年月,父级分类,count(订单编号) as 总订单量
                                                    FROM (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') dg
                                                    GROUP BY dg.币种,dg.团队,dg.年月
                                                    ) cx2  ON cx.币种 = cx2.币种 AND  cx.团队 = cx2.团队 AND  cx.年月 = cx2.年月
                                        GROUP BY cx.`币种`,cx.`团队`, cx.`年月`, cx.`是否改派`, cx.`父级分类`
                                        ORDER BY cx.`币种`,cx.`团队`, cx.`年月`, cx.`是否改派` DESC,总订单 DESC
        							) s1
        						GROUP BY s1.`币种`,s1.`团队`, s1.`年月`, s1.`是否改派`, s1.`父级分类`
        					   	with rollup
        				) s2 
        		) s
                ORDER BY FIELD(s.`币种`,'台湾','香港','总计'),
        				FIELD(s.`团队`,'神龙家族-港澳台','火凤凰-港澳台','红杉家族-港澳台','红杉家族-港澳台2','金狮-港澳台','总计'),
        				FIELD(s.`年月`, DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 3 MONTH),'%Y%m'),'总计'),
        				FIELD(s.`是否改派`,'直发','改派','总计'),
        				FIELD(s.`父级分类`,'总计'),
        				s.总订单 DESC;'''.format(team, day_last)
        df5 = pd.read_sql_query(sql=sql5, con=self.engine1)
        listT.append(df5)

        # 产品
        sql6 = '''SELECT * 
				    FROM ( SELECT   IFNULL( cx.`币种`,'总计') as 币种,
                                    IFNULL( cx.`团队`,'总计') as 家族,
                                    IFNULL( cx.`年月`,'总计') as 年月,
                                    IFNULL( cx.`产品id`,'总计') as 产品id,
                                    cx.`产品名称`,
							        cx.`父级分类`,
                                    count(订单编号) as 总订单,
                                    SUM(IF(最终状态 = "已签收",1,0)) / SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 完成签收,
                                    SUM(IF(最终状态 = "已签收",1,0)) /  count(订单编号) as 总计签收,
                                    SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) / count(订单编号) as 完成占比,
                                    count(订单编号) /总订单2 单量占比,
                                    SUM(IF(是否改派 = '直发',1,0)) as 直发总订单,
                                    IFNULL(SUM(IF(最终状态 = "已签收" AND 是否改派 = '直发',1,0)) / SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件") AND 是否改派 = '直发',1,0)),0) as 直发完成签收,
                                    IFNULL(SUM(IF(最终状态 = "已签收" AND 是否改派 = '直发',1,0)) / SUM(IF(是否改派 = '直发',1,0)),0) as 直发总计签收,
                                    IFNULL(SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件") AND 是否改派 = '直发',1,0)) / SUM(IF(是否改派 = '直发',1,0)),0) as 直发完成占比,
                                    IFNULL(SUM(IF(是否改派 = '直发',1,0))  / 直发总订单2,0) as 直发单量占比,
                                    SUM(IF(是否改派 = '改派',1,0)) as 改派总订单,
                                    IFNULL(SUM(IF(最终状态 = "已签收" AND 是否改派 = '改派',1,0)) / SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件") AND 是否改派 = '改派',1,0)),0) as 改派完成签收,
                                    IFNULL(SUM(IF(最终状态 = "已签收" AND 是否改派 = '改派',1,0)) / SUM(IF(是否改派 = '改派',1,0)),0) as 改派总计签收,
                                    IFNULL(SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件") AND 是否改派 = '改派',1,0)) / SUM(IF(是否改派 = '改派',1,0)),0) as 改派完成占比,
                                    IFNULL(SUM(IF(是否改派 = '改派',1,0)) / 改派总订单2,0) 改派单量占比
                            FROM (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') cx
                            LEFT JOIN  (SELECT 币种,团队,年月,count(订单编号) as 总订单2 , 
											    SUM(IF(是否改派 = '直发',1,0)) as 直发总订单2 , 
												SUM(IF(是否改派 = '改派',1,0)) as 改派总订单2 
										FROM (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') da GROUP BY da.币种,da.团队,da.年月
									) cx2  ON cx.币种 = cx2.币种 AND cx.团队 = cx2.团队 AND cx.年月 = cx2.年月
                            GROUP BY cx.币种,cx.团队,cx.年月,`产品id`
	                        with rollup
					) s1
	                ORDER BY	FIELD(s1.`币种`,'台湾','香港','总计'),
								FIELD(s1.`团队`,'神龙家族-港澳台','火凤凰-港澳台','红杉家族-港澳台','红杉家族-港澳台2','金狮-港澳台','总计'),
								FIELD(s1.`年月`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'),'总计'),
								总订单 DESC;'''.format(team, day_yesterday)
        df6 = pd.read_sql_query(sql=sql6, con=self.engine1)
        listT.append(df6)

        # 产品明细-台湾
        sql7 = '''SELECT 币种,团队 家族,年月,产品id,产品名称,父级分类,订单量,
			            concat(ROUND(IF(SUBSTRING_INDEX(订单量 / 品类总订单,'.',1) > 1 ,1,订单量 / 品类总订单 ) * 100,2),'%')  AS 订单品类占比,
			            concat(ROUND(速派签收量 / 速派完成量 * 100,2),'%')  AS 速派完成签收,
			            concat(ROUND(速派签收量 / 速派单量 * 100,2),'%')  AS 速派总计签收,
			            concat(ROUND(速派完成量 / 速派单量 * 100,2),'%')  AS 速派完成占比,
			            concat(ROUND(速派单量 / 订单量 * 100,2),'%')  AS 速派单量占比,
			            concat(ROUND(711签收量 / 711完成量 * 100,2),'%')  AS 711完成签收,
			            concat(ROUND(711签收量 / 711单量 * 100,2),'%')  AS 711总计签收,
			            concat(ROUND(711完成量 / 711单量 * 100,2),'%')  AS 711完成占比,
			            concat(ROUND(711单量 / 订单量 * 100,2),'%')  AS 711单量占比,
			            concat(ROUND(天马签收量 / 天马完成量 * 100,2),'%')  AS 天马完成签收,
			            concat(ROUND(天马签收量 / 天马单量 * 100,2),'%')  AS 天马总计签收,
			            concat(ROUND(天马完成量 / 天马单量 * 100,2),'%')  AS 天马完成占比,
			            concat(ROUND(天马单量 / 订单量 * 100,2),'%')  AS 天马单量占比,
			            concat(ROUND(易速配签收量 / 易速配完成量 * 100,2),'%')  AS 易速配完成签收,
			            concat(ROUND(易速配签收量 / 易速配单量 * 100,2),'%')  AS 易速配总计签收,
			            concat(ROUND(易速配完成量 / 易速配单量 * 100,2),'%')  AS 易速配完成占比,
			            concat(ROUND(易速配单量 / 订单量 * 100,2),'%')  AS 易速配单量占比,
			            concat(ROUND(森鸿签收量 / 森鸿完成量 * 100,2),'%')  AS 森鸿完成签收,
			            concat(ROUND(森鸿签收量 / 森鸿单量 * 100,2),'%')  AS 森鸿总计签收,
			            concat(ROUND(森鸿完成量 / 森鸿单量 * 100,2),'%')  AS 森鸿完成占比,
			            concat(ROUND(森鸿单量 / 订单量 * 100,2),'%')  AS 森鸿单量占比
                FROM (SELECT IFNULL(s1.`币种`, '总计') AS 币种,
							IFNULL(s1.`团队`, '总计') AS 团队,
				            IFNULL(s1.`年月`, '总计') AS 年月,
				            IFNULL(s1.`产品id`, '总计') AS 产品id, 
							IFNULL(s1.`产品名称`, '总计') AS 产品名称,
							IFNULL(s1.`父级分类`, '总计') AS 父级分类,
				            SUM(订单量) 订单量, 
							品类总订单,
				            SUM(速派单量) 速派单量,  SUM(速派签收量) 速派签收量,  SUM(速派完成量) 速派完成量,
				            SUM(711单量) 711单量,  SUM(711签收量) 711签收量,  SUM(711完成量) 711完成量,
				            SUM(天马单量) 天马单量,  SUM(天马签收量) 天马签收量,  SUM(天马完成量) 天马完成量,
				            SUM(易速配单量) 易速配单量,  SUM(易速配签收量) 易速配签收量,  SUM(易速配完成量) 易速配完成量,
				            SUM(森鸿单量) 森鸿单量,  SUM(森鸿签收量) 森鸿签收量,  SUM(森鸿完成量) 森鸿完成量
                    FROM ( SELECT cx.`币种`, cx.`团队`, cx.`年月`, cx.`产品id`,  cx.`产品名称`,	cx.`父级分类`, 
								count(订单编号) AS 订单量,
								SUM(IF(cx.物流方式 = "台湾-速派-新竹" ,1,0)) AS 速派单量,
								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 = "已签收",1,0)) as 速派签收量,
								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 IN ( "已签收", "拒收", "已退货", "理赔", "自发头程丢件") ,1,0)) as  速派完成量,
								SUM(IF(cx.物流方式 = "台湾-速派-711超商" ,1,0)) AS '711单量',
								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 = "已签收",1,0)) as  '711签收量',
								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 IN ( "已签收", "拒收", "已退货", "理赔", "自发头程丢件") ,1,0)) as  '711完成量',
								SUM(IF(cx.物流方式 = "台湾-天马-新竹" ,1,0)) AS 天马单量,
								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 = "已签收",1,0)) as  天马签收量,
								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 IN ( "已签收", "拒收", "已退货", "理赔", "自发头程丢件") ,1,0)) as  天马完成量,
								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" ,1,0)) AS 易速配单量,
								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 = "已签收",1,0)) as  易速配签收量,
							    SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 IN ( "已签收", "拒收", "已退货", "理赔", "自发头程丢件") ,1,0)) as  易速配完成量,
								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" ,1,0)) AS 森鸿单量,
								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "已签收",1,0)) as  森鸿签收量,
								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 IN ( "已签收", "拒收", "已退货", "理赔", "自发头程丢件") ,1,0)) as  森鸿完成量
	                        FROM  (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') cx 
	                        WHERE cx.`是否改派` = '直发'
	                        GROUP BY cx.币种,cx.团队,cx.年月,cx.`产品id` 
	                    ) s1
                    LEFT JOIN 
						(SELECT cx.`币种`,  cx.`团队`, cx.`年月`,  cx.`父级分类`,  count(订单编号) AS 品类总订单 
	                    FROM  (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') cx 
	                    WHERE  cx.`是否改派` = '直发'
	                    GROUP BY	cx.币种,	cx.团队,	cx.年月,cx.`父级分类` 	
	                    ) s2 ON s1.`币种` = s2.`币种` AND s1.`团队` = s2.`团队` AND s1.`年月` = s2.`年月` AND s1.`父级分类` = s2.`父级分类` 	
                    GROUP BY	s1.币种,	s1.团队,	s1.年月,	s1.产品id
                    WITH ROLLUP 	
                ) s WHERE s.`币种` = '台湾'
                ORDER BY FIELD( s.`币种`, '台湾', '香港', '总计' ),
	                    FIELD( s.`团队`, '神龙家族-港澳台', '火凤凰-港澳台', '红杉家族-港澳台', '红杉家族-港澳台2','金狮-港澳台','总计'),
	                    FIELD(s.`年月`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'),'总计'),
	                    s.`订单量` DESC;'''.format(team, day_yesterday)
        df7 = pd.read_sql_query(sql=sql7, con=self.engine1)
        listT.append(df7)
        # 产品明细-香港
        sql8 = '''SELECT 币种,团队 家族,年月,产品id,产品名称,父级分类,订单量,
			            concat(ROUND(IF(SUBSTRING_INDEX(订单量 / 品类总订单,'.',1) > 1 ,1,订单量 / 品类总订单 ) * 100,2),'%')  AS 订单品类占比,
						concat(ROUND(立邦签收量 / 立邦完成量 * 100,2),'%')  AS 立邦完成签收,
						concat(ROUND(立邦签收量 / 立邦单量 * 100,2),'%')  AS 立邦总计签收,
						concat(ROUND(立邦完成量 / 立邦单量 * 100,2),'%')  AS 立邦完成占比,
						concat(ROUND(立邦单量 / 订单量 * 100,2),'%')  AS 立邦单量占比,
						concat(ROUND(森鸿SF签收量 / 森鸿SF完成量 * 100,2),'%')  AS 森鸿SF完成签收,
						concat(ROUND(森鸿SF签收量 / 森鸿SF单量 * 100,2),'%')  AS 森鸿SF总计签收,
						concat(ROUND(森鸿SF完成量 / 森鸿SF单量 * 100,2),'%')  AS 森鸿SF完成占比,
						concat(ROUND(森鸿SF单量 / 订单量 * 100,2),'%')  AS 森鸿SF单量占比,
					    concat(ROUND(森鸿SH签收量 / 森鸿SH完成量 * 100,2),'%')  AS 森鸿SH完成签收,
					    concat(ROUND(森鸿SH签收量 / 森鸿SH单量 * 100,2),'%')  AS 森鸿SH总计签收,
					    concat(ROUND(森鸿SH完成量 / 森鸿SH单量 * 100,2),'%')  AS 森鸿SH完成占比,
					    concat(ROUND(森鸿SH单量 / 订单量 * 100,2),'%')  AS 森鸿SH单量占比,
					    concat(ROUND(易速配SF签收量 / 易速配SF完成量 * 100,2),'%')  AS 易速配SF完成签收,
					    concat(ROUND(易速配SF签收量 / 易速配SF单量 * 100,2),'%')  AS 易速配SF总计签收,
					    concat(ROUND(易速配SF完成量 / 易速配SF单量 * 100,2),'%')  AS 易速配SF完成占比,
					    concat(ROUND(易速配SF单量 / 订单量 * 100,2),'%')  AS 易速配SF单量占比,
					    concat(ROUND(易速配YC签收量 / 易速配YC完成量 * 100,2),'%')  AS 易速配YC完成签收,
					    concat(ROUND(易速配YC签收量 / 易速配YC单量 * 100,2),'%')  AS 易速配YC总计签收,
					    concat(ROUND(易速配YC完成量 / 易速配YC单量 * 100,2),'%')  AS 易速配YC完成占比,
					    concat(ROUND(易速配YC单量 / 订单量 * 100,2),'%')  AS 易速配YC单量占比
                FROM (SELECT IFNULL(s1.`币种`, '总计') AS 币种,
							IFNULL(s1.`团队`, '总计') AS 团队,
				            IFNULL(s1.`年月`, '总计') AS 年月,
				            IFNULL(s1.`产品id`, '总计') AS 产品id, 
							IFNULL(s1.`产品名称`, '总计') AS 产品名称,
							IFNULL(s1.`父级分类`, '总计') AS 父级分类,
				            SUM(订单量) 订单量, 
							品类总订单,
							SUM(立邦单量) 立邦单量,  SUM(立邦签收量) 立邦签收量,  SUM(立邦完成量) 立邦完成量,
				            SUM(森鸿SF单量) 森鸿SF单量,  SUM(森鸿SF签收量) 森鸿SF签收量,  SUM(森鸿SF完成量) 森鸿SF完成量,
				            SUM(森鸿SH单量) 森鸿SH单量,  SUM(森鸿SH签收量) 森鸿SH签收量,  SUM(森鸿SH完成量) 森鸿SH完成量,					
				            SUM(易速配SF单量) 易速配SF单量,  SUM(易速配SF签收量) 易速配SF签收量,  SUM(易速配SF完成量) 易速配SF完成量,				
				            SUM(易速配YC单量) 易速配YC单量,  SUM(易速配YC签收量) 易速配YC签收量,  SUM(易速配YC完成量) 易速配YC完成量
                    FROM ( SELECT cx.`币种`, cx.`团队`, cx.`年月`, cx.`产品id`,  cx.`产品名称`,	cx.`父级分类`, 
								count(订单编号) AS 订单量,
								SUM(IF(cx.物流方式 = "香港-立邦-顺丰" ,1,0)) AS 立邦单量,
								SUM(IF(cx.物流方式 = "香港-立邦-顺丰" AND 最终状态 = "已签收",1,0)) as 立邦签收量,
								SUM(IF(cx.物流方式 = "香港-立邦-顺丰" AND 最终状态 IN ( "已签收", "拒收", "已退货", "理赔", "自发头程丢件") ,1,0)) as  立邦完成量,

								SUM(IF(cx.物流方式 = "香港-森鸿-顺丰渠道" ,1,0)) AS 森鸿SF单量,
								SUM(IF(cx.物流方式 = "香港-森鸿-顺丰渠道" AND 最终状态 = "已签收",1,0)) as  森鸿SF签收量,
								SUM(IF(cx.物流方式 = "香港-森鸿-顺丰渠道" AND 最终状态 IN ( "已签收", "拒收", "已退货", "理赔", "自发头程丢件") ,1,0)) as  森鸿SF完成量,

								SUM(IF(cx.物流方式 = "香港-森鸿-SH渠道" ,1,0)) AS 森鸿SH单量,
								SUM(IF(cx.物流方式 = "香港-森鸿-SH渠道" AND 最终状态 = "已签收",1,0)) as  森鸿SH签收量,
								SUM(IF(cx.物流方式 = "香港-森鸿-SH渠道" AND 最终状态 IN ( "已签收", "拒收", "已退货", "理赔", "自发头程丢件") ,1,0)) as  森鸿SH完成量,

								SUM(IF(cx.物流方式 = "香港-易速配-顺丰" ,1,0)) AS 易速配SF单量,
								SUM(IF(cx.物流方式 = "香港-易速配-顺丰" AND 最终状态 = "已签收",1,0)) as  易速配SF签收量,
							    SUM(IF(cx.物流方式 = "香港-易速配-顺丰" AND 最终状态 IN ( "已签收", "拒收", "已退货", "理赔", "自发头程丢件") ,1,0)) as  易速配SF完成量,

								SUM(IF(cx.物流方式 = "香港-易速配-顺丰YC" ,1,0)) AS 易速配YC单量,
								SUM(IF(cx.物流方式 = "香港-易速配-顺丰YC" AND 最终状态 = "已签收",1,0)) as  易速配YC签收量,
								SUM(IF(cx.物流方式 = "香港-易速配-顺丰YC" AND 最终状态 IN ( "已签收", "拒收", "已退货", "理赔", "自发头程丢件") ,1,0)) as  易速配YC完成量
	                        FROM  (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') cx 
	                        WHERE cx.`是否改派` = '直发'
	                        GROUP BY cx.币种,cx.团队,cx.年月,cx.`产品id` 
	                    ) s1
                    LEFT JOIN 
						(SELECT cx.`币种`,  cx.`团队`, cx.`年月`,  cx.`父级分类`,  count(订单编号) AS 品类总订单 
	                    FROM  (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') cx 
	                    WHERE  cx.`是否改派` = '直发'
	                    GROUP BY	cx.币种,	cx.团队,	cx.年月,cx.`父级分类` 	
	                    ) s2 ON s1.`币种` = s2.`币种` AND s1.`团队` = s2.`团队` AND s1.`年月` = s2.`年月` AND s1.`父级分类` = s2.`父级分类` 	
                    GROUP BY	s1.币种,	s1.团队,	s1.年月,	s1.产品id
                    WITH ROLLUP 	
                ) s WHERE s.`币种` = '香港'
                ORDER BY FIELD( s.`币种`, '台湾', '香港', '总计' ),
	                    FIELD( s.`团队`, '神龙家族-港澳台', '火凤凰-港澳台', '红杉家族-港澳台', '红杉家族-港澳台2','金狮-港澳台','总计'),
	                    FIELD(s.`年月`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'),'总计'),
	                    s.`订单量` DESC;'''.format(team, day_yesterday)
        df8 = pd.read_sql_query(sql=sql8, con=self.engine1)
        listT.append(df8)

        print('正在写入excel…………')
        today = datetime.date.today().strftime('%Y.%m.%d')
        for wbbook in ['神龙', '火凤凰', '红杉', '金狮']:
            file_path = 'D:\\Users\\Administrator\\Desktop\\输出文件\\{} {}-签收率.xlsx'.format(today, wbbook)
            sheet_name = ['每日', '总表', '总表上月', '物流', '物流上月', '品类', '品类上月', '产品', '产品明细台湾', '产品明细香港']
            df0 = pd.DataFrame([])  # 创建空的dataframe数据框
            df0.to_excel(file_path, index=False)  # 备用：可以向不同的sheet写入数据（创建新的工作表并进行写入）
            writer = pd.ExcelWriter(file_path, engine='openpyxl')  # 初始化写入对象
            book = load_workbook(file_path)  # 可以向不同的sheet写入数据（对现有工作表的追加）
            writer.book = book  # 将数据写入excel中的sheet2表,sheet_name改变后即是新增一个sheet
            for i in range(len(listT)):
                listT[i].to_excel(excel_writer=writer, sheet_name=sheet_name[i], index=False)
            if 'Sheet1' in book.sheetnames:  # 删除新建文档时的第一个工作表
                del book['Sheet1']
            writer.save()
            writer.close()
            # print('正在运行' + wbbook + '表宏…………')
            # app = xlwings.App(visible=False, add_book=False)  # 运行宏调整
            # app.display_alerts = False
            # wbsht = app.books.open('D:/Users/Administrator/Desktop/新版-格式转换(工具表).xlsm')
            # wbsht1 = app.books.open(file_path)
            # wbsht.macro('py_sl_总运行')()
            # wbsht1.save()
            # wbsht1.close()
            # wbsht.close()
            # app.quit()
        print('----已写入excel ')
    # 获取各团队近两个月的物流数据
    def qsb_report_T(self, team, day_yesterday, day_last):
        match = {'gat': '港台'}
        filePath = []
        listT = []  # 查询sql的结果 存放池
        print('正在获取---' + match[team] + '---签收率…………')
        # 物流
        sql3 = '''SELECT s2.币种,s2.团队 家族,s2.年月,s2.是否改派,s2.物流方式,
						s2.总订单,
						concat(ROUND(s2.签收 / s2.完成 * 100,2),'%') 完成签收,
						concat(ROUND(s2.签收 / s2.总订单 * 100,2),'%') 总计签收,
						concat(ROUND(s2.完成 / s2.总订单 * 100,2),'%') 完成占比,
						concat(ROUND(s2.总订单 / s2.总订单量 * 100,2),'%') 单量占比,
						concat(ROUND(s2.签收金额 / s2.总计金额 * 100,2),'%') '总计签收(金额)',
						concat(ROUND(IF(s2.物流方式 like "总计", IF(@i > 1 ,@i - SUBSTRING_INDEX(@i,'.',1), @i), IF((@i:=@i + 总订单 / 总订单量) >1, @i - SUBSTRING_INDEX(@i,'.',1),  @i)) * 100,2),'%')  as 累计占比,
			            null 序号
				 FROM ( SELECT  IFNULL(s1.币种,'总计') as 币种,
                                IFNULL(s1.团队,'总计') as 团队,
                                IFNULL(s1.年月,'总计') as 年月,
                                IFNULL(s1.是否改派,'总计') as 是否改派,
                                IFNULL(s1.物流方式,'总计') as 物流方式,
								SUM(s1.签收) 签收,
								SUM(s1.完成) 完成,
								SUM(s1.总订单) 总订单,
								SUM(s1.总订单量) 总订单量,
								SUM(s1.签收金额) 签收金额,
								SUM(s1.总计金额) 总计金额
                        FROM (	SELECT cx.币种 as 币种,
								        cx.团队 as 团队,
								        cx.年月 as 年月,
								        cx.是否改派 as 是否改派,
								        cx.物流方式 as 物流方式,
                                        SUM(IF(最终状态 = "已签收",1,0)) as 签收,
                                        SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 完成,
                                        count(订单编号) as 总订单,
                                        总订单量,
                                        @i:=0,
										SUM(IF(最终状态 = "已签收",`价格RMB`,0)) as 签收金额,
										SUM(`价格RMB`) as 总计金额
                                FROM (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') cx
                                LEFT JOIN (SELECT 币种,团队,年月,物流方式,count(订单编号) as 总订单量
                                            FROM (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') dg
                                            GROUP BY dg.币种,dg.团队,dg.年月
                                          ) cx2 ON cx.币种 = cx2.币种 AND  cx.团队 = cx2.团队 AND  cx.年月 = cx2.年月
                                GROUP BY cx.`币种`,cx.`团队`, cx.`年月`, cx.`是否改派`, cx.`物流方式`
                                ORDER BY cx.`币种`,cx.`团队`, cx.`年月`, cx.`是否改派` DESC,总订单 DESC
						    ) s1
						    GROUP BY s1.`币种`,s1.`团队`, s1.`年月`, s1.`是否改派`, s1.`物流方式`
					   	    with rollup
					    ) s2
                ORDER BY    FIELD(s2.`币种`,'台湾','香港','总计'),
							FIELD(s2.`团队`,'神龙家族-港澳台','火凤凰-港澳台','红杉家族-港澳台','红杉家族-港澳台2','金狮-港澳台','总计'),
							FIELD(s2.`年月`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'),'总计'),
							FIELD(s2.`是否改派`,'直发','改派','总计'),
							FIELD(s2.`物流方式`,'总计'),
							s2.总订单 DESC;'''.format(team, day_yesterday)
        df3 = pd.read_sql_query(sql=sql3, con=self.engine1)
        listT.append(df3)
        # 物流-上月
        sql4 = '''SELECT 币种,团队 家族,年月,是否改派,物流方式,总订单,完成签收,总计签收,完成占比,单量占比,总计签收金额 as '总计签收(金额)',累计占比, @rownum:=@rownum+1 AS 序号
		        FROM ( SELECT s2.币种,
        							s2.团队,
        							s2.年月,
        							s2.是否改派,
        							s2.物流方式,
        							s2.总订单,
        							concat(ROUND(s2.签收 / s2.完成 * 100,2),'%') 完成签收,
        							concat(ROUND(s2.签收 / s2.总订单 * 100,2),'%') 总计签收,
        							concat(ROUND(s2.完成 / s2.总订单 * 100,2),'%') 完成占比,
        							concat(ROUND(s2.总订单 / s2.总订单量 * 100,2),'%') 单量占比,
        							concat(ROUND(s2.签收金额 / s2.总计金额 * 100,2),'%') 总计签收金额,
        							concat(ROUND(IF(s2.物流方式 like "总计", IF(@i > 1 ,@i - SUBSTRING_INDEX(@i,'.',1), @i), IF((@i:=@i + 总订单 / 总订单量) >1, @i - SUBSTRING_INDEX(@i,'.',1),  @i)) * 100,2),'%')  as 累计占比
        				 FROM ( SELECT  IFNULL(s1.币种,'总计') as 币种,
                                        IFNULL(s1.团队,'总计') as 团队,
                                        IFNULL(s1.年月,'总计') as 年月,
                                        IFNULL(s1.是否改派,'总计') as 是否改派,
                                        IFNULL(s1.物流方式,'总计') as 物流方式,
        								SUM(s1.签收) 签收,
        								SUM(s1.完成) 完成,
        								SUM(s1.总订单) 总订单,
        								SUM(s1.总订单量) 总订单量,
        								SUM(s1.签收金额) 签收金额,
        								SUM(s1.总计金额) 总计金额
                                FROM (	SELECT cx.币种 as 币种,
        								        cx.团队 as 团队,
        								        cx.年月 as 年月,
        								        cx.是否改派 as 是否改派,
        								        cx.物流方式 as 物流方式,
                                                SUM(IF(最终状态 = "已签收",1,0)) as 签收,
                                                SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 完成,
                                                count(订单编号) as 总订单,
                                                总订单量,
                                                @i:=0,
        										SUM(IF(最终状态 = "已签收",`价格RMB`,0)) as 签收金额,
        										SUM(`价格RMB`) as 总计金额
                                        FROM (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') cx
                                        LEFT JOIN (SELECT 币种,团队,年月,物流方式,count(订单编号) as 总订单量
                                                    FROM (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') dg
                                                    GROUP BY dg.币种,dg.团队,dg.年月
                                                ) cx2 ON cx.币种 = cx2.币种 AND  cx.团队 = cx2.团队 AND  cx.年月 = cx2.年月
                                        GROUP BY cx.`币种`,cx.`团队`, cx.`年月`, cx.`是否改派`, cx.`物流方式`
                                        ORDER BY cx.`币种`,cx.`团队`, cx.`年月`, cx.`是否改派` DESC,总订单 DESC
        						    ) s1
        						    GROUP BY s1.`币种`,s1.`团队`, s1.`年月`, s1.`是否改派`, s1.`物流方式`
        					   	    with rollup
        					    ) s2
                        ) s
                        ORDER BY FIELD(s.`币种`,'台湾','香港','总计'),
        					    FIELD(s.`团队`,'神龙家族-港澳台','火凤凰-港澳台','红杉家族-港澳台','红杉家族-港澳台2','金狮-港澳台','总计'),
        					    FIELD(s.`年月`, DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 3 MONTH),'%Y%m'),'总计'),
        					    FIELD(s.`是否改派`,'直发','改派','总计'),
        					    FIELD(s.`物流方式`,'总计'),
        					    s.总订单 DESC;'''.format(team, day_last)
        df4 = pd.read_sql_query(sql=sql4, con=self.engine1)
        listT.append(df4)

        # 品类
        sql5 = '''SELECT s2.币种,
								s2.团队 家族,
								s2.年月,
								s2.是否改派,
								s2.父级分类,
								s2.总订单,
								concat(ROUND(s2.签收 / s2.完成 * 100,2),'%') 完成签收,
								concat(ROUND(s2.签收 / s2.总订单 * 100,2),'%') 总计签收,
								concat(ROUND(s2.完成 / s2.总订单 * 100,2),'%') 完成占比,
								concat(ROUND(s2.总订单 / s2.总订单量 * 100,2),'%') 单量占比,
								concat(ROUND(s2.签收金额 / s2.总计金额 * 100,2),'%') '总计签收(金额)',
                                concat(ROUND(IF(s2.父级分类 like "总计", IF(@i > 1 ,@i - SUBSTRING_INDEX(@i,'.',1), @i), IF((@i:=@i + 总订单 / 总订单量) >1, @i - SUBSTRING_INDEX(@i,'.',1),  @i)) * 100,2),'%')  as 累计占比,
			                    null 序号
				 FROM (
                        SELECT  IFNULL(s1.币种,'总计') as 币种,
                                IFNULL(s1.团队,'总计') as 团队,
                                IFNULL(s1.年月,'总计') as 年月,
                                IFNULL(s1.是否改派,'总计') as 是否改派,
                                IFNULL(s1.父级分类,'总计') as 父级分类,
								SUM(s1.签收) 签收,
								SUM(s1.完成) 完成,
								SUM(s1.总订单) 总订单,
								SUM(s1.总订单量) 总订单量,
								SUM(s1.签收金额) 签收金额,
								SUM(s1.总计金额) 总计金额
                        FROM (	 SELECT cx.币种 as 币种,
								        cx.团队 as 团队,
								        cx.年月 as 年月,
								        cx.是否改派 as 是否改派,
								        cx.父级分类 as 父级分类,
                                        SUM(IF(最终状态 = "已签收",1,0)) as 签收,
                                        SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 完成,
                                        count(订单编号) as 总订单,
                                        总订单量,
                                        @i:=0,
										SUM(IF(最终状态 = "已签收",`价格RMB`,0)) as 签收金额,
										SUM(`价格RMB`) as 总计金额
                                FROM (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') cx
                                LEFT JOIN  (SELECT 币种,团队,年月,父级分类,count(订单编号) as 总订单量
                                            FROM (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') dg
                                            GROUP BY dg.币种,dg.团队,dg.年月
                                            ) cx2  ON cx.币种 = cx2.币种 AND  cx.团队 = cx2.团队 AND  cx.年月 = cx2.年月
                                GROUP BY cx.`币种`,cx.`团队`, cx.`年月`, cx.`是否改派`, cx.`父级分类`
                                ORDER BY cx.`币种`,cx.`团队`, cx.`年月`, cx.`是否改派` DESC,总订单 DESC
							) s1
						GROUP BY s1.`币种`,s1.`团队`, s1.`年月`, s1.`是否改派`, s1.`父级分类`
					   	with rollup
				 ) s2
				 ORDER BY	FIELD(s2.`币种`,'台湾','香港','总计'),
							FIELD(s2.`团队`,'神龙家族-港澳台','火凤凰-港澳台','红杉家族-港澳台','红杉家族-港澳台2','金狮-港澳台','总计'),
							FIELD(s2.`年月`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'),'总计'),
							FIELD(s2.`是否改派`,'直发','改派','总计'),
							FIELD(s2.`父级分类`,'总计'),
							s2.总订单 DESC;'''.format(team, day_yesterday)
        df5 = pd.read_sql_query(sql=sql5, con=self.engine1)
        listT.append(df5)
        # 品类-上月
        sql5 = '''SELECT 币种,团队 家族,年月,是否改派,父级分类,总订单,完成签收,总计签收,完成占比,单量占比,总计签收金额 as `总计签收(金额)`,累计占比, @rownum:=@rownum+1 AS 序号
		        FROM (SELECT s2.币种,
        								s2.团队,
        								s2.年月,
        								s2.是否改派,
        								s2.父级分类,
        								s2.总订单,
        								concat(ROUND(s2.签收 / s2.完成 * 100,2),'%') 完成签收,
        								concat(ROUND(s2.签收 / s2.总订单 * 100,2),'%') 总计签收,
        								concat(ROUND(s2.完成 / s2.总订单 * 100,2),'%') 完成占比,
        								concat(ROUND(s2.总订单 / s2.总订单量 * 100,2),'%') 单量占比,
        								concat(ROUND(s2.签收金额 / s2.总计金额 * 100,2),'%') 总计签收金额,
                                        concat(ROUND(IF(s2.父级分类 like "总计", IF(@i > 1 ,@i - SUBSTRING_INDEX(@i,'.',1), @i), IF((@i:=@i + 总订单 / 总订单量) >1, @i - SUBSTRING_INDEX(@i,'.',1),  @i)) * 100,2),'%')  as 累计占比
        				 FROM (
                                SELECT  IFNULL(s1.币种,'总计') as 币种,
                                        IFNULL(s1.团队,'总计') as 团队,
                                        IFNULL(s1.年月,'总计') as 年月,
                                        IFNULL(s1.是否改派,'总计') as 是否改派,
                                        IFNULL(s1.父级分类,'总计') as 父级分类,
        								SUM(s1.签收) 签收,
        								SUM(s1.完成) 完成,
        								SUM(s1.总订单) 总订单,
        								SUM(s1.总订单量) 总订单量,
        								SUM(s1.签收金额) 签收金额,
        								SUM(s1.总计金额) 总计金额
                                FROM (	 SELECT cx.币种 as 币种,
        								        cx.团队 as 团队,
        								        cx.年月 as 年月,
        								        cx.是否改派 as 是否改派,
        								        cx.父级分类 as 父级分类,
                                                SUM(IF(最终状态 = "已签收",1,0)) as 签收,
                                                SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 完成,
                                                count(订单编号) as 总订单,
                                                总订单量,
                                                @i:=0,
        										SUM(IF(最终状态 = "已签收",`价格RMB`,0)) as 签收金额,
        										SUM(`价格RMB`) as 总计金额
                                        FROM (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') cx
                                        LEFT JOIN  (SELECT 币种,团队,年月,父级分类,count(订单编号) as 总订单量
                                                    FROM (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') dg
                                                    GROUP BY dg.币种,dg.团队,dg.年月
                                                    ) cx2  ON cx.币种 = cx2.币种 AND  cx.团队 = cx2.团队 AND  cx.年月 = cx2.年月
                                        GROUP BY cx.`币种`,cx.`团队`, cx.`年月`, cx.`是否改派`, cx.`父级分类`
                                        ORDER BY cx.`币种`,cx.`团队`, cx.`年月`, cx.`是否改派` DESC,总订单 DESC
        							) s1
        						GROUP BY s1.`币种`,s1.`团队`, s1.`年月`, s1.`是否改派`, s1.`父级分类`
        					   	with rollup
        				) s2 
        		) s
                ORDER BY FIELD(s.`币种`,'台湾','香港','总计'),
        				FIELD(s.`团队`,'神龙家族-港澳台','火凤凰-港澳台','红杉家族-港澳台','红杉家族-港澳台2','金狮-港澳台','总计'),
        				FIELD(s.`年月`, DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 3 MONTH),'%Y%m'),'总计'),
        				FIELD(s.`是否改派`,'直发','改派','总计'),
        				FIELD(s.`父级分类`,'总计'),
        				s.总订单 DESC;'''.format(team, day_last)
        df5 = pd.read_sql_query(sql=sql5, con=self.engine1)
        listT.append(df5)

        # 产品
        sql6 = '''SELECT * 
				    FROM ( SELECT   IFNULL( cx.`币种`,'总计') as 币种,
                                    IFNULL( cx.`团队`,'总计') as 家族,
                                    IFNULL( cx.`年月`,'总计') as 年月,
                                    IFNULL( cx.`产品id`,'总计') as 产品id,
                                    cx.`产品名称`,
							        cx.`父级分类`,
                                    count(订单编号) as 总订单,
                                    SUM(IF(最终状态 = "已签收",1,0)) / SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 完成签收,
                                    SUM(IF(最终状态 = "已签收",1,0)) /  count(订单编号) as 总计签收,
                                    SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) / count(订单编号) as 完成占比,
                                    count(订单编号) /总订单2 单量占比,
                                    SUM(IF(是否改派 = '直发',1,0)) as 直发总订单,
                                    IFNULL(SUM(IF(最终状态 = "已签收" AND 是否改派 = '直发',1,0)) / SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件") AND 是否改派 = '直发',1,0)),0) as 直发完成签收,
                                    IFNULL(SUM(IF(最终状态 = "已签收" AND 是否改派 = '直发',1,0)) / SUM(IF(是否改派 = '直发',1,0)),0) as 直发总计签收,
                                    IFNULL(SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件") AND 是否改派 = '直发',1,0)) / SUM(IF(是否改派 = '直发',1,0)),0) as 直发完成占比,
                                    IFNULL(SUM(IF(是否改派 = '直发',1,0))  / 直发总订单2,0) as 直发单量占比,
                                    SUM(IF(是否改派 = '改派',1,0)) as 改派总订单,
                                    IFNULL(SUM(IF(最终状态 = "已签收" AND 是否改派 = '改派',1,0)) / SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件") AND 是否改派 = '改派',1,0)),0) as 改派完成签收,
                                    IFNULL(SUM(IF(最终状态 = "已签收" AND 是否改派 = '改派',1,0)) / SUM(IF(是否改派 = '改派',1,0)),0) as 改派总计签收,
                                    IFNULL(SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件") AND 是否改派 = '改派',1,0)) / SUM(IF(是否改派 = '改派',1,0)),0) as 改派完成占比,
                                    IFNULL(SUM(IF(是否改派 = '改派',1,0)) / 改派总订单2,0) 改派单量占比
                            FROM (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') cx
                            LEFT JOIN  (SELECT 币种,团队,年月,count(订单编号) as 总订单2 , 
											    SUM(IF(是否改派 = '直发',1,0)) as 直发总订单2 , 
												SUM(IF(是否改派 = '改派',1,0)) as 改派总订单2 
										FROM (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') da GROUP BY da.币种,da.团队,da.年月
									) cx2  ON cx.币种 = cx2.币种 AND cx.团队 = cx2.团队 AND cx.年月 = cx2.年月
                            GROUP BY cx.币种,cx.团队,cx.年月,`产品id`
	                        with rollup
					) s1
	                ORDER BY	FIELD(s1.`币种`,'台湾','香港','总计'),
								FIELD(s1.`团队`,'神龙家族-港澳台','火凤凰-港澳台','红杉家族-港澳台','红杉家族-港澳台2','金狮-港澳台','总计'),
								FIELD(s1.`年月`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'),'总计'),
								总订单 DESC;'''.format(team, day_yesterday)
        df6 = pd.read_sql_query(sql=sql6, con=self.engine1)
        listT.append(df6)

        # 产品明细-台湾
        sql7 = '''SELECT 币种,团队 家族,年月,产品id,产品名称,父级分类,订单量,
			            concat(ROUND(IF(SUBSTRING_INDEX(订单量 / 品类总订单,'.',1) > 1 ,1,订单量 / 品类总订单 ) * 100,2),'%')  AS 订单品类占比,
			            concat(ROUND(速派签收量 / 速派完成量 * 100,2),'%')  AS 速派完成签收,
			            concat(ROUND(速派签收量 / 速派单量 * 100,2),'%')  AS 速派总计签收,
			            concat(ROUND(速派完成量 / 速派单量 * 100,2),'%')  AS 速派完成占比,
			            concat(ROUND(速派单量 / 订单量 * 100,2),'%')  AS 速派单量占比,
			            concat(ROUND(711签收量 / 711完成量 * 100,2),'%')  AS 711完成签收,
			            concat(ROUND(711签收量 / 711单量 * 100,2),'%')  AS 711总计签收,
			            concat(ROUND(711完成量 / 711单量 * 100,2),'%')  AS 711完成占比,
			            concat(ROUND(711单量 / 订单量 * 100,2),'%')  AS 711单量占比,
			            concat(ROUND(天马签收量 / 天马完成量 * 100,2),'%')  AS 天马完成签收,
			            concat(ROUND(天马签收量 / 天马单量 * 100,2),'%')  AS 天马总计签收,
			            concat(ROUND(天马完成量 / 天马单量 * 100,2),'%')  AS 天马完成占比,
			            concat(ROUND(天马单量 / 订单量 * 100,2),'%')  AS 天马单量占比,
			            concat(ROUND(易速配签收量 / 易速配完成量 * 100,2),'%')  AS 易速配完成签收,
			            concat(ROUND(易速配签收量 / 易速配单量 * 100,2),'%')  AS 易速配总计签收,
			            concat(ROUND(易速配完成量 / 易速配单量 * 100,2),'%')  AS 易速配完成占比,
			            concat(ROUND(易速配单量 / 订单量 * 100,2),'%')  AS 易速配单量占比,
			            concat(ROUND(森鸿签收量 / 森鸿完成量 * 100,2),'%')  AS 森鸿完成签收,
			            concat(ROUND(森鸿签收量 / 森鸿单量 * 100,2),'%')  AS 森鸿总计签收,
			            concat(ROUND(森鸿完成量 / 森鸿单量 * 100,2),'%')  AS 森鸿完成占比,
			            concat(ROUND(森鸿单量 / 订单量 * 100,2),'%')  AS 森鸿单量占比
                FROM (SELECT IFNULL(s1.`币种`, '总计') AS 币种,
							IFNULL(s1.`团队`, '总计') AS 团队,
				            IFNULL(s1.`年月`, '总计') AS 年月,
				            IFNULL(s1.`产品id`, '总计') AS 产品id, 
							IFNULL(s1.`产品名称`, '总计') AS 产品名称,
							IFNULL(s1.`父级分类`, '总计') AS 父级分类,
				            SUM(订单量) 订单量, 
							品类总订单,
				            SUM(速派单量) 速派单量,  SUM(速派签收量) 速派签收量,  SUM(速派完成量) 速派完成量,
				            SUM(711单量) 711单量,  SUM(711签收量) 711签收量,  SUM(711完成量) 711完成量,
				            SUM(天马单量) 天马单量,  SUM(天马签收量) 天马签收量,  SUM(天马完成量) 天马完成量,
				            SUM(易速配单量) 易速配单量,  SUM(易速配签收量) 易速配签收量,  SUM(易速配完成量) 易速配完成量,
				            SUM(森鸿单量) 森鸿单量,  SUM(森鸿签收量) 森鸿签收量,  SUM(森鸿完成量) 森鸿完成量
                    FROM ( SELECT cx.`币种`, cx.`团队`, cx.`年月`, cx.`产品id`,  cx.`产品名称`,	cx.`父级分类`, 
								count(订单编号) AS 订单量,
								SUM(IF(cx.物流方式 = "台湾-速派-新竹" ,1,0)) AS 速派单量,
								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 = "已签收",1,0)) as 速派签收量,
								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 IN ( "已签收", "拒收", "已退货", "理赔", "自发头程丢件") ,1,0)) as  速派完成量,
								SUM(IF(cx.物流方式 = "台湾-速派-711超商" ,1,0)) AS '711单量',
								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 = "已签收",1,0)) as  '711签收量',
								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 IN ( "已签收", "拒收", "已退货", "理赔", "自发头程丢件") ,1,0)) as  '711完成量',
								SUM(IF(cx.物流方式 = "台湾-天马-新竹" ,1,0)) AS 天马单量,
								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 = "已签收",1,0)) as  天马签收量,
								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 IN ( "已签收", "拒收", "已退货", "理赔", "自发头程丢件") ,1,0)) as  天马完成量,
								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" ,1,0)) AS 易速配单量,
								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 = "已签收",1,0)) as  易速配签收量,
							    SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 IN ( "已签收", "拒收", "已退货", "理赔", "自发头程丢件") ,1,0)) as  易速配完成量,
								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" ,1,0)) AS 森鸿单量,
								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "已签收",1,0)) as  森鸿签收量,
								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 IN ( "已签收", "拒收", "已退货", "理赔", "自发头程丢件") ,1,0)) as  森鸿完成量
	                        FROM  (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') cx 
	                        WHERE cx.`是否改派` = '直发'
	                        GROUP BY cx.币种,cx.团队,cx.年月,cx.`产品id` 
	                    ) s1
                    LEFT JOIN 
						(SELECT cx.`币种`,  cx.`团队`, cx.`年月`,  cx.`父级分类`,  count(订单编号) AS 品类总订单 
	                    FROM  (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') cx 
	                    WHERE  cx.`是否改派` = '直发'
	                    GROUP BY	cx.币种,	cx.团队,	cx.年月,cx.`父级分类` 	
	                    ) s2 ON s1.`币种` = s2.`币种` AND s1.`团队` = s2.`团队` AND s1.`年月` = s2.`年月` AND s1.`父级分类` = s2.`父级分类` 	
                    GROUP BY	s1.币种,	s1.团队,	s1.年月,	s1.产品id
                    WITH ROLLUP 	
                ) s WHERE s.`币种` = '台湾'
                ORDER BY FIELD( s.`币种`, '台湾', '香港', '总计' ),
	                    FIELD( s.`团队`, '神龙家族-港澳台', '火凤凰-港澳台', '红杉家族-港澳台', '红杉家族-港澳台2','金狮-港澳台','总计'),
	                    FIELD(s.`年月`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'),'总计'),
	                    s.`订单量` DESC;'''.format(team, day_yesterday)
        df7 = pd.read_sql_query(sql=sql7, con=self.engine1)
        listT.append(df7)
        # 产品明细-香港
        sql8 = '''SELECT 币种,团队 家族,年月,产品id,产品名称,父级分类,订单量,
			            concat(ROUND(IF(SUBSTRING_INDEX(订单量 / 品类总订单,'.',1) > 1 ,1,订单量 / 品类总订单 ) * 100,2),'%')  AS 订单品类占比,
						concat(ROUND(立邦签收量 / 立邦完成量 * 100,2),'%')  AS 立邦完成签收,
						concat(ROUND(立邦签收量 / 立邦单量 * 100,2),'%')  AS 立邦总计签收,
						concat(ROUND(立邦完成量 / 立邦单量 * 100,2),'%')  AS 立邦完成占比,
						concat(ROUND(立邦单量 / 订单量 * 100,2),'%')  AS 立邦单量占比,
						concat(ROUND(森鸿SF签收量 / 森鸿SF完成量 * 100,2),'%')  AS 森鸿SF完成签收,
						concat(ROUND(森鸿SF签收量 / 森鸿SF单量 * 100,2),'%')  AS 森鸿SF总计签收,
						concat(ROUND(森鸿SF完成量 / 森鸿SF单量 * 100,2),'%')  AS 森鸿SF完成占比,
						concat(ROUND(森鸿SF单量 / 订单量 * 100,2),'%')  AS 森鸿SF单量占比,
					    concat(ROUND(森鸿SH签收量 / 森鸿SH完成量 * 100,2),'%')  AS 森鸿SH完成签收,
					    concat(ROUND(森鸿SH签收量 / 森鸿SH单量 * 100,2),'%')  AS 森鸿SH总计签收,
					    concat(ROUND(森鸿SH完成量 / 森鸿SH单量 * 100,2),'%')  AS 森鸿SH完成占比,
					    concat(ROUND(森鸿SH单量 / 订单量 * 100,2),'%')  AS 森鸿SH单量占比,
					    concat(ROUND(易速配SF签收量 / 易速配SF完成量 * 100,2),'%')  AS 易速配SF完成签收,
					    concat(ROUND(易速配SF签收量 / 易速配SF单量 * 100,2),'%')  AS 易速配SF总计签收,
					    concat(ROUND(易速配SF完成量 / 易速配SF单量 * 100,2),'%')  AS 易速配SF完成占比,
					    concat(ROUND(易速配SF单量 / 订单量 * 100,2),'%')  AS 易速配SF单量占比,
					    concat(ROUND(易速配YC签收量 / 易速配YC完成量 * 100,2),'%')  AS 易速配YC完成签收,
					    concat(ROUND(易速配YC签收量 / 易速配YC单量 * 100,2),'%')  AS 易速配YC总计签收,
					    concat(ROUND(易速配YC完成量 / 易速配YC单量 * 100,2),'%')  AS 易速配YC完成占比,
					    concat(ROUND(易速配YC单量 / 订单量 * 100,2),'%')  AS 易速配YC单量占比
                FROM (SELECT IFNULL(s1.`币种`, '总计') AS 币种,
							IFNULL(s1.`团队`, '总计') AS 团队,
				            IFNULL(s1.`年月`, '总计') AS 年月,
				            IFNULL(s1.`产品id`, '总计') AS 产品id, 
							IFNULL(s1.`产品名称`, '总计') AS 产品名称,
							IFNULL(s1.`父级分类`, '总计') AS 父级分类,
				            SUM(订单量) 订单量, 
							品类总订单,
							SUM(立邦单量) 立邦单量,  SUM(立邦签收量) 立邦签收量,  SUM(立邦完成量) 立邦完成量,
				            SUM(森鸿SF单量) 森鸿SF单量,  SUM(森鸿SF签收量) 森鸿SF签收量,  SUM(森鸿SF完成量) 森鸿SF完成量,
				            SUM(森鸿SH单量) 森鸿SH单量,  SUM(森鸿SH签收量) 森鸿SH签收量,  SUM(森鸿SH完成量) 森鸿SH完成量,					
				            SUM(易速配SF单量) 易速配SF单量,  SUM(易速配SF签收量) 易速配SF签收量,  SUM(易速配SF完成量) 易速配SF完成量,				
				            SUM(易速配YC单量) 易速配YC单量,  SUM(易速配YC签收量) 易速配YC签收量,  SUM(易速配YC完成量) 易速配YC完成量
                    FROM ( SELECT cx.`币种`, cx.`团队`, cx.`年月`, cx.`产品id`,  cx.`产品名称`,	cx.`父级分类`, 
								count(订单编号) AS 订单量,
								SUM(IF(cx.物流方式 = "香港-立邦-顺丰" ,1,0)) AS 立邦单量,
								SUM(IF(cx.物流方式 = "香港-立邦-顺丰" AND 最终状态 = "已签收",1,0)) as 立邦签收量,
								SUM(IF(cx.物流方式 = "香港-立邦-顺丰" AND 最终状态 IN ( "已签收", "拒收", "已退货", "理赔", "自发头程丢件") ,1,0)) as  立邦完成量,

								SUM(IF(cx.物流方式 = "香港-森鸿-顺丰渠道" ,1,0)) AS 森鸿SF单量,
								SUM(IF(cx.物流方式 = "香港-森鸿-顺丰渠道" AND 最终状态 = "已签收",1,0)) as  森鸿SF签收量,
								SUM(IF(cx.物流方式 = "香港-森鸿-顺丰渠道" AND 最终状态 IN ( "已签收", "拒收", "已退货", "理赔", "自发头程丢件") ,1,0)) as  森鸿SF完成量,

								SUM(IF(cx.物流方式 = "香港-森鸿-SH渠道" ,1,0)) AS 森鸿SH单量,
								SUM(IF(cx.物流方式 = "香港-森鸿-SH渠道" AND 最终状态 = "已签收",1,0)) as  森鸿SH签收量,
								SUM(IF(cx.物流方式 = "香港-森鸿-SH渠道" AND 最终状态 IN ( "已签收", "拒收", "已退货", "理赔", "自发头程丢件") ,1,0)) as  森鸿SH完成量,

								SUM(IF(cx.物流方式 = "香港-易速配-顺丰" ,1,0)) AS 易速配SF单量,
								SUM(IF(cx.物流方式 = "香港-易速配-顺丰" AND 最终状态 = "已签收",1,0)) as  易速配SF签收量,
							    SUM(IF(cx.物流方式 = "香港-易速配-顺丰" AND 最终状态 IN ( "已签收", "拒收", "已退货", "理赔", "自发头程丢件") ,1,0)) as  易速配SF完成量,

								SUM(IF(cx.物流方式 = "香港-易速配-顺丰YC" ,1,0)) AS 易速配YC单量,
								SUM(IF(cx.物流方式 = "香港-易速配-顺丰YC" AND 最终状态 = "已签收",1,0)) as  易速配YC签收量,
								SUM(IF(cx.物流方式 = "香港-易速配-顺丰YC" AND 最终状态 IN ( "已签收", "拒收", "已退货", "理赔", "自发头程丢件") ,1,0)) as  易速配YC完成量
	                        FROM  (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') cx 
	                        WHERE cx.`是否改派` = '直发'
	                        GROUP BY cx.币种,cx.团队,cx.年月,cx.`产品id` 
	                    ) s1
                    LEFT JOIN 
						(SELECT cx.`币种`,  cx.`团队`, cx.`年月`,  cx.`父级分类`,  count(订单编号) AS 品类总订单 
	                    FROM  (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') cx 
	                    WHERE  cx.`是否改派` = '直发'
	                    GROUP BY	cx.币种,	cx.团队,	cx.年月,cx.`父级分类` 	
	                    ) s2 ON s1.`币种` = s2.`币种` AND s1.`团队` = s2.`团队` AND s1.`年月` = s2.`年月` AND s1.`父级分类` = s2.`父级分类` 	
                    GROUP BY	s1.币种,	s1.团队,	s1.年月,	s1.产品id
                    WITH ROLLUP 	
                ) s WHERE s.`币种` = '香港'
                ORDER BY FIELD( s.`币种`, '台湾', '香港', '总计' ),
	                    FIELD( s.`团队`, '神龙家族-港澳台', '火凤凰-港澳台', '红杉家族-港澳台', '红杉家族-港澳台2','金狮-港澳台','总计'),
	                    FIELD(s.`年月`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'),'总计'),
	                    s.`订单量` DESC;'''.format(team, day_yesterday)
        df8 = pd.read_sql_query(sql=sql8, con=self.engine1)
        listT.append(df8)

        print('正在写入excel…………')
        today = datetime.date.today().strftime('%Y.%m.%d')
        for wbbook in ['神龙', '火凤凰', '红杉', '金狮']:
            file_path = 'D:\\Users\\Administrator\\Desktop\\输出文件\\{} {}-签收率.xlsx'.format(today, wbbook)
            sheet_name = ['每日', '总表', '总表上月', '物流', '物流上月', '品类', '品类上月', '产品', '产品明细台湾', '产品明细香港']
            df0 = pd.DataFrame([])  # 创建空的dataframe数据框
            df0.to_excel(file_path, index=False)  # 备用：可以向不同的sheet写入数据（创建新的工作表并进行写入）
            writer = pd.ExcelWriter(file_path, engine='openpyxl')  # 初始化写入对象
            book = load_workbook(file_path)  # 可以向不同的sheet写入数据（对现有工作表的追加）
            writer.book = book  # 将数据写入excel中的sheet2表,sheet_name改变后即是新增一个sheet
            for i in range(len(listT)):
                listT[i].to_excel(excel_writer=writer, sheet_name=sheet_name[i], index=False)
            if 'Sheet1' in book.sheetnames:  # 删除新建文档时的第一个工作表
                del book['Sheet1']
            writer.save()
            writer.close()
            # print('正在运行' + wbbook + '表宏…………')
            # app = xlwings.App(visible=False, add_book=False)  # 运行宏调整
            # app.display_alerts = False
            # wbsht = app.books.open('D:/Users/Administrator/Desktop/新版-格式转换(工具表).xlsm')
            # wbsht1 = app.books.open(file_path)
            # wbsht.macro('py_sl_总运行')()
            # wbsht1.save()
            # wbsht1.close()
            # wbsht.close()
            # app.quit()
        print('----已写入excel ')

if __name__ == '__main__':
    m = QueryUpdate()
    start: datetime = datetime.datetime.now()
    match1 = {'slgat': '神龙-港台',
              'slgat_hfh': '火凤凰-港台',
              'slgat_hs': '红杉-港台',
              'slgat_js': '金狮-港台',
              'gat': '港台'}
    team = 'gat'
    # -----------------------------------------------手动导入状态运行（一）-----------------------------------------
    write = '本期'
    # write = '上期'
    last_time = '2021-07-20'
    m.readFormHost(team, write, last_time)       #  更新签收表---港澳台（一）


    # m.makeSql(team)         # 转存表---港澳台（一.附表）

    # m.EportOrder(team)      #  导出需要更新的签收表---港澳台(二)


    # m.EportOrderBook(team)  #  导出总的签收表---港澳台(三)


    # m.qsb_report(team, '2021-06-26', '2021-05-26')

    m.qsb_new('gat')         #  获取报表
    print('耗时：', datetime.datetime.now() - start)