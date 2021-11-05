import pandas as pd
import os
import datetime
import xlwings

import requests
import json
import sys
from sso_updata import QueryTwo
from queue import Queue
from dateutil.relativedelta import relativedelta
from threading import Thread #  使用 threading 模块创建线程
import pandas.io.formats.excel

from sqlalchemy import create_engine
from settings import Settings
from emailControl import EmailControl
from openpyxl import load_workbook  # 可以向不同的sheet写入数据
from openpyxl.styles import Font, Border, Side, PatternFill, colors, Alignment  # 设置字体风格为Times New Roman，大小为16，粗体、斜体，颜色蓝色

from mysqlControl import MysqlControl
# -*- coding:utf-8 -*-
class QueryUpdate(Settings):
    def __init__(self):
        Settings.__init__(self)
        self.session = requests.session()  # 实例化session，维持会话,可以让我们在跨请求时保存某些参数
        self.q = Queue()  # 多线程调用的函数不能用return返回值，用来保存返回值
        self.engine1 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql1['user'],
                                                                                    self.mysql1['password'],
                                                                                    self.mysql1['host'],
                                                                                    self.mysql1['port'],
                                                                                    self.mysql1['datebase']))
        self.engine2 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql2['user'],
                                                                                    self.mysql2['password'],
                                                                                    self.mysql2['host'],
                                                                                    self.mysql2['port'],
                                                                                    self.mysql2['datebase']))
        self.engine20 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql20['user'],
                                                                                    self.mysql20['password'],
                                                                                    self.mysql20['host'],
                                                                                    self.mysql20['port'],
                                                                                    self.mysql20['datebase']))
        self.engine3 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql3['user'],
                                                                                    self.mysql3['password'],
                                                                                    self.mysql3['host'],
                                                                                    self.mysql3['port'],
                                                                                    self.mysql3['datebase']))
        self.e = EmailControl()
        self.m = MysqlControl()
        # self.sso = QueryTwo()
    def reSetEngine(self):
        self.engine1 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql1['user'],
                                                                                    self.mysql1['password'],
                                                                                    self.mysql1['host'],
                                                                                    self.mysql1['port'],
                                                                                    self.mysql1['datebase']))
        self.engine2 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql2['user'],
                                                                                    self.mysql2['password'],
                                                                                    self.mysql2['host'],
                                                                                    self.mysql2['port'],
                                                                                    self.mysql2['datebase']))

    # 获取签收表内容---港澳台更新签收总表
    def readFormHost(self):
        match = {'换货': '换货表',
                '退货': '退货表',
                '工单收集': '工单收集表'}
        path = r'H:\桌面\需要用到的文件\test'
        dirs = os.listdir(path=path)
        # ---读取execl文件---
        for dir in dirs:
            filePath = os.path.join(path, dir)
            if dir[:2] != '~$':
                print(filePath)
                if '换货' in dir:
                    wb_data = '换货表'
                elif '退货' in dir:
                    wb_data = '退货表'
                elif '工单' in dir:
                    wb_data = '工单收集表'
                else:
                    wb_data = None
                    pass
                self.wbsheetHost(filePath, wb_data)
        print('处理耗时：', datetime.datetime.now() - start)
    # 工作表的订单信息
    def wbsheetHost(self, filePath, wb_data):
        fileType = os.path.splitext(filePath)[1]
        app = xlwings.App(visible=False, add_book=False)
        app.display_alerts = False
        if 'xls' in fileType:
            wb = app.books.open(filePath, update_links=False, read_only=True)
            for sht in wb.sheets:
                try:
                    db = None
                    db = sht.used_range.options(pd.DataFrame, header=1, numbers=int, index=False).value
                    # print(db.columns)
                    if wb_data == '换货表':
                        db = db[['订单编号', '运单号', '物流渠道', '产品Id', '产品名称', '数量', '电话', '反馈方式', '金额', '克隆后金额', '是否上门取货', '反馈问题类型',
                                 '新订单编号', '新产品名称', '支付类型', '登记人', '导入时间', '处理人', '处理时间', '下单时间', '币种', '团队', '包裹到仓']]
                    elif wb_data == '退货表':
                        db = db[['订单编号', '运单号', '物流渠道', '产品Id', '产品名称', '数量', '电话', '反馈方式', '金额', '是否上门取货', '反馈问题类型',
                                 '退款金额', '支付类型', '登记人', '导入时间', '处理人', '处理时间', '下单时间', '币种','团队', '包裹到仓', '站点ID']]
                    elif wb_data == '工单收集表':
                        db = db[['订单编号', '产品Id', '产品名称', '问题类型', '环节问题', '订单金额', '订单状态', '运单号', '物流状态', '签收时间', '所属团队',
                                 '提交形式', '提交时间', '同步模块', '模块进展', '登记人', '币种', '数量']]
                except Exception as e:
                    print('xxxx查看失败：' + sht.name, str(Exception) + str(e))
                if db is not None and len(db) > 0:
                    print('++++正在导入查询：' + sht.name + '表； 共：' + str(len(db)) + '行', 'sheet共：' + str(sht.used_range.last_cell.row) + '行')
                    db.to_sql(wb_data, con=self.engine1, index=False, if_exists='replace')
                    print('++++导入成功：' + sht.name + '--->>>到查询缓存表')
                else:
                    print('----------数据为空导入失败：' + sht.name)
            wb.close()
        app.quit()

    # 写入更新缓存表
    def writeSql(self):
        data_now = (datetime.datetime.now() - relativedelta(months=5)).strftime('%Y%m')
        # data_now = datetime.datetime.now().strftime('%Y%m')
        print(data_now)
        listT = []  # 查询sql的结果 存放池
        print('正在获取 第一部分 信息…………')
        print('正在获取 产品前十（分币种&分家族） 信息…………')
        # sql = '''SELECT 年月,币种, 团队,产品id,产品名称,COUNT(订单编号) as 单量,SUM(数量) as 商品数量
        #         FROM d1_gat d
        #         WHERE d.`年月` ='{0}'
        #         GROUP BY d.`年月`, d.`产品id`
        #         ORDER BY 单量 DESC
        #         LIMIT 10;'''.format(data_now)
        sql = '''(SELECT 年月,币种, null 团队,产品id,产品名称,COUNT(订单编号) as 单量,SUM(数量) as 商品数量
                    FROM d1_gat d
                    WHERE d.`年月` ='{0}' and d.`币种` = '台湾'
                    GROUP BY d.`年月`,d.`币种`, d.`产品id`
                    ORDER BY 单量 DESC
                    LIMIT 10 ) 
                UNION all 
                (SELECT 年月,币种, 团队,产品id,产品名称,COUNT(订单编号) as 单量,SUM(数量) as 商品数量
                    FROM d1_gat d
                    WHERE d.`年月` ='{0}' and d.`币种` = '台湾' and d.`团队` LIKE '神龙%'
                    GROUP BY d.`年月`,d.`币种`,d.`团队`, d.`产品id`
                    ORDER BY 单量 DESC
                    LIMIT 10 ) 
                UNION all 
                (SELECT 年月,币种, 团队,产品id,产品名称,COUNT(订单编号) as 单量,SUM(数量) as 商品数量
                    FROM d1_gat d
                    WHERE d.`年月` ='{0}' and d.`币种` = '台湾' and d.`团队` LIKE '火凤凰%'
                    GROUP BY d.`年月`,d.`币种`,d.`团队`, d.`产品id`
                    ORDER BY 单量 DESC
                    LIMIT 10 ) 
                UNION all 
                (SELECT 年月,币种, 团队,产品id,产品名称,COUNT(订单编号) as 单量,SUM(数量) as 商品数量
                    FROM d1_gat d
                    WHERE d.`年月` ='{0}' and d.`币种` = '台湾' and d.`团队` LIKE '金鹏%'
                    GROUP BY d.`年月`,d.`币种`,d.`团队`, d.`产品id`
                    ORDER BY 单量 DESC
                    LIMIT 10) 
                UNION all 
                (SELECT 年月,币种, 团队,产品id,产品名称,COUNT(订单编号) as 单量,SUM(数量) as 商品数量
                    FROM d1_gat d
                    WHERE d.`年月` ='{0}' and d.`币种` = '台湾' and d.`团队` LIKE '金狮%'
                    GROUP BY d.`年月`,d.`币种`,d.`团队`, d.`产品id`
                    ORDER BY 单量 DESC
                    LIMIT 10) 
                UNION all 
                (SELECT 年月,币种, null 团队,产品id,产品名称,COUNT(订单编号) as 单量,SUM(数量) as 商品数量
                    FROM d1_gat d
                    WHERE d.`年月` ='{0}' and d.`币种` = '香港'
                    GROUP BY d.`年月`,d.`币种`, d.`产品id`
                    ORDER BY 单量 DESC
                    LIMIT 10) 
                UNION all 
                (SELECT 年月,币种, 团队,产品id,产品名称,COUNT(订单编号) as 单量,SUM(数量) as 商品数量
                    FROM d1_gat d
                    WHERE d.`年月` ='{0}' and d.`币种` = '香港' and d.`团队` LIKE '神龙%'
                    GROUP BY d.`年月`,d.`币种`,d.`团队`, d.`产品id`
                    ORDER BY 单量 DESC
                    LIMIT 10) 
                UNION all 
                (SELECT 年月,币种, 团队,产品id,产品名称,COUNT(订单编号) as 单量,SUM(数量) as 商品数量
                    FROM d1_gat d
                    WHERE d.`年月` ='{0}' and d.`币种` = '香港' and d.`团队` LIKE '火凤凰%'
                    GROUP BY d.`年月`,d.`币种`,d.`团队`, d.`产品id`
                    ORDER BY 单量 DESC
                    LIMIT 10) 
                UNION all 
                (SELECT 年月,币种, 团队,产品id,产品名称,COUNT(订单编号) as 单量,SUM(数量) as 商品数量
                    FROM d1_gat d
                    WHERE d.`年月` ='{0}' and d.`币种` = '香港' and d.`团队` LIKE '金狮%'
                    GROUP BY d.`年月`,d.`币种`,d.`团队`, d.`产品id`
                    ORDER BY 单量 DESC
                    LIMIT 10);'''.format(data_now)
        df = pd.read_sql_query(sql=sql, con=self.engine1)
        df.to_sql('product_info', con=self.engine1, index=False, if_exists='replace')

        print('正在获取 上月产品前十（分币种） 信息…………')
        sql1 = '''SELECT ss.年月, ss.币种, ss.团队, ss.产品id, ss.产品名称, 
                        IF(ss.单量 = 0,NULL,ss.单量) as 单量,
                        IF(ss.商品数量 = 0,NULL,ss.商品数量) as 商品数量,
                        IF(换货单量 = 0,NULL,换货单量) as 换货单量,
                        IF(换货数量 = 0,NULL,换货数量) as 换货数量,
                        IF(退货单量 = 0,NULL,退货单量) as 退货单量,
                        IF(退货数量 = 0,NULL,退货数量) as 退货数量,
                        IF(SUM(换货单量 + 退货单量) = 0,NULL,SUM(换货单量 + 退货单量)) as 已处理量,
                        IF(工单单量 = 0,NULL,工单单量) as 工单单量,
                        IF(工单数量 = 0,NULL,工单数量) as 工单数量,
                        concat(ROUND(IF(SUM(换货单量 + 退货单量) = 0,NULL,SUM(换货单量 + 退货单量))  / 工单单量 * 100,2),'%') as 处理占比,
                        IF(下错订单 = 0,NULL,下错订单) 下错订单,concat(ROUND(IF(下错订单 = 0,NULL,下错订单) / SUM(换货单量 + 退货单量) * 100,2),'%') as 占比,
                        IF(重复订单 = 0,NULL,重复订单) 重复订单,concat(ROUND(IF(重复订单 = 0,NULL,重复订单) / SUM(换货单量 + 退货单量) * 100,2),'%') as 占比,
                        IF(尺寸不合 = 0,NULL,尺寸不合) 尺寸不合,concat(ROUND(IF(尺寸不合 = 0,NULL,尺寸不合) / SUM(换货单量 + 退货单量) * 100,2),'%') as 占比,
                        IF(尺码偏大 = 0,NULL,尺码偏大) 尺码偏大,concat(ROUND(IF(尺码偏大 = 0,NULL,尺码偏大) / SUM(换货单量 + 退货单量) * 100,2),'%') as 占比,
                        IF(尺码偏小 = 0,NULL,尺码偏小) 尺码偏小,concat(ROUND(IF(尺码偏小 = 0,NULL,尺码偏小) / SUM(换货单量 + 退货单量) * 100,2),'%') as 占比,
                        IF(价格较高 = 0,NULL,价格较高) 价格较高,concat(ROUND(IF(价格较高 = 0,NULL,价格较高) / SUM(换货单量 + 退货单量) * 100,2),'%') as 占比,
                        IF(产品质量不合格 = 0,NULL,产品质量不合格) 产品质量不合格,concat(ROUND(IF(产品质量不合格 = 0,NULL,产品质量不合格) / SUM(换货单量 + 退货单量) * 100,2),'%') as 占比,
                        IF(产品瑕疵 = 0,NULL,产品瑕疵) 产品瑕疵,concat(ROUND(IF(产品瑕疵 = 0,NULL,产品瑕疵) / SUM(换货单量 + 退货单量) * 100,2),'%') as 占比,
                        IF(漏发错发 = 0,NULL,漏发错发) 漏发错发,concat(ROUND(IF(漏发错发 = 0,NULL,漏发错发) / SUM(换货单量 + 退货单量) * 100,2),'%') as 占比,
                        IF(产品不符合客户预期 = 0,NULL,产品不符合客户预期) 产品不符合客户预期,concat(ROUND(IF(产品不符合客户预期 = 0,NULL,产品不符合客户预期) / SUM(换货单量 + 退货单量) * 100,2),'%') as 占比,
                        IF(与网站不符 = 0,NULL,与网站不符) 与网站不符,concat(ROUND(IF(与网站不符 = 0,NULL,与网站不符) / SUM(换货单量 + 退货单量) * 100,2),'%') as 占比,
                        IF(中国制造 = 0,NULL,中国制造) 中国制造,concat(ROUND(IF(中国制造 = 0,NULL,中国制造) / SUM(换货单量 + 退货单量) * 100,2),'%') as 占比,
                        IF(非正品拒收 = 0,NULL,非正品拒收) 非正品拒收,concat(ROUND(IF(非正品拒收 = 0,NULL,非正品拒收) / SUM(换货单量 + 退货单量) * 100,2),'%') as 占比,
                        IF(产品到货无法使用 = 0,NULL,产品到货无法使用) 产品到货无法使用,concat(ROUND(IF(产品到货无法使用 = 0,NULL,产品到货无法使用) / SUM(换货单量 + 退货单量) * 100,2),'%') as 占比,
                        IF(客户不会使用该产品 = 0,NULL,客户不会使用该产品) 客户不会使用该产品,concat(ROUND(IF(客户不会使用该产品 = 0,NULL,客户不会使用该产品) / SUM(换货单量 + 退货单量) * 100,2),'%') as 占比,
                        IF(客户自身原因 = 0,NULL,客户自身原因) 客户自身原因,concat(ROUND(IF(客户自身原因 = 0,NULL,客户自身原因) / SUM(换货单量 + 退货单量) * 100,2),'%') as 占比,
                        IF(没有产品说明书 = 0,NULL,没有产品说明书) 没有产品说明书,concat(ROUND(IF(没有产品说明书 = 0,NULL,没有产品说明书) / SUM(换货单量 + 退货单量) * 100,2),'%') as 占比,
                        IF(不喜欢 = 0,NULL,不喜欢) 不喜欢,concat(ROUND(IF(不喜欢 = 0,NULL,不喜欢) / SUM(换货单量 + 退货单量) * 100,2),'%') as 占比,
                        IF(无订购 = 0,NULL,无订购) 无订购,concat(ROUND(IF(无订购 = 0,NULL,无订购) / SUM(换货单量 + 退货单量) * 100,2),'%') as 占比,
                        IF(无理由拒收退货 = 0,NULL,无理由拒收退货) 无理由拒收退货,concat(ROUND(IF(无理由拒收退货 = 0,NULL,无理由拒收退货) / SUM(换货单量 + 退货单量) * 100,2),'%') as 占比,
                        IF(已在其他地方购买 = 0,NULL,已在其他地方购买) 已在其他地方购买,concat(ROUND(IF(已在其他地方购买 = 0,NULL,已在其他地方购买) / SUM(换货单量 + 退货单量) * 100,2),'%') as 占比,
                        IF(其他 = 0,NULL,其他) 其他,concat(ROUND(IF(其他 = 0,NULL,其他) / SUM(换货单量 + 退货单量) * 100,2),'%') as 占比
                FROM product_info ss
                LEFT JOIN
                (SELECT DATE_FORMAT(导入时间,'%Y%m') as 年月,币种,'合计' 团队,产品id,COUNT(订单编号) 换货单量,SUM(数量) as 换货数量
                    FROM 换货表 th
                    GROUP BY DATE_FORMAT(导入时间,'%Y%m'),币种, 产品id
                    ORDER BY DATE_FORMAT(导入时间,'%Y%m') DESC,币种, 换货单量 DESC
                    ) cx ON ss.年月 = cx.年月 AND ss.币种 = cx.币种 AND ss.团队 = cx.团队 AND ss.产品id = cx.产品id
                LEFT JOIN
                (SELECT DATE_FORMAT(导入时间,'%Y%m') as 年月,币种,'合计' 团队,产品id,COUNT(订单编号) 退货单量,SUM(数量) as 退货数量
                    FROM 退货表 th
                    GROUP BY DATE_FORMAT(导入时间,'%Y%m'),币种, 产品id
                    ORDER BY DATE_FORMAT(导入时间,'%Y%m') DESC,币种, 退货单量 DESC
                ) cx2 ON ss.年月 = cx2.年月 AND ss.币种 = cx2.币种 AND ss.团队 = cx2.团队 AND ss.产品id = cx2.产品id
                LEFT JOIN
                (SELECT DATE_FORMAT(提交时间,'%Y%m') as 年月,币种,'合计' 所属团队,产品id,COUNT(订单编号) 工单单量,SUM(数量) as 工单数量,
                        SUM(IF(`问题类型` = '下错订单',1,0)) AS 下错订单,
                        SUM(IF(`问题类型` = '重复订单',1,0)) AS 重复订单,
                        SUM(IF(`问题类型` = '尺寸不合',1,0)) AS 尺寸不合,
                        SUM(IF(`问题类型` = '尺码偏大',1,0)) AS 尺码偏大,
                        SUM(IF(`问题类型` = '尺码偏小',1,0)) AS 尺码偏小,
                        SUM(IF(`问题类型` = '价格较高',1,0)) AS 价格较高,
                        SUM(IF(`问题类型` = '产品质量不合格',1,0)) AS 产品质量不合格,
                        SUM(IF(`问题类型` = '产品瑕疵',1,0)) AS 产品瑕疵,
                        SUM(IF(`问题类型` = '漏发错发',1,0)) AS 漏发错发,
                        SUM(IF(`问题类型` = '产品不符合客户预期',1,0)) AS 产品不符合客户预期,
                        SUM(IF(`问题类型` = '与网站不符',1,0)) AS 与网站不符,
                        SUM(IF(`问题类型` = '中国制造',1,0)) AS 中国制造,
                        SUM(IF(`问题类型` = '非正品拒收',1,0)) AS 非正品拒收,
                        SUM(IF(`问题类型` = '产品到货无法使用',1,0)) AS 产品到货无法使用,
                        SUM(IF(`问题类型` = '客户不会使用该产品',1,0)) AS 客户不会使用该产品,
                        SUM(IF(`问题类型` = '客户自身原因',1,0)) AS 客户自身原因,
                        SUM(IF(`问题类型` = '没有产品说明书',1,0)) AS 没有产品说明书,
                        SUM(IF(`问题类型` = '不喜欢',1,0)) AS 不喜欢,
                        SUM(IF(`问题类型` = '无订购',1,0)) AS 无订购,
                        SUM(IF(`问题类型` = '无理由拒收退货',1,0)) AS 无理由拒收退货,
                        SUM(IF(`问题类型` = '已在其他地方购买',1,0)) AS 已在其他地方购买,
                        SUM(IF(`问题类型` = '其他',1,0)) AS 其他
                    FROM 工单收集表 th
                        GROUP BY DATE_FORMAT(提交时间,'%Y%m'),币种, 产品id
                        ORDER BY DATE_FORMAT(提交时间,'%Y%m') DESC,币种, 工单单量 DESC
                ) cx3 ON ss.年月 = cx3.年月 AND ss.币种 = cx3.币种 AND ss.团队 = cx3.所属团队 AND ss.产品id = cx3.产品id
                WHERE ss.团队 = '合计'
                GROUP BY ss.年月,ss.币种,ss.团队,ss.产品id;'''
        df1 = pd.read_sql_query(sql=sql1, con=self.engine1)
        listT.append(df1)
        print('正在获取 上月产品前十（分币种-团队） 信息…………')
        sql2 = '''SELECT ss.年月, ss.币种, ss.团队, ss.产品id, ss.产品名称, 
                        IF(ss.单量 = 0,NULL,ss.单量) as 单量,
                        IF(ss.商品数量 = 0,NULL,ss.商品数量) as 商品数量,
                        IF(换货单量 = 0,NULL,换货单量) as 换货单量,
                        IF(换货数量 = 0,NULL,换货数量) as 换货数量,
                        IF(退货单量 = 0,NULL,退货单量) as 退货单量,
                        IF(退货数量 = 0,NULL,退货数量) as 退货数量,
                        IF(SUM(换货单量 + 退货单量) = 0,NULL,SUM(换货单量 + 退货单量)) as 已处理量,
                        IF(工单单量 = 0,NULL,工单单量) as 工单单量,
                        IF(工单数量 = 0,NULL,工单数量) as 工单数量,
                        concat(ROUND(IF(SUM(换货单量 + 退货单量) = 0,NULL,SUM(换货单量 + 退货单量))  / 工单单量 * 100,2),'%') as 处理占比,
                        IF(下错订单 = 0,NULL,下错订单) 下错订单,concat(ROUND(IF(下错订单 = 0,NULL,下错订单) / SUM(换货单量 + 退货单量) * 100,2),'%') as 占比,
                        IF(重复订单 = 0,NULL,重复订单) 重复订单,concat(ROUND(IF(重复订单 = 0,NULL,重复订单) / SUM(换货单量 + 退货单量) * 100,2),'%') as 占比,
                        IF(尺寸不合 = 0,NULL,尺寸不合) 尺寸不合,concat(ROUND(IF(尺寸不合 = 0,NULL,尺寸不合) / SUM(换货单量 + 退货单量) * 100,2),'%') as 占比,
                        IF(尺码偏大 = 0,NULL,尺码偏大) 尺码偏大,concat(ROUND(IF(尺码偏大 = 0,NULL,尺码偏大) / SUM(换货单量 + 退货单量) * 100,2),'%') as 占比,
                        IF(尺码偏小 = 0,NULL,尺码偏小) 尺码偏小,concat(ROUND(IF(尺码偏小 = 0,NULL,尺码偏小) / SUM(换货单量 + 退货单量) * 100,2),'%') as 占比,
                        IF(价格较高 = 0,NULL,价格较高) 价格较高,concat(ROUND(IF(价格较高 = 0,NULL,价格较高) / SUM(换货单量 + 退货单量) * 100,2),'%') as 占比,
                        IF(产品质量不合格 = 0,NULL,产品质量不合格) 产品质量不合格,concat(ROUND(IF(产品质量不合格 = 0,NULL,产品质量不合格) / SUM(换货单量 + 退货单量) * 100,2),'%') as 占比,
                        IF(产品瑕疵 = 0,NULL,产品瑕疵) 产品瑕疵,concat(ROUND(IF(产品瑕疵 = 0,NULL,产品瑕疵) / SUM(换货单量 + 退货单量) * 100,2),'%') as 占比,
                        IF(漏发错发 = 0,NULL,漏发错发) 漏发错发,concat(ROUND(IF(漏发错发 = 0,NULL,漏发错发) / SUM(换货单量 + 退货单量) * 100,2),'%') as 占比,
                        IF(产品不符合客户预期 = 0,NULL,产品不符合客户预期) 产品不符合客户预期,concat(ROUND(IF(产品不符合客户预期 = 0,NULL,产品不符合客户预期) / SUM(换货单量 + 退货单量) * 100,2),'%') as 占比,
                        IF(与网站不符 = 0,NULL,与网站不符) 与网站不符,concat(ROUND(IF(与网站不符 = 0,NULL,与网站不符) / SUM(换货单量 + 退货单量) * 100,2),'%') as 占比,
                        IF(中国制造 = 0,NULL,中国制造) 中国制造,concat(ROUND(IF(中国制造 = 0,NULL,中国制造) / SUM(换货单量 + 退货单量) * 100,2),'%') as 占比,
                        IF(非正品拒收 = 0,NULL,非正品拒收) 非正品拒收,concat(ROUND(IF(非正品拒收 = 0,NULL,非正品拒收) / SUM(换货单量 + 退货单量) * 100,2),'%') as 占比,
                        IF(产品到货无法使用 = 0,NULL,产品到货无法使用) 产品到货无法使用,concat(ROUND(IF(产品到货无法使用 = 0,NULL,产品到货无法使用) / SUM(换货单量 + 退货单量) * 100,2),'%') as 占比,
                        IF(客户不会使用该产品 = 0,NULL,客户不会使用该产品) 客户不会使用该产品,concat(ROUND(IF(客户不会使用该产品 = 0,NULL,客户不会使用该产品) / SUM(换货单量 + 退货单量) * 100,2),'%') as 占比,
                        IF(客户自身原因 = 0,NULL,客户自身原因) 客户自身原因,concat(ROUND(IF(客户自身原因 = 0,NULL,客户自身原因) / SUM(换货单量 + 退货单量) * 100,2),'%') as 占比,
                        IF(没有产品说明书 = 0,NULL,没有产品说明书) 没有产品说明书,concat(ROUND(IF(没有产品说明书 = 0,NULL,没有产品说明书) / SUM(换货单量 + 退货单量) * 100,2),'%') as 占比,
                        IF(不喜欢 = 0,NULL,不喜欢) 不喜欢,concat(ROUND(IF(不喜欢 = 0,NULL,不喜欢) / SUM(换货单量 + 退货单量) * 100,2),'%') as 占比,
                        IF(无订购 = 0,NULL,无订购) 无订购,concat(ROUND(IF(无订购 = 0,NULL,无订购) / SUM(换货单量 + 退货单量) * 100,2),'%') as 占比,
                        IF(无理由拒收退货 = 0,NULL,无理由拒收退货) 无理由拒收退货,concat(ROUND(IF(无理由拒收退货 = 0,NULL,无理由拒收退货) / SUM(换货单量 + 退货单量) * 100,2),'%') as 占比,
                        IF(已在其他地方购买 = 0,NULL,已在其他地方购买) 已在其他地方购买,concat(ROUND(IF(已在其他地方购买 = 0,NULL,已在其他地方购买) / SUM(换货单量 + 退货单量) * 100,2),'%') as 占比,
                        IF(其他 = 0,NULL,其他) 其他,concat(ROUND(IF(其他 = 0,NULL,其他) / SUM(换货单量 + 退货单量) * 100,2),'%') as 占比
                FROM product_info ss
                LEFT JOIN
                (SELECT DATE_FORMAT(导入时间,'%Y%m') as 年月,币种,团队,产品id,COUNT(订单编号) 换货单量,SUM(数量) as 换货数量
                    FROM 换货表 th
                    GROUP BY DATE_FORMAT(导入时间,'%Y%m'),币种, 团队, 产品id
                    ORDER BY DATE_FORMAT(导入时间,'%Y%m') DESC,币种, 团队 , 换货单量 DESC
                    ) cx ON ss.年月 = cx.年月 AND ss.币种 = cx.币种 AND ss.团队 = cx.团队 AND ss.产品id = cx.产品id
                LEFT JOIN
                (SELECT DATE_FORMAT(导入时间,'%Y%m') as 年月,币种,团队,产品id,COUNT(订单编号) 退货单量,SUM(数量) as 退货数量
                    FROM 退货表 th
                    GROUP BY DATE_FORMAT(导入时间,'%Y%m'),币种, 团队, 产品id
                    ORDER BY DATE_FORMAT(导入时间,'%Y%m') DESC,币种, 团队 , 退货单量 DESC
                ) cx2 ON ss.年月 = cx2.年月 AND ss.币种 = cx2.币种 AND ss.团队 = cx2.团队 AND ss.产品id = cx2.产品id
                LEFT JOIN
                (SELECT DATE_FORMAT(提交时间,'%Y%m') as 年月,币种,所属团队,产品id,COUNT(订单编号) 工单单量,SUM(数量) as 工单数量,
                        SUM(IF(`问题类型` = '下错订单',1,0)) AS 下错订单,
                        SUM(IF(`问题类型` = '重复订单',1,0)) AS 重复订单,
                        SUM(IF(`问题类型` = '尺寸不合',1,0)) AS 尺寸不合,
                        SUM(IF(`问题类型` = '尺码偏大',1,0)) AS 尺码偏大,
                        SUM(IF(`问题类型` = '尺码偏小',1,0)) AS 尺码偏小,
                        SUM(IF(`问题类型` = '价格较高',1,0)) AS 价格较高,
                        SUM(IF(`问题类型` = '产品质量不合格',1,0)) AS 产品质量不合格,
                        SUM(IF(`问题类型` = '产品瑕疵',1,0)) AS 产品瑕疵,
                        SUM(IF(`问题类型` = '漏发错发',1,0)) AS 漏发错发,
                        SUM(IF(`问题类型` = '产品不符合客户预期',1,0)) AS 产品不符合客户预期,
                        SUM(IF(`问题类型` = '与网站不符',1,0)) AS 与网站不符,
                        SUM(IF(`问题类型` = '中国制造',1,0)) AS 中国制造,
                        SUM(IF(`问题类型` = '非正品拒收',1,0)) AS 非正品拒收,
                        SUM(IF(`问题类型` = '产品到货无法使用',1,0)) AS 产品到货无法使用,
                        SUM(IF(`问题类型` = '客户不会使用该产品',1,0)) AS 客户不会使用该产品,
                        SUM(IF(`问题类型` = '客户自身原因',1,0)) AS 客户自身原因,
                        SUM(IF(`问题类型` = '没有产品说明书',1,0)) AS 没有产品说明书,
                        SUM(IF(`问题类型` = '不喜欢',1,0)) AS 不喜欢,
                        SUM(IF(`问题类型` = '无订购',1,0)) AS 无订购,
                        SUM(IF(`问题类型` = '无理由拒收退货',1,0)) AS 无理由拒收退货,
                        SUM(IF(`问题类型` = '已在其他地方购买',1,0)) AS 已在其他地方购买,
                        SUM(IF(`问题类型` = '其他',1,0)) AS 其他
                    FROM 工单收集表 th
                        GROUP BY DATE_FORMAT(提交时间,'%Y%m'),币种, 所属团队, 产品id
                        ORDER BY DATE_FORMAT(提交时间,'%Y%m') DESC,币种, 所属团队 , 工单单量 DESC
                ) cx3 ON ss.年月 = cx3.年月 AND ss.币种 = cx3.币种 AND ss.团队 = cx3.所属团队 AND ss.产品id = cx3.产品id
                WHERE ss.团队 != '合计'
                GROUP BY ss.年月,ss.币种,ss.团队,ss.产品id;'''
        df2= pd.read_sql_query(sql=sql2, con=self.engine1)
        listT.append(df2)

        print('正在获取 第二部分 信息…………')
        print('正在获取 单量 信息…………')
        sql5 = '''SELECT s1.*,s2.`0%单量`,s2.`<10%单量`,s2.`<20%单量`, s2.`<30%单量`, s2.`<40%单量`, s2.`<50%单量`, s2.`>=50%单量`,s2.不全款单量,s2.退货单量,
                        s1.`换货单量` + s2.退货单量 as 退换补单量,工单单量
                FROM (SELECT DATE_FORMAT(导入时间,'%Y%m') as 年月,币种,团队,
                            SUM(IF(`占比` = '0%',1,0)) AS '0%单量',
                            SUM(IF(`占比` = '<10%',1,0)) AS '<10%单量',
                            SUM(IF(`占比` = '<20%',1,0)) AS '<20%单量',
                            SUM(IF(`占比` = '<30%',1,0)) AS '<30%单量',
                            SUM(IF(`占比` = '<40%',1,0)) AS '<40%单量',
                            SUM(IF(`占比` = '<50%',1,0)) AS '<50%单量',
                            SUM(IF(`占比` = '>=50%',1,0)) AS '>=50%单量',
                            SUM(IF(`占比` != '0%',1,0)) AS '非换补单量',
                            COUNT(订单编号) 换货单量
                        FROM ( SELECT *,IF(克隆后金额/金额 = 0,'0%',
                                        IF(克隆后金额/金额 > 0 AND 克隆后金额/金额 <= 0.1,'<10%',
                                        IF(克隆后金额/金额 > 0.1 AND 克隆后金额/金额 <= 0.2,'<20%',
                                        IF(克隆后金额/金额 > 0.2 AND 克隆后金额/金额 <= 0.3,'<30%',
                                        IF(克隆后金额/金额 > 0.3 AND 克隆后金额/金额 <= 0.4,'<40%',
                                        IF(克隆后金额/金额 > 0.4 AND 克隆后金额/金额 <= 0.5,'<50%',
                                        IF(克隆后金额/金额 > 0.5,'>=50%',克隆后金额/金额))))))) as 占比
                                FROM 换货表
                            ) th
                        GROUP BY DATE_FORMAT(导入时间,'%Y%m'),币种, 团队
                ) s1
                LEFT JOIN
                (SELECT DATE_FORMAT(导入时间,'%Y%m') as 年月,币种,团队,
                        SUM(IF(`占比` = '0%',1,0)) AS '0%单量',
                        SUM(IF(`占比` = '<10%',1,0)) AS '<10%单量',
                        SUM(IF(`占比` = '<20%',1,0)) AS '<20%单量',
                        SUM(IF(`占比` = '<30%',1,0)) AS '<30%单量',
                        SUM(IF(`占比` = '<40%',1,0)) AS '<40%单量',
                        SUM(IF(`占比` = '<50%',1,0)) AS '<50%单量',
                        SUM(IF(`占比` = '>=50%',1,0)) AS '>=50%单量',
                        SUM(IF(`占比` != '0%',1,0)) AS '不全款单量',
                        COUNT(订单编号) 退货单量
                FROM ( SELECT *,IF(退款金额/金额 = 0,'0%',
                                IF(退款金额/金额 > 0 AND 退款金额/金额 <= 0.1,'<10%',
                                IF(退款金额/金额 > 0.1 AND 退款金额/金额 <= 0.2,'<20%',
                                IF(退款金额/金额 > 0.2 AND 退款金额/金额 <= 0.3,'<30%',
                                IF(退款金额/金额 > 0.3 AND 退款金额/金额 <= 0.4,'<40%',
                                IF(退款金额/金额 > 0.4 AND 退款金额/金额 <= 0.5,'<50%',
                                IF(退款金额/金额 > 0.5,'>=50%',退款金额/金额))))))) as 占比
                        FROM 退货表
                    ) th
                GROUP BY DATE_FORMAT(导入时间,'%Y%m'),币种, 团队
                ) s2 ON s1.年月 = s2.年月 AND s1.币种 = s2.币种 AND s1.团队 = s2.团队
                LEFT JOIN
                ( SELECT DATE_FORMAT(提交时间,'%Y%m') as 年月,币种,所属团队,COUNT(订单编号) 工单单量
                    FROM 工单收集表
                    GROUP BY DATE_FORMAT(提交时间,'%Y%m'),币种, 所属团队
                ) s3 ON s1.年月 = s3.年月 AND s1.币种 = s3.币种 AND s1.团队 = s3.所属团队;'''
        df5 = pd.read_sql_query(sql=sql5, con=self.engine1)
        listT.append(df5)
        print('正在获取 克隆金额 信息…………')
        sql6 = '''SELECT s1.*,s2.`0%单量`,s2.`<10%单量`,s2.`<20%单量`, s2.`<30%单量`, s2.`<40%单量`, s2.`<50%单量`, s2.`>=50%单量`,s2.不全款单量,s2.退货单量,
                        s1.`换货单量` + s2.退货单量 as 退换补单量,工单单量
                FROM (SELECT DATE_FORMAT(导入时间,'%Y%m') as 年月,币种,团队,
                            SUM(IF(`占比` = '0%',克隆后金额,0)) AS '0%单量',
                            SUM(IF(`占比` = '<10%',克隆后金额,0)) AS '<10%单量',
                            SUM(IF(`占比` = '<20%',克隆后金额,0)) AS '<20%单量',
                            SUM(IF(`占比` = '<30%',克隆后金额,0)) AS '<30%单量',
                            SUM(IF(`占比` = '<40%',克隆后金额,0)) AS '<40%单量',
                            SUM(IF(`占比` = '<50%',克隆后金额,0)) AS '<50%单量',
                            SUM(IF(`占比` = '>=50%',克隆后金额,0)) AS '>=50%单量',
                            SUM(IF(`占比` != '0%',克隆后金额,0)) AS '非换补单量',
                            SUM(订单编号) 换货单量
                        FROM ( SELECT *,IF(克隆后金额/金额 = 0,'0%',
                                        IF(克隆后金额/金额 > 0 AND 克隆后金额/金额 <= 0.1,'<10%',
                                        IF(克隆后金额/金额 > 0.1 AND 克隆后金额/金额 <= 0.2,'<20%',
                                        IF(克隆后金额/金额 > 0.2 AND 克隆后金额/金额 <= 0.3,'<30%',
                                        IF(克隆后金额/金额 > 0.3 AND 克隆后金额/金额 <= 0.4,'<40%',
                                        IF(克隆后金额/金额 > 0.4 AND 克隆后金额/金额 <= 0.5,'<50%',
                                        IF(克隆后金额/金额 > 0.5,'>=50%',克隆后金额/金额))))))) as 占比
                                FROM 换货表
                            ) th
                        GROUP BY DATE_FORMAT(导入时间,'%Y%m'),币种, 团队
                ) s1
                LEFT JOIN
                (SELECT DATE_FORMAT(导入时间,'%Y%m') as 年月,币种,团队,
                        SUM(IF(`占比` = '0%',退款金额,0)) AS '0%单量',
                        SUM(IF(`占比` = '<10%',退款金额,0)) AS '<10%单量',
                        SUM(IF(`占比` = '<20%',退款金额,0)) AS '<20%单量',
                        SUM(IF(`占比` = '<30%',退款金额,0)) AS '<30%单量',
                        SUM(IF(`占比` = '<40%',退款金额,0)) AS '<40%单量',
                        SUM(IF(`占比` = '<50%',退款金额,0)) AS '<50%单量',
                        SUM(IF(`占比` = '>=50%',退款金额,0)) AS '>=50%单量',
                        SUM(IF(`占比` != '0%',退款金额,0)) AS '不全款单量',
                        SUM(订单编号) 退货单量
                FROM ( SELECT *,IF(退款金额/金额 = 0,'0%',
                                IF(退款金额/金额 > 0 AND 退款金额/金额 <= 0.1,'<10%',
                                IF(退款金额/金额 > 0.1 AND 退款金额/金额 <= 0.2,'<20%',
                                IF(退款金额/金额 > 0.2 AND 退款金额/金额 <= 0.3,'<30%',
                                IF(退款金额/金额 > 0.3 AND 退款金额/金额 <= 0.4,'<40%',
                                IF(退款金额/金额 > 0.4 AND 退款金额/金额 <= 0.5,'<50%',
                                IF(退款金额/金额 > 0.5,'>=50%',退款金额/金额))))))) as 占比
                        FROM 退货表
                    ) th
                GROUP BY DATE_FORMAT(导入时间,'%Y%m'),币种, 团队
                ) s2 ON s1.年月 = s2.年月 AND s1.币种 = s2.币种 AND s1.团队 = s2.团队
                LEFT JOIN
                ( SELECT DATE_FORMAT(提交时间,'%Y%m') as 年月,币种,所属团队,COUNT(订单编号) 工单单量
                    FROM 工单收集表
                    GROUP BY DATE_FORMAT(提交时间,'%Y%m'),币种, 所属团队
                ) s3 ON s1.年月 = s3.年月 AND s1.币种 = s3.币种 AND s1.团队 = s3.所属团队;'''
        df6= pd.read_sql_query(sql=sql6, con=self.engine1)
        listT.append(df6)
        print('正在获取 金额 信息…………')
        sql7 = '''SELECT s1.*,s2.`0%单量`,s2.`<10%单量`,s2.`<20%单量`, s2.`<30%单量`, s2.`<40%单量`, s2.`<50%单量`, s2.`>=50%单量`,s2.不全款单量,s2.退货单量,
                        s1.`换货单量` + s2.退货单量 as 退换补单量,工单单量
                FROM (SELECT DATE_FORMAT(导入时间,'%Y%m') as 年月,币种,团队,
                            SUM(IF(`占比` = '0%',金额,0)) AS '0%单量',
                            SUM(IF(`占比` = '<10%',金额,0)) AS '<10%单量',
                            SUM(IF(`占比` = '<20%',金额,0)) AS '<20%单量',
                            SUM(IF(`占比` = '<30%',金额,0)) AS '<30%单量',
                            SUM(IF(`占比` = '<40%',金额,0)) AS '<40%单量',
                            SUM(IF(`占比` = '<50%',金额,0)) AS '<50%单量',
                            SUM(IF(`占比` = '>=50%',金额,0)) AS '>=50%单量',
                            SUM(IF(`占比` != '0%',金额,0)) AS '非换补单量',
                            SUM(订单编号) 换货单量
                        FROM ( SELECT *,IF(克隆后金额/金额 = 0,'0%',
                                        IF(克隆后金额/金额 > 0 AND 克隆后金额/金额 <= 0.1,'<10%',
                                        IF(克隆后金额/金额 > 0.1 AND 克隆后金额/金额 <= 0.2,'<20%',
                                        IF(克隆后金额/金额 > 0.2 AND 克隆后金额/金额 <= 0.3,'<30%',
                                        IF(克隆后金额/金额 > 0.3 AND 克隆后金额/金额 <= 0.4,'<40%',
                                        IF(克隆后金额/金额 > 0.4 AND 克隆后金额/金额 <= 0.5,'<50%',
                                        IF(克隆后金额/金额 > 0.5,'>=50%',克隆后金额/金额))))))) as 占比
                                FROM 换货表
                            ) th
                        GROUP BY DATE_FORMAT(导入时间,'%Y%m'),币种, 团队
                ) s1
                LEFT JOIN
                (SELECT DATE_FORMAT(导入时间,'%Y%m') as 年月,币种,团队,
                        SUM(IF(`占比` = '0%',金额,0)) AS '0%单量',
                        SUM(IF(`占比` = '<10%',金额,0)) AS '<10%单量',
                        SUM(IF(`占比` = '<20%',金额,0)) AS '<20%单量',
                        SUM(IF(`占比` = '<30%',金额,0)) AS '<30%单量',
                        SUM(IF(`占比` = '<40%',金额,0)) AS '<40%单量',
                        SUM(IF(`占比` = '<50%',金额,0)) AS '<50%单量',
                        SUM(IF(`占比` = '>=50%',金额,0)) AS '>=50%单量',
                        SUM(IF(`占比` != '0%',金额,0)) AS '不全款单量',
                        SUM(金额) 退货单量
                FROM ( SELECT *,IF(退款金额/金额 = 0,'0%',
                                IF(退款金额/金额 > 0 AND 退款金额/金额 <= 0.1,'<10%',
                                IF(退款金额/金额 > 0.1 AND 退款金额/金额 <= 0.2,'<20%',
                                IF(退款金额/金额 > 0.2 AND 退款金额/金额 <= 0.3,'<30%',
                                IF(退款金额/金额 > 0.3 AND 退款金额/金额 <= 0.4,'<40%',
                                IF(退款金额/金额 > 0.4 AND 退款金额/金额 <= 0.5,'<50%',
                                IF(退款金额/金额 > 0.5,'>=50%',退款金额/金额))))))) as 占比
                        FROM 退货表
                    ) th
                GROUP BY DATE_FORMAT(导入时间,'%Y%m'),币种, 团队
                ) s2 ON s1.年月 = s2.年月 AND s1.币种 = s2.币种 AND s1.团队 = s2.团队
                LEFT JOIN
                ( SELECT DATE_FORMAT(提交时间,'%Y%m') as 年月,币种,所属团队,COUNT(订单编号) 工单单量
                    FROM 工单收集表
                    GROUP BY DATE_FORMAT(提交时间,'%Y%m'),币种, 所属团队
                ) s3 ON s1.年月 = s3.年月 AND s1.币种 = s3.币种 AND s1.团队 = s3.所属团队;'''
        df7= pd.read_sql_query(sql=sql7, con=self.engine1)
        listT.append(df7)

        print('正在写入excel…………')
        today = datetime.date.today().strftime('%Y.%m.%d')
        file_path = 'H:\桌面\需要用到的文件\\{} 客服处理汇总.xlsx'.format(today)
        sheet_name = ['分币种', '分币种分团队', '退换补单量', '退换补返回金额', '退换补金额']
        df0 = pd.DataFrame([])                                      # 创建空的dataframe数据框
        df0.to_excel(file_path, index=False)                        # 备用：可以向不同的sheet写入数据（创建新的工作表并进行写入）
        writer = pd.ExcelWriter(file_path, engine='openpyxl')       # 初始化写入对象
        book = load_workbook(file_path)                             # 可以向不同的sheet写入数据（对现有工作表的追加）
        writer.book = book                                          # 将数据写入excel中的sheet2表,sheet_name改变后即是新增一个sheet
        for i in range(len(listT)):
            listT[i].to_excel(excel_writer=writer, sheet_name=sheet_name[i], index=False)
        if 'Sheet1' in book.sheetnames:                             # 删除新建文档时的第一个工作表
            del book['Sheet1']
        writer.save()
        writer.close()
        # try:
        #     print('正在运行表宏…………')
        #     app = xlwings.App(visible=False, add_book=False)  # 运行宏调整
        #     app.display_alerts = False
        #     wbsht = app.books.open('D:/Users/Administrator/Desktop/新版-格式转换(工具表).xlsm')
        #     wbsht1 = app.books.open(file_path)
        #     wbsht.macro('zl_report_day')()
        #     wbsht1.save()
        #     wbsht1.close()
        #     wbsht.close()
        #     app.quit()
        # except Exception as e:
        #     print('运行失败：', str(Exception) + str(e))
        print('----已写入excel ')


if __name__ == '__main__':
    m = QueryUpdate()
    start: datetime = datetime.datetime.now()
    # -----------------------------------------------手动查询状态运行（一）-----------------------------------------
    # m.readFormHost()
    m.writeSql()

    print('输出耗时：', datetime.datetime.now() - start)

