import pandas as pd
from sqlalchemy import create_engine
from settings import Settings
from queryControl import QueryControl
from emailControl import EmailControl
from bpsControl import BpsControl
from sltemMonitoring import SltemMonitoring

from openpyxl import load_workbook  # 可以向不同的sheet写入数据
from tkinter import messagebox
import os
import zipfile
import xlwings as xl
import xlwings
import datetime
from dateutil.relativedelta import relativedelta
import time


class MysqlControl(Settings):
    def __init__(self):
        Settings.__init__(self)
        self.engine1 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql1['user'],
                                                                                    self.mysql1['password'],
                                                                                    self.mysql1['host'],
                                                                                    self.mysql1['port'],
                                                                                    self.mysql1['datebase']))
        self.engine2 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql2['user'],
                                                                                    self.mysql2['password'],
                                                                                    self.mysql2['host'],
                                                                                    self.mysql2['port'],
                                                                                    self.mysql2['datebase']))
        self.engine20 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql20['user'],
                                                                                    self.mysql20['password'],
                                                                                    self.mysql20['host'],
                                                                                    self.mysql20['port'],
                                                                                    self.mysql20['datebase']))
        self.engine3 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql3['user'],
                                                                                    self.mysql3['password'],
                                                                                    self.mysql3['host'],
                                                                                    self.mysql3['port'],
                                                                                    self.mysql3['datebase']))
        self.e = EmailControl()
        self.d = QueryControl()

    def reSetEngine(self):
        self.engine1 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql1['user'],
                                                                                    self.mysql1['password'],
                                                                                    self.mysql1['host'],
                                                                                    self.mysql1['port'],
                                                                                    self.mysql1['datebase']))
        self.engine2 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql2['user'],
                                                                                    self.mysql2['password'],
                                                                                    self.mysql2['host'],
                                                                                    self.mysql2['port'],
                                                                                    self.mysql2['datebase']))

    def writeSqlReplace(self, dataFrame):
        dataFrame.to_sql('tem', con=self.engine1, index=False, if_exists='replace')

    def replaceInto(self, team, dfColumns):
        columns = list(dfColumns)
        columns = ', '.join(columns)
        if team == 'slrb':
            print(team + '---9')
            sql = 'REPLACE INTO {}({}, 添加时间) SELECT *, NOW() 添加时间 FROM tem; '.format(team, columns)
        elif team == 'slgat':
            print(team + '---909')  # 当天和前天的添加时间比较，判断是否一样数据
            sql = 'INSERT IGNORE INTO {}({}, 添加时间, 更新时间) SELECT *, CURDATE() 添加时间, NOW() 更新时间 FROM tem; '.format(team,
                                                                                                                 columns)
        else:
            print(team)
            sql = 'INSERT IGNORE INTO {}({}, 添加时间) SELECT *, NOW() 添加时间 FROM tem; '.format(team, columns)
            # sql = 'INSERT IGNORE INTO {}_copy({}, 添加时间) SELECT *, NOW() 添加时间 FROM tem; '.format(team, columns)
        try:
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=100)
        except Exception as e:
            print('插入失败：', str(Exception) + str(e))

    def readSql(self, sql):
        db = pd.read_sql_query(sql=sql, con=self.engine1)
        # db = pd.read_sql(sql=sql, con=self.engine1) or team == 'slgat'
        return db

    # 备用查询函数--开始（停用）
    def creatMyOrder(self, team):
        match = {'slgat': '"神龙家族-港澳台"',
                 'sltg': '"神龙家族-泰国"',
                 'slxmt': '"神龙家族-新加坡", "神龙家族-马来西亚"',
                 'slzb': '"神龙家族-直播团队"',
                 'slyn': '"神龙家族-越南"',
                 'slrb': '"神龙家族-日本团队"'}
        # 12-1月的
        if team == 'sltg' or team == 'slxmt' or team == 'slrb' or team == 'slgat':
            yy = int((datetime.datetime.now().replace(day=1) - datetime.timedelta(days=1)).strftime('%Y'))
            mm = int((datetime.datetime.now().replace(day=1) - datetime.timedelta(days=1)).strftime('%m'))
            begin = datetime.date(yy, mm, 1)
            print(begin)
            yy2 = int(datetime.datetime.now().strftime('%Y'))
            mm2 = int(datetime.datetime.now().strftime('%m'))
            dd2 = int(datetime.datetime.now().strftime('%d'))
            end = datetime.date(yy2, mm2, dd2)
            print(end)
        else:
            # 11-12月的
            begin = datetime.date(2020, 11, 1)
            print(begin)
            end = datetime.date(2020, 12, 1)
            print(end)
        for i in range((end - begin).days):  # 按天循环获取订单状态
            day = begin + datetime.timedelta(days=i)
            # print(str(day))
            yesterday = str(day) + ' 23:59:59'
            last_month = str(day)
            sql = '''SELECT a.id, 
                            a.month 年月, 
                            a.month_mid 旬, 
                            a.rq 日期, 
                            dim_area.name 团队, 
                            a.region_code 区域,
                            dim_currency_lang.pname 币种,
                            a.beform 订单来源, 
                            a.order_number 订单编号, 
                            a.ship_phone 电话号码, 
                            UPPER(a.waybill_number) 运单编号,
                            a.order_status 系统订单状态id, 
                            a.logistics_status 系统物流状态id, 
                            IF(a.second=0,'直发','改派') 是否改派, 
                            dim_trans_way.all_name 物流方式,
                            dim_trans_way.simple_name 物流名称,
                            dim_trans_way.remark 运输方式,
                            a.logistics_type 货物类型,
                            IF(a.low_price=0,'否','是') 是否低价, 
                            a.product_id 产品id, 
                            gk_product.name 产品名称, 
                            dim_cate.ppname 父级分类,
                            dim_cate.pname 二级分类,
                            dim_cate.name 三级分类,
                            dim_payment.pay_name 付款方式, 
                            a.amount 价格, 
                            a.addtime 下单时间, 
                            a.verity_time 审核时间, 
                            a.delivery_time 仓储扫描时间, 
                            a.finish_status 完结状态, 
                            a.endtime 完结状态时间, 
                            a.salesRMB 价格RMB, 
                            intervals.intervals 价格区间,
                            a.purchase 成本价, 
                            a.logistics_cost 物流花费, 
                            a.package_cost 打包花费, 
                            a.other_fee 其它花费, 
                            a.weight 包裹重量,
                            a.volume 包裹体积,
                            a.ship_zip 邮编,
                            a.turn_purchase_time 添加物流单号时间,
                            a.del_reason 订单删除原因
                    FROM gk_order a 
                            left join dim_area ON dim_area.id = a.area_id 
                            left join dim_payment on dim_payment.id = a.payment_id
                            left join gk_product on gk_product.id = a.product_id 
                            left join dim_trans_way on dim_trans_way.id = a.logistics_id
                            left join dim_cate on dim_cate.id = gk_product.third_cate_id 
                            left join intervals on intervals.id = a.intervals
                            left join dim_currency_lang on dim_currency_lang.id = a.currency_lang_id
                    WHERE a.rq = '{}' AND a.rq <= '{}'
                        AND dim_area.name IN ({});'''.format(last_month, yesterday, match[team])
            print('正在获取 ' + match[team] + last_month[5:7] + '-' + yesterday[8:10] + ' 号订单…………')
            df = pd.read_sql_query(sql=sql, con=self.engine2)
            # print('----已获取 ' + yesterday[8:10] + ' 号订单…………')
            sql = 'SELECT * FROM dim_order_status;'
            df1 = pd.read_sql_query(sql=sql, con=self.engine1)
            # print('----已获取订单状态')
            print('+++合并订单状态中…………')
            df = pd.merge(left=df, right=df1, left_on='系统订单状态id', right_on='id', how='left')
            # print('----已合并订单状态')
            sql = 'SELECT * FROM dim_logistics_status;'
            df1 = pd.read_sql_query(sql=sql, con=self.engine1)
            # print('----已获取物流状态')
            print('+++合并物流状态中…………')
            df = pd.merge(left=df, right=df1, left_on='系统物流状态id', right_on='id', how='left')
            df = df.drop(labels=['id', 'id_y', '系统订单状态id', '系统物流状态id'], axis=1)
            df.rename(columns={'id_x': 'id', 'name_x': '系统订单状态', 'name_y': '系统物流状态'}, inplace=True)
            # print('----已获取 ' + yesterday[8:10] + ' 号订单与物流状态')
            # self.reSetEngine()
            print('++++++正在将 ' + yesterday[8:10] + ' 号订单写入数据库++++++')
            # 这一句会报错,需要修改my.ini文件中的[mysqld]段中的"max_allowed_packet = 1024M"
            try:
                df.to_sql('sl_order', con=self.engine1, index=False, if_exists='replace')
                sql = 'REPLACE INTO {}_order_list SELECT * FROM sl_order; '.format(team)
                pd.read_sql_query(sql=sql, con=self.engine1, chunksize=1000)
                # df.to_sql(team + '_order_list_copy', con=self.engine1, index=False, if_exists='replace')
            except Exception as e:
                print('插入失败：', str(Exception) + str(e))
            # print(df)
        return '写入完成'
    # 备用查询函数--结束（停用）

    def update_gk_product(self):        # 更新产品id的列表
        yy = int((datetime.datetime.now() - datetime.timedelta(days=5)).strftime('%Y'))
        mm = int((datetime.datetime.now() - datetime.timedelta(days=5)).strftime('%m'))
        dd = int((datetime.datetime.now() - datetime.timedelta(days=5)).strftime('%d'))
        begin = datetime.date(yy, mm, dd)
        # begin = datetime.date(2018, 1, 1)
        print(begin)
        yy2 = int(datetime.datetime.now().strftime('%Y'))
        mm2 = int(datetime.datetime.now().strftime('%m'))
        dd2 = int(datetime.datetime.now().strftime('%d'))
        end = datetime.date(yy2, mm2, dd2)
        # end = datetime.date(2019, 1, 1)
        print(end)
        for i in range((end - begin).days):  # 按天循环获取订单状态
            day = begin + datetime.timedelta(days=i)
            month_last = str(day)
            sql = '''SELECT * FROM  gk_product WHERE gk_product.rq >= '{0}';'''.format(month_last)
            print('正在获取 ' + month_last + ' 号以后的产品详情…………')
            df = pd.read_sql_query(sql=sql, con=self.engine20)
            print('正在写入产品缓存中…………')
            df.to_sql('tem_product', con=self.engine1, index=False, if_exists='replace')
            try:
                print('正在更新中…………')
                sql = 'REPLACE INTO dim_product SELECT *, NOW() 更新时间 FROM tem_product; '
                pd.read_sql_query(sql=sql, con=self.engine1, chunksize=1000)
            except Exception as e:
                print('插入失败：', str(Exception) + str(e))
            print('更新完成…………')

    def creatMyOrderSl(self, team):  # 最近五天的全部订单信息
        match = {'slgat': '"神龙家族-港澳台"',
                 'sltg': '"神龙家族-泰国"',
                 'slxmt': '"神龙家族-新加坡", "神龙家族-马来西亚", "神龙家族-菲律宾"',
                 'slzb': '"神龙家族-直播团队"',
                 'slyn': '"神龙家族-越南"',
                 'slrb': '"神龙家族-日本团队"'}
        # 12-1月的
        if team == 'sltg' or team == 'slxmt' or team == 'slrb' or team == 'slgat':
            # 获取日期时间
            sql = 'SELECT 日期 FROM {0}_order_list WHERE id = (SELECT MAX(id) FROM {0}_order_list);'.format(team)
            rq = pd.read_sql_query(sql=sql, con=self.engine1)
            rq = pd.to_datetime(rq['日期'][0])
            yy = int((rq - datetime.timedelta(days=3)).strftime('%Y'))
            mm = int((rq - datetime.timedelta(days=3)).strftime('%m'))
            dd = int((rq - datetime.timedelta(days=3)).strftime('%d'))
            print(dd)
            # yy = int(datetime.datetime.now().strftime('%Y'))
            # mm = int(datetime.datetime.now().strftime('%m'))
            # # dd = int((datetime.datetime.now() - datetime.timedelta(days=5)).strftime('%d'))
            begin = datetime.date(yy, mm, dd)
            print(begin)
            yy2 = int(datetime.datetime.now().strftime('%Y'))
            mm2 = int(datetime.datetime.now().strftime('%m'))
            dd2 = int(datetime.datetime.now().strftime('%d'))
            end = datetime.date(yy2, mm2, dd2)
            print(end)
        else:
            # 11-12月的
            begin = datetime.date(2021, 1, 4)
            print(begin)
            end = datetime.date(2021, 1, 5)
            print(end)
        for i in range((end - begin).days):  # 按天循环获取订单状态
            day = begin + datetime.timedelta(days=i)
            # print(str(day))
            yesterday = str(day) + ' 23:59:59'
            last_month = str(day)
            # sql = '''SELECT a.id,
            #                 a.month 年月,
            #                 a.month_mid 旬,
            #                 a.rq 日期,
            #                 dim_area.name 团队,
            #                 a.region_code 区域,
            #                 dim_currency_lang.pname 币种,
            #                 a.beform 订单来源,
            #                 a.order_number 订单编号,
            #                 a.qty 数量,
            #                 a.ship_phone 电话号码,
            #                 UPPER(a.waybill_number) 运单编号,
            #                 a.order_status 系统订单状态id,
            #                 a.logistics_status 系统物流状态id,
            #                 IF(a.second=0,'直发','改派') 是否改派,
            #                 dim_trans_way.all_name 物流方式,
            #                 dim_trans_way.simple_name 物流名称,
            #                 dim_trans_way.remark 运输方式,
            #                 a.logistics_type 货物类型,
            #                 IF(a.low_price=0,'否','是') 是否低价,
            #                 a.product_id 产品id,
            #                 gs.product_name 产品名称,
            #                 dim_cate.ppname 父级分类,
            #                 dim_cate.pname 二级分类,
            #                 dim_cate.name 三级分类,
            #                 dim_payment.pay_name 付款方式,
            #                 a.amount 价格,
            #                 a.addtime 下单时间,
            #                 a.verity_time 审核时间,
            #                 a.delivery_time 仓储扫描时间,
            #                 a.finish_status 完结状态,
            #                 a.endtime 完结状态时间,
            #                 a.salesRMB 价格RMB,
            #                 intervals.intervals 价格区间,
            #                 null 成本价,
            #                 a.logistics_cost 物流花费,
            #                 null 打包花费,
            #                 a.other_fee 其它花费,
            #                 a.weight 包裹重量,
            #                 a.volume 包裹体积,
            #                 a.ship_zip 邮编,
            #                 a.turn_purchase_time 添加物流单号时间,
            #                 a.del_reason 订单删除原因
            #         FROM gk_order_kf a
            #                 left join dim_area ON dim_area.id = a.area_id
            #                 left join dim_payment ON dim_payment.id = a.payment_id
            #                 left join (SELECT * FROM gk_sale WHERE id IN (SELECT MAX(id) FROM gk_sale GROUP BY product_id ) ORDER BY id) gs ON gs.product_id = a.product_id
            #                 left join dim_trans_way ON dim_trans_way.id = a.logistics_id
            #                 left join dim_cate ON dim_cate.id = a.third_cate_id
            #                 left join intervals ON intervals.id = a.intervals
            #                 left join dim_currency_lang ON dim_currency_lang.id = a.currency_lang_id
            #         WHERE a.rq = '{0}' AND a.rq <= '{1}'
            #             AND dim_area.name IN ({2});'''.format(last_month, yesterday, match[team])
            # print('正在获取 ' + match[team] + last_month[5:7] + '-' + yesterday[8:10] + ' 号订单…………')
            # df = pd.read_sql_query(sql=sql, con=self.engine2)

            sql = '''SELECT a.id,
                            a.month 年月,
                            a.month_mid 旬,
                            a.rq 日期,
                            sl_dim_area.name 团队,
                            a.region_code 区域,
                            sl_dim_currency_lang.pname 币种,
                            a.beform 订单来源,
                            a.order_number 订单编号,
                            a.qty 数量,
                            a.ship_phone 电话号码,
                            UPPER(a.waybill_number) 运单编号,
                            a.order_status 系统订单状态id,
                            a.logistics_status 系统物流状态id,
                            IF(a.second=0,'直发','改派') 是否改派,
                            sl_dim_trans_way.all_name 物流方式,
                            sl_dim_trans_way.simple_name 物流名称,
                            sl_dim_trans_way.remark 运输方式,
                            a.logistics_type 货物类型,
                            IF(a.low_price=0,'否','是') 是否低价,
                            a.product_id 产品id,
                            gk_product.name 产品名称,
                            sl_dim_cate.ppname 父级分类,
                            sl_dim_cate.pname 二级分类,
                            sl_dim_cate.name 三级分类,
                            sl_dim_payment.pay_name 付款方式,
                            a.amount 价格,
                            a.addtime 下单时间,
                            a.verity_time 审核时间,
                            a.delivery_time 仓储扫描时间,
                            a.finish_status 完结状态,
                            a.endtime 完结状态时间,
                            a.salesRMB 价格RMB,
                            intervals.intervals 价格区间,
                            a.purchase 成本价,
                            a.logistics_cost 物流花费,
                            a.package_cost 打包花费,
                            a.other_fee 其它花费,
                            a.weight 包裹重量,
                            a.volume 包裹体积,
                            a.ship_zip 邮编,
                            a.turn_purchase_time 添加物流单号时间,
                            a.del_reason 订单删除原因
                    FROM gk_order a
                            left join sl_dim_area ON sl_dim_area.id = a.area_id
                            left join sl_dim_payment on sl_dim_payment.id = a.payment_id
                            left join gk_product on gk_product.id = a.product_id
                            left join sl_dim_trans_way on sl_dim_trans_way.id = a.logistics_id
                            left join sl_dim_cate on sl_dim_cate.id = gk_product.third_cate_id
                            left join intervals on intervals.id = a.intervals
                            left join sl_dim_currency_lang on sl_dim_currency_lang.id = a.currency_lang_id
                    WHERE a.rq = '{0}' AND a.rq <= '{1}'
                        AND sl_dim_area.name IN ({2});'''.format(last_month, yesterday, match[team])  # 备用
            print('正在获取 ' + match[team] + last_month[5:7] + '-' + yesterday[8:10] + ' 号订单…………')
            df = pd.read_sql_query(sql=sql, con=self.engine20)

            sql = 'SELECT * FROM dim_order_status;'
            df1 = pd.read_sql_query(sql=sql, con=self.engine1)
            print('+++合并订单状态中…………')
            df = pd.merge(left=df, right=df1, left_on='系统订单状态id', right_on='id', how='left')
            sql = 'SELECT * FROM dim_logistics_status;'
            df1 = pd.read_sql_query(sql=sql, con=self.engine1)
            print('+++合并物流状态中…………')
            df = pd.merge(left=df, right=df1, left_on='系统物流状态id', right_on='id', how='left')
            df = df.drop(labels=['id', 'id_y', '系统订单状态id', '系统物流状态id'], axis=1)
            df.rename(columns={'id_x': 'id', 'name_x': '系统订单状态', 'name_y': '系统物流状态'}, inplace=True)
            print('++++++正在将 ' + yesterday[8:10] + ' 号订单写入数据库++++++')
            # 这一句会报错,需要修改my.ini文件中的[mysqld]段中的"max_allowed_packet = 1024M"
            try:
                df.to_sql('sl_order', con=self.engine1, index=False, if_exists='replace')
                sql = 'REPLACE INTO {}_order_list SELECT *, NOW() 记录时间 FROM sl_order; '.format(team)
                pd.read_sql_query(sql=sql, con=self.engine1, chunksize=1000)
            except Exception as e:
                print('插入失败：', str(Exception) + str(e))
            print('写入完成…………')
        return '写入完成'

    def creatMyOrderSlTWO(self, team):  # 最近两个月的更新订单信息
        match = {'slgat': '"神龙家族-港澳台"',
                 'sltg': '"神龙家族-泰国"',
                 'slxmt': '"神龙家族-新加坡", "神龙家族-马来西亚", "神龙家族-菲律宾"',
                 'slzb': '"神龙家族-直播团队"',
                 'slyn': '"神龙家族-越南"',
                 'slrb': '"神龙家族-日本团队"'}
        today = datetime.date.today().strftime('%Y.%m.%d')
        if team == 'sltg' or team == 'slxmt' or team == 'slrb0' or team == 'slgat0':
            yy = int((datetime.datetime.now().replace(day=1) - datetime.timedelta(days=1)).strftime('%Y'))
            mm = int((datetime.datetime.now().replace(day=1) - datetime.timedelta(days=1)).strftime('%m'))
            begin = datetime.date(yy, mm, 1)
            print(begin)
            yy2 = int(datetime.datetime.now().strftime('%Y'))
            mm2 = int(datetime.datetime.now().strftime('%m'))
            dd2 = int(datetime.datetime.now().strftime('%d'))
            end = datetime.date(yy2, mm2, dd2)
            print(end)
        else:
            begin = datetime.date(2021, 1, 1)
            print(begin)
            end = datetime.date(2021, 3, 7)
            print(end)
        for i in range((end - begin).days):  # 按天循环获取订单状态
            day = begin + datetime.timedelta(days=i)
            # print(str(day))
            yesterday = str(day) + ' 23:59:59'
            last_month = str(day)
            # sql = '''SELECT DISTINCT a.id,
            #                 a.rq 日期,
            #                 dim_currency_lang.pname 币种,
            #                 a.order_number 订单编号,
            #                 a.qty 数量,
            #                 a.ship_phone 电话号码,
            #                 UPPER(a.waybill_number) 运单编号,
            #                 a.order_status 系统订单状态id,
            #                 a.logistics_status 系统物流状态id,
            #                 IF(a.second=0,'直发','改派') 是否改派,
            #                 dim_trans_way.all_name 物流方式,
            #                 dim_trans_way.simple_name 物流名称,
            #                 a.logistics_type 货物类型,
            #                 a.verity_time 审核时间,
            #                 a.delivery_time 仓储扫描时间,
            #                 a.endtime 完结状态时间
            #         FROM gk_order_kf a
            #                 left join dim_area ON dim_area.id = a.area_id
            #                 left join dim_payment on dim_payment.id = a.payment_id
            #                 left join (SELECT * FROM gk_sale WHERE id IN (SELECT MAX(id) FROM gk_sale GROUP BY product_id ) ORDER BY id) gs ON gs.product_id = a.product_id
            #                 left join dim_trans_way on dim_trans_way.id = a.logistics_id
            #                 left join dim_cate on dim_cate.id = a.third_cate_id
            #                 left join intervals on intervals.id = a.intervals
            #                 left join dim_currency_lang on dim_currency_lang.id = a.currency_lang_id
            #         WHERE a.rq = '{}' AND a.rq <= '{}'
            #             AND dim_area.name IN ({});'''.format(last_month, yesterday, match[team])
            # print('正在更新 ' + match[team] + last_month[5:7] + '-' + yesterday[8:10] + ' 号订单信息…………')
            # df = pd.read_sql_query(sql=sql, con=self.engine2)

            sql = '''SELECT a.id,
                            a.rq 日期,
                            sl_dim_currency_lang.pname 币种,
                            a.order_number 订单编号,
                            a.qty 数量,
                            a.ship_phone 电话号码,
                            UPPER(a.waybill_number) 运单编号,
                            a.order_status 系统订单状态id,
                            a.logistics_status 系统物流状态id,
                            IF(a.second=0,'直发','改派') 是否改派,
                            sl_dim_trans_way.all_name 物流方式,
                            sl_dim_trans_way.simple_name 物流名称,
                            a.logistics_type 货物类型,
                            a.verity_time 审核时间,
                            a.delivery_time 仓储扫描时间,
                            a.endtime 完结状态时间
                    FROM gk_order a
                            left join sl_dim_area ON sl_dim_area.id = a.area_id
                            left join sl_dim_payment on sl_dim_payment.id = a.payment_id
                            left join gk_product on gk_product.id = a.product_id
                            left join sl_dim_trans_way on sl_dim_trans_way.id = a.logistics_id
                            left join sl_dim_cate on sl_dim_cate.id = gk_product.third_cate_id
                            left join intervals on intervals.id = a.intervals
                            left join sl_dim_currency_lang on sl_dim_currency_lang.id = a.currency_lang_id
                    WHERE a.rq = '{}' AND a.rq <= '{}'
                        AND sl_dim_area.name IN ({});'''.format(last_month, yesterday, match[team])     # 备用
            print('正在更新 ' + match[team] + last_month[5:7] + '-' + yesterday[8:10] + ' 号订单信息…………')
            df = pd.read_sql_query(sql=sql, con=self.engine20)

            sql = 'SELECT * FROM dim_order_status;'
            df1 = pd.read_sql_query(sql=sql, con=self.engine1)
            print('++++更新订单状态中…………')
            df = pd.merge(left=df, right=df1, left_on='系统订单状态id', right_on='id', how='left')
            sql = 'SELECT * FROM dim_logistics_status;'
            df1 = pd.read_sql_query(sql=sql, con=self.engine1)
            print('++++更新物流状态中…………')
            df = pd.merge(left=df, right=df1, left_on='系统物流状态id', right_on='id', how='left')
            df = df.drop(labels=['id', 'id_y', '系统订单状态id', '系统物流状态id'], axis=1)
            df.rename(columns={'id_x': 'id', 'name_x': '系统订单状态', 'name_y': '系统物流状态'}, inplace=True)
            print('+++++++正在将 ' + yesterday[8:10] + ' 号订单更新到数据库++++++')
            # 这一句会报错,需要修改my.ini文件中的[mysqld]段中的"max_allowed_packet = 1024M"
            df.to_sql('sl_order2', con=self.engine1, index=False, if_exists='replace')
            try:
                sql = '''update {0}_order_list a, sl_order2 b
                        set a.`币种`=b.`币种`,
                            a.`数量`=b.`数量`,
		                    a.`电话号码`=b.`电话号码` ,
		                    a.`运单编号`=b.`运单编号`,
		                    a.`系统订单状态`=b.`系统订单状态`,
		                    a.`系统物流状态`=b.`系统物流状态`,
		                    a.`是否改派`=b.`是否改派`,
		                    a.`物流方式`=b.`物流方式`,
		                    a.`物流名称`=b.`物流名称`,
		                    a.`货物类型`=b.`货物类型`,
		                    a.`审核时间`=b.`审核时间`,
		                    a.`仓储扫描时间`=b.`仓储扫描时间`,
		                    a.`完结状态时间`=b.`完结状态时间`
		                where a.`订单编号`=b.`订单编号`;'''.format(team)
                pd.read_sql_query(sql=sql, con=self.engine1, chunksize=1000)
            except Exception as e:
                print('插入失败：', str(Exception) + str(e))
            print('----更新完成----')
        return '更新完成'

    def connectOrder(self, team):
        match = {'slgat': '港台',
                 'sltg': '泰国',
                 'slxmt': '新马',
                 'slzb': '直播团队',
                 'slyn': '越南',
                 'slrb': '日本'}
        emailAdd = {'slgat': 'giikinliujun@163.com',
                    'sltg': '1845389861@qq.com',
                    'slxmt': 'zhangjing@giikin.com',
                    'slzb': '直播团队',
                    'slyn': '越南',
                    'slrb': 'sunyaru@giikin.com'}
        emailAdd2 = {'sltg': 'zhangjing@giikin.com'}
        if team == 'sltg' or team == 'slxmt' or team == 'slrb0' or team == 'slgat0':
            month_last = (datetime.datetime.now().replace(day=1) - datetime.timedelta(days=1)).strftime('%Y-%m') + '-01'
            month_yesterday = datetime.datetime.now().strftime('%Y-%m-%d')
            month_begin = (datetime.datetime.now() - relativedelta(months=3)).strftime('%Y-%m-%d')
            print(month_begin)
        else:
            month_last = '2021-01-01'
            month_yesterday = '2021-03-07'
            month_begin = '2020-11-01'
        if team == 'slgat':  # 港台查询函数导出
            # sql = '''SELECT 年月, 旬, 日期, 团队,币种, 区域, 订单来源, a.订单编号 订单编号, 电话号码, a.运单编号 运单编号,
            #             IF(出货时间='1990-01-01 00:00:00' or 出货时间='1899-12-30 00:00:00' or 出货时间='0000-00-00 00:00:00', a.仓储扫描时间, 出货时间) 出货时间,
            #             IF(ISNULL(c.标准物流状态), b.物流状态, c.标准物流状态) 物流状态, c.`物流状态代码` 物流状态代码,IF(状态时间='1990-01-01 00:00:00' or 状态时间='1899-12-30 00:00:00' or 状态时间='0000-00-00 00:00:00', '', 状态时间) 状态时间,
            #             IF(上线时间='1990-01-01 00:00:00' or 上线时间='1899-12-30 00:00:00' or 上线时间='0000-00-00 00:00:00', '', 上线时间) 上线时间, 系统订单状态, IF(ISNULL(d.订单编号), 系统物流状态, '已退货') 系统物流状态,
            #             IF(ISNULL(d.订单编号), NULL, '已退货') 退货登记,
            #             IF(ISNULL(d.订单编号), IF(ISNULL(系统物流状态), IF(ISNULL(c.标准物流状态) OR c.标准物流状态 = '未上线', IF(系统订单状态 IN ('已转采购', '待发货'), '未发货', '未上线') , c.标准物流状态), 系统物流状态), '已退货') 最终状态,
            #             是否改派,物流方式,物流名称,运输方式,货物类型,是否低价,付款方式,
            #             产品id,IF(ISNULL(a.产品名称), a.产品名称, f.name) 产品名称,
            #             产品id,产品名称,父级分类,
            #             二级分类,三级分类,下单时间,审核时间,仓储扫描时间,完结状态时间,价格,价格RMB,价格区间,
            #             包裹重量,包裹体积,邮编,IF(ISNULL(b.运单编号), '否', '是') 签收表是否存在,
            #             b.订单编号 签收表订单编号, b.运单编号 签收表运单编号, 原运单号, b.物流状态 签收表物流状态, b.添加时间, a.成本价, a.物流花费, a.打包花费, a.其它花费, a.添加物流单号时间,数量
            #         FROM {0}_order_list a
            #             LEFT JOIN (SELECT * FROM {0} WHERE id IN (SELECT MAX(id) FROM {0} WHERE {0}.添加时间 > '{1}' GROUP BY 运单编号) ORDER BY id) b ON a.`运单编号` = b.`运单编号`
            #             LEFT JOIN {0}_logisitis_match c ON b.物流状态 = c.签收表物流状态
            #             LEFT JOIN {0}_return d ON a.订单编号 = d.订单编号
            #             LEFT JOIN dim_product f ON f.id = a.产品id
            #         WHERE a.日期 >= '{2}' AND a.日期 <= '{3}'
            #             AND a.系统订单状态 IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)','已退货(物流)', '已退货(不拆包物流)')
            #         ORDER BY a.`下单时间`;'''.format(team, month_begin, month_last, month_yesterday)

            sql = '''SELECT 年月, 旬, 日期, 团队,币种, 区域, 订单来源, a.订单编号 订单编号, 电话号码, a.运单编号 运单编号,
                        IF(出货时间='1990-01-01 00:00:00' or 出货时间='1899-12-30 00:00:00' or 出货时间='0000-00-00 00:00:00', a.仓储扫描时间, 出货时间) 出货时间,
                        IF(ISNULL(c.标准物流状态), b.物流状态, c.标准物流状态) 物流状态, c.`物流状态代码` 物流状态代码,IF(状态时间='1990-01-01 00:00:00' or 状态时间='1899-12-30 00:00:00' or 状态时间='0000-00-00 00:00:00', '', 状态时间) 状态时间,
                        IF(上线时间='1990-01-01 00:00:00' or 上线时间='1899-12-30 00:00:00' or 上线时间='0000-00-00 00:00:00', '', 上线时间) 上线时间, 系统订单状态, IF(ISNULL(d.订单编号), 系统物流状态, '已退货') 系统物流状态,
                        IF(ISNULL(d.订单编号), NULL, '已退货') 退货登记,
                        IF(ISNULL(d.订单编号), IF(ISNULL(系统物流状态), IF(ISNULL(c.标准物流状态) OR c.标准物流状态 = '未上线', IF(系统订单状态 IN ('已转采购', '待发货'), '未发货', '未上线') , c.标准物流状态), 系统物流状态), '已退货') 最终状态,
                        是否改派,物流方式,物流名称,运输方式,货物类型,是否低价,付款方式,产品id,产品名称,父级分类,
                        二级分类,三级分类,下单时间,审核时间,仓储扫描时间,完结状态时间,价格,价格RMB,价格区间,
                        包裹重量,包裹体积,邮编,IF(ISNULL(b.运单编号), '否', '是') 签收表是否存在,
                        b.订单编号 签收表订单编号, b.运单编号 签收表运单编号, 原运单号, b.物流状态 签收表物流状态, b.添加时间, a.成本价, a.物流花费, a.打包花费, a.其它花费, a.添加物流单号时间,数量
                    FROM {0}_order_list a
                        LEFT JOIN (SELECT * FROM {0} WHERE id IN (SELECT MAX(id) FROM {0} WHERE {0}.添加时间 > '{1}' GROUP BY 运单编号) ORDER BY id) b ON a.`运单编号` = b.`运单编号`
                        LEFT JOIN {0}_logisitis_match c ON b.物流状态 = c.签收表物流状态
                        LEFT JOIN {0}_return d ON a.订单编号 = d.订单编号
                    WHERE a.日期 >= '{2}' AND a.日期 <= '{3}'
                        AND a.系统订单状态 IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)','已退货(物流)', '已退货(不拆包物流)')
                    ORDER BY a.`下单时间`;'''.format(team, month_begin, month_last, month_yesterday)
        elif team == 'slxmt':  # 新马物流查询函数导出
            sql = '''SELECT 年月, 旬, 日期, 团队,币种, 区域, 订单来源, a.订单编号 订单编号, 电话号码, a.运单编号 运单编号,
                        IF(ISNULL(b.出货时间) or b.出货时间='0000-00-00 00:00:00' or b.状态时间='1990-01-01 00:00:00', g.出货时间, b.出货时间) 出货时间, IF(ISNULL(c.标准物流状态), b.物流状态, c.标准物流状态) 物流状态, c.`物流状态代码` 物流状态代码,
                        IF(b.状态时间='1990-01-01 00:00:00' or b.状态时间='1899-12-30 00:00:00' or b.状态时间='0000-00-00 00:00:00', '', b.状态时间) 状态时间, 上线时间, 系统订单状态, IF(ISNULL(d.订单编号), 系统物流状态, '已退货') 系统物流状态, IF(ISNULL(d.订单编号), NULL, '已退货') 退货登记,
                        IF(ISNULL(d.订单编号), IF(ISNULL(系统物流状态), IF(ISNULL(c.标准物流状态) OR c.标准物流状态 = '未上线', IF(系统订单状态 IN ('已转采购', '待发货'), '未发货', '未上线') , c.标准物流状态), 系统物流状态), '已退货') 最终状态,
                        是否改派,物流方式,物流名称,运输方式,货物类型,是否低价,付款方式,产品id,产品名称,父级分类,
                        二级分类,三级分类,下单时间,审核时间,仓储扫描时间,完结状态时间,价格,价格RMB,价格区间, 
                        包裹重量,包裹体积,邮编,IF(ISNULL(b.运单编号), '否', '是') 签收表是否存在,
                        b.订单编号 签收表订单编号, b.运单编号 签收表运单编号, 原运单号, b.物流状态 签收表物流状态, b.添加时间, a.成本价, a.物流花费, a.打包花费, a.其它花费, a.添加物流单号时间,数量
                    FROM {0}_order_list a  
                        LEFT JOIN (SELECT * FROM {0} WHERE id IN (SELECT MAX(id) FROM {0} WHERE {0}.添加时间 > '{1}' GROUP BY 运单编号) ORDER BY id) b ON a.`运单编号` = b.`运单编号`
                        LEFT JOIN {0}_logisitis_match c ON b.物流状态 = c.签收表物流状态 
                        LEFT JOIN {0}_return d ON a.订单编号 = d.订单编号
                        LEFT JOIN (SELECT * FROM {0}wl WHERE id IN (SELECT MAX(id) FROM {0}wl  WHERE {0}wl.添加时间 > '{1}' GROUP BY 运单编号) ORDER BY id) g ON a.运单编号 = g.运单编号
                    WHERE a.日期 >= '{2}' AND a.日期 <= '{3}'
                        AND a.系统订单状态 IN ('已审核', '待发货', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)', '待发货转审核') 
                    ORDER BY a.`下单时间`;'''.format(team, month_begin, month_last, month_yesterday)
        elif team == 'sltg':
            sql = '''SELECT 年月, 旬, 日期, 团队,币种, 区域, 订单来源, a.订单编号 订单编号, 电话号码, a.运单编号 运单编号,
                            IF(出货时间='1990-01-01 00:00:00' or 出货时间='1899-12-30 00:00:00' or 出货时间='0000-00-00 00:00:00', null, 出货时间) 出货时间,
                            IF(ISNULL(c.标准物流状态), b.物流状态, c.标准物流状态) 物流状态, c.`物流状态代码` 物流状态代码,IF(状态时间='1990-01-01 00:00:00' or 状态时间='1899-12-30 00:00:00' or 状态时间='0000-00-00 00:00:00', '', 状态时间) 状态时间,
                            IF(上线时间='1990-01-01 00:00:00' or 上线时间='1899-12-30 00:00:00' or 上线时间='0000-00-00 00:00:00', '', 上线时间) 上线时间, 系统订单状态, IF(ISNULL(d.订单编号), 系统物流状态, '已退货') 系统物流状态,
                            IF(ISNULL(d.订单编号), NULL, '已退货') 退货登记,
                            IF(ISNULL(d.订单编号), IF(ISNULL(系统物流状态), IF(ISNULL(c.标准物流状态) OR c.标准物流状态 = '未上线', IF(系统订单状态 IN ('已转采购', '待发货'), '未发货', '未上线') , c.标准物流状态), 系统物流状态), '已退货') 最终状态,
                            是否改派,物流方式,物流名称,运输方式,货物类型,是否低价,付款方式,产品id,产品名称,父级分类,
                            二级分类,三级分类,下单时间,审核时间,仓储扫描时间,完结状态时间,价格,价格RMB,价格区间,
                            包裹重量,包裹体积,邮编,IF(ISNULL(b.运单编号), '否', '是') 签收表是否存在,
                            b.订单编号 签收表订单编号, b.运单编号 签收表运单编号, 原运单号, b.物流状态 签收表物流状态, b.添加时间, a.成本价, a.物流花费, a.打包花费, a.其它花费, a.添加物流单号时间,数量
                    FROM {0}_order_list a  
                        LEFT JOIN (SELECT * FROM {0} WHERE id IN (SELECT MAX(id) FROM {0} WHERE {0}.添加时间 > '{1}' GROUP BY 运单编号) ORDER BY id) b ON a.`运单编号` = b.`运单编号`
                        LEFT JOIN {0}_logisitis_match c ON b.物流状态 = c.签收表物流状态
                        LEFT JOIN {0}_return d ON a.订单编号 = d.订单编号
                    WHERE a.日期 >= '{2}' AND a.日期 <= '{3}'
                        AND a.系统订单状态 IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)')
                    ORDER BY a.`下单时间`;'''.format(team, month_begin, month_last, month_yesterday)
        else:
            sql = '''SELECT 年月, 旬, 日期, 团队,币种, 区域, 订单来源, a.订单编号 订单编号, 电话号码, a.运单编号 运单编号,
                        IF(出货时间='1990-01-01 00:00:00' or 出货时间='1899-12-30 00:00:00' or 出货时间='0000-00-00 00:00:00', null, 出货时间) 出货时间,
                        IF(ISNULL(c.标准物流状态), b.物流状态, c.标准物流状态) 物流状态, c.`物流状态代码` 物流状态代码,IF(状态时间='1990-01-01 00:00:00' or 状态时间='1899-12-30 00:00:00' or 状态时间='0000-00-00 00:00:00', '', 状态时间) 状态时间,
                        IF(上线时间='1990-01-01 00:00:00' or 上线时间='1899-12-30 00:00:00' or 上线时间='0000-00-00 00:00:00', '', 上线时间) 上线时间, 系统订单状态, IF(ISNULL(d.订单编号), 系统物流状态, '已退货') 系统物流状态,
                        IF(ISNULL(d.订单编号), NULL, '已退货') 退货登记,
                        IF(ISNULL(d.订单编号), IF(ISNULL(系统物流状态), IF(ISNULL(c.标准物流状态) OR c.标准物流状态 = '未上线', IF(系统订单状态 IN ('已转采购', '待发货'), '未发货', '未上线') , c.标准物流状态), 系统物流状态), '已退货') 最终状态,
                        是否改派,物流方式,物流名称,运输方式,货物类型,是否低价,付款方式,产品id,产品名称,父级分类,
                        二级分类,三级分类,下单时间,审核时间,仓储扫描时间,完结状态时间,价格,价格RMB,价格区间,
                        包裹重量,包裹体积,邮编,IF(ISNULL(b.运单编号), '否', '是') 签收表是否存在,
                        b.订单编号 签收表订单编号, b.运单编号 签收表运单编号, 原运单号, b.物流状态 签收表物流状态,b.添加时间, a.成本价, a.物流花费, a.打包花费, a.其它花费, a.添加物流单号时间,数量
                    FROM {0}_order_list a  
                        LEFT JOIN (SELECT * FROM {0} WHERE id IN (SELECT MAX(id) FROM {0} WHERE {0}.添加时间 > '{1}' GROUP BY 运单编号) ORDER BY id) b ON a.`运单编号` = b.`运单编号`
                        LEFT JOIN {0}_logisitis_match c ON b.物流状态 = c.签收表物流状态
                        LEFT JOIN {0}_return d ON a.订单编号 = d.订单编号
                    WHERE a.日期 >= '{2}' AND a.日期 <= '{3}'
                        AND a.系统订单状态 IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)')
                    ORDER BY a.`下单时间`;'''.format(team, month_begin, month_last, month_yesterday)
        print('正在获取---' + match[team] + ' ---全部导出数据内容…………')
        df = pd.read_sql_query(sql=sql, con=self.engine1)
        print('正在写入---' + match[team] + ' ---临时缓存…………')

        # 备用临时缓存表
        if team == 'sl':
            df.to_sql('d1', con=self.engine1, index=False, if_exists='replace')
        else:
            df.to_sql('d1_{0}'.format(team), con=self.engine1, index=False, if_exists='replace')
        today = datetime.date.today().strftime('%Y.%m.%d')
        print('正在写入excel…………')
        df.to_excel('D:\\Users\\Administrator\\Desktop\\输出文件\\{} 神龙{}签收表.xlsx'.format(today, match[team]),
                    sheet_name=match[team], index=False)
        print('----已写入excel')
        filePath = ['D:\\Users\\Administrator\\Desktop\\输出文件\\{} 神龙{}签收表.xlsx'.format(today, match[team])]
        print('输出文件成功…………')
        # 文件太大无法发送的
        if team == 'slgat':
            print('---' + match[team] + ' 不发送邮件')
        else:
            self.e.send('{} 神龙{}签收表.xlsx'.format(today, match[team]), filePath,
                        emailAdd[team])
        if team == 'sltg':  # 补充发送一份
            self.e.send('{} 神龙{}签收表.xlsx'.format(today, match[team]), filePath,
                        emailAdd2[team])

        # 导入签收率表中和输出物流时效（不包含全部的订单状态）
        print('正在打印' + match[team] + ' 物流时效…………')
        if team == 'slgat0':
            print('---' + match[team] + ' 不打印文件')
        else:
            # pass
            self.data_wl(team)
        print('正在写入' + match[team] + ' 全部签收表中…………')
        if team == 'slrb':
            sql = 'REPLACE INTO {0}_zqsb_rb SELECT *, NOW() 更新时间 FROM d1_{0};'.format(team)
        else:
            sql = 'REPLACE INTO {0}_zqsb SELECT *, NOW() 更新时间 FROM d1_{0};'.format(team)
        pd.read_sql_query(sql=sql, con=self.engine1, chunksize=1000)
        print('----已写入' + match[team] + '全部签收表中')

    # 物流时效
    def data_wl(self, team):  # 获取各团队近两个月的物流数据
        match = {'slgat': ['台湾', '香港'],
                 'sltg': ['泰国'],
                 'slxmt': ['新加坡', '马来西亚', '菲律宾'],
                 'slrb': ['日本']}
        emailAdd = {'台湾': 'giikinliujun@163.com',
                    '香港': 'giikinliujun@163.com',
                    '新加坡': 'zhangjing@giikin.com',
                    '马来西亚': 'zhangjing@giikin.com',
                    '菲律宾': 'zhangjing@giikin.com',
                    '泰国': '1845389861@qq.com',
                    '日本': 'sunyaru@giikin.com'}
        emailAdd2 = {'泰国': 'zhangjing@giikin.com'}
        if team == 'sltg0' or team == 'slxmt0' or team == 'slrb0' or team == 'slgat0':
            month_last = (datetime.datetime.now().replace(day=1)).strftime('%Y-%m-%d')
        else:
            pass
        for tem in match[team]:
            filePath = []
            listT = []  # 查询sql的结果 存放池
            print('正在获取---' + tem + '---物流时效…………')
            # 总月
            sql = '''SELECT 年月,币种,物流方式,IF(s.天数=90,NULL,s.天数) AS 天数,总计 ,签收量,完成量,签收率完成,签收率总计,累计完成占比
                    FROM (SELECT IFNULL(年月,'总计') AS 年月,
								IFNULL(币种,'总计') AS 币种,
                                IFNULL(物流方式,'总计') AS 物流方式,
				                IFNULL(天数,'总计') AS 天数,
				                SUM(总计) AS 总计 ,
				                IFNULL(SUM(签收量),0) AS 签收量,
				                IFNULL(SUM(完成量),0) AS 完成量,
				                SUM(签收量) / SUM(完成量) AS '签收率完成',
				                SUM(签收量) / SUM(总计) AS '签收率总计',
				                '' AS 累计完成占比
                        FROM(SELECT gat_z.年月,gat_z.币种,gat_z.物流方式,IF(ISNULL(gat_z.下单出库时), 90, gat_z.下单出库时) AS 天数, 订单量 总计,签收量,完成量
                            FROM (SELECT  年月,币种,物流方式,DATEDIFF(`仓储扫描时间`,`下单时间`) AS 下单出库时,COUNT(`订单编号`) AS 订单量
                                FROM  d1_{0} cx
                                WHERE cx.`币种` = '{1}'	
                                    AND  cx.`是否改派` = '直发'
    			                    AND  cx.系统订单状态 IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)')
                                GROUP BY 年月,币种,物流方式,下单出库时
                                ORDER BY 年月,币种,物流方式,下单出库时
                                ) gat_z
                            LEFT JOIN  
                                (SELECT  年月,币种,物流方式,DATEDIFF(`仓储扫描时间`,`下单时间`) AS 下单出库时,COUNT(`订单编号`) AS 签收量
                                FROM  d1_{0} cx
                                WHERE cx.`币种` = '{1}'		
                                    AND  cx.`是否改派` = '直发'
                                    AND  cx.系统订单状态 IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)')
                                    AND  cx.`最终状态` = '已签收'
                                GROUP BY 年月,币种,物流方式,下单出库时
                                ORDER BY 年月,币种,物流方式,下单出库时
                                ) gat_yqs
                            ON gat_z.`年月` = gat_yqs.`年月` 
                                AND gat_z.`币种` = gat_yqs.`币种` 
    	                        AND gat_z.`物流方式` = gat_yqs.`物流方式`
    	                        AND gat_z.`下单出库时` = gat_yqs.`下单出库时`
                            LEFT JOIN 
                                (SELECT  年月,币种,物流方式,DATEDIFF(`仓储扫描时间`,`下单时间`) AS 下单出库时,COUNT(`订单编号`) AS 完成量
                                FROM  d1_{0} cx
                                WHERE cx.`币种` = '{1}'	
                                    AND  cx.`是否改派` = '直发'
    			                    AND  cx.系统订单状态 IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)')
                                    AND  cx.`最终状态` IN ('已签收','拒收','理赔','已退货')
                                GROUP BY 年月,币种,物流方式,下单出库时
                                ORDER BY 年月,币种,物流方式,下单出库时
                                ) gat_wc
                            ON gat_z.`年月` = gat_wc.`年月` 
                                AND gat_z.`币种` = gat_wc.`币种` 
    	                        AND gat_z.`物流方式` = gat_wc.`物流方式`
    	                        AND gat_z.`下单出库时` = gat_wc.`下单出库时`
                        )	sl
                        GROUP BY 年月,币种,物流方式,sl.天数
                        with rollup
                    ) s WHERE s.`币种` != '总计'  AND s.`年月` != '总计';'''.format(team, tem)
            df = pd.read_sql_query(sql=sql, con=self.engine1)
            listT.append(df)
            sql2 = '''SELECT 年月,币种,物流方式,IF(s.天数=90,NULL,s.天数) AS 天数,总计 ,签收量,完成量,签收率完成,签收率总计,累计完成占比
                    FROM (SELECT IFNULL(年月,'总计') AS 年月,
                                IFNULL(币种,'总计') AS 币种,
                                IFNULL(物流方式,'总计') AS 物流方式,
				                IFNULL(天数,'总计') AS 天数,
				                SUM(总计) AS 总计 ,
				                IFNULL(SUM(签收量),0) AS 签收量,
				                IFNULL(SUM(完成量),0) AS 完成量,
				                SUM(签收量) / SUM(完成量) AS '签收率完成',
				                SUM(签收量) / SUM(总计) AS '签收率总计',
				                '' AS 累计完成占比
                        FROM(SELECT gat_z.年月,gat_z.币种,gat_z.物流方式,IF(ISNULL(gat_z.出库完成时), 90, gat_z.出库完成时) AS 天数, 订单量 总计,签收量,完成量
                            FROM (SELECT  年月,币种,物流方式,DATEDIFF(IFNULL(`完结状态时间`,`状态时间`),`仓储扫描时间`) AS 出库完成时,COUNT(`订单编号`) AS 订单量
                                FROM  d1_{0} cx
                                WHERE cx.`币种` = '{1}'	
                                    AND  cx.`是否改派` = '直发'
    			                          AND  cx.系统订单状态 IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)')
                                GROUP BY 年月,币种,物流方式,出库完成时
                                ORDER BY 年月,币种,物流方式,出库完成时
                                ) gat_z
                            LEFT JOIN  
                                (SELECT  年月,币种,物流方式,DATEDIFF(IFNULL(`完结状态时间`,`状态时间`),`仓储扫描时间`) AS 出库完成时,COUNT(`订单编号`) AS 签收量
                                FROM  d1_{0} cx
                                WHERE cx.`币种` = '{1}'
                                    AND  cx.`是否改派` = '直发'
                                    AND  cx.系统订单状态 IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)')
                                    AND  cx.`最终状态` = '已签收'
                                GROUP BY 年月,币种,物流方式,出库完成时
                                ORDER BY 年月,币种,物流方式,出库完成时
                                ) gat_yqs
                            ON gat_z.`年月` = gat_yqs.`年月` 
                                AND gat_z.`币种` = gat_yqs.`币种` 
    	                        AND gat_z.`物流方式` = gat_yqs.`物流方式`
    	                        AND gat_z.`出库完成时` = gat_yqs.`出库完成时`
                            LEFT JOIN 
                                (SELECT  年月,币种,物流方式,DATEDIFF(IFNULL(`完结状态时间`,`状态时间`),`仓储扫描时间`) AS 出库完成时,COUNT(`订单编号`) AS 完成量
                                FROM  d1_{0} cx
                                WHERE cx.`币种` = '{1}'	
                                    AND  cx.`是否改派` = '直发'
    			                    AND  cx.系统订单状态 IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)')
                                    AND  cx.`最终状态` IN ('已签收','拒收','理赔','已退货')
                                GROUP BY 年月,币种,物流方式,出库完成时
                                ORDER BY 年月,币种,物流方式,出库完成时
                                ) gat_wc
                            ON gat_z.`年月` = gat_wc.`年月` 
                                AND gat_z.`币种` = gat_wc.`币种` 
    	                        AND gat_z.`物流方式` = gat_wc.`物流方式`
    	                        AND gat_z.`出库完成时` = gat_wc.`出库完成时`
                        )	sl
                        GROUP BY 年月,币种,物流方式,sl.天数
                        with rollup
                    ) s WHERE s.`币种` != '总计'  AND s.`年月` != '总计';'''.format(team, tem)
            df2 = pd.read_sql_query(sql=sql2, con=self.engine1)
            listT.append(df2)
            sql3 = '''SELECT 年月,币种,物流方式,IF(s.天数=90,NULL,s.天数) AS 天数,总计 ,签收量,完成量,签收率完成,签收率总计,累计完成占比
                    FROM (SELECT IFNULL(年月,'总计') AS 年月,
								IFNULL(币种,'总计') AS 币种,
                                IFNULL(物流方式,'总计') AS 物流方式,
				                IFNULL(天数,'总计') AS 天数,
				                SUM(总计) AS 总计 ,
				                IFNULL(SUM(签收量),0) AS 签收量,
				                IFNULL(SUM(完成量),0) AS 完成量,
				                SUM(签收量) / SUM(完成量) AS '签收率完成',
				                SUM(签收量) / SUM(总计) AS '签收率总计',
				                '' AS 累计完成占比
                        FROM(SELECT gat_z.年月,gat_z.币种,gat_z.物流方式,IF(ISNULL(gat_z.下单完成时), 90, gat_z.下单完成时) AS 天数, 订单量 总计,签收量,完成量
                            FROM (SELECT  年月,币种,物流方式,DATEDIFF(IFNULL(`完结状态时间`,`状态时间`),`下单时间`) AS 下单完成时,COUNT(`订单编号`) AS 订单量
                                FROM  d1_{0} cx
                                WHERE cx.`币种` = '{1}'	
                                    AND  cx.`是否改派` = '直发'
    			                    AND  cx.系统订单状态 IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)')
                                GROUP BY 年月,币种,物流方式,下单完成时
                                ORDER BY 年月,币种,物流方式,下单完成时
                                ) gat_z
                            LEFT JOIN  
                                (SELECT  年月,币种,物流方式,DATEDIFF(IFNULL(`完结状态时间`,`状态时间`),`下单时间`) AS 下单完成时,COUNT(`订单编号`) AS 签收量
                                FROM  d1_{0} cx
                                WHERE cx.`币种` = '{1}' 	
                                    AND  cx.`是否改派` = '直发'
                                    AND  cx.系统订单状态 IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)')
                                    AND  cx.`最终状态` = '已签收'
                                GROUP BY 年月,币种,物流方式,下单完成时
                                ORDER BY 年月,币种,物流方式,下单完成时
                                ) gat_yqs
                            ON gat_z.`年月` = gat_yqs.`年月` 
                                AND gat_z.`币种` = gat_yqs.`币种` 
    	                        AND gat_z.`物流方式` = gat_yqs.`物流方式`
    	                        AND gat_z.`下单完成时` = gat_yqs.`下单完成时`
                            LEFT JOIN 
                                (SELECT  年月,币种,物流方式,DATEDIFF(IFNULL(`完结状态时间`,`状态时间`),`下单时间`) AS 下单完成时,COUNT(`订单编号`) AS 完成量
                                FROM  d1_{0} cx
                                WHERE cx.`币种` = '{1}'	
                                    AND  cx.`是否改派` = '直发'
    			                    AND  cx.系统订单状态 IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)')
                                    AND  cx.`最终状态` IN ('已签收','拒收','理赔','已退货')
                                GROUP BY 年月,币种,物流方式,下单完成时
                                ORDER BY 年月,币种,物流方式,下单完成时
                                ) gat_wc
                            ON gat_z.`年月` = gat_wc.`年月` 
                                AND gat_z.`币种` = gat_wc.`币种` 
    	                        AND gat_z.`物流方式` = gat_wc.`物流方式`
    	                        AND gat_z.`下单完成时` = gat_wc.`下单完成时`
                        )	sl
                        GROUP BY 年月,币种,物流方式,sl.天数
                        with rollup
                    ) s WHERE s.`币种` != '总计'  AND s.`年月` != '总计';'''.format(team, tem)
            df3 = pd.read_sql_query(sql=sql3, con=self.engine1)
            listT.append(df3)
            sql4 = '''SELECT 年月,币种,物流方式,IF(s.天数=90,NULL,s.天数) AS 天数,总计 ,签收量,完成量,签收率完成,签收率总计,累计完成占比
                    FROM (SELECT IFNULL(年月,'总计') AS 年月,
												IFNULL(币种,'总计') AS 币种,
                        IFNULL(物流方式,'总计') AS 物流方式,
				                IFNULL(天数,'总计') AS 天数,
				                SUM(总计) AS 总计 ,
				                IFNULL(SUM(签收量),0) AS 签收量,
				                IFNULL(SUM(完成量),0) AS 完成量,
				                SUM(签收量) / SUM(完成量) AS '签收率完成',
				                SUM(签收量) / SUM(总计) AS '签收率总计',
				                '' AS 累计完成占比
                        FROM(SELECT gat_z.年月,gat_z.币种,gat_z.物流方式,IF(ISNULL(gat_z.下单完成时), 90, gat_z.下单完成时) AS 天数, 订单量 总计,签收量,完成量
                            FROM (SELECT  年月,币种,物流方式,DATEDIFF(IFNULL(`完结状态时间`,`状态时间`),`下单时间`) AS 下单完成时,COUNT(`订单编号`) AS 订单量
                                FROM  d1_{0} cx
                                WHERE cx.`币种` = '{1}'	
                                    AND  cx.`是否改派` = '改派'
    			                    AND  cx.系统订单状态 IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)')
                                GROUP BY 年月,币种,物流方式,下单完成时
                                ORDER BY 年月,币种,物流方式,下单完成时
                                ) gat_z
                            LEFT JOIN  
                                (SELECT  年月,币种,物流方式,DATEDIFF(IFNULL(`完结状态时间`,`状态时间`),`下单时间`) AS 下单完成时,COUNT(`订单编号`) AS 签收量
                                FROM  d1_{0} cx
                                WHERE cx.`币种` = '{1}' 	
                                    AND  cx.`是否改派` = '改派'
                                    AND  cx.系统订单状态 IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)')
                                    AND  cx.`最终状态` = '已签收'
                                GROUP BY 年月,币种,物流方式,下单完成时
                                ORDER BY 年月,币种,物流方式,下单完成时
                                ) gat_yqs
                            ON gat_z.`年月` = gat_yqs.`年月` 
                                AND gat_z.`币种` = gat_yqs.`币种` 
    	                        AND gat_z.`物流方式` = gat_yqs.`物流方式`
    	                        AND gat_z.`下单完成时` = gat_yqs.`下单完成时`
                            LEFT JOIN 
                                (SELECT  年月,币种,物流方式,DATEDIFF(IFNULL(`完结状态时间`,`状态时间`),`下单时间`) AS 下单完成时,COUNT(`订单编号`) AS 完成量
                                FROM  d1_{0} cx
                                WHERE cx.`币种` = '{1}'	
                                    AND  cx.`是否改派` = '改派'
    			                    AND  cx.系统订单状态 IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)')
                                    AND  cx.`最终状态` IN ('已签收','拒收','理赔','已退货')
                                GROUP BY 年月,币种,物流方式,下单完成时
                                ORDER BY 年月,币种,物流方式,下单完成时
                                ) gat_wc
                            ON gat_z.`年月` = gat_wc.`年月` 
                                AND gat_z.`币种` = gat_wc.`币种` 
    	                        AND gat_z.`物流方式` = gat_wc.`物流方式`
    	                        AND gat_z.`下单完成时` = gat_wc.`下单完成时`
                        )	sl
                        GROUP BY 年月,币种,物流方式,sl.天数
                        with rollup
                    ) s WHERE s.`币种` != '总计'  AND s.`年月` != '总计';'''.format(team, tem)
            df4 = pd.read_sql_query(sql=sql4, con=self.engine1)
            listT.append(df4)
            # 分旬
            print('正在获取---' + tem + '---物流分旬时效…………')
            sql10 = '''SELECT 年月,币种,物流方式,旬,IF(s.天数=90,NULL,s.天数) AS 天数,总计 ,签收量,完成量,签收率完成,签收率总计,累计完成占比
                                FROM (SELECT IFNULL(年月,'总计') AS 年月,
            								IFNULL(币种,'总计') AS 币种,
                                            IFNULL(物流方式,'总计') AS 物流方式,
                                            IFNULL(旬,'总计') AS 旬,
            				                IFNULL(天数,'总计') AS 天数,
            				                SUM(总计) AS 总计 ,
            				                IFNULL(SUM(签收量),0) AS 签收量,
            				                IFNULL(SUM(完成量),0) AS 完成量,
            				                SUM(签收量) / SUM(完成量) AS '签收率完成',
            				                SUM(签收量) / SUM(总计) AS '签收率总计',
            				                '' AS 累计完成占比
                                    FROM(SELECT gat_z.年月,gat_z.币种,gat_z.物流方式,gat_z.旬,IF(ISNULL(gat_z.下单出库时), 90, gat_z.下单出库时) AS 天数, 订单量 总计,签收量,完成量
                                        FROM (SELECT  年月,币种,物流方式,旬,DATEDIFF(`仓储扫描时间`,`下单时间`) AS 下单出库时,COUNT(`订单编号`) AS 订单量
                                            FROM  d1_{0} cx
                                            WHERE cx.`币种` = '{1}'	
                                                AND  cx.`是否改派` = '直发'
                			                    AND  cx.系统订单状态 IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)')
                                            GROUP BY 年月,币种,物流方式,旬,下单出库时
                                            ORDER BY 年月,币种,物流方式,旬,下单出库时
                                            ) gat_z
                                        LEFT JOIN  
                                            (SELECT  年月,币种,物流方式,旬,DATEDIFF(`仓储扫描时间`,`下单时间`) AS 下单出库时,COUNT(`订单编号`) AS 签收量
                                            FROM  d1_{0} cx
                                            WHERE cx.`币种` = '{1}' 	
                                                AND  cx.`是否改派` = '直发'
                                                AND  cx.系统订单状态 IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)')
                                                AND  cx.`最终状态` = '已签收'
                                            GROUP BY 年月,币种,物流方式,旬,下单出库时
                                            ORDER BY 年月,币种,物流方式,旬,下单出库时
                                            ) gat_yqs
                                        ON gat_z.`年月` = gat_yqs.`年月` 
                                            AND gat_z.`币种` = gat_yqs.`币种` 
                	                        AND gat_z.`物流方式` = gat_yqs.`物流方式`
                	                        AND gat_z.`旬` = gat_yqs.`旬`
                	                        AND gat_z.`下单出库时` = gat_yqs.`下单出库时`
                                        LEFT JOIN 
                                            (SELECT  年月,币种,物流方式,旬,DATEDIFF(`仓储扫描时间`,`下单时间`) AS 下单出库时,COUNT(`订单编号`) AS 完成量
                                            FROM  d1_{0} cx
                                            WHERE cx.`币种` = '{1}'	
                                                AND  cx.`是否改派` = '直发'
                			                    AND  cx.系统订单状态 IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)')
                                                AND  cx.`最终状态` IN ('已签收','拒收','理赔','已退货')
                                            GROUP BY 年月,币种,物流方式,旬,下单出库时
                                            ORDER BY 年月,币种,物流方式,旬,下单出库时
                                            ) gat_wc
                                        ON gat_z.`年月` = gat_wc.`年月` 
                                            AND gat_z.`币种` = gat_wc.`币种` 
                	                        AND gat_z.`物流方式` = gat_wc.`物流方式`
                	                        AND gat_z.`旬` = gat_wc.`旬`
                	                        AND gat_z.`下单出库时` = gat_wc.`下单出库时`
                                    )	sl
                                    GROUP BY 年月,币种,物流方式,旬,sl.天数
                                    with rollup
                                ) s WHERE s.`币种` != '总计'  AND s.`年月` != '总计';'''.format(team, tem)
            df10 = pd.read_sql_query(sql=sql10, con=self.engine1)
            listT.append(df10)
            sql20 = '''SELECT 年月,币种,物流方式,旬,IF(s.天数=90,NULL,s.天数) AS 天数,总计 ,签收量,完成量,签收率完成,签收率总计,累计完成占比
                                FROM (SELECT IFNULL(年月,'总计') AS 年月,
                                            IFNULL(币种,'总计') AS 币种,
                                            IFNULL(物流方式,'总计') AS 物流方式,
                                            IFNULL(旬,'总计') AS 旬,
            				                IFNULL(天数,'总计') AS 天数,
            				                SUM(总计) AS 总计 ,
            				                IFNULL(SUM(签收量),0) AS 签收量,
            				                IFNULL(SUM(完成量),0) AS 完成量,
            				                SUM(签收量) / SUM(完成量) AS '签收率完成',
            				                SUM(签收量) / SUM(总计) AS '签收率总计',
            				                '' AS 累计完成占比
                                    FROM(SELECT gat_z.年月,gat_z.币种,gat_z.物流方式,gat_z.旬,IF(ISNULL(gat_z.出库完成时), 90, gat_z.出库完成时) AS 天数, 订单量 总计,签收量,完成量
                                        FROM (SELECT  年月,币种,物流方式,旬,DATEDIFF(IFNULL(`完结状态时间`,`状态时间`),`仓储扫描时间`) AS 出库完成时,COUNT(`订单编号`) AS 订单量
                                            FROM  d1_{0} cx
                                            WHERE cx.`币种` = '{1}'	
                                                AND  cx.`是否改派` = '直发'
                			                          AND  cx.系统订单状态 IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)')
                                            GROUP BY 年月,币种,物流方式,旬,出库完成时
                                            ORDER BY 年月,币种,物流方式,旬,出库完成时
                                            ) gat_z
                                        LEFT JOIN  
                                            (SELECT  年月,币种,物流方式,旬,DATEDIFF(IFNULL(`完结状态时间`,`状态时间`),`仓储扫描时间`) AS 出库完成时,COUNT(`订单编号`) AS 签收量
                                            FROM  d1_{0} cx
                                            WHERE cx.`币种` = '{1}' 	
                                                AND  cx.`是否改派` = '直发'
                                                AND  cx.系统订单状态 IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)')
                                                AND  cx.`最终状态` = '已签收'
                                            GROUP BY 年月,币种,物流方式,旬,出库完成时
                                            ORDER BY 年月,币种,物流方式,旬,出库完成时
                                            ) gat_yqs
                                        ON gat_z.`年月` = gat_yqs.`年月` 
                                            AND gat_z.`币种` = gat_yqs.`币种` 
                	                        AND gat_z.`物流方式` = gat_yqs.`物流方式`
                	                        AND gat_z.`旬` = gat_yqs.`旬`
                	                        AND gat_z.`出库完成时` = gat_yqs.`出库完成时`
                                        LEFT JOIN 
                                            (SELECT  年月,币种,物流方式,旬,DATEDIFF(IFNULL(`完结状态时间`,`状态时间`),`仓储扫描时间`) AS 出库完成时,COUNT(`订单编号`) AS 完成量
                                            FROM  d1_{0} cx
                                            WHERE cx.`币种` = '{1}'	
                                                AND  cx.`是否改派` = '直发'
                			                    AND  cx.系统订单状态 IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)')
                                                AND  cx.`最终状态` IN ('已签收','拒收','理赔','已退货')
                                            GROUP BY 年月,币种,物流方式,旬,出库完成时
                                            ORDER BY 年月,币种,物流方式,旬,出库完成时
                                            ) gat_wc
                                        ON gat_z.`年月` = gat_wc.`年月` 
                                            AND gat_z.`币种` = gat_wc.`币种` 
                	                        AND gat_z.`物流方式` = gat_wc.`物流方式`
                	                        AND gat_z.`旬` = gat_wc.`旬`
                	                        AND gat_z.`出库完成时` = gat_wc.`出库完成时`
                                    )	sl
                                    GROUP BY 年月,币种,物流方式,旬,sl.天数
                                    with rollup
                                ) s WHERE s.`币种` != '总计'  AND s.`年月` != '总计';'''.format(team, tem)
            df20 = pd.read_sql_query(sql=sql20, con=self.engine1)
            listT.append(df20)
            sql30 = '''SELECT 年月,币种,物流方式,旬,IF(s.天数=90,NULL,s.天数) AS 天数,总计 ,签收量,完成量,签收率完成,签收率总计,累计完成占比
                                FROM (SELECT IFNULL(年月,'总计') AS 年月,
            								IFNULL(币种,'总计') AS 币种,
                                            IFNULL(物流方式,'总计') AS 物流方式,
                                            IFNULL(旬,'总计') AS 旬,
            				                IFNULL(天数,'总计') AS 天数,
            				                SUM(总计) AS 总计 ,
            				                IFNULL(SUM(签收量),0) AS 签收量,
            				                IFNULL(SUM(完成量),0) AS 完成量,
            				                SUM(签收量) / SUM(完成量) AS '签收率完成',
            				                SUM(签收量) / SUM(总计) AS '签收率总计',
            				                '' AS 累计完成占比
                                    FROM(SELECT gat_z.年月,gat_z.币种,gat_z.物流方式,gat_z.旬,IF(ISNULL(gat_z.下单完成时), 90, gat_z.下单完成时) AS 天数, 订单量 总计,签收量,完成量
                                        FROM (SELECT  年月,币种,物流方式,旬,DATEDIFF(IFNULL(`完结状态时间`,`状态时间`),`下单时间`) AS 下单完成时,COUNT(`订单编号`) AS 订单量
                                            FROM  d1_{0} cx
                                            WHERE cx.`币种` = '{1}'	
                                                AND  cx.`是否改派` = '直发'
                			                    AND  cx.系统订单状态 IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)')
                                            GROUP BY 年月,币种,物流方式,旬,下单完成时
                                            ORDER BY 年月,币种,物流方式,旬,下单完成时
                                            ) gat_z
                                        LEFT JOIN  
                                            (SELECT  年月,币种,物流方式,旬,DATEDIFF(IFNULL(`完结状态时间`,`状态时间`),`下单时间`) AS 下单完成时,COUNT(`订单编号`) AS 签收量
                                            FROM  d1_{0} cx
                                            WHERE cx.`币种` = '{1}' 	
                                                AND  cx.`是否改派` = '直发'
                                                AND  cx.系统订单状态 IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)')
                                                AND  cx.`最终状态` = '已签收'
                                            GROUP BY 年月,币种,物流方式,旬,下单完成时
                                            ORDER BY 年月,币种,物流方式,旬,下单完成时
                                            ) gat_yqs
                                        ON gat_z.`年月` = gat_yqs.`年月` 
                                            AND gat_z.`币种` = gat_yqs.`币种` 
                	                        AND gat_z.`物流方式` = gat_yqs.`物流方式`
                	                        AND gat_z.`旬` = gat_yqs.`旬`
                	                        AND gat_z.`下单完成时` = gat_yqs.`下单完成时`
                                        LEFT JOIN 
                                            (SELECT  年月,币种,物流方式,旬,DATEDIFF(IFNULL(`完结状态时间`,`状态时间`),`下单时间`) AS 下单完成时,COUNT(`订单编号`) AS 完成量
                                            FROM  d1_{0} cx
                                            WHERE cx.`币种` = '{1}'	
                                                AND  cx.`是否改派` = '直发'
                			                    AND  cx.系统订单状态 IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)')
                                                AND  cx.`最终状态` IN ('已签收','拒收','理赔','已退货')
                                            GROUP BY 年月,币种,物流方式,旬,下单完成时
                                            ORDER BY 年月,币种,物流方式,旬,下单完成时
                                            ) gat_wc
                                        ON gat_z.`年月` = gat_wc.`年月` 
                                            AND gat_z.`币种` = gat_wc.`币种` 
                	                        AND gat_z.`物流方式` = gat_wc.`物流方式`
                	                        AND gat_z.`旬` = gat_wc.`旬`
                	                        AND gat_z.`下单完成时` = gat_wc.`下单完成时`
                                    )	sl
                                    GROUP BY 年月,币种,物流方式,旬,sl.天数
                                    with rollup
                                ) s WHERE s.`币种` != '总计'  AND s.`年月` != '总计';'''.format(team, tem)
            df30 = pd.read_sql_query(sql=sql30, con=self.engine1)
            listT.append(df30)
            sql40 = '''SELECT 年月,币种,物流方式,旬,IF(s.天数=90,NULL,s.天数) AS 天数,总计 ,签收量,完成量,签收率完成,签收率总计,累计完成占比
                                FROM (SELECT IFNULL(年月,'总计') AS 年月,
            												IFNULL(币种,'总计') AS 币种,
                                    IFNULL(物流方式,'总计') AS 物流方式,
                                    IFNULL(旬,'总计') AS 旬,
            				                IFNULL(天数,'总计') AS 天数,
            				                SUM(总计) AS 总计 ,
            				                IFNULL(SUM(签收量),0) AS 签收量,
            				                IFNULL(SUM(完成量),0) AS 完成量,
            				                SUM(签收量) / SUM(完成量) AS '签收率完成',
            				                SUM(签收量) / SUM(总计) AS '签收率总计',
            				                '' AS 累计完成占比
                                    FROM(SELECT gat_z.年月,gat_z.币种,gat_z.物流方式,gat_z.旬,IF(ISNULL(gat_z.下单完成时), 90, gat_z.下单完成时) AS 天数, 订单量 总计,签收量,完成量
                                        FROM (SELECT  年月,币种,物流方式,旬,DATEDIFF(IFNULL(`完结状态时间`,`状态时间`),`下单时间`) AS 下单完成时,COUNT(`订单编号`) AS 订单量
                                            FROM  d1_{0} cx
                                            WHERE cx.`币种` = '{1}'	
                                                AND  cx.`是否改派` = '改派'
                			                    AND  cx.系统订单状态 IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)')
                                            GROUP BY 年月,币种,物流方式,旬,下单完成时
                                            ORDER BY 年月,币种,物流方式,旬,下单完成时
                                            ) gat_z
                                        LEFT JOIN  
                                            (SELECT  年月,币种,物流方式,旬,DATEDIFF(IFNULL(`完结状态时间`,`状态时间`),`下单时间`) AS 下单完成时,COUNT(`订单编号`) AS 签收量
                                            FROM  d1_{0} cx
                                            WHERE cx.`币种` = '{1}' 	
                                                AND  cx.`是否改派` = '改派'
                                                AND  cx.系统订单状态 IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)')
                                                AND  cx.`最终状态` = '已签收'
                                            GROUP BY 年月,币种,物流方式,旬,下单完成时
                                            ORDER BY 年月,币种,物流方式,旬,下单完成时
                                            ) gat_yqs
                                        ON gat_z.`年月` = gat_yqs.`年月` 
                                            AND gat_z.`币种` = gat_yqs.`币种` 
                	                        AND gat_z.`物流方式` = gat_yqs.`物流方式`
                	                        AND gat_z.`旬` = gat_yqs.`旬`
                	                        AND gat_z.`下单完成时` = gat_yqs.`下单完成时`
                                        LEFT JOIN 
                                            (SELECT  年月,币种,物流方式,旬,DATEDIFF(IFNULL(`完结状态时间`,`状态时间`),`下单时间`) AS 下单完成时,COUNT(`订单编号`) AS 完成量
                                            FROM  d1_{0} cx
                                            WHERE cx.`币种` = '{1}'	
                                                AND  cx.`是否改派` = '改派'
                			                    AND  cx.系统订单状态 IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)')
                                                AND  cx.`最终状态` IN ('已签收','拒收','理赔','已退货')
                                            GROUP BY 年月,币种,物流方式,旬,下单完成时
                                            ORDER BY 年月,币种,物流方式,旬,下单完成时
                                            ) gat_wc
                                        ON gat_z.`年月` = gat_wc.`年月` 
                                            AND gat_z.`币种` = gat_wc.`币种` 
                	                        AND gat_z.`物流方式` = gat_wc.`物流方式`
                	                        AND gat_z.`旬` = gat_wc.`旬`
                	                        AND gat_z.`下单完成时` = gat_wc.`下单完成时`
                                    )	sl
                                    GROUP BY 年月,币种,物流方式,旬,sl.天数
                                    with rollup
                                ) s WHERE s.`币种` != '总计'  AND s.`年月` != '总计';'''.format(team, tem)
            df40 = pd.read_sql_query(sql=sql40, con=self.engine1)
            listT.append(df40)
            print('正在写入excel…………')
            today = datetime.date.today().strftime('%Y.%m.%d')
            file_path = 'D:\\Users\\Administrator\\Desktop\\输出文件\\{} {}物流时效.xlsx'.format(today, tem)
            sheet_name = ['下单出库时', '出库完成时', '下单完成时', '改派下单完成时', '下单出库(分旬)', '出库完成(分旬)', '下单完成(分旬)', '改派下单完成(分旬)']
            df0 = pd.DataFrame([])                       # 创建空的dataframe数据框
            df0.to_excel(file_path, index=False)         # 备用：可以向不同的sheet写入数据（创建新的工作表并进行写入）
            writer = pd.ExcelWriter(file_path, engine='openpyxl')  # 初始化写入对象
            book = load_workbook(file_path)              # 可以向不同的sheet写入数据（对现有工作表的追加）
            writer.book = book                           # 将数据写入excel中的sheet2表,sheet_name改变后即是新增一个sheet
            for i in range(len(listT)):
                listT[i]['签收率完成'] = listT[i]['签收率完成'].fillna(value=0)
                listT[i]['签收率总计'] = listT[i]['签收率总计'].fillna(value=0)
                listT[i]['签收率完成'] = listT[i]['签收率完成'].apply(lambda x: format(x, '.2%'))
                listT[i]['签收率总计'] = listT[i]['签收率总计'].apply(lambda x: format(x, '.2%'))
                listT[i].to_excel(excel_writer=writer, sheet_name=sheet_name[i], index=False)
            if 'Sheet1' in book.sheetnames:              # 删除新建文档时的第一个工作表
                del book['Sheet1']
            writer.save()
            writer.close()
            print('正在运行宏…………')
            app = xl.App(visible=False, add_book=False)  # 运行宏调整
            app.display_alerts = False
            wbsht = app.books.open('D:/Users/Administrator/Desktop/新版-格式转换(工具表).xlsm')
            wbsht1 = app.books.open(file_path)
            wbsht.macro('sltem物流时效')()
            wbsht1.save()
            wbsht1.close()
            wbsht.close()
            app.quit()
            print('----已写入excel ')
            filePath.append(file_path)
            self.e.send('{} {}物流时效.xlsx'.format(today, tem), filePath,
                        emailAdd[tem])
            if team == 'sltg':
                self.e.send('{} {}物流时效.xlsx'.format(today, tem), filePath,
                            emailAdd2[tem])

    # 无运单号查询
    def noWaybillNumber(self, team):
        match1 = {'slgat': '港台',
                  'sltg': '泰国',
                  'slxmt': '新马',
                  'slzb': '直播团队',
                  'slyn': '越南',
                  'slrb': '日本'}
        match = {'slgat': '"神龙家族-港澳台"',
                 'sltg': '"神龙家族-泰国"',
                 'slxmt': '"神龙家族-新加坡", "神龙家族-马来西亚", "神龙家族-菲律宾"',
                 'slzb': '"神龙家族-直播团队"',
                 'slyn': '"神龙家族-越南"',
                 'slrb': '"神龙家族-日本团队"'}
        emailAdd = {'slgat': 'giikinliujun@163.com',
                    'sltg': '1845389861@qq.com',
                    'slxmt': 'zhangjing@giikin.com',
                    'slzb': '直播团队',
                    'slyn': '越南',
                    'slrb': 'sunyaru@giikin.com'}
        emailAdd2 = {'sltg': 'libin@giikin.com'}
        print('正在查询{}无运单订单列表…………'.format(match[team]))
        yesterday = (datetime.datetime.now() - datetime.timedelta(days=1)).strftime('%Y-%m-%d')
        last_month = (datetime.datetime.now().replace(day=1) - datetime.timedelta(days=1)).strftime('%Y-%m') + '-01'
        # last_month = '2020-07-21'
        sql = '''SELECT a.rq 日期, 
                        dim_area.name 团队, 
                        a.order_number 订单编号,
                        a.waybill_number 运单编号,
                        a.order_status 系统订单状态id, 
                        IF(a.second=0,'直发','改派') 是否改派, 
                        dim_trans_way.all_name 物流方式, 
                        a.addtime 下单时间, 
                        a.verity_time 审核时间
                FROM gk_order_kf a 
                    left join dim_area ON dim_area.id = a.area_id
                    left join dim_trans_way on dim_trans_way.id = a.logistics_id
                WHERE 
                    a.rq >= '{}' AND a.rq <= '{}'
                    AND dim_area.name IN ({})
                    AND ISNULL(waybill_number) 
                    AND order_status NOT IN (1, 8, 11, 14, 16)
                ORDER BY a.rq;'''.format(last_month, yesterday, match[team])
        df = pd.read_sql_query(sql=sql, con=self.engine2)
        sql = 'SELECT * FROM dim_order_status;'
        df1 = pd.read_sql_query(sql=sql, con=self.engine1)
        print('正在合并订单状态…………')
        df = pd.merge(left=df, right=df1, left_on='系统订单状态id', right_on='id', how='left')
        df = df.drop(labels=['id', '系统订单状态id', ], axis=1)
        df.rename(columns={'name': '系统订单状态'}, inplace=True)
        today = datetime.date.today().strftime('%Y.%m.%d')
        df.to_excel('D:\\Users\\Administrator\\Desktop\\输出文件\\{} 神龙{}无运单号列表.xlsx'.format(today, match1[team]),
                    sheet_name=match1[team], index=False)
        filePath = ['D:\\Users\\Administrator\\Desktop\\输出文件\\{} 神龙{}无运单号列表.xlsx'.format(today, match1[team])]
        print('输出文件成功…………')
        self.e.send('{} 神龙{}无运单号列表.xlsx'.format(today, match1[team]), filePath,
                    emailAdd[team])
        if team == 'sltg':
            self.e.send('{} 神龙{}无运单号列表.xlsx'.format(today, match1[team]), filePath,
                        emailAdd2[team])

    # 产品花费表
    def orderCost(self, team):
        if datetime.datetime.now().day >= 9:
            endDate = datetime.datetime.now().replace(day=1).strftime('%Y-%m-%d')
            startDate = (datetime.datetime.now().replace(day=1) - datetime.timedelta(days=1)).strftime('%Y-%m') + '-01'
            endDate = [endDate, datetime.datetime.now().strftime('%Y-%m-%d')]
            startDate = [startDate, datetime.datetime.now().replace(day=1).strftime('%Y-%m-%d')]
        else:
            endDate = datetime.datetime.now().replace(day=1).strftime('%Y-%m-%d')
            startDate = (datetime.datetime.now().replace(day=1) - datetime.timedelta(days=1)).strftime('%Y-%m') + '-01'
            endDate = [endDate]
            startDate = [startDate]
        match = {'SG': '新加坡',
                 'MY': '马来西亚',
                 'PH': '菲律宾',
                 'JP': '日本',
                 'HK': '香港',
                 'TW': '台湾',
                 'TH': '泰国'}
        match2 = {'SG': 'slxmt_zqsb',
                  'MY': 'slxmt_zqsb',
                  'PH': 'slxmt_zqsb',
                  'JP': 'slrb_zqsb_rb',
                  'HK': 'slgat_zqsb',
                  'TW': 'slgat_zqsb',
                  'TH': 'sltg_zqsb'}
        emailAdd = {'SG': 'zhangjing@giikin.com',
                    'MY': 'zhangjing@giikin.com',
                    'PH': 'zhangjing@giikin.com',
                    'JP': 'sunyaru@giikin.com',
                    'HK': 'giikinliujun@163.com',
                    'TW': 'giikinliujun@163.com',
                    'TH': '1845389861@qq.com'}
        # filePath = []
        for i in range(len(endDate)):
            print('正在查询' + match[team] + '产品花费表…………')
            sql = '''SELECT s1.`month` AS '月份',
                         s1.area AS '地区',
                         s1.leader AS '负责人',
                         s1.pid AS '产品ID',
                         s1.pname AS '产品名称',
                         s1.cate1 AS '一级品类',
                         s1.cate2 AS '二级品类',
                         s1.cate3 AS '三级品类',
                         s1.orders AS '订单量',
                         s1.orders - s1.gps AS '直发订单量',
                         s1.gps AS '改派订单量',
                         s1.salesRMB / s1.orders AS '客单价',
                         s1.salesRMB / s1.adcost AS 'ROI',
                         
                         s1.orders / s2.orders AS '订单品类占比',
                         s1.cgcost_zf / s1.salesRMB AS '直发采购/销售额',
                         s1.adcost / s1.salesRMB AS '花费占比',
                         s1.wlcost / s1.salesRMB AS '运费占比',
                         s1.qtcost / s1.salesRMB AS '手续费占比',
                         ( s1.cgcost_zf + s1.adcost + s1.wlcost + s1.qtcost ) / s1.salesRMB AS '总成本占比',
                         s1.salesRMB_yqs / (s1.salesRMB_yqs + s1.salesRMB_yjs) AS '金额签收/完成',
                         s1.salesRMB_yqs / s1.salesRMB AS '金额签收/总计',
                         (s1.salesRMB_yqs + s1.salesRMB_yjs) / s1.salesRMB AS '金额完成占比',
                         s1.yqs / (s1.yqs + s1.yjs) AS '数量签收/完成',
                         (s1.yqs + s1.yjs) / s1.orders AS '数量完成占比',
                         s3.orders AS '昨日订单量'
            FROM (
                        SELECT DISTINCT EXTRACT(YEAR_MONTH FROM a.rq) AS `month`,
                                     b.pname AS area,
                                     c.uname AS leader,
                                     a.product_id AS pid,
                                     e.`product_name` AS pname,
                                     d.ppname AS cate1,
                                     d.pname AS cate2,
                                     d.`name` AS cate3,
            -- 						 GROUP_CONCAT(DISTINCT a.low_price) AS low_price,
                                     SUM(a.orders) AS orders,
                                     SUM(a.yqs) AS yqs,
                                     SUM(a.yjs) AS yjs,
                                     SUM(a.salesRMB) AS salesRMB,
                                     SUM(a.salesRMB_yqs) AS salesRMB_yqs,
                                     SUM(a.salesRMB_yjs) AS salesRMB_yjs,
                                     SUM(a.gps) AS gps,
            --                         SUM(a.cgcost_zf) AS cgcost_zf,
            --                         SUM(a.adcost) AS adcost,
                                    null cgcost_zf,
                                    null adcost,
                                     SUM(a.wlcost) AS wlcost,
                                     SUM(a.qtcost) AS qtcost		 
                        FROM gk_order_day_kf a
                            LEFT JOIN dim_currency_lang b ON a.currency_lang_id = b.id
                            LEFT JOIN dim_area c on c.id = a.area_id
                            LEFT JOIN dim_cate d on d.id = a.third_cate_id
                            LEFT JOIN (SELECT * FROM gk_sale WHERE id IN (SELECT MAX(id) FROM gk_sale GROUP BY product_id ) ORDER BY id) e on e.product_id = a.product_id
                        WHERE a.rq >= '{startDate}'
                            AND a.rq < '{endDate}'
                            AND b.pcode = '{team}'
                            AND c.uname = '王冰'
                            AND a.beform <> 'mf'
                            AND c.uid <> 10099  -- 过滤翼虎
                        GROUP BY EXTRACT(YEAR_MONTH FROM a.rq), b.pname, c.uname, a.product_id
                     ) s1
                     
                      LEFT JOIN
                        
                     (
                        SELECT EXTRACT(YEAR_MONTH FROM a.rq) AS `month`,
                                     b.pname AS area,
                                     c.uname AS leader,
                                     d.ppname AS cate1,
                                     SUM(a.orders) AS orders
                        FROM gk_order_day_kf a
                            LEFT JOIN dim_currency_lang b ON a.currency_lang_id = b.id
                            LEFT JOIN dim_area c on c.id = a.area_id
                            LEFT JOIN dim_cate d on d.id = a.third_cate_id
                        WHERE a.rq >= '{startDate}'
                            AND a.rq < '{endDate}'
                            AND b.pcode = '{team}'
                            AND c.uname = '王冰'
                            AND a.beform <> 'mf'
                            AND c.uid <> 10099  
                        GROUP BY EXTRACT(YEAR_MONTH FROM a.rq), b.pname, c.uname, d.ppname
                     ) s2  ON s1.`month`=s2.`month` AND s1.area=s2.area AND s1.leader=s2.leader AND s1.cate1=s2.cate1
                    
                      LEFT JOIN
                    
                     (
                      SELECT EXTRACT(YEAR_MONTH FROM a.rq) AS `month`,
                                     b.pname AS area,
                                     c.uname AS leader,
                                     a.product_id AS pid,
                                     SUM(a.orders) AS orders
                        FROM gk_order_day_kf a
                            LEFT JOIN dim_currency_lang b ON a.currency_lang_id = b.id
                            LEFT JOIN dim_area c on c.id = a.area_id
                        WHERE a.rq = DATE_SUB(CURDATE(), INTERVAL 1 DAY)
                                    AND b.pcode = '{team}'
                                    AND c.uname = '王冰'
                                    AND a.beform <> 'mf'
                                    AND c.uid <> 10099  
                        GROUP BY EXTRACT(YEAR_MONTH FROM a.rq), b.pname, c.uname, a.product_id
                     ) s3 ON s1.area=s3.area AND s1.leader=s3.leader AND s1.pid=s3.pid
            WHERE s1.orders > 0
            ORDER BY s1.orders DESC'''.format(team=team, startDate=startDate[i], endDate=endDate[i])
            df = pd.read_sql_query(sql=sql, con=self.engine2)
            print('正在输出' + match[team] + '产品花费表…………')
            columns = ['订单品类占比', '直发采购/销售额', '花费占比', '运费占比', '手续费占比', '总成本占比',
                       '金额签收/完成', '金额签收/总计', '金额完成占比', '数量签收/完成', '数量完成占比']
            for column in columns:
                df[column] = df[column].fillna(value=0)
                df[column] = df[column].apply(lambda x: format(x, '.2%'))
            today = datetime.date.today().strftime('%Y.%m.%d')
            if i == 0:
                df.to_excel('D:\\Users\\Administrator\\Desktop\\输出文件\\{} {}上月产品花费表.xlsx'.format(today, match[team]),
                            sheet_name=match[team], index=False)
                # filePath.append('D:\\Users\\Administrator\\Desktop\\输出文件\\{} {}上月产品花费表.xlsx'.format(today, match[team]))
            else:
                df.to_excel('D:\\Users\\Administrator\\Desktop\\输出文件\\{} {}本月产品花费表.xlsx'.format(today, match[team]),
                            sheet_name=match[team], index=False)
                # filePath.append('D:\\Users\\Administrator\\Desktop\\输出文件\\{} {}本月产品花费表.xlsx'.format(today, match[team]))
        # self.e.send(match[team] + '产品花费表', filePath,
        #             emailAdd[team])
        self.d.sl_tem_cost(match2[team], match[team])

    # 各团队(泰国)全部订单表-函数
    def tgOrderQuan(self, team):  # 3天内的
        match1 = {'slgat': '港台',
                  'sltg': '泰国',
                  'slxmt': '新马',
                  'slzb': '直播团队',
                  'slyn': '越南',
                  'slrb': '日本'}
        match = {'slgat': '"神龙家族-港澳台"',
                 'sltg': '"神龙家族-泰国"',
                 'slxmt': '"神龙家族-新加坡", "神龙家族-马来西亚", "神龙家族-菲律宾"',
                 'slzb': '"神龙家族-直播团队"',
                 'slyn': '"神龙家族-越南"',
                 'slrb': '"神龙家族-日本团队"'}
        print('正在获取' + match1[team] + '最近 10 天订单…………')
        yesterday = (datetime.datetime.now()).strftime('%Y-%m-%d')
        # yesterday = (datetime.datetime.now().replace(month=1, day=10)).strftime('%Y-%m-%d')
        print(yesterday)
        last_month = (datetime.datetime.now() - datetime.timedelta(days=10)).strftime('%Y-%m-%d')
        # last_month = (datetime.datetime.now().replace(month=1, day=5)).strftime('%Y-%m-%d')
        print(last_month)
        sql = '''SELECT a.id,
                        a.订单编号 order_number,
                        a.团队 area_id,
                        '' main_id,
                        a.电话号码 ship_phone,
                        a.邮编 ship_zip,
                        a.价格 amount,
                        a.系统订单状态 order_status,
                        UPPER(a.运单编号) waybill_number,
                        a.付款方式 pay_type,
                        a.下单时间 addtime,
                        a.审核时间 update_time,
                        a.产品id goods_id, 
                        '' quantity,
                        a.物流方式 logistics_id,
                        '' op_id,
                        CONCAT(a.产品id,'#' ,a.产品名称) goods_name, 
                        a.是否改派 secondsend_status,
                        a.是否低价 low_price
                FROM {}_order_list a 
                WHERE a.日期 >= '{}' AND a.日期 <= '{}';'''.format(team, last_month, yesterday)
        try:
            df = pd.read_sql_query(sql=sql, con=self.engine1)
            print(df)
            print('正在写入缓存表中…………')
            df.to_sql('tem_sl', con=self.engine3, index=False, if_exists='replace')
            print('++++更新缓存完成++++')
        except Exception as e:
            print('更新缓存失败：', str(Exception) + str(e))
        print('正在写入 ' + match1[team] + ' 全部订单表中…………')
        sql = 'REPLACE INTO 全部订单_{} SELECT *, NOW() 添加时间 FROM tem_sl;'.format(team)
        pd.read_sql_query(sql=sql, con=self.engine3, chunksize=100)
        # 获取订单明细（泰国）
        print('======正在启动查询订单程序>>>>>')
        b = BpsControl('nixiumin@giikin.com', 'nixiumin123@.')
        match = {'slgat': '港台',
                 'sltg': '泰国',
                 'slxmt': '新马',
                 'slzb': '直播团队',
                 'slyn': '越南',
                 'slrb': '日本'}
        team = 'sltg'
        searchType = '订单号'  # 运单号，订单号；查询切换
        b.getNumberT(team, searchType)
        print('查询耗时：', datetime.datetime.now() - start)
        time.sleep(10)
        b.getNumberAdd(team, searchType)
        print('补充耗时：', datetime.datetime.now() - start)


if __name__ == '__main__':
    #  messagebox.showinfo("提示！！！", "当前查询已完成--->>> 请前往（ 输出文件 ）查看")200
    m = MysqlControl()
    start = datetime.datetime.now()

    # 更新产品id的列表
    m.update_gk_product()

    for team in ['sltg', 'slgat', 'slrb', 'slxmt']:  # 无运单号查询200
        m.noWaybillNumber(team)

    match = {'SG': '新加坡',
             'MY': '马来西亚',
             'PH': '菲律宾',
             'JP': '日本',
             'HK': '香港',
             'TW': '台湾',
             'TH': '泰国'}
    # match = {'JP': '日本'}
    for team in match.keys():  # 产品花费表200
        m.orderCost(team)

    sm = SltemMonitoring()
    for team in ['菲律宾', '新加坡', '马来西亚', '日本', '香港', '台湾', '泰国']:  # 成本查询
        sm.costWaybill(team)

    # （泰国）全部订单表200
    m.tgOrderQuan('sltg')

    # 测试物流时效
    # team = 'slgat'
    # m.data_wl(team)
    # for team in ['sltg', 'slgat', 'slrb', 'slxmt']:
    #     m.data_wl(team)



