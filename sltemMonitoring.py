import pandas as pd
import os
import xlwings as xl
import pandas.io.formats.excel
from sqlalchemy import create_engine
from settings import Settings
from emailControl import EmailControl
from openpyxl import load_workbook  # 可以向不同的sheet写入数据
import datetime
import xlwings


# -*- coding:utf-8 -*-
class SltemMonitoring(Settings):
    def __init__(self):
        Settings.__init__(self)
        self.engine1 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql1['user'],
                                                                                    self.mysql1['password'],
                                                                                    self.mysql1['host'],
                                                                                    self.mysql1['port'],
                                                                                    self.mysql1['datebase']))
        self.engine2 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql2['user'],
                                                                                    self.mysql2['password'],
                                                                                    self.mysql2['host'],
                                                                                    self.mysql2['port'],
                                                                                    self.mysql2['datebase']))
        self.engine3 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql3['user'],
                                                                                    self.mysql3['password'],
                                                                                    self.mysql3['host'],
                                                                                    self.mysql3['port'],
                                                                                    self.mysql3['datebase']))
        self.e = EmailControl()

    def reSetEngine(self):
        self.engine1 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql1['user'],
                                                                                    self.mysql1['password'],
                                                                                    self.mysql1['host'],
                                                                                    self.mysql1['port'],
                                                                                    self.mysql1['datebase']))
        self.engine2 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql2['user'],
                                                                                    self.mysql2['password'],
                                                                                    self.mysql2['host'],
                                                                                    self.mysql2['port'],
                                                                                    self.mysql2['datebase']))

    def check_time(self, team):
        match = {'新加坡': 'slxmt',
                 '马来西亚': 'slxmt',
                 '菲律宾': 'slxmt',
                 '日本': 'slrb',
                 '香港': 'slgat',
                 '台湾': 'slgat',
                 '泰国': 'sltg'}
        Time_day = []
        for i in range(1, datetime.datetime.now().month + 1):  # 获取当年当前的月份时间
            try:
                daytime = (datetime.datetime.now().replace(month=i)).strftime('%Y-%m') + (
                    (datetime.datetime.now()).strftime('-%d'))
                Time_day.append(daytime)
            except Exception as e:
                print('xxxx时间配置出错,已手动调整：' + str(i) + '月份', str(Exception) + str(e))
                Time_day.append(
                    str(int(datetime.datetime.now().year)) + '-' + str(i) + (datetime.datetime.now().strftime('-%d')))
        for i in range(datetime.datetime.now().month + 1, 13):  # 获取往年当前的月份时间
            try:
                daytime = str(int(datetime.datetime.now().year) -1) + (datetime.datetime.now().replace(month=i)).strftime('-%m') + (
                    (datetime.datetime.now()).strftime('-%d'))
                Time_day.append(daytime)
            except Exception as e:
                print('xxxx时间配置出错失败00：' + str(i) + '月份', str(Exception) + str(e))
                Time_day.append(str(int(datetime.datetime.now().year) - 1) + '-' + str(i) + (
                    datetime.datetime.now().strftime('-%d')))
        # 对时间数组进行排序  list.sort(cmp=None, key=None, reverse=False)；reverse -- 排序规则，reverse = True 降序， reverse = False 升序（默认）
        Time_day.sort()
        print('正在获取本次同期比较需要的---具体时间......')
        print(Time_day[11])
        print(Time_day[10])
        # 获取监控表是否有同期上传时间的数据
        rq_day = (datetime.datetime.now().replace(day=1) - datetime.timedelta(days=1)).strftime('%Y-%m') + '-01'
        print('正在检查监控表是否有需要的---具体日期......')
        sql = '''SELECT distinct qsb.`记录时间` FROM qsb_{0} qsb WHERE qsb.`记录时间`>='{1}';'''.format(match[team], rq_day)
        rq = pd.read_sql_query(sql=sql, con=self.engine1)
        df = rq['记录时间'].values              # datafram转为数组
        info = ''
        for r in df:
            # print(type(r.strftime('%Y')))
            if Time_day[10] == r.strftime('%Y-%m-%d'):
                print(r)
                info = '---已确认，可以进行同期数据对比'
                break
            else:
                info = '---需要手动上传需要时间的数据'

        if info == '---已确认，可以进行同期数据对比':
            print('++++++完成时间确认++++++')
            print('===>>>已确认' + team + '团队开始运行<<<===')
            self.order_Monitoring(team)  # 各月缓存
            self.data_Monitoring(team)  # 两月数据
            self.sl_Monitoring(team)  # 输出数据
            self.sl_Monitoring_two(team)  # 输出数据
            print('===>>>' + team + '团队运行结束<<<===')
        else:
            print(info)

    def order_Monitoring(self, team):        # 获取各团队各月的签收表数据 和 成本数据内容
        match = {'新加坡': 'slxmt',
                 '马来西亚': 'slxmt',
                 '菲律宾': 'slxmt',
                 '日本': 'slrb',
                 '香港': 'slgat',
                 '台湾': 'slgat',
                 '泰国': 'sltg'}
        match2 = {'新加坡': 'SG',
                  '马来西亚': 'MY',
                  '菲律宾': 'PH',
                  '日本': 'JP',
                  '香港': 'HK',
                  '台湾': 'TW',
                  '泰国': 'TH'}
        start: datetime = datetime.datetime.now()
        print('正在获取' + team + '每月（全部）缓存签收数据…………')
        Time_da = []
        for i in range(datetime.datetime.now().month + 1, 13):
            try:
                daytime = str(int(datetime.datetime.now().year) - 1) + (
                    datetime.datetime.now().replace(month=i)).strftime('-%m') + (
                              datetime.datetime.now().strftime('-%d'))
                Time_da.append(daytime)
            except Exception as e:
                print('xxxx时间配置出错失败：' + str(i) + '月份', str(Exception) + str(e))
                Time_da.append(str(int(datetime.datetime.now().year) - 1) + '-' + str(i) + (
                    datetime.datetime.now().strftime('-%d')))
        month_last = Time_da[6]
        print(month_last)
        sql = '''SELECT 年月, 旬, 日期, 币种, 订单来源, a.订单编号 订单编号, 
                        IF(出货时间='1990-01-01 00:00:00' or 出货时间='1899-12-30 00:00:00' or 出货时间='0000-00-00 00:00:00', '', 出货时间) 出货时间,
                        IF(状态时间='1990-01-01 00:00:00' or 状态时间='1899-12-30 00:00:00' or 状态时间='0000-00-00 00:00:00', '', 状态时间) 状态时间,
                        IF(上线时间='1990-01-01 00:00:00' or 上线时间='1899-12-30 00:00:00' or 上线时间='0000-00-00 00:00:00', '', 上线时间) 上线时间, 
                        系统订单状态, 
                        IF(ISNULL(d.订单编号), 系统物流状态, '已退货') 系统物流状态,
                        IF(ISNULL(d.订单编号), NULL, '已退货') 退货登记,
                        IF(ISNULL(d.订单编号), IF(ISNULL(系统物流状态), IF(ISNULL(c.标准物流状态) OR c.标准物流状态 = '未上线', IF(系统订单状态 IN ('已转采购', '待发货'), '未发货', '未上线') , c.标准物流状态), 系统物流状态), '已退货') 最终状态,
                        是否改派,物流方式,物流名称,运输方式,是否低价,产品id,产品名称,父级分类,二级分类,三级分类,下单时间,
                        审核时间,仓储扫描时间,完结状态时间,价格区间,价格RMB
                    FROM {0}_order_list a 
                    LEFT JOIN (SELECT * FROM {0} WHERE id IN (SELECT MAX(id) FROM {0} GROUP BY 运单编号) ORDER BY id) b
                        ON a.`运单编号` = b.`运单编号`
                    LEFT JOIN {0}_logisitis_match c 
                        ON b.物流状态 = c.签收表物流状态
                    LEFT JOIN {0}_return d 
                        ON a.订单编号 = d.订单编号
                    WHERE a.日期 >= '{1}' AND a.币种 = '{2}'
                        AND a.系统订单状态 IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)')
                    ORDER BY a.`下单时间`;'''.format(match[team], month_last, team)
        df = pd.read_sql_query(sql=sql, con=self.engine1)
        df.to_sql('qsb_缓存_month', con=self.engine1, index=False, if_exists='replace')
        print('----已写入' + team + '每月（全部）缓存签收表中')
        # print('正在获取' + team + '每月（全部）缓存成本…………'.format(team))
        # yesterday = (datetime.datetime.now() - datetime.timedelta(days=1)).strftime('%Y-%m-%d') + ' 23:59:59'
        # sql = '''SELECT EXTRACT(YEAR_MONTH FROM a.rq) AS 年月,
        #                     b.pname AS 团队,
        #                     c.uname AS leader,
        #                     d.ppname AS 品类,
        #                     SUM(a.orders) AS 订单量,
        #                     COUNT(DISTINCT a.product_id) AS 活跃产品数,
        #                     SUM(a.yqs) AS 签收量,
        #                     SUM(a.yjs + a.yth) AS 拒收量,
        #                     SUM(a.salesRMB) AS 销售额,
        #                     SUM(a.salesRMB_yqs) AS 签收额,
        #                     SUM(a.salesRMB_yjs + a.salesRMB_yth) AS 拒收额,
        #                     SUM(a.gps) AS 改派订单量,
        #                     SUM(a.cgcost) AS 总采购额,
        #                     SUM(a.cgcost_zf) AS 直发采购额,
        #                     SUM(a.adcost) AS 广告成本,
        #                     SUM(a.wlcost) AS 物流成本,
        #                     SUM(a.qtcost) AS 手续费
        #             FROM gk_order_day a
        #                 LEFT JOIN dim_currency_lang b ON a.currency_lang_id = b.id
        #                 LEFT JOIN dim_area c on c.id = a.area_id
        #                 LEFT JOIN dim_cate d on d.id = a.third_cate_id
        #                 LEFT JOIN gk_product e on e.id = a.product_id
        #             WHERE a.rq >= '{0}'
        #                 AND a.rq < '{1}'
        #                 AND b.pcode = '{2}'
        #                 AND c.uname = '王冰'
        #                 AND a.beform <> 'mf'
        #                 AND c.uid <> 10099  -- 过滤翼虎
        #             GROUP BY b.pname, c.uname,EXTRACT(YEAR_MONTH FROM a.rq)
        #             ORDER BY EXTRACT(YEAR_MONTH FROM a.rq) desc ;'''.format(month_last, yesterday, match2[team])
        # df = pd.read_sql_query(sql=sql, con=self.engine2)
        # df.to_sql('zg_cost_缓存_month', con=self.engine1, index=False, if_exists='replace')
        # print('已导入' + team + '每月（全部）缓存成本表中+++')
        print('缓存耗时：', datetime.datetime.now() - start)
    def data_Monitoring(self, team):     # 获取各团队近两个月的签收表数据
        match3 = {'新加坡': 'slxmt',
                  '马来西亚': 'slxmt',
                  '菲律宾': 'slxmt',
                  '日本': 'slrb',
                  '香港': 'slgat',
                  '台湾': 'slgat',
                  '泰国': 'sltg'}
        start: datetime = datetime.datetime.now()
        month_last = (datetime.datetime.now().replace(day=1) - datetime.timedelta(days=1)).strftime('%Y-%m') + '-01'
        sql = '''SELECT 年月, 旬, 日期, 币种, 订单来源, a.订单编号 订单编号,
                        IF(出货时间='1990-01-01 00:00:00' or 出货时间='1899-12-30 00:00:00' or 出货时间='0000-00-00 00:00:00', '', 出货时间) 出货时间,
                        IF(状态时间='1990-01-01 00:00:00' or 状态时间='1899-12-30 00:00:00' or 状态时间='0000-00-00 00:00:00', '', 状态时间) 状态时间,
                        IF(上线时间='1990-01-01 00:00:00' or 上线时间='1899-12-30 00:00:00' or 上线时间='0000-00-00 00:00:00', '', 上线时间) 上线时间,
                        系统订单状态,
                        IF(ISNULL(d.订单编号), 系统物流状态, '已退货') 系统物流状态,
                        IF(ISNULL(d.订单编号), NULL, '已退货') 退货登记,
                        IF(ISNULL(d.订单编号), IF(ISNULL(系统物流状态), IF(ISNULL(c.标准物流状态) OR c.标准物流状态 = '未上线', IF(系统订单状态 IN ('已转采购', '待发货'), '未发货', '未上线') , c.标准物流状态), 系统物流状态), '已退货') 最终状态,
                        是否改派,物流方式,物流名称,运输方式,是否低价,产品id,产品名称,父级分类,二级分类,三级分类,下单时间,
                        审核时间,仓储扫描时间,完结状态时间,价格区间,价格RMB
                    FROM {0}_order_list a
                    LEFT JOIN (SELECT * FROM {0} WHERE id IN (SELECT MAX(id) FROM {0} GROUP BY 运单编号) ORDER BY id) b
                        ON a.`运单编号` = b.`运单编号`
                    LEFT JOIN {0}_logisitis_match c
                        ON b.物流状态 = c.签收表物流状态
                    LEFT JOIN {0}_return d
                        ON a.订单编号 = d.订单编号
                    WHERE a.日期 >= '{1}' AND a.币种 = '{2}'
                        AND a.系统订单状态 IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)')
                    ORDER BY a.`下单时间`;'''.format(match3[team], month_last, team)
        print('正在获取---' + team + '---最近两个月监控数据…………')
        df = pd.read_sql_query(sql=sql, con=self.engine1)
        today = datetime.date.today().strftime('%Y.%m.%d')
        print('正在写入excel…………')
        df.to_excel('F:\\查询\\订单数据\\{} {}监控数据源表.xlsx'.format(today, team),
                    sheet_name=team, index=False)
        print('----已写入excel ')
        print('正在写入-数据源-缓存中…………')
        df.to_sql('qsb_缓存', con=self.engine1, index=False, if_exists='replace')
        print('正在写入 ' + team + ' 监控数据表中…………')
        columns = list(df.columns)
        columns = ', '.join(columns)
        sql = '''INSERT IGNORE INTO qsb_{}({}, 记录时间) SELECT *, CURDATE() 记录时间 FROM qsb_缓存; '''.format(match3[team], columns)
        try:
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=2000)
        except Exception as e:
            print('插入失败：', str(Exception) + str(e))
        print('----已写入' + team + '近两月监控-签收表中')
        print('获取耗时：', datetime.datetime.now() - start)
    def costWaybill(self, team):        # 获取各团队近两个月的成本数据内容
        match = {'新加坡': 'SG',
                 '马来西亚': 'MY',
                 '菲律宾': 'PH',
                 '日本': 'JP',
                 '香港': 'HK',
                 '台湾': 'TW',
                 '泰国': 'TH'}
        start: datetime = datetime.datetime.now()
        print('正在查询 {} 品类总成本…………'.format(team))
        yesterday = (datetime.datetime.now() - datetime.timedelta(days=1)).strftime('%Y-%m-%d') + ' 23:59:59'
        print(yesterday)
        last_month = (datetime.datetime.now().replace(day=1) - datetime.timedelta(days=1)).strftime('%Y-%m') + '-01'
        print(last_month)
        sql = '''SELECT s1.年月, s1.团队, s1.品类,
                        s1.销售额,
                        s1.订单量,
                        s1.订单量 - s1.改派订单量 AS '直发订单量',
                        (s1.订单量 - s1.改派订单量) / s1.订单量 AS '直发占比',
                        s1.改派订单量,
                        s1.改派订单量 / s1.订单量 AS '改派占比',
                        s1.订单量 / s2.订单品类量 AS '订单品类占比',
                        s1.销售额 / s1.订单量 AS '客单价',
                        s1.销售额 / s1.广告成本 AS 'ROI',
                        s1.活跃产品数,
                        s1.订单量 / s1.活跃产品数 AS 产能,
                        s1.直发采购额 / s1.销售额 AS '直发采购占比',
                        s1.广告成本 / s1.销售额 AS '广告占比',
                        s1.物流成本 / s1.销售额 AS '运费占比',
                        s1.手续费 / s1.销售额 AS '手续费占比',
                        ( s1.直发采购额 + s1.广告成本 + s1.物流成本 + s1.手续费 ) / s1.销售额 AS '总成本占比',
                        s1.签收额 / (s1.签收额 + s1.拒收额) AS '金额签收完成',
                        s1.签收额 / s1.销售额 AS '金额签收总计',
                        (s1.签收额 + s1.拒收额) / s1.销售额 AS '金额完成占比',
                        s1.签收量 / (s1.签收量 + s1.拒收量) AS '数量签收完成',
                        s1.签收量 / s1.订单量 AS '数量签收总计',
                        (s1.签收量 + s1.拒收量) / s1.订单量 AS '数量完成占比',
                        s1.签收额 / (s1.签收额 + s1.拒收额) -( s1.直发采购额 + s1.广告成本 + s1.物流成本 + s1.手续费 ) / s1.销售额 AS 利润率,
                        (s1.签收额 / (s1.签收额 + s1.拒收额) -( s1.直发采购额 + s1.广告成本 + s1.物流成本 + s1.手续费 ) /s1.销售额) * (s1.销售额 / s1.订单量) AS 利润值
                FROM ( SELECT EXTRACT(YEAR_MONTH FROM a.rq) AS 年月,
                            b.pname AS 团队,
                            c.uname AS leader,
                            d.ppname AS 品类,
                            SUM(a.orders) AS 订单量,
                            COUNT(DISTINCT a.product_id) AS 活跃产品数,
                            SUM(a.yqs) AS 签收量,
                            SUM(a.yjs) AS 拒收量,
                            SUM(a.salesRMB) AS 销售额,
                            SUM(a.salesRMB_yqs) AS 签收额,
                            SUM(a.salesRMB_yjs) AS 拒收额,
                            SUM(a.gps) AS 改派订单量,
                --          SUM(a.cgcost) AS 总采购额,
                --          SUM(a.cgcost_zf) AS 直发采购额,
                --          SUM(a.adcost) AS 广告成本,
                            null 总采购额,
                            null 直发采购额,
                            null 广告成本,
                            SUM(a.wlcost) AS 物流成本,
                            SUM(a.qtcost) AS 手续费
                    FROM gk_order_day_kf a
                        LEFT JOIN dim_currency_lang b ON a.currency_lang_id = b.id
                        LEFT JOIN dim_area c on c.id = a.area_id
                        LEFT JOIN dim_cate d on d.id = a.third_cate_id
                        LEFT JOIN (SELECT * FROM gk_sale WHERE id IN (SELECT MAX(id) FROM gk_sale GROUP BY product_id ) ORDER BY id) e on e.product_id = a.product_id
                    WHERE a.rq >= '{0}'
                        AND a.rq < '{1}'
                        AND b.pcode = '{2}'
                        AND c.uname = '王冰'
                        AND a.beform <> 'mf'
                        AND c.uid <> 10099  -- 过滤翼虎
                    GROUP BY b.pname, c.uname, d.ppname
                    ORDER BY a.product_id
                ) s1
                LEFT JOIN
                ( SELECT EXTRACT(YEAR_MONTH FROM a.rq) AS 年月,
                            b.pname AS 团队,
                            c.uname AS leader,
                            d.ppname AS 品类,
                            SUM(a.orders) AS 订单品类量
                FROM gk_order_day_kf a
                    LEFT JOIN dim_currency_lang b ON a.currency_lang_id = b.id
                    LEFT JOIN dim_area c on c.id = a.area_id
                    LEFT JOIN dim_cate d on d.id = a.third_cate_id
                    WHERE a.rq >= '{0}'
                        AND a.rq < '{1}'
                        AND b.pcode = '{2}'
                        AND c.uname = '王冰'
                        AND a.beform <> 'mf'
                        AND c.uid <> 10099
                    GROUP BY b.pname, c.uname
                ) s2
                    ON s1.团队=s2.团队 AND s1.leader=s2.leader
                    WHERE s1.订单量 > 0
                    GROUP BY s1.年月 ,s1.团队, s1.品类
                UNION ALL
                SELECT s3.年月, s3.团队, s3.品类,
                        s3.销售额,
                        s3.订单量,
                        s3.订单量 - s3.改派订单量 AS 直发订单量,
                        (s3.订单量 - s3.改派订单量) / s3.订单量 直发占比,
                        s3.改派订单量,
                        s3.改派订单量 / s3.订单量 改派占比,
                        '1' 订单品类占比,
                        s3.销售额 / s3.订单量 客单价,
                        s3.销售额 / s3.广告成本 ROI,
                        S3.活跃产品数,
                        s3.订单量 / S3.活跃产品数 AS 产能,
                        s3.直发采购额 / s3.销售额 '直发采购占比',
                        s3.广告成本 / s3.销售额 '广告占比',
                        s3.物流成本 / s3.销售额 '运费占比',
                        s3.手续费 / s3.销售额 '手续费占比',
                        s3.广告成本 + s3.物流成本 + s3.手续费 + s3.直发采购额 / s3.销售额 '总成本占比',
                        s3.签收额 / (s3.拒收额 + s3.拒收额) '金额签收完成',
                        s3.签收额 / s3.拒收额 +s3.销售额 '金额签收总计',
                        (s3.签收额 + s3.拒收额) / s3.销售额 '金额完成占比',
                        s3.签收量 / (s3.拒收量 + s3.签收量) '数量签收完成',
                        s3.签收量 / s3.订单量 '数量签收总计',
                        (s3.拒收量 + s3.签收量) / s3.订单量 '数量完成占比',
                        s3.签收额 / (s3.签收额 + s3.拒收额) -( s3.直发采购额 + s3.广告成本 + s3.物流成本 + s3.手续费 ) / s3.销售额 AS 利润率,
                        (s3.签收额 / (s3.签收额 + s3.拒收额) -( s3.直发采购额 + s3.广告成本 + s3.物流成本 + s3.手续费 ) / s3.销售额) * (s3.销售额 / s3.订单量) AS 利润值
                FROM ( SELECT EXTRACT(YEAR_MONTH FROM a.rq) AS 年月,
                            b.pname AS 团队,
                            '合计' AS 品类,
                            SUM(a.salesRMB) 销售额,
                            SUM(a.orders) AS 订单量,
                            COUNT(DISTINCT a.product_id) AS 活跃产品数,
                            SUM(a.yqs) AS 签收量,
                            SUM(a.yjs) AS 拒收量,
                            SUM(a.salesRMB_yqs) AS 签收额,
                            SUM(a.salesRMB_yjs) AS 拒收额,
                            SUM(a.gps) AS 改派订单量,
                --          SUM(a.cgcost_zf) AS 直发采购额,
                --          SUM(a.adcost) AS 广告成本,
                            null 直发采购额,
                            null 广告成本,
                            SUM(a.wlcost) AS 物流成本,
                            SUM(a.qtcost) AS 手续费
                    FROM gk_order_day_kf a
                        LEFT JOIN dim_currency_lang b ON a.currency_lang_id = b.id
                        LEFT JOIN dim_area c on c.id = a.area_id
                        LEFT JOIN dim_cate d on d.id = a.third_cate_id
                        LEFT JOIN (SELECT * FROM gk_sale WHERE id IN (SELECT MAX(id) FROM gk_sale GROUP BY product_id ) ORDER BY id) e on e.product_id = a.product_id
                    WHERE a.rq >= '{0}'
                        AND a.rq < '{1}'
                        AND b.pcode = '{2}'
                        AND c.uname = '王冰'
                        AND a.beform <> 'mf'
                        AND c.uid <> 10099  -- 过滤翼虎
                    GROUP BY b.pname, c.uname
                ) s3
                ORDER BY 订单量;'''.format(last_month, yesterday, match[team])
        df = pd.read_sql_query(sql=sql, con=self.engine2)
        columns = list(df)
        columns = ', '.join(columns)  # 插入mysql的标题使用，否则无法导入更新
        print('正在缓存…………')
        df.to_sql('zg_cost_缓存', con=self.engine1, index=False, if_exists='replace')
        sql = 'INSERT IGNORE INTO zg_cost_sltem({}, 记录时间)  SELECT *, CURDATE() 记录时间  FROM zg_cost_缓存;'.format(columns)
        pd.read_sql_query(sql=sql, con=self.engine1, chunksize=100)
        print('已导入成本总表中+++')
        # print('正在查询 ' + team + ' 品类近两个月成本…………')
        # sql = '''SELECT *
        #         FROM (SELECT a.rq AS 年月,
        #                     b.pname AS 团队,
        #                     c.uname AS leader,
        #                     d.ppname AS 品类,
        #                     SUM(a.orders) AS 订单量,
        #                     COUNT(DISTINCT a.product_id) AS 活跃产品数,
        #                     SUM(a.yqs) AS 签收量,
        #                     SUM(a.yjs + a.yth) AS 拒收量,
        #                     SUM(a.salesRMB) AS 销售额,
        #                     SUM(a.salesRMB_yqs) AS 签收额,
        #                     SUM(a.salesRMB_yjs + a.salesRMB_yth) AS 拒收额,
        #                     SUM(a.gps) AS 改派订单量,
        #                     SUM(a.cgcost) AS 总采购额,
        #                     SUM(a.cgcost_zf) AS 直发采购额,
        #                     SUM(a.adcost) AS 广告成本,
        #                     SUM(a.wlcost) AS 物流成本,
        #                     SUM(a.qtcost) AS 手续费
        #             FROM gk_order_day a
        #                 LEFT JOIN dim_currency_lang b ON a.currency_lang_id = b.id
        #                 LEFT JOIN dim_area c on c.id = a.area_id
        #                 LEFT JOIN dim_cate d on d.id = a.third_cate_id
        #                 LEFT JOIN gk_product e on e.id = a.product_id
        #             WHERE a.rq >= '{0}'
        #                 AND a.rq <= '{1}'
        #                 AND b.pcode = '{2}'
        #                 AND c.uname = '王冰'
        #                 AND a.beform <> 'mf'
        #                 AND c.uid <> 10099  -- 过滤翼虎
        #             GROUP BY b.pname, c.uname, a.rq, a.cate_id
        #             UNION ALL
        #             SELECT '' AS 年月,
        #                     b.pname AS 团队,
        #                     c.uname AS leader,
        #                     EXTRACT(YEAR_MONTH FROM a.rq) AS 品类,
        #                     SUM(a.orders) AS 订单量,
        #                     COUNT(DISTINCT a.product_id) AS 活跃产品数,
        #                     SUM(a.yqs) AS 签收量,
        #                     SUM(a.yjs + a.yth) AS 拒收量,
        #                     SUM(a.salesRMB) AS 销售额,
        #                     SUM(a.salesRMB_yqs) AS 签收额,
        #                     SUM(a.salesRMB_yjs + a.salesRMB_yth) AS 拒收额,
        #                     SUM(a.gps) AS 改派订单量,
        #                     SUM(a.cgcost) AS 总采购额,
        #                     SUM(a.cgcost_zf) AS 直发采购额,
        #                     SUM(a.adcost) AS 广告成本,
        #                     SUM(a.wlcost) AS 物流成本,
        #                     SUM(a.qtcost) AS 手续费
        #             FROM gk_order_day a
        #                 LEFT JOIN dim_currency_lang b ON a.currency_lang_id = b.id
        #                 LEFT JOIN dim_area c on c.id = a.area_id
        #                 LEFT JOIN dim_cate d on d.id = a.third_cate_id
        #                 LEFT JOIN gk_product e on e.id = a.product_id
        #             WHERE a.rq >= '{0}'
        #                 AND a.rq <= '{1}'
        #                 AND b.pcode = '{2}'
        #                 AND c.uname = '王冰'
        #                 AND a.beform <> 'mf'
        #                 AND c.uid <> 10099  -- 过滤翼虎
        #             GROUP BY b.pname, c.uname,EXTRACT(YEAR_MONTH FROM a.rq)
        #             ) sl
        #             order by 年月;'''.format(last_month, yesterday, match[team])
        # df = pd.read_sql_query(sql=sql, con=self.engine2)
        # columns = list(df)
        # columns = ', '.join(columns)  # 插入mysql的标题使用，否则无法导入更新
        # print('正在缓存…………')
        # df.to_sql('zg_cost_缓存', con=self.engine1, index=False, if_exists='replace')
        # sql = 'INSERT IGNORE INTO zg_cost_sltem_copy({}, 记录时间)  SELECT *, CURDATE() 记录时间  FROM zg_cost_缓存;'.format(columns)
        # pd.read_sql_query(sql=sql, con=self.engine1, chunksize=100)
        # print('已导入' + team + '成本两月表中+++')
        print('成本耗时：', datetime.datetime.now() - start)
    def sl_Monitoring(self,team):
        match2 = {'新加坡': 'qsb_slxmt',
                  '马来西亚': 'qsb_slxmt',
                  '菲律宾': 'qsb_slxmt',
                  '日本': 'qsb_slrb',
                  '香港': 'qsb_slgat',
                  '台湾': 'qsb_slgat',
                  '泰国': 'qsb_sltg'}
        match3 = {'日本': r'D:\Users\Administrator\Desktop\查询\日本监控',
                  '泰国': r'D:\Users\Administrator\Desktop\查询\泰国监控',
                  '香港': r'D:\Users\Administrator\Desktop\查询\港台监控',
                  '台湾': r'D:\Users\Administrator\Desktop\查询\港台监控',
                  '菲律宾': r'D:\Users\Administrator\Desktop\查询\新马监控',
                  '新加坡': r'D:\Users\Administrator\Desktop\查询\新马监控',
                  '马来西亚': r'D:\Users\Administrator\Desktop\查询\新马监控'}
        emailAdd = {'香港': 'giikinliujun@163.com',
                    '台湾': 'giikinliujun@163.com',
                    '泰国': '1845389861@qq.com',
                    '新加坡': 'zhangjing@giikin.com',
                    '马来西亚': 'zhangjing@giikin.com',
                    '菲律宾': 'zhangjing@giikin.com',
                    '日本': 'sunyaru@giikin.com'}
        emailAdd2 = {'泰国': 'zhangjing@giikin.com'}
        # 筛选最大ID
        # print('正在创建临时监控数据源')
        # today = datetime.date.today().strftime('%Y.%m.%d')
        # sql = '''SELECT * FROM {0} WHERE id IN (SELECT MAX(id) FROM {0} GROUP BY 订单编号);'''.format(match2[team])
        # df = pd.read_sql_query(sql=sql, con=self.engine1)
        # df.to_sql('qsb_临时', con=self.engine1, index=False, if_exists='replace')
        start: datetime = datetime.datetime.now()
        Time_day = []
        for i in range(1, datetime.datetime.now().month + 1):  # 获取当年当前的月份时间
            try:
                daytime = (datetime.datetime.now().replace(month=i)).strftime('%Y-%m') + (
                    (datetime.datetime.now()).strftime('-%d'))
                Time_day.append(daytime)
            except Exception as e:
                print('xxxx时间配置出错,已手动调整：' + str(i) + '月份', str(Exception) + str(e))
                Time_day.append(
                    str(int(datetime.datetime.now().year)) + '-' + str(i) + (datetime.datetime.now().strftime('-%d')))
        for i in range(datetime.datetime.now().month + 1, 13):  # 获取往年当前的月份时间
            try:
                daytime = str(int(datetime.datetime.now().year) -1) + (datetime.datetime.now().replace(month=i)).strftime('-%m') + (
                    (datetime.datetime.now()).strftime('-%d'))
                Time_day.append(daytime)
            except Exception as e:
                print('xxxx时间配置出错失败00：' + str(i) + '月份', str(Exception) + str(e))
                Time_day.append(str(int(datetime.datetime.now().year) - 1) + '-' + str(i) + (
                    datetime.datetime.now().strftime('-%d')))
        # Time_day = ['2021-02-24', '2020-12-19', '2020-12-14', '2020-12-14', '2020-12-14', '2020-12-14', '2020-12-14', '2020-12-14', '2020-12-14', '2020-12-14', '2020-12-14', '2021-01-24']
        # 对时间数组进行排序  list.sort(cmp=None, key=None, reverse=False)；reverse -- 排序规则，reverse = True 降序， reverse = False 升序（默认）
        Time_day.sort()
        print(Time_day[11])
        print(Time_day[10])
        listT = []                              # 查询sql 存放池
        show_name = []                          # 打印进度需要
        # 月签收率（天）---查询
        sqlqsb2 = '''SELECT sl_gat.`币种`,sl_gat.`年月`,sl_gat.父级分类,sl_gat.二级分类,sl_gat.三级分类,sl_gat.物流方式,sl_gat.旬,
            	            sl_gat.`总订单量`,
							sl_gat.`已签收订单量` / sl_gat.`拒收订单量` AS '总签收/完成',
            	            sl_gat.`已签收订单量` / sl_gat.`总订单量` AS '总签收/总计',
							sl_gat.`退货订单量` / sl_gat.`总订单量` AS '退款率',
            	            sl_gat.`拒收订单量` / sl_gat.`总订单量` AS '总完成占比',
            	            sl_gat.`直发订单量` 直发总计,
							sl_gat.`直发已签收订单量` / sl_gat.`直发拒收订单量` AS '直发签收/完成',
            	            sl_gat.`直发已签收订单量` / sl_gat.`直发订单量` AS '直发签收/总计', 
							sl_gat.`直发拒收订单量` / sl_gat.`直发订单量` AS '直发完成占比',
            	            sl_gat.`改派订单量` 改派总计,
							sl_gat.`改派已签收订单量` / sl_gat.`改派拒收订单量` AS '改派签收/完成',
            	            sl_gat.`改派已签收订单量` / sl_gat.`改派订单量` AS '改派签收/总计',
							sl_gat.`改派拒收订单量` / sl_gat.`改派订单量` AS '改派完成占比'
					FROM (SELECT  sl_zong.币种,
					    IFNULL(sl_zong.年月,'合计') 年月,
						IFNULL(sl_zong.父级分类,'合计') 父级分类,
						IFNULL(sl_zong.二级分类,'合计') 二级分类,
                        IFNULL(sl_zong.三级分类,'合计') 三级分类,
                        IFNULL(sl_zong.物流方式,'合计') 物流方式,
						IFNULL(sl_zong.旬,'合计') 旬,
						SUM(总订单量) 总订单量,
                        IFNULL(SUM(直发订单量),0) 直发订单量,
                        IFNULL(SUM(直发已签收订单量),0) 直发已签收订单量,
                        IFNULL(SUM(直发拒收订单量),0) 直发拒收订单量,
                        (SUM(总订单量) - IFNULL(SUM(直发订单量),0)) AS 改派订单量,
                        IFNULL(SUM(改派已签收订单量),0) 改派已签收订单量,
                        IFNULL(SUM(改派拒收订单量),0) 改派拒收订单量,
						IFNULL(SUM(已签收订单量),0) 已签收订单量,
						IFNULL(SUM(拒收订单量),0) 拒收订单量,
						IFNULL(SUM(退货订单量),0) 退货订单量
            		FROM (SELECT  币种,年月,父级分类,二级分类,三级分类,物流方式,旬,COUNT(`订单编号`) 总订单量
            				FROM  {0}	sl_cx
            				WHERE  (sl_cx.`记录时间`= '{2}' AND (sl_cx.`日期` between DATE_SUB('{2}', INTERVAL DAY('{2}')-1 DAY) AND  DATE_SUB('{2}',INTERVAL 1 DAY))
								    OR sl_cx.`记录时间`= '{3}' AND (sl_cx.`日期` between DATE_SUB('{3}', INTERVAL DAY('{3}')-1 DAY) AND  DATE_SUB('{3}',INTERVAL 1 DAY)))
								AND sl_cx.`币种` = '{1}' 
								AND sl_cx.`父级分类` IS NOT NULL  
								AND sl_cx.`系统订单状态` IN {4} 
            				GROUP BY {5}
            				ORDER BY 币种,年月
            				) sl_zong
            		LEFT JOIN
            				(SELECT 币种,年月,父级分类,二级分类,三级分类,物流方式,旬,COUNT(`订单编号`) 直发订单量
            				FROM  {0}	sl_zong_zf
            				WHERE (sl_zong_zf.`记录时间`= '{2}' AND (sl_zong_zf.`日期` between DATE_SUB('{2}', INTERVAL DAY('{2}')-1 DAY) AND  DATE_SUB('{2}',INTERVAL 1 DAY))
								    OR sl_zong_zf.`记录时间`= '{3}' AND (sl_zong_zf.`日期` between DATE_SUB('{3}', INTERVAL DAY('{3}')-1 DAY) AND  DATE_SUB('{3}',INTERVAL 1 DAY)))
							   AND sl_zong_zf.`币种` = '{1}' 
							   AND sl_zong_zf.`父级分类` IS NOT NULL
            				   AND sl_zong_zf.`是否改派` = "直发"
            				   AND sl_zong_zf.`系统订单状态` IN {4} 
            				GROUP BY {5}
            				ORDER BY 币种,年月
            				) sl_zong_zf 
            		ON sl_zong_zf.`币种` = sl_zong.`币种` AND sl_zong_zf.`年月` = sl_zong.`年月`AND sl_zong_zf.`父级分类` = sl_zong.`父级分类` 
            			AND sl_zong_zf.`二级分类` = sl_zong.`二级分类` AND sl_zong_zf.`三级分类` = sl_zong.`三级分类` 
            			AND sl_zong_zf.`物流方式` = sl_zong.`物流方式` AND sl_zong_zf.`旬` = sl_zong.`旬` 							
            		LEFT JOIN
            				(SELECT 币种,年月,父级分类,二级分类,三级分类,物流方式,旬,COUNT(`订单编号`) 直发已签收订单量
            				FROM  {0}	slzf_qs
            				WHERE (slzf_qs.`记录时间`= '{2}' AND (slzf_qs.`日期` between DATE_SUB('{2}', INTERVAL DAY('{2}')-1 DAY) AND  DATE_SUB('{2}',INTERVAL 1 DAY))
								    OR slzf_qs.`记录时间`= '{3}' AND (slzf_qs.`日期` between DATE_SUB('{3}', INTERVAL DAY('{3}')-1 DAY) AND  DATE_SUB('{3}',INTERVAL 1 DAY)))
								AND slzf_qs.`币种` = '{1}' 
								AND slzf_qs.`父级分类` IS NOT NULL
								AND slzf_qs.`是否改派` = "直发" 
								AND slzf_qs.`最终状态` = "已签收" 
								AND slzf_qs.`系统订单状态` IN {4} 
            				GROUP BY {5}
            				ORDER BY 币种,年月
            				) sl_cx_zf_qs
            		ON sl_cx_zf_qs.`币种` = sl_zong.`币种` AND sl_cx_zf_qs.`年月` = sl_zong.`年月` AND sl_cx_zf_qs.`父级分类` = sl_zong.`父级分类` 
            			AND sl_cx_zf_qs.`二级分类` = sl_zong.`二级分类` AND sl_cx_zf_qs.`三级分类` = sl_zong.`三级分类` 
            			AND sl_cx_zf_qs.`物流方式` = sl_zong.`物流方式` AND sl_cx_zf_qs.`旬` = sl_zong.`旬` 		
            		LEFT JOIN
            				(SELECT 币种,年月,父级分类,二级分类,三级分类,物流方式,旬,COUNT(`订单编号`) 直发拒收订单量
            				FROM  {0}	slzf_js
            				WHERE (slzf_js.`记录时间`= '{2}' AND (slzf_js.`日期` between DATE_SUB('{2}', INTERVAL DAY('{2}')-1 DAY) AND  DATE_SUB('{2}',INTERVAL 1 DAY))
								    OR slzf_js.`记录时间`= '{3}' AND (slzf_js.`日期` between DATE_SUB('{3}', INTERVAL DAY('{3}')-1 DAY) AND  DATE_SUB('{3}',INTERVAL 1 DAY)))
								AND slzf_js.`币种` = '{1}' 
								AND slzf_js.`父级分类` IS NOT NULL
            					AND slzf_js.`是否改派` = "直发"  
            					AND slzf_js.`最终状态` IN ('拒收', '理赔', '已签收', '已退货') 
            					AND slzf_js.`系统订单状态`IN {4}
            				GROUP BY {5}
            				ORDER BY 币种,年月
            				) sl_cx_zf_js
            		ON sl_cx_zf_js.`币种` = sl_zong.`币种` AND sl_cx_zf_js.`年月` = sl_zong.`年月` AND sl_cx_zf_js.`父级分类` = sl_zong.`父级分类` 
            			AND sl_cx_zf_js.`二级分类` = sl_zong.`二级分类` AND sl_cx_zf_js.`三级分类` = sl_zong.`三级分类` 
            			AND sl_cx_zf_js.`物流方式` = sl_zong.`物流方式` AND sl_cx_zf_js.`旬` = sl_zong.`旬` 
            		LEFT JOIN
            				(SELECT 币种,年月,父级分类,二级分类,三级分类,物流方式,旬,COUNT(`订单编号`) 改派订单量
            				FROM  {0}	sl_cx_gp
            				WHERE (sl_cx_gp.`记录时间`= '{2}' AND (sl_cx_gp.`日期` between DATE_SUB('{2}', INTERVAL DAY('{2}')-1 DAY) AND  DATE_SUB('{2}',INTERVAL 1 DAY))
								 OR sl_cx_gp.`记录时间`= '{3}' AND (sl_cx_gp.`日期` between DATE_SUB('{3}', INTERVAL DAY('{3}')-1 DAY) AND  DATE_SUB('{3}',INTERVAL 1 DAY)))
								AND sl_cx_gp.`币种` = '{1}' 
								AND sl_cx_gp.`父级分类` IS NOT NULL
            					AND sl_cx_gp.`是否改派` = "改派"
								AND sl_cx_gp.`系统订单状态` IN {4} 
            				GROUP BY {5}
            				ORDER BY 币种,年月
            				) sl_zong_gp
            		ON sl_zong_gp.`币种` = sl_zong.`币种` AND sl_zong_gp.`年月` = sl_zong.`年月` AND sl_zong_gp.`父级分类` = sl_zong.`父级分类` 
            			AND sl_zong_gp.`二级分类` = sl_zong.`二级分类` AND sl_zong_gp.`三级分类` = sl_zong.`三级分类` 
            			AND sl_zong_gp.`物流方式` = sl_zong.`物流方式` AND sl_zong_gp.`旬` = sl_zong.`旬` 
            		LEFT JOIN
            				(SELECT 币种,年月,父级分类,二级分类,三级分类,物流方式,旬,COUNT(`订单编号`) 改派已签收订单量
            				FROM  {0}	slgp_qs
            				WHERE (slgp_qs.`记录时间`= '{2}' AND (slgp_qs.`日期` between DATE_SUB('{2}', INTERVAL DAY('{2}')-1 DAY) AND  DATE_SUB('{2}',INTERVAL 1 DAY))
								 OR slgp_qs.`记录时间`= '{3}' AND (slgp_qs.`日期` between DATE_SUB('{3}', INTERVAL DAY('{3}')-1 DAY) AND  DATE_SUB('{3}',INTERVAL 1 DAY)))
								AND slgp_qs.`币种` = '{1}' 
								AND slgp_qs.`父级分类` IS NOT NULL
            					AND slgp_qs.`是否改派` = "改派"  
            					AND slgp_qs.`最终状态` = "已签收"
            					AND slgp_qs.`系统订单状态`IN {4}
            				GROUP BY {5}
            				ORDER BY 币种,年月
            				) sl_cx_gp_qs
            			ON sl_cx_gp_qs.`币种` = sl_zong.`币种` AND sl_cx_gp_qs.`年月` = sl_zong.`年月` AND sl_cx_gp_qs.`父级分类` = sl_zong.`父级分类` 
            				AND sl_cx_gp_qs.`二级分类` = sl_zong.`二级分类` AND sl_cx_gp_qs.`三级分类` = sl_zong.`三级分类` 
            				AND sl_cx_gp_qs.`物流方式` = sl_zong.`物流方式` AND sl_cx_gp_qs.`旬` = sl_zong.`旬` 
            		LEFT JOIN
            				(SELECT 币种,年月,父级分类,二级分类,三级分类,物流方式,旬,COUNT(`订单编号`) 改派拒收订单量
            				FROM  {0}	sl_cx_gp_jushou
            				WHERE (sl_cx_gp_jushou.`记录时间`= '{2}' AND (sl_cx_gp_jushou.`日期` between DATE_SUB('{2}', INTERVAL DAY('{2}')-1 DAY) AND  DATE_SUB('{2}',INTERVAL 1 DAY))
									OR sl_cx_gp_jushou.`记录时间`= '{3}' AND (sl_cx_gp_jushou.`日期` between DATE_SUB('{3}', INTERVAL DAY('{3}')-1 DAY) AND  DATE_SUB('{3}',INTERVAL 1 DAY)))
								AND sl_cx_gp_jushou.`币种` = '{1}' 
								AND sl_cx_gp_jushou.`父级分类` IS NOT NULL
            					AND sl_cx_gp_jushou.`是否改派` = "改派"  
            					AND sl_cx_gp_jushou.`最终状态` IN ('拒收', '理赔', '已签收', '已退货') 
								AND sl_cx_gp_jushou.`系统订单状态` IN {4}
            				GROUP BY {5}
            				ORDER BY 币种,年月
            				) sl_cx_gp_js
            			ON sl_cx_gp_js.`币种` = sl_zong.`币种` AND sl_cx_gp_js.`年月` = sl_zong.`年月` AND sl_cx_gp_js.`父级分类` = sl_zong.`父级分类` 
            				AND sl_cx_gp_js.`二级分类` = sl_zong.`二级分类` AND sl_cx_gp_js.`三级分类` = sl_zong.`三级分类` 
            				AND sl_cx_gp_js.`物流方式` = sl_zong.`物流方式` AND sl_cx_gp_js.`旬` = sl_zong.`旬` 				
            		LEFT JOIN
            				(SELECT 币种,年月,父级分类,二级分类,三级分类,物流方式,旬,COUNT(`订单编号`) 已签收订单量
            				FROM  {0}	sl_cx_qianshou
            				WHERE (sl_cx_qianshou.`记录时间`= '{2}' AND (sl_cx_qianshou.`日期` between DATE_SUB('{2}', INTERVAL DAY('{2}')-1 DAY) AND  DATE_SUB('{2}',INTERVAL 1 DAY))
									OR sl_cx_qianshou.`记录时间`= '{3}' AND (sl_cx_qianshou.`日期` between DATE_SUB('{3}', INTERVAL DAY('{3}')-1 DAY) AND  DATE_SUB('{3}',INTERVAL 1 DAY)))
								AND sl_cx_qianshou.`币种` = '{1}' 
								AND sl_cx_qianshou.`父级分类` IS NOT NULL
            					AND sl_cx_qianshou.`最终状态` = "已签收"
								AND sl_cx_qianshou.`系统订单状态`IN {4} 
            				GROUP BY {5}
            				ORDER BY 币种,年月
            				) sl_zong_qianshou
            			ON sl_zong_qianshou.`币种` = sl_zong.`币种` AND sl_zong_qianshou.`年月` = sl_zong.`年月` AND sl_zong_qianshou.`父级分类` = sl_zong.`父级分类` 
            				AND sl_zong_qianshou.`二级分类` = sl_zong.`二级分类`  AND sl_zong_qianshou.`三级分类` = sl_zong.`三级分类` 
            				AND sl_zong_qianshou.`物流方式` = sl_zong.`物流方式` AND sl_zong_qianshou.`旬` = sl_zong.`旬`
            		LEFT JOIN
            				(SELECT 币种,年月,父级分类,二级分类,三级分类,物流方式,旬,COUNT(`订单编号`) 拒收订单量
            				FROM  {0}	sl_cx_jushou
            				WHERE (sl_cx_jushou.`记录时间`= '{2}' AND (sl_cx_jushou.`日期` between DATE_SUB('{2}', INTERVAL DAY('{2}')-1 DAY) AND  DATE_SUB('{2}',INTERVAL 1 DAY))
									OR sl_cx_jushou.`记录时间`= '{3}' AND (sl_cx_jushou.`日期` between DATE_SUB('{3}', INTERVAL DAY('{3}')-1 DAY) AND  DATE_SUB('{3}',INTERVAL 1 DAY)))
							    AND sl_cx_jushou.`币种` = '{1}' 
								AND sl_cx_jushou.`父级分类` IS NOT NULL
            					AND sl_cx_jushou.`最终状态` IN ('拒收', '理赔', '已签收', '已退货') 
								AND sl_cx_jushou.`系统订单状态` IN {4} 
            				GROUP BY {5}
            				ORDER BY 币种,年月
            				) sl_zong_jushou
            			ON sl_zong_jushou.`币种` = sl_zong.`币种` AND sl_zong_jushou.`年月` = sl_zong.`年月` AND sl_zong_jushou.`父级分类` = sl_zong.`父级分类` 
            				AND sl_zong_jushou.`二级分类` = sl_zong.`二级分类` AND sl_zong_jushou.`三级分类` = sl_zong.`三级分类` 
            				AND sl_zong_jushou.`物流方式` = sl_zong.`物流方式` AND sl_zong_jushou.`旬` = sl_zong.`旬` 
            		LEFT JOIN
            				(SELECT 币种,年月,父级分类,二级分类,三级分类,物流方式,旬,COUNT(`订单编号`) 退货订单量
            				FROM  {0}	sl_cx_tuihuo
            				WHERE (sl_cx_tuihuo.`记录时间`= '{2}' AND (sl_cx_tuihuo.`日期` between DATE_SUB('{2}', INTERVAL DAY('{2}')-1 DAY) AND  DATE_SUB('{2}',INTERVAL 1 DAY))
									OR sl_cx_tuihuo.`记录时间`= '{3}' AND (sl_cx_tuihuo.`日期` between DATE_SUB('{3}', INTERVAL DAY('{3}')-1 DAY) AND  DATE_SUB('{3}',INTERVAL 1 DAY)))
								AND sl_cx_tuihuo.`币种` = '{1}' 
								AND sl_cx_tuihuo.`父级分类` IS NOT NULL
            					AND sl_cx_tuihuo.`最终状态` = "已退货"
								AND sl_cx_tuihuo.`系统订单状态`IN {4} 
            				GROUP BY {5}
            				ORDER BY 币种,年月
            				) sl_zong_tuihuo
            			ON sl_zong_tuihuo.`币种` = sl_zong.`币种` AND sl_zong_tuihuo.`年月` = sl_zong.`年月` AND sl_zong_tuihuo.`父级分类` = sl_zong.`父级分类` 
            				AND sl_zong_tuihuo.`二级分类` = sl_zong.`二级分类` AND sl_zong_tuihuo.`三级分类` = sl_zong.`三级分类` 
            				AND sl_zong_tuihuo.`物流方式` = sl_zong.`物流方式`AND sl_zong_tuihuo.`旬` = sl_zong.`旬` 
            		GROUP BY sl_zong.年月,sl_zong.父级分类,sl_zong.二级分类,sl_zong.三级分类,sl_zong.物流方式,sl_zong.旬
            		with rollup) sl_gat;'''.format(match2[team], team, Time_day[11], Time_day[10], ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)'), '币种,年月,父级分类,二级分类,三级分类,物流方式,旬')
        listT.append(sqlqsb2)
        show_name.append(' 月（天）签收率_…………')
        # 月签收率（整月）---查询
        sqlqsb3 = '''SELECT sl_gat.`币种`,sl_gat.`年月`,sl_gat.父级分类,sl_gat.二级分类,sl_gat.三级分类,sl_gat.物流方式,sl_gat.旬,
            				sl_gat.`总订单量`,
							sl_gat.`已签收订单量` / sl_gat.`拒收订单量` AS '总签收/完成',
            				sl_gat.`已签收订单量` / sl_gat.`总订单量` AS '总签收/总计',
							sl_gat.`退货订单量` / sl_gat.`总订单量` AS '退款率',
            				sl_gat.`拒收订单量` / sl_gat.`总订单量` AS '总完成占比',
            				sl_gat.`直发订单量` 直发总计,
							sl_gat.`直发已签收订单量` / sl_gat.`直发拒收订单量` AS '直发签收/完成',
            				sl_gat.`直发已签收订单量` / sl_gat.`直发订单量` AS '直发签收/总计', 
							sl_gat.`直发拒收订单量` / sl_gat.`直发订单量` AS '直发完成占比',
            				sl_gat.`改派订单量` 改派总计,
							sl_gat.`改派已签收订单量` / sl_gat.`改派拒收订单量` AS '改派签收/完成',
            				sl_gat.`改派已签收订单量` / sl_gat.`改派订单量` AS '改派签收/总计',
							sl_gat.`改派拒收订单量` / sl_gat.`改派订单量` AS '改派完成占比'							
            FROM (SELECT  sl_zong.币种,
                        IFNULL(sl_zong.年月,'合计') 年月,
                        IFNULL(sl_zong.父级分类,'合计') 父级分类,
                        IFNULL(sl_zong.二级分类,'合计') 二级分类,
                        IFNULL(sl_zong.三级分类,'合计') 三级分类,
                        IFNULL(sl_zong.物流方式,'合计') 物流方式,
                        IFNULL(sl_zong.旬,'合计') 旬,
                        SUM(总订单量) 总订单量,
            			IFNULL(SUM(直发订单量),0) 直发订单量,
						IFNULL(SUM(直发已签收订单量),0) 直发已签收订单量,
						IFNULL(SUM(直发拒收订单量),0) 直发拒收订单量,
            			(SUM(总订单量) - IFNULL(SUM(直发订单量),0)) AS 改派订单量,
						IFNULL(SUM(改派已签收订单量),0) 改派已签收订单量,
						IFNULL(SUM(改派拒收订单量),0) 改派拒收订单量,
						IFNULL(SUM(已签收订单量),0) 已签收订单量,
						IFNULL(SUM(拒收订单量),0) 拒收订单量,
						IFNULL(SUM(退货订单量),0) 退货订单量
            		FROM (SELECT  币种,年月,父级分类,二级分类,三级分类,物流方式,旬,COUNT(`订单编号`) 总订单量
            				FROM  {0}	sl_cx
            				WHERE (sl_cx.`记录时间`= '{2}' AND sl_cx.`日期` < DATE_SUB('{2}', INTERVAL DAY('{2}')-1 DAY) 
									OR sl_cx.`记录时间`= '{3}' AND sl_cx.`日期` < DATE_SUB('{3}', INTERVAL DAY('{3}')-1 DAY) )
								AND sl_cx.`币种` = '{1}' AND sl_cx.`父级分类` IS NOT NULL
								AND sl_cx.`系统订单状态` IN {4} 
            				GROUP BY {5} ORDER BY 币种,年月
            				) sl_zong
            		LEFT JOIN
            				(SELECT 币种,年月,父级分类,二级分类,三级分类,物流方式,旬,COUNT(`订单编号`) 直发订单量
            				FROM  {0}	sl_zong_zf
            				WHERE (sl_zong_zf.`记录时间`= '{2}' AND sl_zong_zf.`日期` < DATE_SUB('{2}', INTERVAL DAY('{2}')-1 DAY) 
									OR sl_zong_zf.`记录时间`= '{3}' AND sl_zong_zf.`日期` < DATE_SUB('{3}', INTERVAL DAY('{3}')-1 DAY))
								AND sl_zong_zf.`币种` = '{1}' AND sl_zong_zf.`父级分类` IS NOT NULL
            		AND sl_zong_zf.`是否改派` = "直发"
								AND sl_zong_zf.`系统订单状态` IN {4}
            				GROUP BY {5} ORDER BY 币种,年月
            				) sl_zong_zf 
            		ON sl_zong_zf.`币种` = sl_zong.`币种` AND sl_zong_zf.`年月` = sl_zong.`年月`AND sl_zong_zf.`父级分类` = sl_zong.`父级分类` 
            			AND sl_zong_zf.`二级分类` = sl_zong.`二级分类` AND sl_zong_zf.`三级分类` = sl_zong.`三级分类` 
            			AND sl_zong_zf.`物流方式` = sl_zong.`物流方式` AND sl_zong_zf.`旬` = sl_zong.`旬` 							
            		LEFT JOIN
            				(SELECT 币种,年月,父级分类,二级分类,三级分类,物流方式,旬,COUNT(`订单编号`) 直发已签收订单量
            				FROM  {0}	slzf_qs
            				WHERE (slzf_qs.`记录时间`= '{2}' AND slzf_qs.`日期` < DATE_SUB('{2}', INTERVAL DAY('{2}')-1 DAY) 
									OR slzf_qs.`记录时间`= '{3}' AND slzf_qs.`日期` < DATE_SUB('{3}', INTERVAL DAY('{3}')-1 DAY) )
								AND slzf_qs.`币种` = '{1}' AND slzf_qs.`父级分类` IS NOT NULL
            					AND slzf_qs.`是否改派` = "直发" AND slzf_qs.`最终状态` = "已签收"
								AND slzf_qs.`系统订单状态` IN {4}
            				GROUP BY {5} ORDER BY 币种,年月
            				) sl_cx_zf_qs
            		ON sl_cx_zf_qs.`币种` = sl_zong.`币种` AND sl_cx_zf_qs.`年月` = sl_zong.`年月` AND sl_cx_zf_qs.`父级分类` = sl_zong.`父级分类` 
            			AND sl_cx_zf_qs.`二级分类` = sl_zong.`二级分类` AND sl_cx_zf_qs.`三级分类` = sl_zong.`三级分类` 
            			AND sl_cx_zf_qs.`物流方式` = sl_zong.`物流方式` AND sl_cx_zf_qs.`旬` = sl_zong.`旬` 		
            		LEFT JOIN
            				(SELECT 币种,年月,父级分类,二级分类,三级分类,物流方式,旬,COUNT(`订单编号`) 直发拒收订单量
            				FROM  {0}	slzf_js
            				WHERE (slzf_js.`记录时间`= '{2}' AND slzf_js.`日期` < DATE_SUB('{2}', INTERVAL DAY('{2}')-1 DAY) 
									OR slzf_js.`记录时间`= '{3}' AND slzf_js.`日期` < DATE_SUB('{3}', INTERVAL DAY('{3}')-1 DAY))
								AND slzf_js.`币种` = '{1}' AND slzf_js.`父级分类` IS NOT NULL
            					AND slzf_js.`是否改派` = "直发" AND slzf_js.`最终状态` IN ('拒收', '理赔', '已签收', '已退货') 
								AND slzf_js.`系统订单状态`IN {4}
            				GROUP BY {5} ORDER BY 币种,年月
            				) sl_cx_zf_js
            		ON sl_cx_zf_js.`币种` = sl_zong.`币种` AND sl_cx_zf_js.`年月` = sl_zong.`年月` AND sl_cx_zf_js.`父级分类` = sl_zong.`父级分类` 
            			AND sl_cx_zf_js.`二级分类` = sl_zong.`二级分类` AND sl_cx_zf_js.`三级分类` = sl_zong.`三级分类` 
            			AND sl_cx_zf_js.`物流方式` = sl_zong.`物流方式` AND sl_cx_zf_js.`旬` = sl_zong.`旬` 
            		LEFT JOIN
            				(SELECT 币种,年月,父级分类,二级分类,三级分类,物流方式,旬,COUNT(`订单编号`) 改派订单量
            				FROM  {0}	sl_cx_gp
            				WHERE (sl_cx_gp.`记录时间`= '{2}' AND sl_cx_gp.`日期` < DATE_SUB('{2}', INTERVAL DAY('{2}')-1 DAY) 
									OR sl_cx_gp.`记录时间`= '{3}' AND sl_cx_gp.`日期` < DATE_SUB('{3}', INTERVAL DAY('{3}')-1 DAY))
								AND sl_cx_gp.`币种` = '{1}' AND sl_cx_gp.`父级分类` IS NOT NULL
            					AND sl_cx_gp.`是否改派` = "改派" AND sl_cx_gp.`系统订单状态` IN {4}
            				GROUP BY {5} ORDER BY 币种,年月
            				) sl_zong_gp
            		ON sl_zong_gp.`币种` = sl_zong.`币种` AND sl_zong_gp.`年月` = sl_zong.`年月` AND sl_zong_gp.`父级分类` = sl_zong.`父级分类` 
            			AND sl_zong_gp.`二级分类` = sl_zong.`二级分类` AND sl_zong_gp.`三级分类` = sl_zong.`三级分类` 
            			AND sl_zong_gp.`物流方式` = sl_zong.`物流方式` AND sl_zong_gp.`旬` = sl_zong.`旬` 
            		LEFT JOIN
            				(SELECT 币种,年月,父级分类,二级分类,三级分类,物流方式,旬,COUNT(`订单编号`) 改派已签收订单量
            				FROM  {0}	slgp_qs
            				WHERE (slgp_qs.`记录时间`= '{2}' AND slgp_qs.`日期` < DATE_SUB('{2}', INTERVAL DAY('{2}')-1 DAY) 
									OR slgp_qs.`记录时间`= '{3}' AND slgp_qs.`日期` < DATE_SUB('{3}', INTERVAL DAY('{3}')-1 DAY))
								AND slgp_qs.`币种` = '{1}' AND slgp_qs.`父级分类` IS NOT NULL
            					AND slgp_qs.`是否改派` = "改派"  AND slgp_qs.`最终状态` = "已签收"
								AND slgp_qs.`系统订单状态`IN {4}
            				GROUP BY {5} ORDER BY 币种,年月
            				) sl_cx_gp_qs
            			ON sl_cx_gp_qs.`币种` = sl_zong.`币种` AND sl_cx_gp_qs.`年月` = sl_zong.`年月` AND sl_cx_gp_qs.`父级分类` = sl_zong.`父级分类` 
            				AND sl_cx_gp_qs.`二级分类` = sl_zong.`二级分类` AND sl_cx_gp_qs.`三级分类` = sl_zong.`三级分类` 
            				AND sl_cx_gp_qs.`物流方式` = sl_zong.`物流方式` AND sl_cx_gp_qs.`旬` = sl_zong.`旬` 
            		LEFT JOIN
            				(SELECT 币种,年月,父级分类,二级分类,三级分类,物流方式,旬,COUNT(`订单编号`) 改派拒收订单量
            				FROM  {0}	sl_cx_gp_jushou
            				WHERE (sl_cx_gp_jushou.`记录时间`= '{2}' AND sl_cx_gp_jushou.`日期` < DATE_SUB('{2}', INTERVAL DAY('{2}')-1 DAY) 
									OR sl_cx_gp_jushou.`记录时间`= '{3}' AND sl_cx_gp_jushou.`日期` < DATE_SUB('{3}', INTERVAL DAY('{3}')-1 DAY))
								AND sl_cx_gp_jushou.`币种` = '{1}' AND sl_cx_gp_jushou.`父级分类` IS NOT NULL
            					AND sl_cx_gp_jushou.`是否改派` = "改派"  AND sl_cx_gp_jushou.`最终状态` IN ('拒收', '理赔', '已签收', '已退货') 
								AND sl_cx_gp_jushou.`系统订单状态` IN {4} 
            				GROUP BY {5} ORDER BY 币种,年月
            				) sl_cx_gp_js
            			ON sl_cx_gp_js.`币种` = sl_zong.`币种` AND sl_cx_gp_js.`年月` = sl_zong.`年月` AND sl_cx_gp_js.`父级分类` = sl_zong.`父级分类` 
            				AND sl_cx_gp_js.`二级分类` = sl_zong.`二级分类` AND sl_cx_gp_js.`三级分类` = sl_zong.`三级分类` 
            				AND sl_cx_gp_js.`物流方式` = sl_zong.`物流方式` AND sl_cx_gp_js.`旬` = sl_zong.`旬` 				
            		LEFT JOIN
            				(SELECT 币种,年月,父级分类,二级分类,三级分类,物流方式,旬,COUNT(`订单编号`) 已签收订单量
            				FROM  {0}	sl_cx_qianshou
            				WHERE (sl_cx_qianshou.`记录时间`= '{2}' AND sl_cx_qianshou.`日期` < DATE_SUB('{2}', INTERVAL DAY('{2}')-1 DAY) 
								   OR sl_cx_qianshou.`记录时间`= '{3}' AND sl_cx_qianshou.`日期` < DATE_SUB('{3}', INTERVAL DAY('{3}')-1 DAY))
								AND sl_cx_qianshou.`币种` = '{1}' AND sl_cx_qianshou.`父级分类` IS NOT NULL
            					AND sl_cx_qianshou.`最终状态` = "已签收"  AND sl_cx_qianshou.`系统订单状态`IN {4}
            				GROUP BY {5} ORDER BY 币种,年月
            				) sl_zong_qianshou
            			ON sl_zong_qianshou.`币种` = sl_zong.`币种` AND sl_zong_qianshou.`年月` = sl_zong.`年月` AND sl_zong_qianshou.`父级分类` = sl_zong.`父级分类` 
            				AND sl_zong_qianshou.`二级分类` = sl_zong.`二级分类`  AND sl_zong_qianshou.`三级分类` = sl_zong.`三级分类` 
            				AND sl_zong_qianshou.`物流方式` = sl_zong.`物流方式` AND sl_zong_qianshou.`旬` = sl_zong.`旬`
            		LEFT JOIN
            				(SELECT 币种,年月,父级分类,二级分类,三级分类,物流方式,旬,COUNT(`订单编号`) 拒收订单量
            				FROM  {0}	sl_cx_jushou
            				WHERE (sl_cx_jushou.`记录时间`= '{2}' AND sl_cx_jushou.`日期` < DATE_SUB('{2}', INTERVAL DAY('{2}')-1 DAY) 
									OR sl_cx_jushou.`记录时间`= '{3}' AND sl_cx_jushou.`日期` < DATE_SUB('{3}', INTERVAL DAY('{3}')-1 DAY))
								AND sl_cx_jushou.`币种` = '{1}' AND sl_cx_jushou.`父级分类` IS NOT NULL
            					AND sl_cx_jushou.`最终状态` IN ('拒收', '理赔', '已签收', '已退货')  AND sl_cx_jushou.`系统订单状态` IN {4}
            				GROUP BY {5} ORDER BY 币种,年月
            				) sl_zong_jushou
            			ON sl_zong_jushou.`币种` = sl_zong.`币种` AND sl_zong_jushou.`年月` = sl_zong.`年月` AND sl_zong_jushou.`父级分类` = sl_zong.`父级分类` 
            				AND sl_zong_jushou.`二级分类` = sl_zong.`二级分类` AND sl_zong_jushou.`三级分类` = sl_zong.`三级分类` 
            				AND sl_zong_jushou.`物流方式` = sl_zong.`物流方式` AND sl_zong_jushou.`旬` = sl_zong.`旬` 
            		LEFT JOIN
            				(SELECT 币种,年月,父级分类,二级分类,三级分类,物流方式,旬,COUNT(`订单编号`) 退货订单量
            				FROM  {0}	sl_cx_tuihuo
            				WHERE (sl_cx_tuihuo.`记录时间`= '{2}' AND sl_cx_tuihuo.`日期` < DATE_SUB('{2}', INTERVAL DAY('{2}')-1 DAY) 
									OR sl_cx_tuihuo.`记录时间`= '{3}' AND sl_cx_tuihuo.`日期` < DATE_SUB('{3}', INTERVAL DAY('{3}')-1 DAY))
								AND sl_cx_tuihuo.`币种` = '{1}' AND sl_cx_tuihuo.`父级分类` IS NOT NULL
            					AND sl_cx_tuihuo.`最终状态` = "已退货" AND sl_cx_tuihuo.`系统订单状态`IN {4}
            				GROUP BY {5} ORDER BY 币种,年月
            				) sl_zong_tuihuo
            			ON sl_zong_tuihuo.`币种` = sl_zong.`币种` AND sl_zong_tuihuo.`年月` = sl_zong.`年月` AND sl_zong_tuihuo.`父级分类` = sl_zong.`父级分类` 
            				AND sl_zong_tuihuo.`二级分类` = sl_zong.`二级分类` AND sl_zong_tuihuo.`三级分类` = sl_zong.`三级分类` 
            				AND sl_zong_tuihuo.`物流方式` = sl_zong.`物流方式`AND sl_zong_tuihuo.`旬` = sl_zong.`旬` 
            		GROUP BY sl_zong.年月,sl_zong.父级分类,sl_zong.二级分类,sl_zong.三级分类,sl_zong.物流方式,sl_zong.旬
            		with rollup) sl_gat;'''.format(match2[team], team, Time_day[11], Time_day[10], ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)'), '币种,年月,父级分类,二级分类,三级分类,物流方式,旬')
        listT.append(sqlqsb3)
        show_name.append(' 月（月）签收率_…………')
        # 月签收率（旬）---查询
        sqlqsb4 = '''SELECT sl_gat.`年月`,sl_gat.旬,sl_gat.`币种`,sl_gat.父级分类,sl_gat.二级分类,sl_gat.三级分类,sl_gat.物流方式,
				            sl_gat.`总订单量`,
				            sl_gat.`已签收订单量` / sl_gat.`拒收订单量` AS '总签收/完成',
				            sl_gat.`已签收订单量` / sl_gat.`总订单量` AS '总签收/总计',
				            sl_gat.`退货订单量` / sl_gat.`总订单量` AS '退款率',
				            sl_gat.`拒收订单量` / sl_gat.`总订单量` AS '总完成占比',
				            sl_gat.`直发订单量` 直发总计,
				            sl_gat.`直发已签收订单量` / sl_gat.`直发拒收订单量` AS '直发签收/完成',
				            sl_gat.`直发已签收订单量` / sl_gat.`直发订单量` AS '直发签收/总计',
                            sl_gat.`直发拒收订单量` / sl_gat.`直发订单量` AS '直发完成占比',
				            sl_gat.`改派订单量` 改派总计,
				            sl_gat.`改派已签收订单量` / sl_gat.`改派拒收订单量` AS '改派签收/完成',
				            sl_gat.`改派已签收订单量` / sl_gat.`改派订单量` AS '改派签收/总计',
				            sl_gat.`改派拒收订单量` / sl_gat.`改派订单量` AS '改派完成占比'
            FROM (SELECT  sl_zong.币种,
                        IFNULL(sl_zong.年月,'合计') 年月,
                        IFNULL(sl_zong.旬,'合计') 旬,
                        IFNULL(sl_zong.父级分类,'合计') 父级分类,
                        IFNULL(sl_zong.二级分类,'合计') 二级分类,
						IFNULL(sl_zong.三级分类,'合计') 三级分类,
						IFNULL(sl_zong.物流方式,'合计') 物流方式,
						SUM(总订单量) 总订单量,
						IFNULL(SUM(直发订单量),0) 直发订单量,
						IFNULL(SUM(直发已签收订单量),0) 直发已签收订单量,
						IFNULL(SUM(直发拒收订单量),0) 直发拒收订单量,
						(SUM(总订单量) - IFNULL(SUM(直发订单量),0)) AS 改派订单量,
						IFNULL(SUM(改派已签收订单量),0) 改派已签收订单量,
						IFNULL(SUM(改派拒收订单量),0) 改派拒收订单量,
						IFNULL(SUM(已签收订单量),0) 已签收订单量,
						IFNULL(SUM(拒收订单量),0) 拒收订单量,
						IFNULL(SUM(退货订单量),0) 退货订单量
		    FROM (SELECT  币种,年月,旬,父级分类,二级分类,三级分类,物流方式,COUNT(`订单编号`) 总订单量
				FROM  {0}	cx
				WHERE  cx.`记录时间`= '{2}'
					AND cx.`币种` = '{1}'
					AND cx.`父级分类` IS NOT NULL
					AND cx.`系统订单状态` IN {3} 
				GROUP BY 币种,年月,旬,父级分类,二级分类,三级分类,物流方式
				ORDER BY 币种,年月
				) sl_zong
		    LEFT JOIN
				(SELECT 币种,年月,父级分类,二级分类,三级分类,物流方式,旬,COUNT(`订单编号`) 直发订单量
				FROM  {0}	cx_zf
				WHERE cx_zf.`记录时间`= '{2}'
					AND cx_zf.`币种` = '{1}' 
					AND cx_zf.`父级分类` IS NOT NULL
					AND cx_zf.`系统订单状态` IN {3} 
				AND cx_zf.`是否改派` = "直发"
				GROUP BY 币种,年月,旬,父级分类,二级分类,三级分类,物流方式
				ORDER BY 币种,年月
				) sl_zong_zf 
		    ON sl_zong_zf.`币种` = sl_zong.`币种` AND sl_zong_zf.`年月` = sl_zong.`年月`AND sl_zong_zf.`父级分类` = sl_zong.`父级分类` 
			    AND sl_zong_zf.`二级分类` = sl_zong.`二级分类` AND sl_zong_zf.`三级分类` = sl_zong.`三级分类` 
			    AND sl_zong_zf.`物流方式` = sl_zong.`物流方式` AND sl_zong_zf.`旬` = sl_zong.`旬` 				
		    LEFT JOIN
				(SELECT 币种,年月,旬,父级分类,二级分类,三级分类,物流方式,COUNT(`订单编号`) 直发已签收订单量
				FROM  {0}	sl_cx_zf_qianshou
				WHERE sl_cx_zf_qianshou.`记录时间`= '{2}'
					AND sl_cx_zf_qianshou.`币种` = '{1}' 
					AND sl_cx_zf_qianshou.`父级分类` IS NOT NULL
					AND sl_cx_zf_qianshou.`是否改派` = "直发"
					AND sl_cx_zf_qianshou.`最终状态` = "已签收"
					AND sl_cx_zf_qianshou.`系统订单状态` IN {3} 
				GROUP BY 币种,年月,旬,父级分类,二级分类,三级分类,物流方式
				ORDER BY 币种,年月
				) sl_cx_zf_qs
		    ON sl_cx_zf_qs.`币种` = sl_zong.`币种` AND sl_cx_zf_qs.`年月` = sl_zong.`年月` AND sl_cx_zf_qs.`父级分类` = sl_zong.`父级分类` 
			    AND sl_cx_zf_qs.`二级分类` = sl_zong.`二级分类` AND sl_cx_zf_qs.`三级分类` = sl_zong.`三级分类` 
			    AND sl_cx_zf_qs.`物流方式` = sl_zong.`物流方式` AND sl_cx_zf_qs.`旬` = sl_zong.`旬` 		
		    LEFT JOIN
				(SELECT 币种,年月,旬,父级分类,二级分类,三级分类,物流方式,COUNT(`订单编号`) 直发拒收订单量
				FROM  {0}	sl_cx_zf_jushou
				WHERE sl_cx_zf_jushou.`记录时间`= '{2}'
					AND sl_cx_zf_jushou.`币种` = '{1}' 
					AND sl_cx_zf_jushou.`父级分类` IS NOT NULL
					AND sl_cx_zf_jushou.`是否改派` = "直发" 
					AND sl_cx_zf_jushou.`最终状态` IN ('拒收', '理赔', '已签收', '已退货') 
					AND sl_cx_zf_jushou.`系统订单状态` IN {3} 
				GROUP BY 币种,年月,旬,父级分类,二级分类,三级分类,物流方式
				ORDER BY 币种,年月
				) sl_cx_zf_js
		    ON sl_cx_zf_js.`币种` = sl_zong.`币种` AND sl_cx_zf_js.`年月` = sl_zong.`年月` AND sl_cx_zf_js.`父级分类` = sl_zong.`父级分类` 
			    AND sl_cx_zf_js.`二级分类` = sl_zong.`二级分类` AND sl_cx_zf_js.`三级分类` = sl_zong.`三级分类` 
			    AND sl_cx_zf_js.`物流方式` = sl_zong.`物流方式` AND sl_cx_zf_js.`旬` = sl_zong.`旬` 
		    LEFT JOIN
				(SELECT 币种,年月,旬,父级分类,二级分类,三级分类,物流方式,COUNT(`订单编号`) 改派订单量
				FROM  {0}	sl_cx_gp
				WHERE sl_cx_gp.`记录时间`= '{2}'
					AND sl_cx_gp.`币种` = '{1}' 
					AND sl_cx_gp.`父级分类` IS NOT NULL
					AND sl_cx_gp.`是否改派` = "改派"
					AND sl_cx_gp.`系统订单状态` IN {3} 
				GROUP BY 币种,年月,旬,父级分类,二级分类,三级分类,物流方式
				ORDER BY 币种,年月
				) sl_zong_gp
		    ON sl_zong_gp.`币种` = sl_zong.`币种` AND sl_zong_gp.`年月` = sl_zong.`年月` AND sl_zong_gp.`父级分类` = sl_zong.`父级分类` 
			    AND sl_zong_gp.`二级分类` = sl_zong.`二级分类` AND sl_zong_gp.`三级分类` = sl_zong.`三级分类` 
			    AND sl_zong_gp.`物流方式` = sl_zong.`物流方式` AND sl_zong_gp.`旬` = sl_zong.`旬` 
		    LEFT JOIN
				(SELECT 币种,年月,旬,父级分类,二级分类,三级分类,物流方式,COUNT(`订单编号`) 改派已签收订单量
				FROM  {0}	sl_cx_gp_qianshou
				WHERE sl_cx_gp_qianshou.`记录时间`= '{2}'
					AND sl_cx_gp_qianshou.`币种` = '{1}' 
					AND sl_cx_gp_qianshou.`父级分类` IS NOT NULL
					AND sl_cx_gp_qianshou.`是否改派` = "改派"
					AND sl_cx_gp_qianshou.`最终状态` = "已签收"
					AND sl_cx_gp_qianshou.`系统订单状态` IN {3} 
				GROUP BY 币种,年月,旬,父级分类,二级分类,三级分类,物流方式
				ORDER BY 币种,年月
				) sl_cx_gp_qs
			ON sl_cx_gp_qs.`币种` = sl_zong.`币种` AND sl_cx_gp_qs.`年月` = sl_zong.`年月` AND sl_cx_gp_qs.`父级分类` = sl_zong.`父级分类` 
				AND sl_cx_gp_qs.`二级分类` = sl_zong.`二级分类` AND sl_cx_gp_qs.`三级分类` = sl_zong.`三级分类` 
				AND sl_cx_gp_qs.`物流方式` = sl_zong.`物流方式` AND sl_cx_gp_qs.`旬` = sl_zong.`旬` 
		    LEFT JOIN
				(SELECT 币种,年月,旬,父级分类,二级分类,三级分类,物流方式,COUNT(`订单编号`) 改派拒收订单量
				FROM  {0}	sl_cx_gp_jushou
				WHERE sl_cx_gp_jushou.`记录时间`= '{2}'
					AND sl_cx_gp_jushou.`币种` = '{1}' 
					AND sl_cx_gp_jushou.`父级分类` IS NOT NULL
					AND sl_cx_gp_jushou.`是否改派` = "改派"
					AND sl_cx_gp_jushou.`最终状态` IN ('拒收', '理赔', '已签收', '已退货') 
					AND sl_cx_gp_jushou.`系统订单状态` IN {3} 
				GROUP BY 币种,年月,旬,父级分类,二级分类,三级分类,物流方式
				ORDER BY 币种,年月
				) sl_cx_gp_js
			ON sl_cx_gp_js.`币种` = sl_zong.`币种` AND sl_cx_gp_js.`年月` = sl_zong.`年月` AND sl_cx_gp_js.`父级分类` = sl_zong.`父级分类` 
				AND sl_cx_gp_js.`二级分类` = sl_zong.`二级分类` AND sl_cx_gp_js.`三级分类` = sl_zong.`三级分类` 
				AND sl_cx_gp_js.`物流方式` = sl_zong.`物流方式` AND sl_cx_gp_js.`旬` = sl_zong.`旬` 
		    LEFT JOIN
				(SELECT 币种,年月,旬,父级分类,二级分类,三级分类,物流方式,COUNT(`订单编号`) 已签收订单量
				FROM  {0}	sl_cx_qianshou
				WHERE sl_cx_qianshou.`记录时间`= '{2}'
					AND sl_cx_qianshou.`币种` = '{1}' 
					AND sl_cx_qianshou.`父级分类` IS NOT NULL
					AND sl_cx_qianshou.`最终状态` = "已签收"
					AND sl_cx_qianshou.`系统订单状态` IN {3} 
				GROUP BY 币种,年月,旬,父级分类,二级分类,三级分类,物流方式
				ORDER BY 币种,年月
				) sl_zong_qianshou
			ON sl_zong_qianshou.`币种` = sl_zong.`币种` AND sl_zong_qianshou.`年月` = sl_zong.`年月` AND sl_zong_qianshou.`父级分类` = sl_zong.`父级分类` 
				AND sl_zong_qianshou.`二级分类` = sl_zong.`二级分类`  AND sl_zong_qianshou.`三级分类` = sl_zong.`三级分类` 
				AND sl_zong_qianshou.`物流方式` = sl_zong.`物流方式` AND sl_zong_qianshou.`旬` = sl_zong.`旬`
		    LEFT JOIN
				(SELECT 币种,年月,旬,父级分类,二级分类,三级分类,物流方式,COUNT(`订单编号`) 拒收订单量
				FROM  {0}	sl_cx_jushou
				WHERE sl_cx_jushou.`记录时间`= '{2}'
					AND sl_cx_jushou.`币种` = '{1}' 
					AND sl_cx_jushou.`父级分类` IS NOT NULL
					AND sl_cx_jushou.`最终状态` IN ('拒收', '理赔', '已签收', '已退货') 
					AND sl_cx_jushou.`系统订单状态` IN {3} 
				GROUP BY 币种,年月,旬,父级分类,二级分类,三级分类,物流方式
				ORDER BY 币种,年月
				) sl_zong_jushou
			ON sl_zong_jushou.`币种` = sl_zong.`币种` AND sl_zong_jushou.`年月` = sl_zong.`年月` AND sl_zong_jushou.`父级分类` = sl_zong.`父级分类` 
				AND sl_zong_jushou.`二级分类` = sl_zong.`二级分类` 
				AND sl_zong_jushou.`三级分类` = sl_zong.`三级分类` AND sl_zong_jushou.`物流方式` = sl_zong.`物流方式` AND sl_zong_jushou.`旬` = sl_zong.`旬` 
		    LEFT JOIN
				(SELECT 币种,年月,旬,父级分类,二级分类,三级分类,物流方式,COUNT(`订单编号`) 退货订单量
				FROM  {0}	sl_cx_tuihuo
				WHERE sl_cx_tuihuo.`记录时间`= '{2}'
					AND sl_cx_tuihuo.`币种` = '{1}' 
					AND sl_cx_tuihuo.`父级分类` IS NOT NULL
					AND sl_cx_tuihuo.`最终状态` = "已退货"
					AND sl_cx_tuihuo.`系统订单状态` IN {3} 
				GROUP BY 币种,年月,旬,父级分类,二级分类,三级分类,物流方式
				ORDER BY 币种,年月
				) sl_zong_tuihuo
			ON sl_zong_tuihuo.`币种` = sl_zong.`币种` AND sl_zong_tuihuo.`年月` = sl_zong.`年月` AND sl_zong_tuihuo.`父级分类` = sl_zong.`父级分类`
				AND sl_zong_tuihuo.`二级分类` = sl_zong.`二级分类` AND sl_zong_tuihuo.`三级分类` = sl_zong.`三级分类` 
				AND sl_zong_tuihuo.`物流方式` = sl_zong.`物流方式`AND sl_zong_tuihuo.`旬` = sl_zong.`旬` 
		    GROUP BY sl_zong.年月,sl_zong.旬,sl_zong.父级分类,sl_zong.二级分类,sl_zong.三级分类,sl_zong.物流方式
	    with rollup) sl_gat;'''.format(match2[team], team, Time_day[11], ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)'))
        listT.append(sqlqsb4)
        show_name.append(' 月（旬）签收率_…………')
        # 月签收率（各月）---查询
        sqlqsb5 = '''SELECT sl_gat.`币种`,sl_gat.`年月`,sl_gat.父级分类,sl_gat.二级分类,sl_gat.三级分类,'' 产品名称,sl_gat.物流方式,sl_gat.旬,
				                sl_gat.`总订单量`,
				                sl_gat.`已签收订单量` / sl_gat.`拒收订单量` AS '总签收/完成',
				                sl_gat.`已签收订单量` / sl_gat.`总订单量` AS '总签收/总计',
				                sl_gat.`退货订单量` / sl_gat.`总订单量` AS '退款率',
				                sl_gat.`拒收订单量` / sl_gat.`总订单量` AS '总完成占比',
				                sl_gat.`直发订单量` 直发总计,
				                sl_gat.`直发已签收订单量` / sl_gat.`直发拒收订单量` AS '直发签收/完成',
				                sl_gat.`直发已签收订单量` / sl_gat.`直发订单量` AS '直发签收/总计', 
				                sl_gat.`直发拒收订单量` / sl_gat.`直发订单量` AS '直发完成占比',
				                sl_gat.`改派订单量` 改派总计,
				                sl_gat.`改派已签收订单量` / sl_gat.`改派拒收订单量` AS '改派签收/完成',
				                sl_gat.`改派已签收订单量` / sl_gat.`改派订单量` AS '改派签收/总计',
				                sl_gat.`改派拒收订单量` / sl_gat.`改派订单量` AS '改派完成占比',
				                sl_gat.`总销售额`,
				                sl_gat.`已签收销售额` / sl_gat.`拒收销售额` AS '总签收/完成(金额)',
				                sl_gat.`已签收销售额` / sl_gat.`总销售额` AS '总签收/总计(金额)',
				                sl_gat.`退货销售额` / sl_gat.`总销售额` AS '退款率(金额)',
				                sl_gat.`拒收销售额` / sl_gat.`总销售额` AS '总完成占比(金额)',
				                sl_gat.`直发销售额`,
				                sl_gat.`直发已签收销售额` / sl_gat.`直发拒收销售额` AS '直发签收/完成(金额)',
				                sl_gat.`直发已签收销售额` / sl_gat.`直发销售额` AS '直发签收/总计(金额)',
				                sl_gat.`直发拒收销售额` / sl_gat.`直发销售额` AS '直发完成占比(金额)',
				                sl_gat.`改派销售额`,
				                sl_gat.`改派已签收销售额` / sl_gat.`改派拒收销售额` AS '改派签收/完成(金额)',
				                sl_gat.`改派已签收销售额` / sl_gat.`改派销售额` AS '改派签收/总计(金额)',
				                sl_gat.`改派拒收销售额` / sl_gat.`改派销售额` AS '改派完成占比(金额)'
                        FROM (SELECT  sl_zong.币种,
                                    IFNULL(sl_zong.年月,'合计') 年月,
                                    IFNULL(sl_zong.父级分类,'合计') 父级分类,
                                    IFNULL(sl_zong.二级分类,'合计') 二级分类,
							        IFNULL(sl_zong.三级分类,'合计') 三级分类,
							        IFNULL(sl_zong.物流方式,'合计') 物流方式,
							        IFNULL(sl_zong.旬,'合计') 旬,
							        SUM(总订单量) 总订单量,
							        SUM(总销售额) 总销售额,
							        IFNULL(SUM(直发订单量),0) 直发订单量,
							        IFNULL(SUM(直发销售额),0) 直发销售额,
							        IFNULL(SUM(直发已签收订单量),0) 直发已签收订单量,
							        IFNULL(SUM(直发已签收销售额),0) 直发已签收销售额,
							        IFNULL(SUM(直发拒收订单量),0) 直发拒收订单量,
							        IFNULL(SUM(直发拒收销售额),0) 直发拒收销售额,
							        (SUM(总订单量) - IFNULL(SUM(直发订单量),0)) AS 改派订单量,
							        (SUM(总销售额) - IFNULL(SUM(直发销售额),0)) AS 改派销售额,
							        IFNULL(SUM(改派已签收订单量),0) 改派已签收订单量,
							        IFNULL(SUM(改派已签收销售额),0) 改派已签收销售额,
							        IFNULL(SUM(改派拒收订单量),0) 改派拒收订单量,
							        IFNULL(SUM(改派拒收销售额),0) 改派拒收销售额,
							        IFNULL(SUM(已签收订单量),0) 已签收订单量,
							        IFNULL(SUM(已签收销售额),0) 已签收销售额,
							        IFNULL(SUM(拒收订单量),0) 拒收订单量,
							        IFNULL(SUM(拒收销售额),0) 拒收销售额,
							        IFNULL(SUM(退货订单量),0) 退货订单量,
							        IFNULL(SUM(退货销售额),0) 退货销售额
		FROM (SELECT  币种,年月,父级分类,二级分类,三级分类,物流方式,旬,COUNT(`订单编号`) 总订单量,SUM(`价格RMB`) 总销售额
				FROM  {0}	sl_cx
				WHERE sl_cx.`币种` = '{1}' 
					AND sl_cx.`父级分类` IS NOT NULL
					AND sl_cx.`系统订单状态` IN {2} 
				GROUP BY 币种,年月,父级分类,二级分类,三级分类,物流方式,旬
				ORDER BY 币种,年月
				) sl_zong
		LEFT JOIN
				(SELECT 币种,年月,父级分类,二级分类,三级分类,物流方式,旬,COUNT(`订单编号`) 直发订单量,SUM(`价格RMB`) 直发销售额
				FROM  {0}	sl_zong_zf
				WHERE sl_zong_zf.`币种` = '{1}' 
					AND sl_zong_zf.`是否改派` = "直发"
					AND sl_zong_zf.`父级分类` IS NOT NULL
					AND sl_zong_zf.`系统订单状态` IN {2}
				GROUP BY 币种,年月,父级分类,二级分类,三级分类,物流方式,旬
				ORDER BY 币种,年月
				) sl_zong_zf 
		ON sl_zong_zf.`币种` = sl_zong.`币种` AND sl_zong_zf.`年月` = sl_zong.`年月`AND sl_zong_zf.`父级分类` = sl_zong.`父级分类` 
			AND sl_zong_zf.`二级分类` = sl_zong.`二级分类` AND sl_zong_zf.`三级分类` = sl_zong.`三级分类` 
			AND sl_zong_zf.`物流方式` = sl_zong.`物流方式` AND sl_zong_zf.`旬` = sl_zong.`旬` 							
		LEFT JOIN
				(SELECT 币种,年月,父级分类,二级分类,三级分类,物流方式,旬,COUNT(`订单编号`) 直发已签收订单量,SUM(`价格RMB`) 直发已签收销售额
				FROM  {0}	slzf_qs
				WHERE slzf_qs.`币种` = '{1}' 
					AND slzf_qs.`是否改派` = "直发"
					AND slzf_qs.`最终状态` = "已签收"
					AND slzf_qs.`父级分类` IS NOT NULL
					AND slzf_qs.`系统订单状态` IN {2} 
				GROUP BY 币种,年月,父级分类,二级分类,三级分类,物流方式,旬
				ORDER BY 币种,年月
				) sl_cx_zf_qs
		ON sl_cx_zf_qs.`币种` = sl_zong.`币种` AND sl_cx_zf_qs.`年月` = sl_zong.`年月` AND sl_cx_zf_qs.`父级分类` = sl_zong.`父级分类` 
			AND sl_cx_zf_qs.`二级分类` = sl_zong.`二级分类` AND sl_cx_zf_qs.`三级分类` = sl_zong.`三级分类` 
			AND sl_cx_zf_qs.`物流方式` = sl_zong.`物流方式` AND sl_cx_zf_qs.`旬` = sl_zong.`旬` 		
		LEFT JOIN
				(SELECT 币种,年月,父级分类,二级分类,三级分类,物流方式,旬,COUNT(`订单编号`) 直发拒收订单量,SUM(`价格RMB`) 直发拒收销售额
				FROM  {0}	slzf_js
				WHERE slzf_js.`币种` = '{1}' 
						AND slzf_js.`是否改派` = "直发" 
						AND slzf_js.`最终状态` IN ('拒收', '理赔', '已签收', '已退货') 
						AND slzf_js.`父级分类` IS NOT NULL 
						AND slzf_js.`系统订单状态` IN {2} 
				GROUP BY 币种,年月,父级分类,二级分类,三级分类,物流方式,旬
				ORDER BY 币种,年月
				) sl_cx_zf_js
		ON sl_cx_zf_js.`币种` = sl_zong.`币种` AND sl_cx_zf_js.`年月` = sl_zong.`年月` AND sl_cx_zf_js.`父级分类` = sl_zong.`父级分类` 
			AND sl_cx_zf_js.`二级分类` = sl_zong.`二级分类` AND sl_cx_zf_js.`三级分类` = sl_zong.`三级分类` 
			AND sl_cx_zf_js.`物流方式` = sl_zong.`物流方式` AND sl_cx_zf_js.`旬` = sl_zong.`旬` 
		LEFT JOIN
				(SELECT 币种,年月,父级分类,二级分类,三级分类,物流方式,旬,COUNT(`订单编号`) 改派订单量,SUM(`价格RMB`) 改派销售额
				FROM  {0}	sl_cx_gp
				WHERE sl_cx_gp.`币种` = '{1}' 
						AND sl_cx_gp.`是否改派` = "改派"
						AND sl_cx_gp.`父级分类` IS NOT NULL
						AND sl_cx_gp.`系统订单状态` IN {2} 
					GROUP BY 币种,年月,父级分类,二级分类,三级分类,物流方式,旬
					ORDER BY 币种,年月
				) sl_zong_gp
		ON sl_zong_gp.`币种` = sl_zong.`币种` AND sl_zong_gp.`年月` = sl_zong.`年月` AND sl_zong_gp.`父级分类` = sl_zong.`父级分类` 
			AND sl_zong_gp.`二级分类` = sl_zong.`二级分类` AND sl_zong_gp.`三级分类` = sl_zong.`三级分类` 
			AND sl_zong_gp.`物流方式` = sl_zong.`物流方式` AND sl_zong_gp.`旬` = sl_zong.`旬` 
		LEFT JOIN
				(SELECT 币种,年月,父级分类,二级分类,三级分类,物流方式,旬,COUNT(`订单编号`) 改派已签收订单量,SUM(`价格RMB`) 改派已签收销售额
				FROM  {0}	slgp_qs
				WHERE slgp_qs.`币种` = '{1}' 
					AND slgp_qs.`是否改派` = "改派" 
					AND slgp_qs.`最终状态` = "已签收"
					AND slgp_qs.`父级分类` IS NOT NULL
					AND slgp_qs.`系统订单状态`IN {2}
				GROUP BY 币种,年月,父级分类,二级分类,三级分类,物流方式,旬
				ORDER BY 币种,年月
				) sl_cx_gp_qs
			ON sl_cx_gp_qs.`币种` = sl_zong.`币种` AND sl_cx_gp_qs.`年月` = sl_zong.`年月` AND sl_cx_gp_qs.`父级分类` = sl_zong.`父级分类` 
				AND sl_cx_gp_qs.`二级分类` = sl_zong.`二级分类` AND sl_cx_gp_qs.`三级分类` = sl_zong.`三级分类` 
				AND sl_cx_gp_qs.`物流方式` = sl_zong.`物流方式` AND sl_cx_gp_qs.`旬` = sl_zong.`旬` 
		LEFT JOIN
				(SELECT 币种,年月,父级分类,二级分类,三级分类,物流方式,旬,COUNT(`订单编号`) 改派拒收订单量,SUM(`价格RMB`) 改派拒收销售额
				FROM  {0}	sl_cx_gp_jushou
				WHERE sl_cx_gp_jushou.`币种` = '{1}' 
						AND sl_cx_gp_jushou.`系统订单状态` IN {2} 
						AND sl_cx_gp_jushou.`父级分类` IS NOT NULL
						AND sl_cx_gp_jushou.`是否改派` = "改派" 
						AND sl_cx_gp_jushou.`最终状态` IN ('拒收', '理赔', '已签收', '已退货') 
				GROUP BY 币种,年月,父级分类,二级分类,三级分类,物流方式,旬
				ORDER BY 币种,年月
				) sl_cx_gp_js
			ON sl_cx_gp_js.`币种` = sl_zong.`币种` AND sl_cx_gp_js.`年月` = sl_zong.`年月` AND sl_cx_gp_js.`父级分类` = sl_zong.`父级分类` 
				AND sl_cx_gp_js.`二级分类` = sl_zong.`二级分类` AND sl_cx_gp_js.`三级分类` = sl_zong.`三级分类` 
				AND sl_cx_gp_js.`物流方式` = sl_zong.`物流方式` AND sl_cx_gp_js.`旬` = sl_zong.`旬` 				
		LEFT JOIN
				(SELECT 币种,年月,父级分类,二级分类,三级分类,物流方式,旬,COUNT(`订单编号`) 已签收订单量,SUM(`价格RMB`) 已签收销售额
				FROM  {0}	sl_cx_qianshou
				WHERE sl_cx_qianshou.`币种` = '{1}' 
					AND sl_cx_qianshou.`系统订单状态`IN {2}
					AND sl_cx_qianshou.`最终状态` = "已签收"
					AND sl_cx_qianshou.`父级分类` IS NOT NULL
				GROUP BY 币种,年月,父级分类,二级分类,三级分类,物流方式,旬
				ORDER BY 币种,年月
				) sl_zong_qianshou
			ON sl_zong_qianshou.`币种` = sl_zong.`币种` AND sl_zong_qianshou.`年月` = sl_zong.`年月` AND sl_zong_qianshou.`父级分类` = sl_zong.`父级分类` 
				AND sl_zong_qianshou.`二级分类` = sl_zong.`二级分类`  AND sl_zong_qianshou.`三级分类` = sl_zong.`三级分类` 
				AND sl_zong_qianshou.`物流方式` = sl_zong.`物流方式` AND sl_zong_qianshou.`旬` = sl_zong.`旬`
		LEFT JOIN
				(SELECT 币种,年月,父级分类,二级分类,三级分类,物流方式,旬,COUNT(`订单编号`) 拒收订单量,SUM(`价格RMB`) 拒收销售额
				FROM  {0}	sl_cx_jushou
				WHERE sl_cx_jushou.`币种` = '{1}' 
						AND sl_cx_jushou.`系统订单状态` IN {2}
						AND sl_cx_jushou.`最终状态` IN ('拒收', '理赔', '已签收', '已退货') 
						AND sl_cx_jushou.`父级分类` IS NOT NULL
				GROUP BY 币种,年月,父级分类,二级分类,三级分类,物流方式,旬
				ORDER BY 币种,年月
				) sl_zong_jushou
			ON sl_zong_jushou.`币种` = sl_zong.`币种` AND sl_zong_jushou.`年月` = sl_zong.`年月` AND sl_zong_jushou.`父级分类` = sl_zong.`父级分类` 
				AND sl_zong_jushou.`二级分类` = sl_zong.`二级分类` AND sl_zong_jushou.`三级分类` = sl_zong.`三级分类` 
				AND sl_zong_jushou.`物流方式` = sl_zong.`物流方式` AND sl_zong_jushou.`旬` = sl_zong.`旬` 
		LEFT JOIN
				(SELECT 币种,年月,父级分类,二级分类,三级分类,物流方式,旬,COUNT(`订单编号`) 退货订单量,SUM(`价格RMB`) 退货销售额
				FROM  {0}	sl_cx_tuihuo
				WHERE sl_cx_tuihuo.`币种` = '{1}' 
						AND sl_cx_tuihuo.`系统订单状态`IN {2}
						AND sl_cx_tuihuo.`最终状态` = "已退货"
						AND sl_cx_tuihuo.`父级分类` IS NOT NULL
				GROUP BY 币种,年月,父级分类,二级分类,三级分类,物流方式,旬
				ORDER BY 币种,年月
				) sl_zong_tuihuo
			ON sl_zong_tuihuo.`币种` = sl_zong.`币种` AND sl_zong_tuihuo.`年月` = sl_zong.`年月` AND sl_zong_tuihuo.`父级分类` = sl_zong.`父级分类` 
				AND sl_zong_tuihuo.`二级分类` = sl_zong.`二级分类` AND sl_zong_tuihuo.`三级分类` = sl_zong.`三级分类` 
				AND sl_zong_tuihuo.`物流方式` = sl_zong.`物流方式`AND sl_zong_tuihuo.`旬` = sl_zong.`旬` 
		GROUP BY sl_zong.年月,sl_zong.父级分类,sl_zong.二级分类,sl_zong.三级分类,sl_zong.物流方式,sl_zong.旬
		with rollup) sl_gat 
		ORDER BY sl_gat.`年月` DESC;'''.format('qsb_缓存_month', team, ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)'))
        listT.append(sqlqsb5)
        show_name.append(' 月（各月）签收率_…………')

        # 月物流（天）---查询
        sqlWl2 = '''SELECT sl_rb.`币种`,
				sl_rb.`年月`,
				sl_rb.物流方式,
				sl_rb.父级分类,
				sl_rb.旬,
				sl_rb.`总订单量` 总订单,
				null AS '总签收/完成',
				null AS '总签收/总计',
				null AS '退款率',
				null AS '总完成占比',
				sl_rb.`直发订单量` 总计,
				sl_rb.`直发已签收订单量` / sl_rb.`直发拒收订单量` AS '直发签收/完成',
				sl_rb.`直发已签收订单量` / sl_rb.`直发订单量` AS '直发签收/总计',
				sl_rb.`直发拒收订单量` / sl_rb.`直发订单量` AS '直发完成占比',
				null AS  改派总计,
				null AS '改派签收/完成',
				null AS '改派签收/总计',
				null AS '改派完成占比'
        FROM (SELECT  sl_zong.币种,
						IFNULL(sl_zong.年月,'合计') 年月,
						IFNULL(sl_zong.物流方式,'合计') 物流方式,
						IFNULL(sl_zong.父级分类,'合计') 父级分类,
						IFNULL(sl_zong.旬,'合计') 旬,
						SUM(总订单量) 总订单量,
						IFNULL(SUM(直发订单量),0) 直发订单量,
						IFNULL(SUM(直发已签收订单量),0) 直发已签收订单量,
						IFNULL(SUM(直发拒收订单量),0) 直发拒收订单量
			FROM ( SELECT 币种,
										年月,
										物流方式,
										父级分类,
										旬,
										COUNT(`订单编号`) 总订单量
						FROM  {0}	sl_cx
						WHERE sl_cx.`币种` = '{1}' 
                                AND (sl_cx.`记录时间`= '{2}' AND (sl_cx.`日期` between DATE_SUB('{2}', INTERVAL DAY('{2}')-1 DAY) AND  DATE_SUB('{2}',INTERVAL 1 DAY))
									OR sl_cx.`记录时间`= '{3}' AND (sl_cx.`日期` between DATE_SUB('{3}', INTERVAL DAY('{3}')-1 DAY) AND  DATE_SUB('{3}',INTERVAL 1 DAY)))
								AND sl_cx.`父级分类` IS NOT NULL
                                AND sl_cx.`是否改派` = "直发"
                                AND sl_cx.`系统订单状态` IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)') 
						GROUP BY 币种,年月,物流方式,父级分类,旬
						ORDER BY 币种,年月
						) sl_zong
			LEFT JOIN
						(SELECT 币种,
										年月,
										物流方式,
										父级分类,
													旬,
										COUNT(`订单编号`) 直发订单量
						FROM  {0}	sl_cx_zf
						WHERE sl_cx_zf.`币种` = '{1}' 
                                AND (sl_cx_zf.`记录时间`= '{2}' AND (sl_cx_zf.`日期` between DATE_SUB('{2}', INTERVAL DAY('{2}')-1 DAY) AND  DATE_SUB('{2}',INTERVAL 1 DAY))
									OR sl_cx_zf.`记录时间`= '{3}' AND (sl_cx_zf.`日期` between DATE_SUB('{3}', INTERVAL DAY('{3}')-1 DAY) AND  DATE_SUB('{3}',INTERVAL 1 DAY)))
								AND sl_cx_zf.`父级分类` IS NOT NULL
                                AND sl_cx_zf.`系统订单状态` IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)')  
                                AND sl_cx_zf.`是否改派` = "直发"
						GROUP BY 币种,年月,物流方式,父级分类,旬
						ORDER BY 币种,年月
					) sl_zong_zf
					 ON sl_zong_zf.`币种` = sl_zong.`币种` 
								AND sl_zong_zf.`年月` = sl_zong.`年月`
								AND sl_zong_zf.`物流方式` = sl_zong.`物流方式`
						  	    AND sl_zong_zf.`父级分类` = sl_zong.`父级分类` 
								AND sl_zong_zf.`旬` = sl_zong.`旬` 
				LEFT JOIN
						(SELECT 币种,
										年月,
										物流方式,
										父级分类,
													旬,
										COUNT(`订单编号`) 直发已签收订单量
						FROM  {0}	sl_cx_zf_qianshou
						WHERE sl_cx_zf_qianshou.`币种` = '{1}' 
                                AND (sl_cx_zf_qianshou.`记录时间`= '{2}' AND (sl_cx_zf_qianshou.`日期` between DATE_SUB('{2}', INTERVAL DAY('{2}')-1 DAY) AND  DATE_SUB('{2}',INTERVAL 1 DAY))
									OR sl_cx_zf_qianshou.`记录时间`= '{3}' AND (sl_cx_zf_qianshou.`日期` between DATE_SUB('{3}', INTERVAL DAY('{3}')-1 DAY) AND  DATE_SUB('{3}',INTERVAL 1 DAY)))
							    AND sl_cx_zf_qianshou.`父级分类` IS NOT NULL
							    AND sl_cx_zf_qianshou.`系统订单状态` IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)') 
								AND sl_cx_zf_qianshou.`是否改派` = "直发"
								AND sl_cx_zf_qianshou.`最终状态` = "已签收"
						GROUP BY 币种,年月,物流方式,父级分类,旬
						ORDER BY 币种,年月
					) sl_cx_zf_qs
					 ON sl_cx_zf_qs.`币种` = sl_zong.`币种` 
								AND sl_cx_zf_qs.`年月` = sl_zong.`年月`
								AND sl_cx_zf_qs.`物流方式` = sl_zong.`物流方式`
								AND sl_cx_zf_qs.`父级分类` = sl_zong.`父级分类`
								AND sl_cx_zf_qs.`旬` = sl_zong.`旬`	 		
				LEFT JOIN
						(SELECT 币种,
										年月,
										物流方式,
										父级分类,
													旬,
										COUNT(`订单编号`) 直发拒收订单量
						FROM  {0}	sl_cx_zf_jushou
						WHERE sl_cx_zf_jushou.`币种` = '{1}' 
                                AND (sl_cx_zf_jushou.`记录时间`= '{2}' AND (sl_cx_zf_jushou.`日期` between DATE_SUB('{2}', INTERVAL DAY('{2}')-1 DAY) AND  DATE_SUB('{2}',INTERVAL 1 DAY))
									OR sl_cx_zf_jushou.`记录时间`= '{3}' AND (sl_cx_zf_jushou.`日期` between DATE_SUB('{3}', INTERVAL DAY('{3}')-1 DAY) AND  DATE_SUB('{3}',INTERVAL 1 DAY)))
							    AND sl_cx_zf_jushou.`父级分类` IS NOT NULL
							    AND sl_cx_zf_jushou.`系统订单状态` IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)') 
								AND sl_cx_zf_jushou.`是否改派` = "直发"
								AND sl_cx_zf_jushou.`最终状态` IN ('拒收', '理赔', '已签收', '已退货') 
						GROUP BY 币种,年月,物流方式,父级分类,旬
						ORDER BY 币种,年月
					) sl_cx_zf_js
					 ON sl_cx_zf_js.`币种` = sl_zong.`币种` 
								AND sl_cx_zf_js.`年月` = sl_zong.`年月`
								AND sl_cx_zf_js.`物流方式` = sl_zong.`物流方式`
								AND sl_cx_zf_js.`父级分类` = sl_zong.`父级分类` 
								AND sl_cx_zf_js.`旬` = sl_zong.`旬` 											
				GROUP BY sl_zong.年月,sl_zong.物流方式,sl_zong.父级分类,sl_zong.旬
				with rollup) sl_rb;'''.format(match2[team], team, Time_day[11], Time_day[10])
        listT.append(sqlWl2)
        show_name.append(' 月（天）物流…………')
        # 月物流（月）---查询
        sqlWl3= '''SELECT sl_rb.`币种`,
                    	sl_rb.`年月`,
                    	sl_rb.物流方式,
                    	sl_rb.父级分类,
                    	sl_rb.旬,
                    	sl_rb.`总订单量` 总订单,
                    	null AS '总签收/完成',
                    	null AS '总签收/总计',
                    	null AS '退款率',
                    	null AS '总完成占比',
                    	sl_rb.`直发订单量` 总计,
                    	sl_rb.`直发已签收订单量` / sl_rb.`直发拒收订单量` AS '直发签收/完成',
                    	sl_rb.`直发已签收订单量` / sl_rb.`直发订单量` AS '直发签收/总计',
                    	sl_rb.`直发拒收订单量` / sl_rb.`直发订单量` AS '直发完成占比',
                    	null AS  改派总计,
                    	null AS '改派签收/完成',
                    	null AS '改派签收/总计',
                    	null AS '改派完成占比'
        FROM (SELECT  sl_zong.币种,
                    						IFNULL(sl_zong.年月,'合计') 年月,
                    						IFNULL(sl_zong.物流方式,'合计') 物流方式,
                    						IFNULL(sl_zong.父级分类,'合计') 父级分类,
                    						IFNULL(sl_zong.旬,'合计') 旬,
                    						SUM(总订单量) 总订单量,
                    						IFNULL(SUM(直发订单量),0) 直发订单量,
                    						IFNULL(SUM(直发已签收订单量),0) 直发已签收订单量,
                    						IFNULL(SUM(直发拒收订单量),0) 直发拒收订单量
                    		FROM ( SELECT 币种,
                    					年月,
                    					物流方式,
                    					父级分类,
                    					旬,
                    					COUNT(`订单编号`) 总订单量
                    			    FROM  {0}	sl_cx
                    					WHERE sl_cx.`币种` = '{1}' 
        									AND (sl_cx.`记录时间`= '{2}' AND sl_cx.`日期` < DATE_SUB('{2}', INTERVAL DAY('{2}')-1 DAY) 
        										OR sl_cx.`记录时间`= '{3}' AND sl_cx.`日期` < DATE_SUB('{3}', INTERVAL DAY('{3}')-1 DAY))
                    						AND sl_cx.`父级分类` IS NOT NULL
                    						AND sl_cx.`是否改派` = "直发"
                                            AND sl_cx.`系统订单状态` IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)') 
                    				GROUP BY 币种,年月,物流方式,父级分类,旬
                    				ORDER BY 币种,年月
                    			) sl_zong
                    		LEFT JOIN
                    			(SELECT 币种,
                    					年月,
                    					物流方式,
                    					父级分类,
                    					旬,
                    					COUNT(`订单编号`) 直发订单量
                    				FROM  {0}	sl_cx_zf
                    				WHERE sl_cx_zf.`币种` = '{1}' 
        								AND (sl_cx_zf.`记录时间`= '{2}' AND sl_cx_zf.`日期` < DATE_SUB('{2}', INTERVAL DAY('{2}')-1 DAY) 
        									OR sl_cx_zf.`记录时间`= '{3}' AND sl_cx_zf.`日期` < DATE_SUB('{3}', INTERVAL DAY('{3}')-1 DAY))
                    					AND sl_cx_zf.`父级分类` IS NOT NULL
                                        AND sl_cx_zf.`系统订单状态` IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)')  
                                        AND sl_cx_zf.`是否改派` = "直发"
                    				GROUP BY 币种,年月,物流方式,父级分类,旬
                    				ORDER BY 币种,年月
                    			) sl_zong_zf
                    		ON sl_zong_zf.`币种` = sl_zong.`币种` 
                    					AND sl_zong_zf.`年月` = sl_zong.`年月`
                    					AND sl_zong_zf.`物流方式` = sl_zong.`物流方式`
                    					AND sl_zong_zf.`父级分类` = sl_zong.`父级分类` 
                    					AND sl_zong_zf.`旬` = sl_zong.`旬` 
                    		LEFT JOIN
                    			(SELECT 币种,
                    					年月,
                    					物流方式,
                    					父级分类,
                    					旬,
                    					COUNT(`订单编号`) 直发已签收订单量
                    			FROM  {0}	sl_cx_zf_qianshou
                    			WHERE sl_cx_zf_qianshou.`币种` = '{1}' 
        							AND (sl_cx_zf_qianshou.`记录时间`= '{2}' AND sl_cx_zf_qianshou.`日期` < DATE_SUB('{2}', INTERVAL DAY('{2}')-1 DAY) 
        								OR sl_cx_zf_qianshou.`记录时间`= '{3}' AND sl_cx_zf_qianshou.`日期` < DATE_SUB('{3}', INTERVAL DAY('{3}')-1 DAY))
                    				AND sl_cx_zf_qianshou.`父级分类` IS NOT NULL
                    				AND sl_cx_zf_qianshou.`系统订单状态` IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)') 
                    				AND sl_cx_zf_qianshou.`是否改派` = "直发"
                    				AND sl_cx_zf_qianshou.`最终状态` = "已签收"
                    			GROUP BY 币种,年月,物流方式,父级分类,旬
                    			ORDER BY 币种,年月
                    		    ) sl_cx_zf_qs
                    		ON sl_cx_zf_qs.`币种` = sl_zong.`币种` 
                    				AND sl_cx_zf_qs.`年月` = sl_zong.`年月`
                    				AND sl_cx_zf_qs.`物流方式` = sl_zong.`物流方式`
                    				AND sl_cx_zf_qs.`父级分类` = sl_zong.`父级分类`
                    				AND sl_cx_zf_qs.`旬` = sl_zong.`旬`	 		
                    		LEFT JOIN
                    				(SELECT 币种,
                    						年月,
                    						物流方式,
                    						父级分类,
                    						旬,
                    						COUNT(`订单编号`) 直发拒收订单量
                    					FROM  {0}	sl_cx_zf_jushou
                    					WHERE sl_cx_zf_jushou.`币种` = '{1}' 
        									AND (sl_cx_zf_jushou.`记录时间`= '{2}' AND sl_cx_zf_jushou.`日期` < DATE_SUB('{2}', INTERVAL DAY('{2}')-1 DAY) 
        										OR sl_cx_zf_jushou.`记录时间`= '{3}' AND sl_cx_zf_jushou.`日期` < DATE_SUB('{3}', INTERVAL DAY('{3}')-1 DAY))
                    						AND sl_cx_zf_jushou.`父级分类` IS NOT NULL
                    						AND sl_cx_zf_jushou.`系统订单状态` IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)') 
                    						AND sl_cx_zf_jushou.`是否改派` = "直发"
                    						AND sl_cx_zf_jushou.`最终状态` IN ('拒收', '理赔', '已签收', '已退货') 
                    					GROUP BY 币种,年月,物流方式,父级分类,旬
                    					ORDER BY 币种,年月
                    				) sl_cx_zf_js
                            ON sl_cx_zf_js.`币种` = sl_zong.`币种` 
                    			AND sl_cx_zf_js.`年月` = sl_zong.`年月`
                    			AND sl_cx_zf_js.`物流方式` = sl_zong.`物流方式`
                    			AND sl_cx_zf_js.`父级分类` = sl_zong.`父级分类` 
                    			AND sl_cx_zf_js.`旬` = sl_zong.`旬` 												
                    		GROUP BY sl_zong.年月,sl_zong.物流方式,sl_zong.父级分类,sl_zong.旬
                    with rollup) sl_rb;'''.format(match2[team], team, Time_day[11], Time_day[10])
        listT.append(sqlWl3)
        show_name.append(' 月（各月）物流…………')
        # # 成本（天）---查询 临时使用
        # sqlcost20 = '''SELECT ct.团队,
        #                     '' 新月,
        #                     ct.`年月`,
        #                     ct.品类,
        #                     ct.销售额,
		# 	                ct.销售额 / ct.订单量 AS 客单价,
		# 	                ct.订单量,
		# 	                ct.改派订单量,
		# 	                ct.改派订单量 / ct.订单量 改派占比,
		# 	                ct.销售额 / ct.广告成本 ROI,
		# 	                ct.活跃产品数,
		# 	                ct.订单量 / ct.活跃产品数 产能,
		# 	                ct.总采购额 / ct.销售额 AS 采购占比,
		# 	                ct.广告成本 / ct.销售额 AS 花费占比,
		# 	                ct.物流成本 / ct.销售额 AS 运费占比,
		# 	                ct.手续费 / ct.销售额 AS 手续费占比,
		# 	                (ct.总采购额 + ct.广告成本 + ct.物流成本 + ct.手续费) / ct.销售额 AS 总成本占比,
		# 	                ct.签收量 / (ct.签收量 + ct.拒收量) AS '签收/完成',
		# 	                ct.签收量 / ct.订单量 AS '签收/总计',
		# 	                (ct.签收量 + ct.拒收量) / ct.订单量 AS '完成占比'
		# 	    FROM ( SELECT	EXTRACT(YEAR_MONTH FROM ct2.`年月`) 年月,
		# 	                                团队,
		# 	                                品类,
		# 																	SUM(订单量) 订单量,
		# 	                                SUM(活跃产品数) 活跃产品数,
		# 	                                SUM(签收量) 签收量,
		# 	                                SUM(拒收量) 拒收量,
		# 	                                SUM(销售额) 销售额,
		# 	                                SUM(签收额) 签收额,
		# 	                                SUM(拒收额) 拒收额,
		# 	                                SUM(改派订单量) 改派订单量,
		# 	                                SUM(总采购额) 总采购额,
		# 	                                SUM(直发采购额) 直发采购额,
		# 	                                SUM(广告成本) 广告成本,
		# 	                                SUM(物流成本) 物流成本,
		# 	                                SUM(手续费) 手续费
		# 			       FROM {0} ct2
		# 			       WHERE ct2.`品类` <> '未知'
		# 									AND ct2.`团队` = '{1}'
		# 									AND ct2.`记录时间` in ('{2}','{3}')
		# 									AND ct2.`年月` <> '0000-00-00'
        #                     GROUP BY EXTRACT(YEAR_MONTH FROM ct2.`年月`)
		# 		) ct
		# 		order by ct.`年月` desc '''.format('zg_cost_sltem_copy', team, Time_day[0], Time_day[11])
        # listT.append(sqlcost20)
        # show_name.append('（天）成本…………')
        # # 月成本（天）---查询
        # sqlcost2 = '''SELECT ct.团队,
        #                             EXTRACT(YEAR_MONTH FROM ct.记录时间) 新月,
        #                             ct.品类 年月,
        #                             ct.销售额,
        # 			                ct.销售额 / ct.订单量 AS 客单价,
        # 			                ct.订单量,
        # 			                ct.改派订单量,
        # 			                ct.改派订单量 / ct.订单量 改派占比,
        # 			                ct.销售额 / ct.广告成本 ROI,
        # 			                '' 活跃产品数,
        # 			                '' 产能,
        # 			                ct.总采购额 / ct.销售额 AS 采购占比,
        # 			                ct.广告成本 / ct.销售额 AS 花费占比,
        # 			                ct.物流成本 / ct.销售额 AS 运费占比,
        # 			                ct.手续费 / ct.销售额 AS 手续费占比,
        # 			                (ct.总采购额 + ct.广告成本 + ct.物流成本 + ct.手续费) / ct.销售额 AS 总成本占比,
        # 			                ct.签收量 / (ct.签收量 + ct.拒收量) AS '签收/完成',
        # 			                ct.签收量 / ct.订单量 AS '签收/总计',
        # 			                (ct.签收量 + ct.拒收量) / ct.订单量 AS '完成占比'
        # 			    FROM {0} ct
        # 			    WHERE ct.`年月` = '0000-00-00' AND ct.`团队` = '{1}' AND (ct.`记录时间` = '{2}' or  ct.`记录时间` ='{3}')
        #                 GROUP BY 品类,团队
        #                 order by 品类 desc ;'''.format('zg_cost_sltem_copy', team, Time_day[0], Time_day[11])
        # # listT.append(sqlcost2)
        # # show_name.append(' 月（天）成本…………')
        # # 月成本（月）---查询
        # sqlcost3 = '''SELECT ct.团队,
        #                     ct.新月,
        #                     IFNULL(ct.品类,'合计') 品类,
        #                     ct.销售额 / ct.订单量 AS 客单价,
        #                     ct.订单量,
        #                     ct.订单量 / ct.总订单量 AS 订单品类占比,
        #                     ct.直发采购额/ct.销售额 '直发采购额/销售额',
        #                     ct.广告成本 / ct.销售额 AS 花费占比,
        #                     ct.物流成本 / ct.销售额 AS 运费占比,
        #                     ct.手续费 / ct.销售额 AS 手续费占比,
        #                     (ct.广告成本 + ct.物流成本 + ct.手续费  + ct.直发采购额) / ct.销售额 AS 总成本,
        #                     ct.签收额 /(ct.签收额 + ct.拒收额) AS '金额签收/完成',
        #                     ct.签收额 /ct.销售额 AS '金额签收/总计',
        #                     (ct.签收额 + ct.拒收额) / ct.销售额 AS '金额完成占比',
        #                     ct.签收额 /(ct.签收额 + ct.拒收额) - (ct.广告成本 + ct.物流成本 + ct.手续费  + ct.直发采购额) / ct.销售额  AS 利润率,
        #                     (ct.签收额 /(ct.签收额 + ct.拒收额) -(ct.广告成本 + ct.物流成本 + ct.手续费  + ct.直发采购额) / ct.销售额) * (ct.销售额 / ct.订单量) AS 利润值
        #         FROM  ( SELECT 团队,
        #                         新月,
        #                         品类,
        #                         SUM(订量) 订单量,
        #                         SUM(活跃产数) 活跃产品数,SUM(签量) 签收量,SUM(拒量) 拒收量,
        #                         SUM(销额) 销售额,
        #                         SUM(签额) 签收额,
        #                         SUM(拒额) 拒收额,
        #                         SUM(改派订量) 改派订单量,SUM(总采额) 总采购额,
        #                         SUM(直发采额) 直发采购额,
        #                         SUM(广告本) 广告成本,
        #                         SUM(物流本) 物流成本,
        #                         SUM(手费) 手续费,
        #                         总订单量
        #                 FROM( SELECT EXTRACT(YEAR_MONTH FROM ct1.`年月`) 新月,
		# 	                                团队,
		# 	                                品类,
		# 	                                SUM(订单量) 订量,
		# 	                                SUM(活跃产品数) 活跃产数,
		# 	                                SUM(签收量) 签量,
		# 	                                SUM(拒收量) 拒量,
		# 	                                SUM(销售额) 销额,
		# 	                                SUM(签收额) 签额,
		# 	                                SUM(拒收额) 拒额,
		# 	                                SUM(改派订单量) 改派订量,
		# 	                                SUM(总采购额) 总采额,
		# 	                                SUM(直发采购额) 直发采额,
		# 	                                SUM(广告成本) 广告本,
		# 	                                SUM(物流成本) 物流本,
		# 	                                SUM(手续费) 手费
        #                     FROM {0} ct1
        #                     WHERE ct1.`记录时间` = CURDATE()
		#                                 AND ct1.`团队` = '{1}'
		#                                 AND ct1.`品类` <> '未知' AND ct1.`年月` <> '0000-00-00'
        #                     GROUP BY 新月,团队,品类
        #                     ) cs1
        #                 LEFT JOIN
        #                     (SELECT EXTRACT(YEAR_MONTH FROM ct2.`年月`) 新年月,
		# 	                                团队 新团队,
		# 	                                SUM(订单量) 总订单量
        #                     FROM {0} ct2
        #                     WHERE ct2.`记录时间` >= CURDATE()
		#                                 AND ct2.`团队` = '{1}'
		#                                 AND ct2.`品类` <> '未知' AND ct2.`年月` <> '0000-00-00'
        #                     GROUP BY 新年月,新团队
        #                     ORDER BY 新年月,新团队
        #                     ) cs2
        #                 ON cs1.`新月`=cs2.`新年月` AND cs1.`团队`=cs2.`新团队`
        #                 GROUP BY cs1.新月,cs1.品类
        #                 with rollup
        #         ) ct
        #         WHERE ct.新月<>'';'''.format('zg_cost_sltem_copy', team)
        # listT.append(sqlcost3)
        # show_name.append(' 月（月）成本…………')
        # # 月成本（各月）---查询
        # sqlcost4 = '''SELECT ct.团队,
        #                     '' 新月,
        #                     EXTRACT(YEAR_MONTH FROM ct.`年月`) 年月,
        #                     ct.销售额,
		# 	                ct.销售额 / ct.订单量 AS 客单价,
		# 	                ct.订单量,
		# 	                ct.改派订单量,
		# 	                ct.改派订单量 / ct.订单量 改派占比,
		# 	                ct.销售额 / ct.广告成本 ROI,
		# 	                ct.活跃产品数,
		# 	                ct.订单量 / ct.活跃产品数 产能,
		# 	                ct.总采购额 / ct.销售额 AS 采购占比,
		# 	                ct.广告成本 / ct.销售额 AS 花费占比,
		# 	                ct.物流成本 / ct.销售额 AS 运费占比,
		# 	                ct.手续费 / ct.销售额 AS 手续费占比,
		# 	                (ct.总采购额 + ct.广告成本 + ct.物流成本 + ct.手续费) / ct.销售额 AS 总成本占比,
		# 	                ct.签收量 / (ct.签收量 + ct.拒收量) AS '签收/完成',
		# 	                ct.签收量 / ct.订单量 AS '签收/总计',
		# 	                (ct.签收量 + ct.拒收量) / ct.订单量 AS '完成占比'
		# 	    FROM {0} ct
        #         WHERE ct.`品类` <> '未知' AND ct.`团队` = '{1}'
        #         GROUP BY 年月,团队
        #         order by 年月 desc ;'''.format('zg_cost_缓存_month', team)
        # listT.append(sqlcost4)
        # show_name.append(' 月（各月）成本…………')


        # 月时效（天）---查询
        sqltime2 = '''SELECT sl_rb.`币种`,
				sl_rb.`年月`,
				sl_rb.`物流方式`,
				sl_rb.`父级分类`,
				sl_rb.`旬`,
				sl_rb.`总单量`,
				sl_rb.`直发下单出库单量`,
				IFNULL(sl_rb.`直发下单出库时效`,0) 下单出库时效,
				sl_rb.`直发出库完成单量`,
				IFNULL(sl_rb.`直发出库完成时效`,0) 出库完成时效,
				sl_rb.`直发下单完成单量`,
				IFNULL(sl_rb.`直发下单完成时效`,0) 下单完成时效,
				null AS 改派下单完成单量,
				null AS 改派下单完成时效,
				sl_rb.`直发已签收订单量` / sl_rb.`直发下单完成单量` AS '签收/完成',
				sl_rb.`直发已签收订单量`/ sl_rb.`直发下单出库单量` AS '签收/总计'
            FROM( SELECT sl_zong.币种 币种,
						IFNULL(sl_zong.年月,'合计') 年月,
						IFNULL(sl_zong.物流方式,'合计') 物流方式,
						IFNULL(sl_zong.父级分类,'合计') 父级分类,
						IFNULL(sl_zong.旬,'合计') 旬,
						SUM(sl_zong.`总订单量`) 总单量,
						SUM(IFNULL(sl_cx_zf_qs.`直发已签收订单量`,0)) 直发已签收订单量,
						SUM(IFNULL(sl_zong_zf.`直发订单量`,0)) 直发下单出库单量,
						SUM(IFNULL(sl_zong_zf.`直发下单-出库时`,0)) / SUM(IFNULL(sl_zong_zf.`直发订单量`,0)) 直发下单出库时效,
						SUM(IFNULL(sl_cx_zf_wc.`直发出库完成量`,0)) 直发出库完成单量,
						SUM(IFNULL(sl_cx_zf_wc.`直发出库-完成时`,0)) / SUM(IFNULL(sl_cx_zf_wc.`直发出库完成量`,0)) 直发出库完成时效,
						SUM(IFNULL(sl_cx_zf_wc.`直发下单完成量`,0)) 直发下单完成单量,
						SUM(IFNULL(sl_cx_zf_wc.`直发下单-完成时`,0)) /SUM(IFNULL(sl_cx_zf_wc.`直发下单完成量`,0)) 直发下单完成时效
			    FROM (SELECT  币种,
										年月,
										物流方式,
										父级分类,
										旬,
										COUNT(`订单编号`) 总订单量
						FROM  {0}	sl_cx
						WHERE sl_cx.`币种` = '{1}'
						            AND sl_cx.`是否改派` = "直发"
									AND sl_cx.`记录时间`= '{2}'
									AND sl_cx.`父级分类` IS NOT NULL
									AND sl_cx.`系统订单状态` IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)')
						GROUP BY 币种,年月,物流方式,父级分类,旬
						ORDER BY 币种,年月
						) sl_zong
			    LEFT JOIN
						(SELECT 币种,
										年月,
										物流方式,
										父级分类,
										旬,
										COUNT(`订单编号`) 直发订单量,
										SUM(DATEDIFF(`仓储扫描时间`,`下单时间`)) AS '直发下单-出库时'
						FROM  {0}	sl_cx_zf
						WHERE sl_cx_zf.`币种` = '{1}' 
									AND sl_cx_zf.`记录时间`= '{2}'
									AND sl_cx_zf.`父级分类` IS NOT NULL
									AND sl_cx_zf.`系统订单状态` IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)')
									AND sl_cx_zf.`是否改派` = "直发"
									AND sl_cx_zf.`仓储扫描时间` is not null
						GROUP BY 币种,年月,物流方式,父级分类,旬
						ORDER BY 币种,年月
					) sl_zong_zf
					 ON sl_zong_zf.`币种` = sl_zong.`币种` 
								AND sl_zong_zf.`年月` = sl_zong.`年月`
								AND sl_zong_zf.`物流方式` = sl_zong.`物流方式`
						  	AND sl_zong_zf.`父级分类` = sl_zong.`父级分类` 
								AND sl_zong_zf.`旬` = sl_zong.`旬` 
					LEFT JOIN
						(SELECT 币种,
										年月,
										物流方式,
										父级分类,
										旬,
										COUNT(`订单编号`) 直发已签收订单量
						FROM  {0}	sl_cx_zf_qianshou
						WHERE sl_cx_zf_qianshou.`币种` = '{1}' 
									AND sl_cx_zf_qianshou.`记录时间`= '{2}'
									AND sl_cx_zf_qianshou.`父级分类` IS NOT NULL
									AND sl_cx_zf_qianshou.`系统订单状态` IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)')
									AND sl_cx_zf_qianshou.`是否改派` = "直发"
									AND sl_cx_zf_qianshou.`仓储扫描时间` is not null
									AND sl_cx_zf_qianshou.`最终状态` = "已签收"
						GROUP BY 币种,年月,物流方式,父级分类,旬
						ORDER BY 币种,年月
					) sl_cx_zf_qs
					 ON sl_cx_zf_qs.`币种` = sl_zong.`币种` 
								AND sl_cx_zf_qs.`年月` = sl_zong.`年月`
								AND sl_cx_zf_qs.`物流方式` = sl_zong.`物流方式`
						  	AND sl_cx_zf_qs.`父级分类` = sl_zong.`父级分类`
								AND sl_cx_zf_qs.`旬` = sl_zong.`旬`
				LEFT JOIN
						(SELECT 币种,
										年月,
										物流方式,
										父级分类,
										旬,
										COUNT(`订单编号`) 直发出库完成量,
										SUM(DATEDIFF(IFNULL(`完结状态时间`,`状态时间`),`仓储扫描时间`)) AS '直发出库-完成时',
										COUNT(`订单编号`) 直发下单完成量,
										SUM(DATEDIFF(IFNULL(`完结状态时间`,`状态时间`),`下单时间`)) AS '直发下单-完成时'
						FROM  {0}	sl_cx_zf_wancheng
						WHERE sl_cx_zf_wancheng.`币种` = '{1}'
								AND sl_cx_zf_wancheng.`记录时间`= '{2}'
							  AND sl_cx_zf_wancheng.`父级分类` IS NOT NULL
							  AND sl_cx_zf_wancheng.`系统订单状态` IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)')
								AND sl_cx_zf_wancheng.`是否改派` = "直发"
								AND sl_cx_zf_wancheng.`最终状态`IN ('拒收', '理赔', '已签收', '已退货') 
						GROUP BY 币种,年月,物流方式,父级分类,旬
						ORDER BY 币种,年月
					) sl_cx_zf_wc
					 ON sl_cx_zf_wc.`币种` = sl_zong.`币种` 
								AND sl_cx_zf_wc.`年月` = sl_zong.`年月`
								AND sl_cx_zf_wc.`物流方式` = sl_zong.`物流方式`
								AND sl_cx_zf_wc.`父级分类` = sl_zong.`父级分类`
								AND sl_cx_zf_wc.`旬` = sl_zong.`旬`
				GROUP BY sl_zong.年月,sl_zong.物流方式,sl_zong.旬
				with rollup
			) sl_rb;'''.format(match2[team], team, Time_day[11])
        listT.append(sqltime2)
        show_name.append(' 月（天）时效…………')
        # 月时效（旬）---查询
        sqltime3 = '''SELECT sl_rb.`币种`,
				sl_rb.`年月`,
				sl_rb.`旬`,
				sl_rb.`物流方式`,
				sl_rb.`父级分类`,
				sl_rb.`总单量`,
				sl_rb.`直发下单出库单量`,
				IFNULL(sl_rb.`直发下单出库时效`,0) 下单出库时效,
				sl_rb.`直发出库完成单量`,
				IFNULL(sl_rb.`直发出库完成时效`,0) 出库完成时效,
				sl_rb.`直发下单完成单量`,
				IFNULL(sl_rb.`直发下单完成时效`,0) 下单完成时效,
				null AS 改派下单完成单量,
				null AS 改派下单完成时效
            FROM(SELECT sl_zong.币种 币种,
						IFNULL(sl_zong.年月,'合计') 年月,
						IFNULL(sl_zong.旬,'合计') 旬,
						IFNULL(sl_zong.物流方式,'合计') 物流方式,
						IFNULL(sl_zong.父级分类,'合计') 父级分类,
						SUM(sl_zong.`总订单量`) 总单量,
						SUM(IFNULL(sl_zong_zf.`直发订单量`,0)) 直发下单出库单量,
						SUM(IFNULL(sl_zong_zf.`直发下单-出库时`,0)) / SUM(IFNULL(sl_zong_zf.`直发订单量`,0)) 直发下单出库时效,
						SUM(IFNULL(sl_cx_zf_wc.`直发出库完成量`,0)) 直发出库完成单量,
						SUM(IFNULL(sl_cx_zf_wc.`直发出库-完成时`,0)) / SUM(IFNULL(sl_cx_zf_wc.`直发出库完成量`,0)) 直发出库完成时效,
						SUM(IFNULL(sl_cx_zf_wc.`直发下单完成量`,0)) 直发下单完成单量,
						SUM(IFNULL(sl_cx_zf_wc.`直发下单-完成时`,0)) /SUM(IFNULL(sl_cx_zf_wc.`直发下单完成量`,0)) 直发下单完成时效
			    FROM ( SELECT 币种,
										年月,
										旬,
										物流方式,
										父级分类,
										COUNT(`订单编号`) 总订单量
						FROM  {0}	sl_cx
						WHERE sl_cx.`币种` = '{1}' 
								  AND sl_cx.`记录时间`= '{2}'
									AND sl_cx.`父级分类` IS NOT NULL
									AND sl_cx.`是否改派` = "直发"
									AND sl_cx.`系统订单状态` IN ('已审核', '待发货', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)', '待发货转审核') 
						GROUP BY 币种,年月,旬,物流方式,父级分类
						ORDER BY 币种,年月
						) sl_zong
			    LEFT JOIN
						(SELECT 币种,
										年月,
										旬,
										物流方式,
										父级分类,			
										COUNT(`订单编号`) 直发订单量,
										SUM(DATEDIFF(`仓储扫描时间`,`下单时间`)) AS '直发下单-出库时'
						FROM  {0}	sl_cx_zf
						WHERE sl_cx_zf.`币种` = '{1}'
								  AND sl_cx_zf.`记录时间`= '{2}'
									AND sl_cx_zf.`父级分类` IS NOT NULL
									AND sl_cx_zf.`系统订单状态` IN ('已审核', '待发货', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)', '待发货转审核') 
									AND sl_cx_zf.`是否改派` = "直发"
									AND sl_cx_zf.`仓储扫描时间` is not null
						GROUP BY 币种,年月,旬,物流方式,父级分类
						ORDER BY 币种,年月
					) sl_zong_zf
					 ON sl_zong_zf.`币种` = sl_zong.`币种` 
								AND sl_zong_zf.`年月` = sl_zong.`年月`
								AND sl_zong_zf.`物流方式` = sl_zong.`物流方式`
						  	AND sl_zong_zf.`父级分类` = sl_zong.`父级分类` 
								AND sl_zong_zf.`旬` = sl_zong.`旬` 
				LEFT JOIN
						(SELECT 币种,
										年月,
										旬,
										物流方式,
										父级分类,		
										COUNT(`订单编号`) 直发出库完成量,
										SUM(DATEDIFF(IFNULL(`完结状态时间`,`状态时间`),`仓储扫描时间`)) AS '直发出库-完成时',
										COUNT(`订单编号`) 直发下单完成量,
										SUM(DATEDIFF(IFNULL(`完结状态时间`,`状态时间`),`下单时间`)) AS '直发下单-完成时'
						FROM  {0}	sl_cx_zf_wancheng
						WHERE sl_cx_zf_wancheng.`币种` = '{1}'
                AND sl_cx_zf_wancheng.`记录时间`= '{2}'
							  AND sl_cx_zf_wancheng.`父级分类` IS NOT NULL
								AND sl_cx_zf_wancheng.`系统订单状态` IN ('已审核', '待发货', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)', '待发货转审核') 
								AND sl_cx_zf_wancheng.`是否改派` = "直发"
								AND sl_cx_zf_wancheng.`最终状态`IN ('拒收', '理赔', '已签收', '已退货') 
						GROUP BY 币种,年月,旬,物流方式,父级分类
						ORDER BY 币种,年月
					) sl_cx_zf_wc
					 ON sl_cx_zf_wc.`币种` = sl_zong.`币种` 
								AND sl_cx_zf_wc.`年月` = sl_zong.`年月`
								AND sl_cx_zf_wc.`物流方式` = sl_zong.`物流方式`
								AND sl_cx_zf_wc.`父级分类` = sl_zong.`父级分类` 
								AND sl_cx_zf_wc.`旬` = sl_zong.`旬` 	
				GROUP BY sl_zong.年月,sl_zong.旬,sl_zong.物流方式,sl_zong.父级分类
				with rollup
            ) sl_rb;'''.format(match2[team], team, Time_day[11])
        listT.append(sqltime3)
        show_name.append(' 月（旬）时效…………')
        # 月时效(各月)---查询
        sqltime4 = '''SELECT sl_rb.`币种`,
				sl_rb.`年月`,
				sl_rb.`物流方式`,
				sl_rb.`父级分类`,
				sl_rb.`旬`,
				sl_rb.`总单量`,
				sl_rb.`直发下单出库单量`,
				sl_rb.`直发下单出库时效`,
				sl_rb.`直发出库完成单量`,
				sl_rb.`直发出库完成时效`,
				sl_rb.`直发下单完成时效`,
				sl_rb.`直发下单完成单量`,
				null AS 改派下单完成单量,
				null AS 改派下单完成时效,
				sl_rb.`直发已签收订单量` / sl_rb.`直发下单完成单量` AS '签收/完成',
				sl_rb.`直发已签收订单量`/ sl_rb.`直发下单出库单量` AS '签收/总计'
            FROM (SELECT sl_zong.币种 币种,
						IFNULL(sl_zong.年月,'合计') 年月,
						IFNULL(sl_zong.物流方式,'合计') 物流方式,
						IFNULL(sl_zong.父级分类,'合计') 父级分类,
						IFNULL(sl_zong.旬,'合计') 旬,
						SUM(sl_zong.`总订单量`) 总单量,
						SUM(IFNULL(sl_cx_zf_qs.`直发已签收订单量`,0)) 直发已签收订单量,
						SUM(IFNULL(sl_zong_zf.`直发订单量`,0)) 直发下单出库单量,
						SUM(IFNULL(sl_zong_zf.`直发下单-出库时`,0)) / SUM(IFNULL(sl_zong_zf.`直发订单量`,0)) 直发下单出库时效,
						SUM(IFNULL(sl_cx_zf_wc.`直发出库完成量`,0)) 直发出库完成单量,
						SUM(IFNULL(sl_cx_zf_wc.`直发出库-完成时`,0)) / SUM(IFNULL(sl_cx_zf_wc.`直发出库完成量`,0)) 直发出库完成时效,
						SUM(IFNULL(sl_cx_zf_wc.`直发下单完成量`,0)) 直发下单完成单量,
						SUM(IFNULL(sl_cx_zf_wc.`直发下单-完成时`,0)) /SUM(IFNULL(sl_cx_zf_wc.`直发下单完成量`,0)) 直发下单完成时效
			    FROM (SELECT  币种,
										年月,
										物流方式,
										父级分类,
										旬,
										COUNT(`订单编号`) 总订单量
						FROM  {0}	sl_cx
						WHERE sl_cx.`币种` = '{1}' 
								AND sl_cx.`父级分类` IS NOT NULL
								AND sl_cx.`是否改派` = "直发"
								AND sl_cx.`系统订单状态` IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)') 
						GROUP BY 币种,年月,物流方式,父级分类,旬
						ORDER BY 币种,年月
						) sl_zong
			    LEFT JOIN
						(SELECT 币种,
										年月,
										物流方式,
										父级分类,
													旬,
										COUNT(`订单编号`) 直发订单量,
										SUM(DATEDIFF(`仓储扫描时间`,`下单时间`)) AS '直发下单-出库时'
						FROM  {0}	sl_cx_zf
						WHERE sl_cx_zf.`币种` = '{1}'
								AND sl_cx_zf.`父级分类` IS NOT NULL
								AND sl_cx_zf.`系统订单状态` IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)') 
								AND sl_cx_zf.`是否改派` = "直发"
								AND sl_cx_zf.`仓储扫描时间` is not null
						GROUP BY 币种,年月,物流方式,父级分类,旬
						ORDER BY 币种,年月
					) sl_zong_zf
					 ON sl_zong_zf.`币种` = sl_zong.`币种` 
							AND sl_zong_zf.`年月` = sl_zong.`年月`
							AND sl_zong_zf.`物流方式` = sl_zong.`物流方式`
						  	AND sl_zong_zf.`父级分类` = sl_zong.`父级分类` 
							AND sl_zong_zf.`旬` = sl_zong.`旬` 	
					LEFT JOIN
						(SELECT 币种,
										年月,
										物流方式,
										父级分类,
													旬,
										COUNT(`订单编号`) 直发已签收订单量
						FROM  {0}	sl_cx_zf_qianshou
						WHERE sl_cx_zf_qianshou.`币种` = '{1}' 
								AND sl_cx_zf_qianshou.`父级分类` IS NOT NULL
								AND sl_cx_zf_qianshou.`系统订单状态` IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)') 
								AND sl_cx_zf_qianshou.`是否改派` = "直发"
								AND sl_cx_zf_qianshou.`仓储扫描时间` is not null
								AND sl_cx_zf_qianshou.`最终状态` = "已签收"
						GROUP BY 币种,年月,物流方式,父级分类,旬
						ORDER BY 币种,年月
					) sl_cx_zf_qs
					 ON sl_cx_zf_qs.`币种` = sl_zong.`币种` 
							AND sl_cx_zf_qs.`年月` = sl_zong.`年月`
							AND sl_cx_zf_qs.`物流方式` = sl_zong.`物流方式`
						  	AND sl_cx_zf_qs.`父级分类` = sl_zong.`父级分类` 
							AND sl_cx_zf_qs.`旬` = sl_zong.`旬` 	
				LEFT JOIN
						(SELECT 币种,
									年月,
									物流方式,
									父级分类,
										    旬,
									COUNT(`订单编号`) 直发出库完成量,
									SUM(DATEDIFF(IFNULL(`完结状态时间`,`状态时间`),`仓储扫描时间`)) AS '直发出库-完成时',
									COUNT(`订单编号`) 直发下单完成量,
									SUM(DATEDIFF(IFNULL(`完结状态时间`,`状态时间`),`下单时间`)) AS '直发下单-完成时'
						FROM  {0}	sl_cx_zf_wancheng
						WHERE sl_cx_zf_wancheng.`币种` = '{1}'
							AND sl_cx_zf_wancheng.`父级分类` IS NOT NULL
							AND sl_cx_zf_wancheng.`系统订单状态` IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)') 
							AND sl_cx_zf_wancheng.`是否改派` = "直发"
							AND sl_cx_zf_wancheng.`最终状态`IN ('拒收', '理赔', '已签收', '已退货') 
						GROUP BY 币种,年月,物流方式,父级分类,旬
						ORDER BY 币种,年月
					) sl_cx_zf_wc
					 ON sl_cx_zf_wc.`币种` = sl_zong.`币种` 
						AND sl_cx_zf_wc.`年月` = sl_zong.`年月`
						AND sl_cx_zf_wc.`物流方式` = sl_zong.`物流方式`
						AND sl_cx_zf_wc.`父级分类` = sl_zong.`父级分类` 
						AND sl_cx_zf_wc.`旬` = sl_zong.`旬`
				GROUP BY sl_zong.年月,sl_zong.物流方式,sl_zong.父级分类,sl_zong.旬
				with rollup) sl_rb;'''.format('qsb_缓存_month', team)
        listT.append(sqltime4)
        show_name.append(' 月(各月)时效…………')
        listTValue = []                                # 查询sql的结果 存放池
        for i, sql in enumerate(listT):
            print('正在获取 ' + team + show_name[i])
            df = pd.read_sql_query(sql=sql, con=self.engine1)
            print(df)
            columns = list(df.columns)                 # 获取数据的标题名，转为列表
            columns_value = ['采购/销售额', '直发采购/销售额', '运费占比', '手续费占比', '金额签收/完成', '金额签收/总计', '金额完成占比', '数量签收/完成', '数量完成占比',
                             '签收/完成', '签收/总计', '完成占比', '总签收/完成', '总签收/总计', '退款率', '总完成占比', '直发签收/完成', '直发签收/总计', '直发完成占比',
                             '改派签收/完成', '改派签收/总计', '改派完成占比', '总签收/完成(金额)', '总签收/总计(金额)', '退款率(金额)', '总完成占比(金额)', '直发签收/完成(金额)',
                             '直发签收/总计(金额)', '直发完成占比(金额)', '改派签收/完成(金额)', '改派签收/总计(金额)', '改派完成占比(金额)', '订单品类占比', '直发采购额/销售额',
                             '花费占比', '总成本', '利润率', '改派占比', '采购占比', '广告占比', '总成本占比', '签收/完成', '签收/总计', '完成占比']
            for column_val in columns_value:
                if column_val in columns:
                    try:
                        df[column_val] = df[column_val].fillna(value=0)
                        df[column_val] = df[column_val].apply(lambda x: format(x, '.2%'))
                    except Exception as e:
                        print('修改失败：', str(Exception) + str(e) + df[column_val])
            listTValue.append(df)
        print('查询耗时：', datetime.datetime.now() - start)
        today = datetime.date.today().strftime('%Y.%m.%d')
        # sheet_name = ['签率(天)_', '签率(月)_', '签率(旬)_', '签率(总)_', '物流(天)_', '物流(月)_', '成本(天)_', '成本(月)_', '成本(总)_', '时效(天)_', '时效(旬)_', '时效(总)_']  # 生成的工作表的表名
        sheet_name = ['签率(天)_', '签率(月)_', '签率(旬)_', '签率(总)_', '物流(天)_', '物流(月)_', '时效(天)_', '时效(旬)_', '时效(总)_']  # 生成的工作表的表名
        file_Path = []                                # 发送邮箱文件使用
        filePath = ''
        if team == '日本':
            filePath = 'F:\\查询\\日本监控\\{} {}监控表.xlsx'.format(today, team)
        elif team == '泰国':
            filePath = 'F:\\查询\\泰国监控\\{} {}监控表.xlsx'.format(today, team)
        elif team == '新加坡' or team == '马来西亚' or team == '菲律宾':
            filePath = 'F:\\查询\\新马监控\\{} {}监控表.xlsx'.format(today, team)
        elif team == '香港' or team == '台湾':
            filePath = 'F:\\查询\\港台监控\\{} {}监控表.xlsx'.format(today, team)
        if os.path.exists(filePath):                  # 判断是否有需要的表格
            print("正在使用(上月)文件......")
            filePath = filePath
        else:                                         # 判断是否无需要的表格，进行初始化创建
            print("正在创建文件......")
            df0 = pd.DataFrame([])                    # 创建空的dataframe数据框
            df0.to_excel(filePath, index=False)       # 备用：可以向不同的sheet写入数据（创建新的工作表并进行写入）
            filePath = filePath
        print('正在写入excel…………')
        writer = pd.ExcelWriter(filePath, engine='openpyxl')  # 初始化写入对象
        book = load_workbook(filePath)                # 可以向不同的sheet写入数据（对现有工作表的追加）
        writer.book = book                            # 将数据写入excel中的sheet2表,sheet_name改变后即是新增一个sheet
        for i in range(len(listTValue)):
            listTValue[i].to_excel(excel_writer=writer, sheet_name=sheet_name[i] + team, index=False)
        if 'Sheet1' in book.sheetnames:               # 删除新建文档时的第一个工作表
            del book['Sheet1']
        writer.save()
        writer.close()
        print('正在运行宏…………')
        app = xl.App(visible=False, add_book=False)   # 运行宏调整
        app.display_alerts = False
        wbsht = app.books.open('D:/Users/Administrator/Desktop/新版-格式转换(工具表).xlsm')
        wbsht1 = app.books.open(filePath)
        wbsht.macro('sl_总监控运行')()
        wbsht1.save()
        wbsht1.close()
        wbsht.close()
        app.quit()
        print('输出(监控)文件成功…………')
        file_Path.append(filePath)
        self.e.send('{} {}监控表.xlsx'.format(today, team), file_Path,
                    emailAdd[team])
        if team =='泰国':
            self.e.send('{} {}监控表.xlsx'.format(today, team), file_Path,
                        emailAdd2[team])
        print('处理耗时：', datetime.datetime.now() - start)

        # 单月获取数据对比：
    def sl_Monitoring_two(self,team):
        match2 = {'新加坡': 'qsb_slxmt',
                  '马来西亚': 'qsb_slxmt',
                  '菲律宾': 'qsb_slxmt',
                  '日本': 'qsb_slrb',
                  '香港': 'qsb_slgat',
                  '台湾': 'qsb_slgat',
                  '泰国': 'qsb_sltg'}
        match3 = {'日本': r'D:\Users\Administrator\Desktop\查询\日本监控',
                  '泰国': r'D:\Users\Administrator\Desktop\查询\泰国监控',
                  '香港': r'D:\Users\Administrator\Desktop\查询\港台监控',
                  '台湾': r'D:\Users\Administrator\Desktop\查询\港台监控',
                  '菲律宾': r'D:\Users\Administrator\Desktop\查询\新马监控',
                  '新加坡': r'D:\Users\Administrator\Desktop\查询\新马监控',
                  '马来西亚': r'D:\Users\Administrator\Desktop\查询\新马监控'}
        emailAdd = {'香港': 'giikinliujun@163.com',
                    '台湾': 'giikinliujun@163.com',
                    '泰国': '1845389861@qq.com',
                    '新加坡': 'zhangjing@giikin.com',
                    '马来西亚': 'zhangjing@giikin.com',
                    '菲律宾': 'zhangjing@giikin.com',
                    '日本': 'sunyaru@giikin.com'}
        emailAdd2 = {'泰国': 'zhangjing@giikin.com'}
        today = (datetime.datetime.now() + datetime.timedelta(days=1)).strftime('%Y-%m-%d')
        start: datetime = datetime.datetime.now()
        Time_day = []
        for i in range(1, datetime.datetime.now().month + 1):  # 获取当年当前的月份时间
            try:
                daytime = (datetime.datetime.now().replace(month=i)).strftime('%Y-%m') + (
                    (datetime.datetime.now()).strftime('-%d'))
                Time_day.append(daytime)
            except Exception as e:
                print('xxxx时间配置出错,已手动调整：' + str(i) + '月份', str(Exception) + str(e))
                Time_day.append(
                    str(int(datetime.datetime.now().year)) + '-' + str(i) + (datetime.datetime.now().strftime('-%d')))
        for i in range(datetime.datetime.now().month + 1, 13):  # 获取往年当前的月份时间
            try:
                daytime = str(int(datetime.datetime.now().year) -1) + (datetime.datetime.now().replace(month=i)).strftime('-%m') + (
                    (datetime.datetime.now()).strftime('-%d'))
                Time_day.append(daytime)
            except Exception as e:
                print('xxxx时间配置出错失败00：' + str(i) + '月份', str(Exception) + str(e))
                Time_day.append(str(int(datetime.datetime.now().year) - 1) + '-' + str(i) + (
                    datetime.datetime.now().strftime('-%d')))
        # Time_day = ['2021-02-24', '2020-12-19', '2020-12-14', '2020-12-14', '2020-12-14', '2020-12-14', '2020-12-14', '2020-12-14', '2020-12-14', '2020-12-14', '2020-12-14', '2021-01-24']
        # 对时间数组进行排序  list.sort(cmp=None, key=None, reverse=False)；reverse -- 排序规则，reverse = True 降序， reverse = False 升序（默认）
        Time_day.sort()
        print(Time_day[11])
        print(Time_day[10])
        listT = []                              # 查询sql 存放池
        show_name = []                          # 打印进度需要
        # 月签收率（天）---查询
        sqlqsb2 = '''SELECT sl_gat.`币种`,sl_gat.`年月`,sl_gat.父级分类,sl_gat.二级分类,sl_gat.三级分类,sl_gat.物流方式,sl_gat.旬,
            	            sl_gat.`总订单量`,
							sl_gat.`已签收订单量` / sl_gat.`拒收订单量` AS '总签收/完成',
            	            sl_gat.`已签收订单量` / sl_gat.`总订单量` AS '总签收/总计',
							sl_gat.`退货订单量` / sl_gat.`总订单量` AS '退款率',
            	            sl_gat.`拒收订单量` / sl_gat.`总订单量` AS '总完成占比',
            	            sl_gat.`直发订单量` 直发总计,
							sl_gat.`直发已签收订单量` / sl_gat.`直发拒收订单量` AS '直发签收/完成',
            	            sl_gat.`直发已签收订单量` / sl_gat.`直发订单量` AS '直发签收/总计', 
							sl_gat.`直发拒收订单量` / sl_gat.`直发订单量` AS '直发完成占比',
            	            sl_gat.`改派订单量` 改派总计,
							sl_gat.`改派已签收订单量` / sl_gat.`改派拒收订单量` AS '改派签收/完成',
            	            sl_gat.`改派已签收订单量` / sl_gat.`改派订单量` AS '改派签收/总计',
							sl_gat.`改派拒收订单量` / sl_gat.`改派订单量` AS '改派完成占比'
					FROM (SELECT  sl_zong.币种,
					    IFNULL(sl_zong.年月,'合计') 年月,
						IFNULL(sl_zong.父级分类,'合计') 父级分类,
						IFNULL(sl_zong.二级分类,'合计') 二级分类,
                        IFNULL(sl_zong.三级分类,'合计') 三级分类,
                        IFNULL(sl_zong.物流方式,'合计') 物流方式,
						IFNULL(sl_zong.旬,'合计') 旬,
						SUM(总订单量) 总订单量,
                        IFNULL(SUM(直发订单量),0) 直发订单量,
                        IFNULL(SUM(直发已签收订单量),0) 直发已签收订单量,
                        IFNULL(SUM(直发拒收订单量),0) 直发拒收订单量,
                        (SUM(总订单量) - IFNULL(SUM(直发订单量),0)) AS 改派订单量,
                        IFNULL(SUM(改派已签收订单量),0) 改派已签收订单量,
                        IFNULL(SUM(改派拒收订单量),0) 改派拒收订单量,
						IFNULL(SUM(已签收订单量),0) 已签收订单量,
						IFNULL(SUM(拒收订单量),0) 拒收订单量,
						IFNULL(SUM(退货订单量),0) 退货订单量
            		FROM (SELECT  币种,年月,父级分类,二级分类,三级分类,物流方式,旬,COUNT(`订单编号`) 总订单量
            				FROM  {0}	sl_cx
							WHERE (sl_cx.`记录时间`= '{2}' AND sl_cx.`日期` < DATE_SUB('{2}', INTERVAL DAY('{2}')-1 DAY)
								or sl_cx.`记录时间`= '{3}' AND sl_cx.`日期` < DATE_SUB('{3}', INTERVAL DAY('{3}')-1 DAY))
								AND sl_cx.`币种` = '{1}' 
								AND sl_cx.`父级分类` IS NOT NULL  
								AND sl_cx.`系统订单状态` IN {4} 
            				GROUP BY {5}
            				ORDER BY 币种,年月
            				) sl_zong
            		LEFT JOIN
            				(SELECT 币种,年月,父级分类,二级分类,三级分类,物流方式,旬,COUNT(`订单编号`) 直发订单量
            				FROM  {0}	sl_zong_zf
            				WHERE (sl_zong_zf.`记录时间`= '{2}' AND sl_zong_zf.`日期` < DATE_SUB('{2}', INTERVAL DAY('{2}')-1 DAY)
                                or sl_zong_zf.`记录时间`= '{3}' AND sl_zong_zf.`日期` < DATE_SUB('{3}', INTERVAL DAY('{3}')-1 DAY))
							    AND sl_zong_zf.`币种` = '{1}' 
							    AND sl_zong_zf.`父级分类` IS NOT NULL
            				    AND sl_zong_zf.`是否改派` = "直发"
            				    AND sl_zong_zf.`系统订单状态` IN {4} 
            				GROUP BY {5}
            				ORDER BY 币种,年月
            				) sl_zong_zf 
            		ON sl_zong_zf.`币种` = sl_zong.`币种` AND sl_zong_zf.`年月` = sl_zong.`年月`AND sl_zong_zf.`父级分类` = sl_zong.`父级分类` 
            			AND sl_zong_zf.`二级分类` = sl_zong.`二级分类` AND sl_zong_zf.`三级分类` = sl_zong.`三级分类` 
            			AND sl_zong_zf.`物流方式` = sl_zong.`物流方式` AND sl_zong_zf.`旬` = sl_zong.`旬` 							
            		LEFT JOIN
            				(SELECT 币种,年月,父级分类,二级分类,三级分类,物流方式,旬,COUNT(`订单编号`) 直发已签收订单量
            				FROM  {0}	slzf_qs
            				WHERE (slzf_qs.`记录时间`= '{2}' AND slzf_qs.`日期` < DATE_SUB('{2}', INTERVAL DAY('{2}')-1 DAY)
                                or slzf_qs.`记录时间`= '{3}' AND slzf_qs.`日期` < DATE_SUB('{3}', INTERVAL DAY('{3}')-1 DAY))
								AND slzf_qs.`币种` = '{1}' 
								AND slzf_qs.`父级分类` IS NOT NULL
								AND slzf_qs.`是否改派` = "直发" 
								AND slzf_qs.`最终状态` = "已签收" 
								AND slzf_qs.`系统订单状态` IN {4} 
            				GROUP BY {5}
            				ORDER BY 币种,年月
            				) sl_cx_zf_qs
            		ON sl_cx_zf_qs.`币种` = sl_zong.`币种` AND sl_cx_zf_qs.`年月` = sl_zong.`年月` AND sl_cx_zf_qs.`父级分类` = sl_zong.`父级分类` 
            			AND sl_cx_zf_qs.`二级分类` = sl_zong.`二级分类` AND sl_cx_zf_qs.`三级分类` = sl_zong.`三级分类` 
            			AND sl_cx_zf_qs.`物流方式` = sl_zong.`物流方式` AND sl_cx_zf_qs.`旬` = sl_zong.`旬` 		
            		LEFT JOIN
            				(SELECT 币种,年月,父级分类,二级分类,三级分类,物流方式,旬,COUNT(`订单编号`) 直发拒收订单量
            				FROM  {0}	slzf_js
            				WHERE (slzf_js.`记录时间`= '{2}' AND slzf_js.`日期` < DATE_SUB('{2}', INTERVAL DAY('{2}')-1 DAY)
                                or slzf_js.`记录时间`= '{3}' AND slzf_js.`日期` < DATE_SUB('{3}', INTERVAL DAY('{3}')-1 DAY))
								AND slzf_js.`币种` = '{1}' 
								AND slzf_js.`父级分类` IS NOT NULL
            					AND slzf_js.`是否改派` = "直发"  
            					AND slzf_js.`最终状态` IN ('拒收', '理赔', '已签收', '已退货') 
            					AND slzf_js.`系统订单状态`IN {4}
            				GROUP BY {5}
            				ORDER BY 币种,年月
            				) sl_cx_zf_js
            		ON sl_cx_zf_js.`币种` = sl_zong.`币种` AND sl_cx_zf_js.`年月` = sl_zong.`年月` AND sl_cx_zf_js.`父级分类` = sl_zong.`父级分类` 
            			AND sl_cx_zf_js.`二级分类` = sl_zong.`二级分类` AND sl_cx_zf_js.`三级分类` = sl_zong.`三级分类` 
            			AND sl_cx_zf_js.`物流方式` = sl_zong.`物流方式` AND sl_cx_zf_js.`旬` = sl_zong.`旬` 
            		LEFT JOIN
            				(SELECT 币种,年月,父级分类,二级分类,三级分类,物流方式,旬,COUNT(`订单编号`) 改派订单量
            				FROM  {0}	sl_cx_gp
            				WHERE (sl_cx_gp.`记录时间`= '{2}' AND sl_cx_gp.`日期` < DATE_SUB('{2}', INTERVAL DAY('{2}')-1 DAY)
                                or sl_cx_gp.`记录时间`= '{3}' AND sl_cx_gp.`日期` < DATE_SUB('{3}', INTERVAL DAY('{3}')-1 DAY))
								AND sl_cx_gp.`币种` = '{1}' 
								AND sl_cx_gp.`父级分类` IS NOT NULL
            					AND sl_cx_gp.`是否改派` = "改派"
								AND sl_cx_gp.`系统订单状态` IN {4} 
            				GROUP BY {5}
            				ORDER BY 币种,年月
            				) sl_zong_gp
            		ON sl_zong_gp.`币种` = sl_zong.`币种` AND sl_zong_gp.`年月` = sl_zong.`年月` AND sl_zong_gp.`父级分类` = sl_zong.`父级分类` 
            			AND sl_zong_gp.`二级分类` = sl_zong.`二级分类` AND sl_zong_gp.`三级分类` = sl_zong.`三级分类` 
            			AND sl_zong_gp.`物流方式` = sl_zong.`物流方式` AND sl_zong_gp.`旬` = sl_zong.`旬` 
            		LEFT JOIN
            				(SELECT 币种,年月,父级分类,二级分类,三级分类,物流方式,旬,COUNT(`订单编号`) 改派已签收订单量
            				FROM  {0}	slgp_qs
            				WHERE (slgp_qs.`记录时间`= '{2}' AND slgp_qs.`日期` < DATE_SUB('{2}', INTERVAL DAY('{2}')-1 DAY)
                                or slgp_qs.`记录时间`= '{3}' AND slgp_qs.`日期` < DATE_SUB('{3}', INTERVAL DAY('{3}')-1 DAY))
								AND slgp_qs.`币种` = '{1}' 
								AND slgp_qs.`父级分类` IS NOT NULL
            					AND slgp_qs.`是否改派` = "改派"  
            					AND slgp_qs.`最终状态` = "已签收"
            					AND slgp_qs.`系统订单状态`IN {4}
            				GROUP BY {5}
            				ORDER BY 币种,年月
            				) sl_cx_gp_qs
            			ON sl_cx_gp_qs.`币种` = sl_zong.`币种` AND sl_cx_gp_qs.`年月` = sl_zong.`年月` AND sl_cx_gp_qs.`父级分类` = sl_zong.`父级分类` 
            				AND sl_cx_gp_qs.`二级分类` = sl_zong.`二级分类` AND sl_cx_gp_qs.`三级分类` = sl_zong.`三级分类` 
            				AND sl_cx_gp_qs.`物流方式` = sl_zong.`物流方式` AND sl_cx_gp_qs.`旬` = sl_zong.`旬` 
            		LEFT JOIN
            				(SELECT 币种,年月,父级分类,二级分类,三级分类,物流方式,旬,COUNT(`订单编号`) 改派拒收订单量
            				FROM  {0}	sl_cx_gp_jushou
            				WHERE (sl_cx_gp_jushou.`记录时间`= '{2}' AND sl_cx_gp_jushou.`日期` < DATE_SUB('{2}', INTERVAL DAY('{2}')-1 DAY)
                                or sl_cx_gp_jushou.`记录时间`= '{3}' AND sl_cx_gp_jushou.`日期` < DATE_SUB('{3}', INTERVAL DAY('{3}')-1 DAY))
								AND sl_cx_gp_jushou.`币种` = '{1}' 
								AND sl_cx_gp_jushou.`父级分类` IS NOT NULL
            					AND sl_cx_gp_jushou.`是否改派` = "改派"  
            					AND sl_cx_gp_jushou.`最终状态` IN ('拒收', '理赔', '已签收', '已退货') 
								AND sl_cx_gp_jushou.`系统订单状态` IN {4}
            				GROUP BY {5}
            				ORDER BY 币种,年月
            				) sl_cx_gp_js
            			ON sl_cx_gp_js.`币种` = sl_zong.`币种` AND sl_cx_gp_js.`年月` = sl_zong.`年月` AND sl_cx_gp_js.`父级分类` = sl_zong.`父级分类` 
            				AND sl_cx_gp_js.`二级分类` = sl_zong.`二级分类` AND sl_cx_gp_js.`三级分类` = sl_zong.`三级分类` 
            				AND sl_cx_gp_js.`物流方式` = sl_zong.`物流方式` AND sl_cx_gp_js.`旬` = sl_zong.`旬` 				
            		LEFT JOIN
            				(SELECT 币种,年月,父级分类,二级分类,三级分类,物流方式,旬,COUNT(`订单编号`) 已签收订单量
            				FROM  {0}	sl_cx_qianshou
            				WHERE (sl_cx_qianshou.`记录时间`= '{2}' AND sl_cx_qianshou.`日期` < DATE_SUB('{2}', INTERVAL DAY('{2}')-1 DAY)
                                or sl_cx_qianshou.`记录时间`= '{3}' AND sl_cx_qianshou.`日期` < DATE_SUB('{3}', INTERVAL DAY('{3}')-1 DAY))
								AND sl_cx_qianshou.`币种` = '{1}' 
								AND sl_cx_qianshou.`父级分类` IS NOT NULL
            					AND sl_cx_qianshou.`最终状态` = "已签收"
								AND sl_cx_qianshou.`系统订单状态`IN {4} 
            				GROUP BY {5}
            				ORDER BY 币种,年月
            				) sl_zong_qianshou
            			ON sl_zong_qianshou.`币种` = sl_zong.`币种` AND sl_zong_qianshou.`年月` = sl_zong.`年月` AND sl_zong_qianshou.`父级分类` = sl_zong.`父级分类` 
            				AND sl_zong_qianshou.`二级分类` = sl_zong.`二级分类`  AND sl_zong_qianshou.`三级分类` = sl_zong.`三级分类` 
            				AND sl_zong_qianshou.`物流方式` = sl_zong.`物流方式` AND sl_zong_qianshou.`旬` = sl_zong.`旬`
            		LEFT JOIN
            				(SELECT 币种,年月,父级分类,二级分类,三级分类,物流方式,旬,COUNT(`订单编号`) 拒收订单量
            				FROM  {0}	sl_cx_jushou
            				WHERE (sl_cx_jushou.`记录时间`= '{2}' AND sl_cx_jushou.`日期` < DATE_SUB('{2}', INTERVAL DAY('{2}')-1 DAY)
                                or sl_cx_jushou.`记录时间`= '{3}' AND sl_cx_jushou.`日期` < DATE_SUB('{3}', INTERVAL DAY('{3}')-1 DAY))
							    AND sl_cx_jushou.`币种` = '{1}' 
								AND sl_cx_jushou.`父级分类` IS NOT NULL
            					AND sl_cx_jushou.`最终状态` IN ('拒收', '理赔', '已签收', '已退货') 
								AND sl_cx_jushou.`系统订单状态` IN {4} 
            				GROUP BY {5}
            				ORDER BY 币种,年月
            				) sl_zong_jushou
            			ON sl_zong_jushou.`币种` = sl_zong.`币种` AND sl_zong_jushou.`年月` = sl_zong.`年月` AND sl_zong_jushou.`父级分类` = sl_zong.`父级分类` 
            				AND sl_zong_jushou.`二级分类` = sl_zong.`二级分类` AND sl_zong_jushou.`三级分类` = sl_zong.`三级分类` 
            				AND sl_zong_jushou.`物流方式` = sl_zong.`物流方式` AND sl_zong_jushou.`旬` = sl_zong.`旬` 
            		LEFT JOIN
            				(SELECT 币种,年月,父级分类,二级分类,三级分类,物流方式,旬,COUNT(`订单编号`) 退货订单量
            				FROM  {0}	sl_cx_tuihuo
            				WHERE (sl_cx_tuihuo.`记录时间`= '{2}' AND sl_cx_tuihuo.`日期` < DATE_SUB('{2}', INTERVAL DAY('{2}')-1 DAY)
                                or sl_cx_tuihuo.`记录时间`= '{3}' AND sl_cx_tuihuo.`日期` < DATE_SUB('{3}', INTERVAL DAY('{3}')-1 DAY))
								AND sl_cx_tuihuo.`币种` = '{1}' 
								AND sl_cx_tuihuo.`父级分类` IS NOT NULL
            					AND sl_cx_tuihuo.`最终状态` = "已退货"
								AND sl_cx_tuihuo.`系统订单状态`IN {4} 
            				GROUP BY {5}
            				ORDER BY 币种,年月
            				) sl_zong_tuihuo
            			ON sl_zong_tuihuo.`币种` = sl_zong.`币种` AND sl_zong_tuihuo.`年月` = sl_zong.`年月` AND sl_zong_tuihuo.`父级分类` = sl_zong.`父级分类` 
            				AND sl_zong_tuihuo.`二级分类` = sl_zong.`二级分类` AND sl_zong_tuihuo.`三级分类` = sl_zong.`三级分类` 
            				AND sl_zong_tuihuo.`物流方式` = sl_zong.`物流方式`AND sl_zong_tuihuo.`旬` = sl_zong.`旬` 
            		GROUP BY sl_zong.年月,sl_zong.父级分类,sl_zong.二级分类,sl_zong.三级分类,sl_zong.物流方式,sl_zong.旬
            		with rollup) sl_gat;'''.format(match2[team], team, Time_day[11], Time_day[10], ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)'), '币种,年月,父级分类,二级分类,三级分类,物流方式,旬')
        listT.append(sqlqsb2)
        show_name.append(' 月（天）签收率_…………')
        # 月签收率（整月）---查询
        sqlqsb3 = '''SELECT '' 币种,'' 年月,'' 父级分类,'' 二级分类,'' 三级分类,'' 币种,'' 物流方式,'' 旬,'' 总订单量,
                            '' AS '总签收/完成','' AS '总签收/总计','' 退款率,'' 总完成占比,'' 直发总计,'' AS '直发签收/完成',
                            '' AS '直发签收/总计','' 直发完成占比,'' AS '改派签收/完成','' AS '改派签收/总计','' 改派完成占比
                    FROM {0} cs WHERE  cs.`记录时间` = '{1}'  '''.format(match2[team], today)
        listT.append(sqlqsb3)
        show_name.append(' 月（月）签收率_…………')
        # 月签收率（旬）---查询
        sqlqsb4 = '''SELECT sl_gat.`年月`,sl_gat.旬,sl_gat.`币种`,sl_gat.父级分类,sl_gat.二级分类,sl_gat.三级分类,sl_gat.物流方式,
				            sl_gat.`总订单量`,
				            sl_gat.`已签收订单量` / sl_gat.`拒收订单量` AS '总签收/完成',
				            sl_gat.`已签收订单量` / sl_gat.`总订单量` AS '总签收/总计',
				            sl_gat.`退货订单量` / sl_gat.`总订单量` AS '退款率',
				            sl_gat.`拒收订单量` / sl_gat.`总订单量` AS '总完成占比',
				            sl_gat.`直发订单量` 直发总计,
				            sl_gat.`直发已签收订单量` / sl_gat.`直发拒收订单量` AS '直发签收/完成',
				            sl_gat.`直发已签收订单量` / sl_gat.`直发订单量` AS '直发签收/总计',
                            sl_gat.`直发拒收订单量` / sl_gat.`直发订单量` AS '直发完成占比',
				            sl_gat.`改派订单量` 改派总计,
				            sl_gat.`改派已签收订单量` / sl_gat.`改派拒收订单量` AS '改派签收/完成',
				            sl_gat.`改派已签收订单量` / sl_gat.`改派订单量` AS '改派签收/总计',
				            sl_gat.`改派拒收订单量` / sl_gat.`改派订单量` AS '改派完成占比'
            FROM (SELECT  sl_zong.币种,
                        IFNULL(sl_zong.年月,'合计') 年月,
                        IFNULL(sl_zong.旬,'合计') 旬,
                        IFNULL(sl_zong.父级分类,'合计') 父级分类,
                        IFNULL(sl_zong.二级分类,'合计') 二级分类,
						IFNULL(sl_zong.三级分类,'合计') 三级分类,
						IFNULL(sl_zong.物流方式,'合计') 物流方式,
						SUM(总订单量) 总订单量,
						IFNULL(SUM(直发订单量),0) 直发订单量,
						IFNULL(SUM(直发已签收订单量),0) 直发已签收订单量,
						IFNULL(SUM(直发拒收订单量),0) 直发拒收订单量,
						(SUM(总订单量) - IFNULL(SUM(直发订单量),0)) AS 改派订单量,
						IFNULL(SUM(改派已签收订单量),0) 改派已签收订单量,
						IFNULL(SUM(改派拒收订单量),0) 改派拒收订单量,
						IFNULL(SUM(已签收订单量),0) 已签收订单量,
						IFNULL(SUM(拒收订单量),0) 拒收订单量,
						IFNULL(SUM(退货订单量),0) 退货订单量
		    FROM (SELECT  币种,年月,旬,父级分类,二级分类,三级分类,物流方式,COUNT(`订单编号`) 总订单量
				FROM  {0}	cx
				WHERE  (cx.`记录时间`= '{2}' AND cx.`日期` < DATE_SUB('{2}', INTERVAL DAY('{2}')-1 DAY)
					or cx.`记录时间`= '{4}' AND cx.`日期` < DATE_SUB('{4}', INTERVAL DAY('{4}')-1 DAY))
					AND cx.`币种` = '{1}'
					AND cx.`父级分类` IS NOT NULL
					AND cx.`系统订单状态` IN {3} 
				GROUP BY 币种,年月,旬,父级分类,二级分类,三级分类,物流方式
				ORDER BY 币种,年月
				) sl_zong
		    LEFT JOIN
				(SELECT 币种,年月,父级分类,二级分类,三级分类,物流方式,旬,COUNT(`订单编号`) 直发订单量
				FROM  {0}	cx_zf
				WHERE (cx_zf.`记录时间`= '{2}' AND cx_zf.`日期` < DATE_SUB('{2}', INTERVAL DAY('{2}')-1 DAY)
					or cx_zf.`记录时间`= '{4}' AND cx_zf.`日期` < DATE_SUB('{4}', INTERVAL DAY('{4}')-1 DAY))
					AND cx_zf.`币种` = '{1}' 
					AND cx_zf.`父级分类` IS NOT NULL
					AND cx_zf.`系统订单状态` IN {3} 
				AND cx_zf.`是否改派` = "直发"
				GROUP BY 币种,年月,旬,父级分类,二级分类,三级分类,物流方式
				ORDER BY 币种,年月
				) sl_zong_zf 
		    ON sl_zong_zf.`币种` = sl_zong.`币种` AND sl_zong_zf.`年月` = sl_zong.`年月`AND sl_zong_zf.`父级分类` = sl_zong.`父级分类` 
			    AND sl_zong_zf.`二级分类` = sl_zong.`二级分类` AND sl_zong_zf.`三级分类` = sl_zong.`三级分类` 
			    AND sl_zong_zf.`物流方式` = sl_zong.`物流方式` AND sl_zong_zf.`旬` = sl_zong.`旬` 				
		    LEFT JOIN
				(SELECT 币种,年月,旬,父级分类,二级分类,三级分类,物流方式,COUNT(`订单编号`) 直发已签收订单量
				FROM  {0}	sl_cx_zf_qianshou
				WHERE (sl_cx_zf_qianshou.`记录时间`= '{2}' AND sl_cx_zf_qianshou.`日期` < DATE_SUB('{2}', INTERVAL DAY('{2}')-1 DAY)
					or sl_cx_zf_qianshou.`记录时间`= '{4}' AND sl_cx_zf_qianshou.`日期` < DATE_SUB('{4}', INTERVAL DAY('{4}')-1 DAY))
					AND sl_cx_zf_qianshou.`币种` = '{1}' 
					AND sl_cx_zf_qianshou.`父级分类` IS NOT NULL
					AND sl_cx_zf_qianshou.`是否改派` = "直发"
					AND sl_cx_zf_qianshou.`最终状态` = "已签收"
					AND sl_cx_zf_qianshou.`系统订单状态` IN {3} 
				GROUP BY 币种,年月,旬,父级分类,二级分类,三级分类,物流方式
				ORDER BY 币种,年月
				) sl_cx_zf_qs
		    ON sl_cx_zf_qs.`币种` = sl_zong.`币种` AND sl_cx_zf_qs.`年月` = sl_zong.`年月` AND sl_cx_zf_qs.`父级分类` = sl_zong.`父级分类` 
			    AND sl_cx_zf_qs.`二级分类` = sl_zong.`二级分类` AND sl_cx_zf_qs.`三级分类` = sl_zong.`三级分类` 
			    AND sl_cx_zf_qs.`物流方式` = sl_zong.`物流方式` AND sl_cx_zf_qs.`旬` = sl_zong.`旬` 		
		    LEFT JOIN
				(SELECT 币种,年月,旬,父级分类,二级分类,三级分类,物流方式,COUNT(`订单编号`) 直发拒收订单量
				FROM  {0}	sl_cx_zf_jushou
				WHERE (sl_cx_zf_jushou.`记录时间`= '{2}' AND sl_cx_zf_jushou.`日期` < DATE_SUB('{2}', INTERVAL DAY('{2}')-1 DAY)
					or sl_cx_zf_jushou.`记录时间`= '{4}' AND sl_cx_zf_jushou.`日期` < DATE_SUB('{4}', INTERVAL DAY('{4}')-1 DAY))
					AND sl_cx_zf_jushou.`币种` = '{1}' 
					AND sl_cx_zf_jushou.`父级分类` IS NOT NULL
					AND sl_cx_zf_jushou.`是否改派` = "直发" 
					AND sl_cx_zf_jushou.`最终状态` IN ('拒收', '理赔', '已签收', '已退货') 
					AND sl_cx_zf_jushou.`系统订单状态` IN {3} 
				GROUP BY 币种,年月,旬,父级分类,二级分类,三级分类,物流方式
				ORDER BY 币种,年月
				) sl_cx_zf_js
		    ON sl_cx_zf_js.`币种` = sl_zong.`币种` AND sl_cx_zf_js.`年月` = sl_zong.`年月` AND sl_cx_zf_js.`父级分类` = sl_zong.`父级分类` 
			    AND sl_cx_zf_js.`二级分类` = sl_zong.`二级分类` AND sl_cx_zf_js.`三级分类` = sl_zong.`三级分类` 
			    AND sl_cx_zf_js.`物流方式` = sl_zong.`物流方式` AND sl_cx_zf_js.`旬` = sl_zong.`旬` 
		    LEFT JOIN
				(SELECT 币种,年月,旬,父级分类,二级分类,三级分类,物流方式,COUNT(`订单编号`) 改派订单量
				FROM  {0}	sl_cx_gp
				WHERE (sl_cx_gp.`记录时间`= '{2}' AND sl_cx_gp.`日期` < DATE_SUB('{2}', INTERVAL DAY('{2}')-1 DAY)
					or sl_cx_gp.`记录时间`= '{4}' AND sl_cx_gp.`日期` < DATE_SUB('{4}', INTERVAL DAY('{4}')-1 DAY))
					AND sl_cx_gp.`币种` = '{1}' 
					AND sl_cx_gp.`父级分类` IS NOT NULL
					AND sl_cx_gp.`是否改派` = "改派"
					AND sl_cx_gp.`系统订单状态` IN {3} 
				GROUP BY 币种,年月,旬,父级分类,二级分类,三级分类,物流方式
				ORDER BY 币种,年月
				) sl_zong_gp
		    ON sl_zong_gp.`币种` = sl_zong.`币种` AND sl_zong_gp.`年月` = sl_zong.`年月` AND sl_zong_gp.`父级分类` = sl_zong.`父级分类` 
			    AND sl_zong_gp.`二级分类` = sl_zong.`二级分类` AND sl_zong_gp.`三级分类` = sl_zong.`三级分类` 
			    AND sl_zong_gp.`物流方式` = sl_zong.`物流方式` AND sl_zong_gp.`旬` = sl_zong.`旬` 
		    LEFT JOIN
				(SELECT 币种,年月,旬,父级分类,二级分类,三级分类,物流方式,COUNT(`订单编号`) 改派已签收订单量
				FROM  {0}	sl_cx_gp_qianshou
				WHERE (sl_cx_gp_qianshou.`记录时间`= '{2}' AND sl_cx_gp_qianshou.`日期` < DATE_SUB('{2}', INTERVAL DAY('{2}')-1 DAY)
					or sl_cx_gp_qianshou.`记录时间`= '{4}' AND sl_cx_gp_qianshou.`日期` < DATE_SUB('{4}', INTERVAL DAY('{4}')-1 DAY))
					AND sl_cx_gp_qianshou.`币种` = '{1}' 
					AND sl_cx_gp_qianshou.`父级分类` IS NOT NULL
					AND sl_cx_gp_qianshou.`是否改派` = "改派"
					AND sl_cx_gp_qianshou.`最终状态` = "已签收"
					AND sl_cx_gp_qianshou.`系统订单状态` IN {3} 
				GROUP BY 币种,年月,旬,父级分类,二级分类,三级分类,物流方式
				ORDER BY 币种,年月
				) sl_cx_gp_qs
			ON sl_cx_gp_qs.`币种` = sl_zong.`币种` AND sl_cx_gp_qs.`年月` = sl_zong.`年月` AND sl_cx_gp_qs.`父级分类` = sl_zong.`父级分类` 
				AND sl_cx_gp_qs.`二级分类` = sl_zong.`二级分类` AND sl_cx_gp_qs.`三级分类` = sl_zong.`三级分类` 
				AND sl_cx_gp_qs.`物流方式` = sl_zong.`物流方式` AND sl_cx_gp_qs.`旬` = sl_zong.`旬` 
		    LEFT JOIN
				(SELECT 币种,年月,旬,父级分类,二级分类,三级分类,物流方式,COUNT(`订单编号`) 改派拒收订单量
				FROM  {0}	sl_cx_gp_jushou
				WHERE (sl_cx_gp_jushou.`记录时间`= '{2}' AND sl_cx_gp_jushou.`日期` < DATE_SUB('{2}', INTERVAL DAY('{2}')-1 DAY)
					or sl_cx_gp_jushou.`记录时间`= '{4}' AND sl_cx_gp_jushou.`日期` < DATE_SUB('{4}', INTERVAL DAY('{4}')-1 DAY))
					AND sl_cx_gp_jushou.`币种` = '{1}' 
					AND sl_cx_gp_jushou.`父级分类` IS NOT NULL
					AND sl_cx_gp_jushou.`是否改派` = "改派"
					AND sl_cx_gp_jushou.`最终状态` IN ('拒收', '理赔', '已签收', '已退货') 
					AND sl_cx_gp_jushou.`系统订单状态` IN {3} 
				GROUP BY 币种,年月,旬,父级分类,二级分类,三级分类,物流方式
				ORDER BY 币种,年月
				) sl_cx_gp_js
			ON sl_cx_gp_js.`币种` = sl_zong.`币种` AND sl_cx_gp_js.`年月` = sl_zong.`年月` AND sl_cx_gp_js.`父级分类` = sl_zong.`父级分类` 
				AND sl_cx_gp_js.`二级分类` = sl_zong.`二级分类` AND sl_cx_gp_js.`三级分类` = sl_zong.`三级分类` 
				AND sl_cx_gp_js.`物流方式` = sl_zong.`物流方式` AND sl_cx_gp_js.`旬` = sl_zong.`旬` 
		    LEFT JOIN
				(SELECT 币种,年月,旬,父级分类,二级分类,三级分类,物流方式,COUNT(`订单编号`) 已签收订单量
				FROM  {0}	sl_cx_qianshou
				WHERE (sl_cx_qianshou.`记录时间`= '{2}' AND sl_cx_qianshou.`日期` < DATE_SUB('{2}', INTERVAL DAY('{2}')-1 DAY)
					or sl_cx_qianshou.`记录时间`= '{4}' AND sl_cx_qianshou.`日期` < DATE_SUB('{4}', INTERVAL DAY('{4}')-1 DAY))
					AND sl_cx_qianshou.`币种` = '{1}' 
					AND sl_cx_qianshou.`父级分类` IS NOT NULL
					AND sl_cx_qianshou.`最终状态` = "已签收"
					AND sl_cx_qianshou.`系统订单状态` IN {3} 
				GROUP BY 币种,年月,旬,父级分类,二级分类,三级分类,物流方式
				ORDER BY 币种,年月
				) sl_zong_qianshou
			ON sl_zong_qianshou.`币种` = sl_zong.`币种` AND sl_zong_qianshou.`年月` = sl_zong.`年月` AND sl_zong_qianshou.`父级分类` = sl_zong.`父级分类` 
				AND sl_zong_qianshou.`二级分类` = sl_zong.`二级分类`  AND sl_zong_qianshou.`三级分类` = sl_zong.`三级分类` 
				AND sl_zong_qianshou.`物流方式` = sl_zong.`物流方式` AND sl_zong_qianshou.`旬` = sl_zong.`旬`
		    LEFT JOIN
				(SELECT 币种,年月,旬,父级分类,二级分类,三级分类,物流方式,COUNT(`订单编号`) 拒收订单量
				FROM  {0}	sl_cx_jushou
				WHERE (sl_cx_jushou.`记录时间`= '{2}' AND sl_cx_jushou.`日期` < DATE_SUB('{2}', INTERVAL DAY('{2}')-1 DAY)
					or sl_cx_jushou.`记录时间`= '{4}' AND sl_cx_jushou.`日期` < DATE_SUB('{4}', INTERVAL DAY('{4}')-1 DAY))
					AND sl_cx_jushou.`币种` = '{1}' 
					AND sl_cx_jushou.`父级分类` IS NOT NULL
					AND sl_cx_jushou.`最终状态` IN ('拒收', '理赔', '已签收', '已退货') 
					AND sl_cx_jushou.`系统订单状态` IN {3} 
				GROUP BY 币种,年月,旬,父级分类,二级分类,三级分类,物流方式
				ORDER BY 币种,年月
				) sl_zong_jushou
			ON sl_zong_jushou.`币种` = sl_zong.`币种` AND sl_zong_jushou.`年月` = sl_zong.`年月` AND sl_zong_jushou.`父级分类` = sl_zong.`父级分类` 
				AND sl_zong_jushou.`二级分类` = sl_zong.`二级分类` 
				AND sl_zong_jushou.`三级分类` = sl_zong.`三级分类` AND sl_zong_jushou.`物流方式` = sl_zong.`物流方式` AND sl_zong_jushou.`旬` = sl_zong.`旬` 
		    LEFT JOIN
				(SELECT 币种,年月,旬,父级分类,二级分类,三级分类,物流方式,COUNT(`订单编号`) 退货订单量
				FROM  {0}	sl_cx_tuihuo
				WHERE (sl_cx_tuihuo.`记录时间`= '{2}' AND sl_cx_tuihuo.`日期` < DATE_SUB('{2}', INTERVAL DAY('{2}')-1 DAY)
					or sl_cx_tuihuo.`记录时间`= '{4}' AND sl_cx_tuihuo.`日期` < DATE_SUB('{4}', INTERVAL DAY('{4}')-1 DAY))
					AND sl_cx_tuihuo.`币种` = '{1}' 
					AND sl_cx_tuihuo.`父级分类` IS NOT NULL
					AND sl_cx_tuihuo.`最终状态` = "已退货"
					AND sl_cx_tuihuo.`系统订单状态` IN {3} 
				GROUP BY 币种,年月,旬,父级分类,二级分类,三级分类,物流方式
				ORDER BY 币种,年月
				) sl_zong_tuihuo
			ON sl_zong_tuihuo.`币种` = sl_zong.`币种` AND sl_zong_tuihuo.`年月` = sl_zong.`年月` AND sl_zong_tuihuo.`父级分类` = sl_zong.`父级分类`
				AND sl_zong_tuihuo.`二级分类` = sl_zong.`二级分类` AND sl_zong_tuihuo.`三级分类` = sl_zong.`三级分类` 
				AND sl_zong_tuihuo.`物流方式` = sl_zong.`物流方式`AND sl_zong_tuihuo.`旬` = sl_zong.`旬` 
		    GROUP BY sl_zong.年月,sl_zong.旬,sl_zong.父级分类,sl_zong.二级分类,sl_zong.三级分类,sl_zong.物流方式
	    with rollup) sl_gat;'''.format(match2[team], team, Time_day[11], ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)'), Time_day[10])
        listT.append(sqlqsb4)
        show_name.append(' 月（旬）签收率_…………')
        # 月签收率（各月）---查询
        sqlqsb5 = '''SELECT sl_gat.`币种`,sl_gat.`年月`,sl_gat.父级分类,sl_gat.二级分类,sl_gat.三级分类,'' 产品名称,sl_gat.物流方式,sl_gat.旬,
				                sl_gat.`总订单量`,
				                sl_gat.`已签收订单量` / sl_gat.`拒收订单量` AS '总签收/完成',
				                sl_gat.`已签收订单量` / sl_gat.`总订单量` AS '总签收/总计',
				                sl_gat.`退货订单量` / sl_gat.`总订单量` AS '退款率',
				                sl_gat.`拒收订单量` / sl_gat.`总订单量` AS '总完成占比',
				                sl_gat.`直发订单量` 直发总计,
				                sl_gat.`直发已签收订单量` / sl_gat.`直发拒收订单量` AS '直发签收/完成',
				                sl_gat.`直发已签收订单量` / sl_gat.`直发订单量` AS '直发签收/总计', 
				                sl_gat.`直发拒收订单量` / sl_gat.`直发订单量` AS '直发完成占比',
				                sl_gat.`改派订单量` 改派总计,
				                sl_gat.`改派已签收订单量` / sl_gat.`改派拒收订单量` AS '改派签收/完成',
				                sl_gat.`改派已签收订单量` / sl_gat.`改派订单量` AS '改派签收/总计',
				                sl_gat.`改派拒收订单量` / sl_gat.`改派订单量` AS '改派完成占比',
				                sl_gat.`总销售额`,
				                sl_gat.`已签收销售额` / sl_gat.`拒收销售额` AS '总签收/完成(金额)',
				                sl_gat.`已签收销售额` / sl_gat.`总销售额` AS '总签收/总计(金额)',
				                sl_gat.`退货销售额` / sl_gat.`总销售额` AS '退款率(金额)',
				                sl_gat.`拒收销售额` / sl_gat.`总销售额` AS '总完成占比(金额)',
				                sl_gat.`直发销售额`,
				                sl_gat.`直发已签收销售额` / sl_gat.`直发拒收销售额` AS '直发签收/完成(金额)',
				                sl_gat.`直发已签收销售额` / sl_gat.`直发销售额` AS '直发签收/总计(金额)',
				                sl_gat.`直发拒收销售额` / sl_gat.`直发销售额` AS '直发完成占比(金额)',
				                sl_gat.`改派销售额`,
				                sl_gat.`改派已签收销售额` / sl_gat.`改派拒收销售额` AS '改派签收/完成(金额)',
				                sl_gat.`改派已签收销售额` / sl_gat.`改派销售额` AS '改派签收/总计(金额)',
				                sl_gat.`改派拒收销售额` / sl_gat.`改派销售额` AS '改派完成占比(金额)'
                        FROM (SELECT  sl_zong.币种,
                                    IFNULL(sl_zong.年月,'合计') 年月,
                                    IFNULL(sl_zong.父级分类,'合计') 父级分类,
                                    IFNULL(sl_zong.二级分类,'合计') 二级分类,
							        IFNULL(sl_zong.三级分类,'合计') 三级分类,
							        IFNULL(sl_zong.物流方式,'合计') 物流方式,
							        IFNULL(sl_zong.旬,'合计') 旬,
							        SUM(总订单量) 总订单量,
							        SUM(总销售额) 总销售额,
							        IFNULL(SUM(直发订单量),0) 直发订单量,
							        IFNULL(SUM(直发销售额),0) 直发销售额,
							        IFNULL(SUM(直发已签收订单量),0) 直发已签收订单量,
							        IFNULL(SUM(直发已签收销售额),0) 直发已签收销售额,
							        IFNULL(SUM(直发拒收订单量),0) 直发拒收订单量,
							        IFNULL(SUM(直发拒收销售额),0) 直发拒收销售额,
							        (SUM(总订单量) - IFNULL(SUM(直发订单量),0)) AS 改派订单量,
							        (SUM(总销售额) - IFNULL(SUM(直发销售额),0)) AS 改派销售额,
							        IFNULL(SUM(改派已签收订单量),0) 改派已签收订单量,
							        IFNULL(SUM(改派已签收销售额),0) 改派已签收销售额,
							        IFNULL(SUM(改派拒收订单量),0) 改派拒收订单量,
							        IFNULL(SUM(改派拒收销售额),0) 改派拒收销售额,
							        IFNULL(SUM(已签收订单量),0) 已签收订单量,
							        IFNULL(SUM(已签收销售额),0) 已签收销售额,
							        IFNULL(SUM(拒收订单量),0) 拒收订单量,
							        IFNULL(SUM(拒收销售额),0) 拒收销售额,
							        IFNULL(SUM(退货订单量),0) 退货订单量,
							        IFNULL(SUM(退货销售额),0) 退货销售额
		FROM (SELECT  币种,年月,父级分类,二级分类,三级分类,物流方式,旬,COUNT(`订单编号`) 总订单量,SUM(`价格RMB`) 总销售额
				FROM  {0}	sl_cx
				WHERE sl_cx.`币种` = '{1}' 
					AND sl_cx.`父级分类` IS NOT NULL
					AND sl_cx.`系统订单状态` IN {2} 
				GROUP BY 币种,年月,父级分类,二级分类,三级分类,物流方式,旬
				ORDER BY 币种,年月
				) sl_zong
		LEFT JOIN
				(SELECT 币种,年月,父级分类,二级分类,三级分类,物流方式,旬,COUNT(`订单编号`) 直发订单量,SUM(`价格RMB`) 直发销售额
				FROM  {0}	sl_zong_zf
				WHERE sl_zong_zf.`币种` = '{1}' 
					AND sl_zong_zf.`是否改派` = "直发"
					AND sl_zong_zf.`父级分类` IS NOT NULL
					AND sl_zong_zf.`系统订单状态` IN {2}
				GROUP BY 币种,年月,父级分类,二级分类,三级分类,物流方式,旬
				ORDER BY 币种,年月
				) sl_zong_zf 
		ON sl_zong_zf.`币种` = sl_zong.`币种` AND sl_zong_zf.`年月` = sl_zong.`年月`AND sl_zong_zf.`父级分类` = sl_zong.`父级分类` 
			AND sl_zong_zf.`二级分类` = sl_zong.`二级分类` AND sl_zong_zf.`三级分类` = sl_zong.`三级分类` 
			AND sl_zong_zf.`物流方式` = sl_zong.`物流方式` AND sl_zong_zf.`旬` = sl_zong.`旬` 							
		LEFT JOIN
				(SELECT 币种,年月,父级分类,二级分类,三级分类,物流方式,旬,COUNT(`订单编号`) 直发已签收订单量,SUM(`价格RMB`) 直发已签收销售额
				FROM  {0}	slzf_qs
				WHERE slzf_qs.`币种` = '{1}' 
					AND slzf_qs.`是否改派` = "直发"
					AND slzf_qs.`最终状态` = "已签收"
					AND slzf_qs.`父级分类` IS NOT NULL
					AND slzf_qs.`系统订单状态` IN {2} 
				GROUP BY 币种,年月,父级分类,二级分类,三级分类,物流方式,旬
				ORDER BY 币种,年月
				) sl_cx_zf_qs
		ON sl_cx_zf_qs.`币种` = sl_zong.`币种` AND sl_cx_zf_qs.`年月` = sl_zong.`年月` AND sl_cx_zf_qs.`父级分类` = sl_zong.`父级分类` 
			AND sl_cx_zf_qs.`二级分类` = sl_zong.`二级分类` AND sl_cx_zf_qs.`三级分类` = sl_zong.`三级分类` 
			AND sl_cx_zf_qs.`物流方式` = sl_zong.`物流方式` AND sl_cx_zf_qs.`旬` = sl_zong.`旬` 		
		LEFT JOIN
				(SELECT 币种,年月,父级分类,二级分类,三级分类,物流方式,旬,COUNT(`订单编号`) 直发拒收订单量,SUM(`价格RMB`) 直发拒收销售额
				FROM  {0}	slzf_js
				WHERE slzf_js.`币种` = '{1}' 
						AND slzf_js.`是否改派` = "直发" 
						AND slzf_js.`最终状态` IN ('拒收', '理赔', '已签收', '已退货') 
						AND slzf_js.`父级分类` IS NOT NULL 
						AND slzf_js.`系统订单状态` IN {2} 
				GROUP BY 币种,年月,父级分类,二级分类,三级分类,物流方式,旬
				ORDER BY 币种,年月
				) sl_cx_zf_js
		ON sl_cx_zf_js.`币种` = sl_zong.`币种` AND sl_cx_zf_js.`年月` = sl_zong.`年月` AND sl_cx_zf_js.`父级分类` = sl_zong.`父级分类` 
			AND sl_cx_zf_js.`二级分类` = sl_zong.`二级分类` AND sl_cx_zf_js.`三级分类` = sl_zong.`三级分类` 
			AND sl_cx_zf_js.`物流方式` = sl_zong.`物流方式` AND sl_cx_zf_js.`旬` = sl_zong.`旬` 
		LEFT JOIN
				(SELECT 币种,年月,父级分类,二级分类,三级分类,物流方式,旬,COUNT(`订单编号`) 改派订单量,SUM(`价格RMB`) 改派销售额
				FROM  {0}	sl_cx_gp
				WHERE sl_cx_gp.`币种` = '{1}' 
						AND sl_cx_gp.`是否改派` = "改派"
						AND sl_cx_gp.`父级分类` IS NOT NULL
						AND sl_cx_gp.`系统订单状态` IN {2} 
					GROUP BY 币种,年月,父级分类,二级分类,三级分类,物流方式,旬
					ORDER BY 币种,年月
				) sl_zong_gp
		ON sl_zong_gp.`币种` = sl_zong.`币种` AND sl_zong_gp.`年月` = sl_zong.`年月` AND sl_zong_gp.`父级分类` = sl_zong.`父级分类` 
			AND sl_zong_gp.`二级分类` = sl_zong.`二级分类` AND sl_zong_gp.`三级分类` = sl_zong.`三级分类` 
			AND sl_zong_gp.`物流方式` = sl_zong.`物流方式` AND sl_zong_gp.`旬` = sl_zong.`旬` 
		LEFT JOIN
				(SELECT 币种,年月,父级分类,二级分类,三级分类,物流方式,旬,COUNT(`订单编号`) 改派已签收订单量,SUM(`价格RMB`) 改派已签收销售额
				FROM  {0}	slgp_qs
				WHERE slgp_qs.`币种` = '{1}' 
					AND slgp_qs.`是否改派` = "改派" 
					AND slgp_qs.`最终状态` = "已签收"
					AND slgp_qs.`父级分类` IS NOT NULL
					AND slgp_qs.`系统订单状态`IN {2}
				GROUP BY 币种,年月,父级分类,二级分类,三级分类,物流方式,旬
				ORDER BY 币种,年月
				) sl_cx_gp_qs
			ON sl_cx_gp_qs.`币种` = sl_zong.`币种` AND sl_cx_gp_qs.`年月` = sl_zong.`年月` AND sl_cx_gp_qs.`父级分类` = sl_zong.`父级分类` 
				AND sl_cx_gp_qs.`二级分类` = sl_zong.`二级分类` AND sl_cx_gp_qs.`三级分类` = sl_zong.`三级分类` 
				AND sl_cx_gp_qs.`物流方式` = sl_zong.`物流方式` AND sl_cx_gp_qs.`旬` = sl_zong.`旬` 
		LEFT JOIN
				(SELECT 币种,年月,父级分类,二级分类,三级分类,物流方式,旬,COUNT(`订单编号`) 改派拒收订单量,SUM(`价格RMB`) 改派拒收销售额
				FROM  {0}	sl_cx_gp_jushou
				WHERE sl_cx_gp_jushou.`币种` = '{1}' 
						AND sl_cx_gp_jushou.`系统订单状态` IN {2} 
						AND sl_cx_gp_jushou.`父级分类` IS NOT NULL
						AND sl_cx_gp_jushou.`是否改派` = "改派" 
						AND sl_cx_gp_jushou.`最终状态` IN ('拒收', '理赔', '已签收', '已退货') 
				GROUP BY 币种,年月,父级分类,二级分类,三级分类,物流方式,旬
				ORDER BY 币种,年月
				) sl_cx_gp_js
			ON sl_cx_gp_js.`币种` = sl_zong.`币种` AND sl_cx_gp_js.`年月` = sl_zong.`年月` AND sl_cx_gp_js.`父级分类` = sl_zong.`父级分类` 
				AND sl_cx_gp_js.`二级分类` = sl_zong.`二级分类` AND sl_cx_gp_js.`三级分类` = sl_zong.`三级分类` 
				AND sl_cx_gp_js.`物流方式` = sl_zong.`物流方式` AND sl_cx_gp_js.`旬` = sl_zong.`旬` 				
		LEFT JOIN
				(SELECT 币种,年月,父级分类,二级分类,三级分类,物流方式,旬,COUNT(`订单编号`) 已签收订单量,SUM(`价格RMB`) 已签收销售额
				FROM  {0}	sl_cx_qianshou
				WHERE sl_cx_qianshou.`币种` = '{1}' 
					AND sl_cx_qianshou.`系统订单状态`IN {2}
					AND sl_cx_qianshou.`最终状态` = "已签收"
					AND sl_cx_qianshou.`父级分类` IS NOT NULL
				GROUP BY 币种,年月,父级分类,二级分类,三级分类,物流方式,旬
				ORDER BY 币种,年月
				) sl_zong_qianshou
			ON sl_zong_qianshou.`币种` = sl_zong.`币种` AND sl_zong_qianshou.`年月` = sl_zong.`年月` AND sl_zong_qianshou.`父级分类` = sl_zong.`父级分类` 
				AND sl_zong_qianshou.`二级分类` = sl_zong.`二级分类`  AND sl_zong_qianshou.`三级分类` = sl_zong.`三级分类` 
				AND sl_zong_qianshou.`物流方式` = sl_zong.`物流方式` AND sl_zong_qianshou.`旬` = sl_zong.`旬`
		LEFT JOIN
				(SELECT 币种,年月,父级分类,二级分类,三级分类,物流方式,旬,COUNT(`订单编号`) 拒收订单量,SUM(`价格RMB`) 拒收销售额
				FROM  {0}	sl_cx_jushou
				WHERE sl_cx_jushou.`币种` = '{1}' 
						AND sl_cx_jushou.`系统订单状态` IN {2}
						AND sl_cx_jushou.`最终状态` IN ('拒收', '理赔', '已签收', '已退货') 
						AND sl_cx_jushou.`父级分类` IS NOT NULL
				GROUP BY 币种,年月,父级分类,二级分类,三级分类,物流方式,旬
				ORDER BY 币种,年月
				) sl_zong_jushou
			ON sl_zong_jushou.`币种` = sl_zong.`币种` AND sl_zong_jushou.`年月` = sl_zong.`年月` AND sl_zong_jushou.`父级分类` = sl_zong.`父级分类` 
				AND sl_zong_jushou.`二级分类` = sl_zong.`二级分类` AND sl_zong_jushou.`三级分类` = sl_zong.`三级分类` 
				AND sl_zong_jushou.`物流方式` = sl_zong.`物流方式` AND sl_zong_jushou.`旬` = sl_zong.`旬` 
		LEFT JOIN
				(SELECT 币种,年月,父级分类,二级分类,三级分类,物流方式,旬,COUNT(`订单编号`) 退货订单量,SUM(`价格RMB`) 退货销售额
				FROM  {0}	sl_cx_tuihuo
				WHERE sl_cx_tuihuo.`币种` = '{1}' 
						AND sl_cx_tuihuo.`系统订单状态`IN {2}
						AND sl_cx_tuihuo.`最终状态` = "已退货"
						AND sl_cx_tuihuo.`父级分类` IS NOT NULL
				GROUP BY 币种,年月,父级分类,二级分类,三级分类,物流方式,旬
				ORDER BY 币种,年月
				) sl_zong_tuihuo
			ON sl_zong_tuihuo.`币种` = sl_zong.`币种` AND sl_zong_tuihuo.`年月` = sl_zong.`年月` AND sl_zong_tuihuo.`父级分类` = sl_zong.`父级分类` 
				AND sl_zong_tuihuo.`二级分类` = sl_zong.`二级分类` AND sl_zong_tuihuo.`三级分类` = sl_zong.`三级分类` 
				AND sl_zong_tuihuo.`物流方式` = sl_zong.`物流方式`AND sl_zong_tuihuo.`旬` = sl_zong.`旬` 
		GROUP BY sl_zong.年月,sl_zong.父级分类,sl_zong.二级分类,sl_zong.三级分类,sl_zong.物流方式,sl_zong.旬
		with rollup) sl_gat 
		ORDER BY sl_gat.`年月` DESC;'''.format('qsb_缓存_month', team, ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)'))
        listT.append(sqlqsb5)
        show_name.append(' 月（各月）签收率_…………')

        # 月物流（天）---查询
        sqlWl2 = '''SELECT sl_rb.`币种`,
				sl_rb.`年月`,
				sl_rb.物流方式,
				sl_rb.父级分类,
				sl_rb.旬,
				sl_rb.`总订单量` 总订单,
				null AS '总签收/完成',
				null AS '总签收/总计',
				null AS '退款率',
				null AS '总完成占比',
				sl_rb.`直发订单量` 总计,
				sl_rb.`直发已签收订单量` / sl_rb.`直发拒收订单量` AS '直发签收/完成',
				sl_rb.`直发已签收订单量` / sl_rb.`直发订单量` AS '直发签收/总计',
				sl_rb.`直发拒收订单量` / sl_rb.`直发订单量` AS '直发完成占比',
				null AS  改派总计,
				null AS '改派签收/完成',
				null AS '改派签收/总计',
				null AS '改派完成占比'
        FROM (SELECT  sl_zong.币种,
						IFNULL(sl_zong.年月,'合计') 年月,
						IFNULL(sl_zong.物流方式,'合计') 物流方式,
						IFNULL(sl_zong.父级分类,'合计') 父级分类,
						IFNULL(sl_zong.旬,'合计') 旬,
						SUM(总订单量) 总订单量,
						IFNULL(SUM(直发订单量),0) 直发订单量,
						IFNULL(SUM(直发已签收订单量),0) 直发已签收订单量,
						IFNULL(SUM(直发拒收订单量),0) 直发拒收订单量
			FROM ( SELECT 币种,
										年月,
										物流方式,
										父级分类,
										旬,
										COUNT(`订单编号`) 总订单量
						FROM  {0}	sl_cx
						WHERE sl_cx.`币种` = '{1}' 
                            AND (sl_cx.`记录时间`= '{2}' AND sl_cx.`日期` < DATE_SUB('{2}', INTERVAL DAY('{2}')-1 DAY)
								or sl_cx.`记录时间`= '{3}' AND sl_cx.`日期` < DATE_SUB('{3}', INTERVAL DAY('{3}')-1 DAY))
							AND sl_cx.`父级分类` IS NOT NULL
                            AND sl_cx.`是否改派` = "直发"
                            AND sl_cx.`系统订单状态` IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)') 
						GROUP BY 币种,年月,物流方式,父级分类,旬
						ORDER BY 币种,年月
						) sl_zong
			LEFT JOIN
						(SELECT 币种,
										年月,
										物流方式,
										父级分类,
													旬,
										COUNT(`订单编号`) 直发订单量
						FROM  {0}	sl_cx_zf
						WHERE sl_cx_zf.`币种` = '{1}' 
                            AND (sl_cx_zf.`记录时间`= '{2}' AND sl_cx_zf.`日期` < DATE_SUB('{2}', INTERVAL DAY('{2}')-1 DAY)
								or sl_cx_zf.`记录时间`= '{3}' AND sl_cx_zf.`日期` < DATE_SUB('{3}', INTERVAL DAY('{3}')-1 DAY))
							AND sl_cx_zf.`父级分类` IS NOT NULL
                            AND sl_cx_zf.`系统订单状态` IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)')  
                            AND sl_cx_zf.`是否改派` = "直发"
						GROUP BY 币种,年月,物流方式,父级分类,旬
						ORDER BY 币种,年月
					) sl_zong_zf
					 ON sl_zong_zf.`币种` = sl_zong.`币种` 
						AND sl_zong_zf.`年月` = sl_zong.`年月`
						AND sl_zong_zf.`物流方式` = sl_zong.`物流方式`
						AND sl_zong_zf.`父级分类` = sl_zong.`父级分类` 
						AND sl_zong_zf.`旬` = sl_zong.`旬` 
				LEFT JOIN
						(SELECT 币种,
										年月,
										物流方式,
										父级分类,
													旬,
										COUNT(`订单编号`) 直发已签收订单量
						FROM  {0}	sl_cx_zf_qianshou
						WHERE sl_cx_zf_qianshou.`币种` = '{1}' 
                            AND (sl_cx_zf_qianshou.`记录时间`= '{2}' AND sl_cx_zf_qianshou.`日期` < DATE_SUB('{2}', INTERVAL DAY('{2}')-1 DAY)
								or sl_cx_zf_qianshou.`记录时间`= '{3}' AND sl_cx_zf_qianshou.`日期` < DATE_SUB('{3}', INTERVAL DAY('{3}')-1 DAY))
							AND sl_cx_zf_qianshou.`父级分类` IS NOT NULL
							AND sl_cx_zf_qianshou.`系统订单状态` IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)') 
							AND sl_cx_zf_qianshou.`是否改派` = "直发"
							AND sl_cx_zf_qianshou.`最终状态` = "已签收"
						GROUP BY 币种,年月,物流方式,父级分类,旬
						ORDER BY 币种,年月
					) sl_cx_zf_qs
					 ON sl_cx_zf_qs.`币种` = sl_zong.`币种` 
						AND sl_cx_zf_qs.`年月` = sl_zong.`年月`
						AND sl_cx_zf_qs.`物流方式` = sl_zong.`物流方式`
						AND sl_cx_zf_qs.`父级分类` = sl_zong.`父级分类`
						AND sl_cx_zf_qs.`旬` = sl_zong.`旬`	 		
				LEFT JOIN
						(SELECT 币种,
										年月,
										物流方式,
										父级分类,
													旬,
										COUNT(`订单编号`) 直发拒收订单量
						FROM  {0}	sl_cx_zf_jushou
						WHERE sl_cx_zf_jushou.`币种` = '{1}' 
                            AND (sl_cx_zf_jushou.`记录时间`= '{2}' AND sl_cx_zf_jushou.`日期` < DATE_SUB('{2}', INTERVAL DAY('{2}')-1 DAY)
								or sl_cx_zf_jushou.`记录时间`= '{3}' AND sl_cx_zf_jushou.`日期` < DATE_SUB('{3}', INTERVAL DAY('{3}')-1 DAY))
							AND sl_cx_zf_jushou.`父级分类` IS NOT NULL
							AND sl_cx_zf_jushou.`系统订单状态` IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)') 
							AND sl_cx_zf_jushou.`是否改派` = "直发"
							AND sl_cx_zf_jushou.`最终状态` IN ('拒收', '理赔', '已签收', '已退货') 
						GROUP BY 币种,年月,物流方式,父级分类,旬
						ORDER BY 币种,年月
					) sl_cx_zf_js
					 ON sl_cx_zf_js.`币种` = sl_zong.`币种` 
						AND sl_cx_zf_js.`年月` = sl_zong.`年月`
						AND sl_cx_zf_js.`物流方式` = sl_zong.`物流方式`
						AND sl_cx_zf_js.`父级分类` = sl_zong.`父级分类` 
						AND sl_cx_zf_js.`旬` = sl_zong.`旬` 											
				GROUP BY sl_zong.年月,sl_zong.物流方式,sl_zong.父级分类,sl_zong.旬
				with rollup) sl_rb;'''.format(match2[team], team, Time_day[11], Time_day[10])
        listT.append(sqlWl2)
        show_name.append(' 月（天）物流…………')
        # 月物流（月）---查询
        sqlWl3= '''SELECT sl_rb.`币种`,
                    	sl_rb.`年月`,
                    	sl_rb.物流方式,
                    	sl_rb.父级分类,
                    	sl_rb.旬,
                    	sl_rb.`总订单量` 总订单,
                    	null AS '总签收/完成',
                    	null AS '总签收/总计',
                    	null AS '退款率',
                    	null AS '总完成占比',
                    	sl_rb.`直发订单量` 总计,
                    	sl_rb.`直发已签收订单量` / sl_rb.`直发拒收订单量` AS '直发签收/完成',
                    	sl_rb.`直发已签收订单量` / sl_rb.`直发订单量` AS '直发签收/总计',
                    	sl_rb.`直发拒收订单量` / sl_rb.`直发订单量` AS '直发完成占比',
                    	null AS  改派总计,
                    	null AS '改派签收/完成',
                    	null AS '改派签收/总计',
                    	null AS '改派完成占比'
        FROM (SELECT  sl_zong.币种,
                    IFNULL(sl_zong.年月,'合计') 年月,
                    IFNULL(sl_zong.物流方式,'合计') 物流方式,
                    IFNULL(sl_zong.父级分类,'合计') 父级分类,
                    IFNULL(sl_zong.旬,'合计') 旬,
                    SUM(总订单量) 总订单量,
                    IFNULL(SUM(直发订单量),0) 直发订单量,
                    IFNULL(SUM(直发已签收订单量),0) 直发已签收订单量,
                    IFNULL(SUM(直发拒收订单量),0) 直发拒收订单量
                    		FROM ( SELECT 币种,
                    					年月,
                    					物流方式,
                    					父级分类,
                    					旬,
                    					COUNT(`订单编号`) 总订单量
                    			    FROM  {0}	sl_cx
                    					WHERE sl_cx.`币种` = '{1}' 
        									AND (sl_cx.`记录时间`= '{2}' AND sl_cx.`日期` < DATE_SUB('{2}', INTERVAL DAY('{2}')-1 DAY) 
        										OR sl_cx.`记录时间`= '{3}' AND sl_cx.`日期` < DATE_SUB('{3}', INTERVAL DAY('{3}')-1 DAY))
                    						AND sl_cx.`父级分类` IS NOT NULL
                    						AND sl_cx.`是否改派` = "直发"
                                            AND sl_cx.`系统订单状态` IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)') 
                    				GROUP BY 币种,年月,物流方式,父级分类,旬
                    				ORDER BY 币种,年月
                    			) sl_zong
                    		LEFT JOIN
                    			(SELECT 币种,
                    					年月,
                    					物流方式,
                    					父级分类,
                    					旬,
                    					COUNT(`订单编号`) 直发订单量
                    				FROM  {0}	sl_cx_zf
                    				WHERE sl_cx_zf.`币种` = '{1}' 
        								AND (sl_cx_zf.`记录时间`= '{2}' AND sl_cx_zf.`日期` < DATE_SUB('{2}', INTERVAL DAY('{2}')-1 DAY) 
        									OR sl_cx_zf.`记录时间`= '{3}' AND sl_cx_zf.`日期` < DATE_SUB('{3}', INTERVAL DAY('{3}')-1 DAY))
                    					AND sl_cx_zf.`父级分类` IS NOT NULL
                                        AND sl_cx_zf.`系统订单状态` IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)')  
                                        AND sl_cx_zf.`是否改派` = "直发"
                    				GROUP BY 币种,年月,物流方式,父级分类,旬
                    				ORDER BY 币种,年月
                    			) sl_zong_zf
                    		ON sl_zong_zf.`币种` = sl_zong.`币种` 
                    					AND sl_zong_zf.`年月` = sl_zong.`年月`
                    					AND sl_zong_zf.`物流方式` = sl_zong.`物流方式`
                    					AND sl_zong_zf.`父级分类` = sl_zong.`父级分类` 
                    					AND sl_zong_zf.`旬` = sl_zong.`旬` 
                    		LEFT JOIN
                    			(SELECT 币种,
                    					年月,
                    					物流方式,
                    					父级分类,
                    					旬,
                    					COUNT(`订单编号`) 直发已签收订单量
                    			FROM  {0}	sl_cx_zf_qianshou
                    			WHERE sl_cx_zf_qianshou.`币种` = '{1}' 
        							AND (sl_cx_zf_qianshou.`记录时间`= '{2}' AND sl_cx_zf_qianshou.`日期` < DATE_SUB('{2}', INTERVAL DAY('{2}')-1 DAY) 
        								OR sl_cx_zf_qianshou.`记录时间`= '{3}' AND sl_cx_zf_qianshou.`日期` < DATE_SUB('{3}', INTERVAL DAY('{3}')-1 DAY))
                    				AND sl_cx_zf_qianshou.`父级分类` IS NOT NULL
                    				AND sl_cx_zf_qianshou.`系统订单状态` IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)') 
                    				AND sl_cx_zf_qianshou.`是否改派` = "直发"
                    				AND sl_cx_zf_qianshou.`最终状态` = "已签收"
                    			GROUP BY 币种,年月,物流方式,父级分类,旬
                    			ORDER BY 币种,年月
                    		    ) sl_cx_zf_qs
                    		ON sl_cx_zf_qs.`币种` = sl_zong.`币种` 
                    				AND sl_cx_zf_qs.`年月` = sl_zong.`年月`
                    				AND sl_cx_zf_qs.`物流方式` = sl_zong.`物流方式`
                    				AND sl_cx_zf_qs.`父级分类` = sl_zong.`父级分类`
                    				AND sl_cx_zf_qs.`旬` = sl_zong.`旬`	 		
                    		LEFT JOIN
                    				(SELECT 币种,
                    						年月,
                    						物流方式,
                    						父级分类,
                    						旬,
                    						COUNT(`订单编号`) 直发拒收订单量
                    					FROM  {0}	sl_cx_zf_jushou
                    					WHERE sl_cx_zf_jushou.`币种` = '{1}' 
        									AND (sl_cx_zf_jushou.`记录时间`= '{2}' AND sl_cx_zf_jushou.`日期` < DATE_SUB('{2}', INTERVAL DAY('{2}')-1 DAY) 
        										OR sl_cx_zf_jushou.`记录时间`= '{3}' AND sl_cx_zf_jushou.`日期` < DATE_SUB('{3}', INTERVAL DAY('{3}')-1 DAY))
                    						AND sl_cx_zf_jushou.`父级分类` IS NOT NULL
                    						AND sl_cx_zf_jushou.`系统订单状态` IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)') 
                    						AND sl_cx_zf_jushou.`是否改派` = "直发"
                    						AND sl_cx_zf_jushou.`最终状态` IN ('拒收', '理赔', '已签收', '已退货') 
                    					GROUP BY 币种,年月,物流方式,父级分类,旬
                    					ORDER BY 币种,年月
                    				) sl_cx_zf_js
                            ON sl_cx_zf_js.`币种` = sl_zong.`币种` 
                    			AND sl_cx_zf_js.`年月` = sl_zong.`年月`
                    			AND sl_cx_zf_js.`物流方式` = sl_zong.`物流方式`
                    			AND sl_cx_zf_js.`父级分类` = sl_zong.`父级分类` 
                    			AND sl_cx_zf_js.`旬` = sl_zong.`旬` 												
                    		GROUP BY sl_zong.年月,sl_zong.物流方式,sl_zong.父级分类,sl_zong.旬
                    with rollup) sl_rb;'''.format(match2[team], team, Time_day[11], Time_day[10])
        listT.append(sqlWl3)
        show_name.append(' 月（各月）物流…………')

        # 月时效（天）---查询
        sqltime2 = '''SELECT sl_rb.`币种`,
				sl_rb.`年月`,
				sl_rb.`物流方式`,
				sl_rb.`父级分类`,
				sl_rb.`旬`,
				sl_rb.`总单量`,
				sl_rb.`直发下单出库单量`,
				IFNULL(sl_rb.`直发下单出库时效`,0) 下单出库时效,
				sl_rb.`直发出库完成单量`,
				IFNULL(sl_rb.`直发出库完成时效`,0) 出库完成时效,
				sl_rb.`直发下单完成单量`,
				IFNULL(sl_rb.`直发下单完成时效`,0) 下单完成时效,
				null AS 改派下单完成单量,
				null AS 改派下单完成时效,
				IFNULL(sl_rb.`直发已签收订单量` / sl_rb.`直发下单完成单量`,0)  AS '签收/完成',
				IFNULL(sl_rb.`直发已签收订单量` / sl_rb.`直发下单出库单量`,0) AS '签收/总计'
            FROM( SELECT sl_zong.币种 币种,
						IFNULL(sl_zong.年月,'合计') 年月,
						IFNULL(sl_zong.物流方式,'合计') 物流方式,
						IFNULL(sl_zong.父级分类,'合计') 父级分类,
						IFNULL(sl_zong.旬,'合计') 旬,
						SUM(sl_zong.`总订单量`) 总单量,
						SUM(IFNULL(sl_cx_zf_qs.`直发已签收订单量`,0)) 直发已签收订单量,
						SUM(IFNULL(sl_zong_zf.`直发订单量`,0)) 直发下单出库单量,
						SUM(IFNULL(sl_zong_zf.`直发下单-出库时`,0)) / SUM(IFNULL(sl_zong_zf.`直发订单量`,0)) 直发下单出库时效,
						SUM(IFNULL(sl_cx_zf_wc.`直发出库完成量`,0)) 直发出库完成单量,
						SUM(IFNULL(sl_cx_zf_wc.`直发出库-完成时`,0)) / SUM(IFNULL(sl_cx_zf_wc.`直发出库完成量`,0)) 直发出库完成时效,
						SUM(IFNULL(sl_cx_zf_wc.`直发下单完成量`,0)) 直发下单完成单量,
						SUM(IFNULL(sl_cx_zf_wc.`直发下单-完成时`,0)) /SUM(IFNULL(sl_cx_zf_wc.`直发下单完成量`,0)) 直发下单完成时效
			    FROM (SELECT  币种,
										年月,
										物流方式,
										父级分类,
										旬,
										COUNT(`订单编号`) 总订单量
						FROM  {0}	sl_cx
						WHERE sl_cx.`币种` = '{1}'
								AND (sl_cx.`记录时间`= '{2}' AND sl_cx.`日期` < DATE_SUB('{2}', INTERVAL DAY('{2}')-1 DAY)
								    or sl_cx.`记录时间`= '{3}' AND sl_cx.`日期` < DATE_SUB('{3}', INTERVAL DAY('{3}')-1 DAY))				
						        AND sl_cx.`是否改派` = "直发"
								AND sl_cx.`父级分类` IS NOT NULL
								AND sl_cx.`系统订单状态` IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)')
						GROUP BY 币种,年月,物流方式,父级分类,旬
						ORDER BY 币种,年月
						) sl_zong
			    LEFT JOIN
						(SELECT 币种,
										年月,
										物流方式,
										父级分类,
										旬,
										COUNT(`订单编号`) 直发订单量,
										SUM(DATEDIFF(`仓储扫描时间`,`下单时间`)) AS '直发下单-出库时'
						FROM  {0}	sl_cx_zf
						WHERE sl_cx_zf.`币种` = '{1}' 
								AND (sl_cx_zf.`记录时间`= '{2}' AND sl_cx_zf.`日期` < DATE_SUB('{2}', INTERVAL DAY('{2}')-1 DAY)
								    or sl_cx_zf.`记录时间`= '{3}' AND sl_cx_zf.`日期` < DATE_SUB('{3}', INTERVAL DAY('{3}')-1 DAY))
								AND sl_cx_zf.`父级分类` IS NOT NULL
								AND sl_cx_zf.`系统订单状态` IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)')
								AND sl_cx_zf.`是否改派` = "直发"
								AND sl_cx_zf.`仓储扫描时间` is not null
						GROUP BY 币种,年月,物流方式,父级分类,旬
						ORDER BY 币种,年月
					) sl_zong_zf
					 ON sl_zong_zf.`币种` = sl_zong.`币种` 
							AND sl_zong_zf.`年月` = sl_zong.`年月`
							AND sl_zong_zf.`物流方式` = sl_zong.`物流方式`
						  	AND sl_zong_zf.`父级分类` = sl_zong.`父级分类` 
							AND sl_zong_zf.`旬` = sl_zong.`旬` 
					LEFT JOIN
						(SELECT 币种,
										年月,
										物流方式,
										父级分类,
										旬,
										COUNT(`订单编号`) 直发已签收订单量
						FROM  {0}	sl_cx_zf_qianshou
						WHERE sl_cx_zf_qianshou.`币种` = '{1}' 
								AND (sl_cx_zf_qianshou.`记录时间`= '{2}' AND sl_cx_zf_qianshou.`日期` < DATE_SUB('{2}', INTERVAL DAY('{2}')-1 DAY)
								    or sl_cx_zf_qianshou.`记录时间`= '{3}' AND sl_cx_zf_qianshou.`日期` < DATE_SUB('{3}', INTERVAL DAY('{3}')-1 DAY))
								AND sl_cx_zf_qianshou.`父级分类` IS NOT NULL
								AND sl_cx_zf_qianshou.`系统订单状态` IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)')
								AND sl_cx_zf_qianshou.`是否改派` = "直发"
								AND sl_cx_zf_qianshou.`仓储扫描时间` is not null
								AND sl_cx_zf_qianshou.`最终状态` = "已签收"
						GROUP BY 币种,年月,物流方式,父级分类,旬
						ORDER BY 币种,年月
					) sl_cx_zf_qs
					 ON sl_cx_zf_qs.`币种` = sl_zong.`币种` 
							AND sl_cx_zf_qs.`年月` = sl_zong.`年月`
							AND sl_cx_zf_qs.`物流方式` = sl_zong.`物流方式`
						  	AND sl_cx_zf_qs.`父级分类` = sl_zong.`父级分类`
							AND sl_cx_zf_qs.`旬` = sl_zong.`旬`
				LEFT JOIN
						(SELECT 币种,
										年月,
										物流方式,
										父级分类,
										旬,
										COUNT(`订单编号`) 直发出库完成量,
										SUM(DATEDIFF(IFNULL(`完结状态时间`,`状态时间`),`仓储扫描时间`)) AS '直发出库-完成时',
										COUNT(`订单编号`) 直发下单完成量,
										SUM(DATEDIFF(IFNULL(`完结状态时间`,`状态时间`),`下单时间`)) AS '直发下单-完成时'
						FROM  {0}	sl_cx_zf_wancheng
						WHERE sl_cx_zf_wancheng.`币种` = '{1}'
							AND (sl_cx_zf_wancheng.`记录时间`= '{2}' AND sl_cx_zf_wancheng.`日期` < DATE_SUB('{2}', INTERVAL DAY('{2}')-1 DAY)
								or sl_cx_zf_wancheng.`记录时间`= '{3}' AND sl_cx_zf_wancheng.`日期` < DATE_SUB('{3}', INTERVAL DAY('{3}')-1 DAY))
							AND sl_cx_zf_wancheng.`父级分类` IS NOT NULL
							AND sl_cx_zf_wancheng.`系统订单状态` IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)')
							AND sl_cx_zf_wancheng.`是否改派` = "直发"
							AND sl_cx_zf_wancheng.`最终状态`IN ('拒收', '理赔', '已签收', '已退货') 
						GROUP BY 币种,年月,物流方式,父级分类,旬
						ORDER BY 币种,年月
					) sl_cx_zf_wc
					 ON sl_cx_zf_wc.`币种` = sl_zong.`币种` 
						AND sl_cx_zf_wc.`年月` = sl_zong.`年月`
						AND sl_cx_zf_wc.`物流方式` = sl_zong.`物流方式`
						AND sl_cx_zf_wc.`父级分类` = sl_zong.`父级分类`
						AND sl_cx_zf_wc.`旬` = sl_zong.`旬`
				GROUP BY sl_zong.年月,sl_zong.物流方式,sl_zong.旬
				with rollup
			) sl_rb;'''.format(match2[team], team, Time_day[11], Time_day[10])
        listT.append(sqltime2)
        show_name.append(' 月（天）时效…………')
        # 月时效（旬）---查询
        sqltime3 = '''SELECT sl_rb.`币种`,
				sl_rb.`年月`,
				sl_rb.`旬`,
				sl_rb.`物流方式`,
				sl_rb.`父级分类`,
				sl_rb.`总单量`,
				sl_rb.`直发下单出库单量`,
				IFNULL(sl_rb.`直发下单出库时效`,0) 下单出库时效,
				sl_rb.`直发出库完成单量`,
				IFNULL(sl_rb.`直发出库完成时效`,0) 出库完成时效,
				sl_rb.`直发下单完成单量`,
				IFNULL(sl_rb.`直发下单完成时效`,0) 下单完成时效,
				null AS 改派下单完成单量,
				null AS 改派下单完成时效
            FROM(SELECT sl_zong.币种 币种,
						IFNULL(sl_zong.年月,'合计') 年月,
						IFNULL(sl_zong.旬,'合计') 旬,
						IFNULL(sl_zong.物流方式,'合计') 物流方式,
						IFNULL(sl_zong.父级分类,'合计') 父级分类,
						SUM(sl_zong.`总订单量`) 总单量,
						SUM(IFNULL(sl_zong_zf.`直发订单量`,0)) 直发下单出库单量,
						SUM(IFNULL(sl_zong_zf.`直发下单-出库时`,0)) / SUM(IFNULL(sl_zong_zf.`直发订单量`,0)) 直发下单出库时效,
						SUM(IFNULL(sl_cx_zf_wc.`直发出库完成量`,0)) 直发出库完成单量,
						SUM(IFNULL(sl_cx_zf_wc.`直发出库-完成时`,0)) / SUM(IFNULL(sl_cx_zf_wc.`直发出库完成量`,0)) 直发出库完成时效,
						SUM(IFNULL(sl_cx_zf_wc.`直发下单完成量`,0)) 直发下单完成单量,
						SUM(IFNULL(sl_cx_zf_wc.`直发下单-完成时`,0)) /SUM(IFNULL(sl_cx_zf_wc.`直发下单完成量`,0)) 直发下单完成时效
			    FROM ( SELECT 币种,
										年月,
										旬,
										物流方式,
										父级分类,
										COUNT(`订单编号`) 总订单量
						FROM  {0}	sl_cx
						WHERE sl_cx.`币种` = '{1}' 
							AND (sl_cx.`记录时间`= '{2}' AND sl_cx.`日期` < DATE_SUB('{2}', INTERVAL DAY('{2}')-1 DAY)
								or sl_cx.`记录时间`= '{3}' AND sl_cx.`日期` < DATE_SUB('{3}', INTERVAL DAY('{3}')-1 DAY))
							AND sl_cx.`父级分类` IS NOT NULL
							AND sl_cx.`是否改派` = "直发"
							AND sl_cx.`系统订单状态` IN ('已审核', '待发货', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)', '待发货转审核') 
						GROUP BY 币种,年月,旬,物流方式,父级分类
						ORDER BY 币种,年月
						) sl_zong
			    LEFT JOIN
						(SELECT 币种,
										年月,
										旬,
										物流方式,
										父级分类,			
										COUNT(`订单编号`) 直发订单量,
										SUM(DATEDIFF(`仓储扫描时间`,`下单时间`)) AS '直发下单-出库时'
						FROM  {0}	sl_cx_zf
						WHERE sl_cx_zf.`币种` = '{1}'
							AND (sl_cx_zf.`记录时间`= '{2}' AND sl_cx_zf.`日期` < DATE_SUB('{2}', INTERVAL DAY('{2}')-1 DAY)
                                or sl_cx_zf.`记录时间`= '{3}' AND sl_cx_zf.`日期` < DATE_SUB('{3}', INTERVAL DAY('{3}')-1 DAY))
							AND sl_cx_zf.`父级分类` IS NOT NULL
							AND sl_cx_zf.`系统订单状态` IN ('已审核', '待发货', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)', '待发货转审核') 
							AND sl_cx_zf.`是否改派` = "直发"
							AND sl_cx_zf.`仓储扫描时间` is not null
						GROUP BY 币种,年月,旬,物流方式,父级分类
						ORDER BY 币种,年月
					) sl_zong_zf
					 ON sl_zong_zf.`币种` = sl_zong.`币种` 
							AND sl_zong_zf.`年月` = sl_zong.`年月`
							AND sl_zong_zf.`物流方式` = sl_zong.`物流方式`
						  	AND sl_zong_zf.`父级分类` = sl_zong.`父级分类` 
							AND sl_zong_zf.`旬` = sl_zong.`旬` 
				LEFT JOIN
						(SELECT 币种,
										年月,
										旬,
										物流方式,
										父级分类,		
										COUNT(`订单编号`) 直发出库完成量,
										SUM(DATEDIFF(IFNULL(`完结状态时间`,`状态时间`),`仓储扫描时间`)) AS '直发出库-完成时',
										COUNT(`订单编号`) 直发下单完成量,
										SUM(DATEDIFF(IFNULL(`完结状态时间`,`状态时间`),`下单时间`)) AS '直发下单-完成时'
						FROM  {0}	sl_cx_zf_wancheng
						WHERE sl_cx_zf_wancheng.`币种` = '{1}'
							AND (sl_cx_zf_wancheng.`记录时间`= '{2}' AND sl_cx_zf_wancheng.`日期` < DATE_SUB('{2}', INTERVAL DAY('{2}')-1 DAY)
                                or sl_cx_zf_wancheng.`记录时间`= '{3}' AND sl_cx_zf_wancheng.`日期` < DATE_SUB('{3}', INTERVAL DAY('{3}')-1 DAY))
							AND sl_cx_zf_wancheng.`父级分类` IS NOT NULL
							AND sl_cx_zf_wancheng.`系统订单状态` IN ('已审核', '待发货', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)', '待发货转审核') 
							AND sl_cx_zf_wancheng.`是否改派` = "直发"
							AND sl_cx_zf_wancheng.`最终状态`IN ('拒收', '理赔', '已签收', '已退货') 
						GROUP BY 币种,年月,旬,物流方式,父级分类
						ORDER BY 币种,年月
					) sl_cx_zf_wc
					 ON sl_cx_zf_wc.`币种` = sl_zong.`币种` 
							AND sl_cx_zf_wc.`年月` = sl_zong.`年月`
							AND sl_cx_zf_wc.`物流方式` = sl_zong.`物流方式`
							AND sl_cx_zf_wc.`父级分类` = sl_zong.`父级分类` 
							AND sl_cx_zf_wc.`旬` = sl_zong.`旬` 	
				GROUP BY sl_zong.年月,sl_zong.旬,sl_zong.物流方式,sl_zong.父级分类
				with rollup
            ) sl_rb;'''.format(match2[team], team, Time_day[11], Time_day[10])
        listT.append(sqltime3)
        show_name.append(' 月（旬）时效…………')
        # 月时效(各月)---查询
        sqltime4 = '''SELECT sl_rb.`币种`,
				sl_rb.`年月`,
				sl_rb.`物流方式`,
				sl_rb.`父级分类`,
				sl_rb.`旬`,
				sl_rb.`总单量`,
				sl_rb.`直发下单出库单量`,
				sl_rb.`直发下单出库时效`,
				sl_rb.`直发出库完成单量`,
				sl_rb.`直发出库完成时效`,
				sl_rb.`直发下单完成时效`,
				sl_rb.`直发下单完成单量`,
				null AS 改派下单完成单量,
				null AS 改派下单完成时效,
				sl_rb.`直发已签收订单量` / sl_rb.`直发下单完成单量` AS '签收/完成',
				sl_rb.`直发已签收订单量` / sl_rb.`直发下单出库单量` AS '签收/总计'
            FROM (SELECT sl_zong.币种 币种,
						IFNULL(sl_zong.年月,'合计') 年月,
						IFNULL(sl_zong.物流方式,'合计') 物流方式,
						IFNULL(sl_zong.父级分类,'合计') 父级分类,
						IFNULL(sl_zong.旬,'合计') 旬,
						SUM(sl_zong.`总订单量`) 总单量,
						SUM(IFNULL(sl_cx_zf_qs.`直发已签收订单量`,0)) 直发已签收订单量,
						SUM(IFNULL(sl_zong_zf.`直发订单量`,0)) 直发下单出库单量,
						SUM(IFNULL(sl_zong_zf.`直发下单-出库时`,0)) / SUM(IFNULL(sl_zong_zf.`直发订单量`,0)) 直发下单出库时效,
						SUM(IFNULL(sl_cx_zf_wc.`直发出库完成量`,0)) 直发出库完成单量,
						SUM(IFNULL(sl_cx_zf_wc.`直发出库-完成时`,0)) / SUM(IFNULL(sl_cx_zf_wc.`直发出库完成量`,0)) 直发出库完成时效,
						SUM(IFNULL(sl_cx_zf_wc.`直发下单完成量`,0)) 直发下单完成单量,
						SUM(IFNULL(sl_cx_zf_wc.`直发下单-完成时`,0)) /SUM(IFNULL(sl_cx_zf_wc.`直发下单完成量`,0)) 直发下单完成时效
			    FROM (SELECT  币种,
										年月,
										物流方式,
										父级分类,
										旬,
										COUNT(`订单编号`) 总订单量
						FROM  {0}	sl_cx
						WHERE sl_cx.`币种` = '{1}' 
								AND sl_cx.`父级分类` IS NOT NULL
								AND sl_cx.`是否改派` = "直发"
								AND sl_cx.`系统订单状态` IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)') 
						GROUP BY 币种,年月,物流方式,父级分类,旬
						ORDER BY 币种,年月
						) sl_zong
			    LEFT JOIN
						(SELECT 币种,
										年月,
										物流方式,
										父级分类,
													旬,
										COUNT(`订单编号`) 直发订单量,
										SUM(DATEDIFF(`仓储扫描时间`,`下单时间`)) AS '直发下单-出库时'
						FROM  {0}	sl_cx_zf
						WHERE sl_cx_zf.`币种` = '{1}'
								AND sl_cx_zf.`父级分类` IS NOT NULL
								AND sl_cx_zf.`系统订单状态` IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)') 
								AND sl_cx_zf.`是否改派` = "直发"
								AND sl_cx_zf.`仓储扫描时间` is not null
						GROUP BY 币种,年月,物流方式,父级分类,旬
						ORDER BY 币种,年月
					) sl_zong_zf
					 ON sl_zong_zf.`币种` = sl_zong.`币种` 
							AND sl_zong_zf.`年月` = sl_zong.`年月`
							AND sl_zong_zf.`物流方式` = sl_zong.`物流方式`
						  	AND sl_zong_zf.`父级分类` = sl_zong.`父级分类` 
							AND sl_zong_zf.`旬` = sl_zong.`旬` 	
					LEFT JOIN
						(SELECT 币种,
										年月,
										物流方式,
										父级分类,
													旬,
										COUNT(`订单编号`) 直发已签收订单量
						FROM  {0}	sl_cx_zf_qianshou
						WHERE sl_cx_zf_qianshou.`币种` = '{1}' 
								AND sl_cx_zf_qianshou.`父级分类` IS NOT NULL
								AND sl_cx_zf_qianshou.`系统订单状态` IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)') 
								AND sl_cx_zf_qianshou.`是否改派` = "直发"
								AND sl_cx_zf_qianshou.`仓储扫描时间` is not null
								AND sl_cx_zf_qianshou.`最终状态` = "已签收"
						GROUP BY 币种,年月,物流方式,父级分类,旬
						ORDER BY 币种,年月
					) sl_cx_zf_qs
					 ON sl_cx_zf_qs.`币种` = sl_zong.`币种` 
							AND sl_cx_zf_qs.`年月` = sl_zong.`年月`
							AND sl_cx_zf_qs.`物流方式` = sl_zong.`物流方式`
						  	AND sl_cx_zf_qs.`父级分类` = sl_zong.`父级分类` 
							AND sl_cx_zf_qs.`旬` = sl_zong.`旬` 	
				LEFT JOIN
						(SELECT 币种,
									年月,
									物流方式,
									父级分类,
										    旬,
									COUNT(`订单编号`) 直发出库完成量,
									SUM(DATEDIFF(IFNULL(`完结状态时间`,`状态时间`),`仓储扫描时间`)) AS '直发出库-完成时',
									COUNT(`订单编号`) 直发下单完成量,
									SUM(DATEDIFF(IFNULL(`完结状态时间`,`状态时间`),`下单时间`)) AS '直发下单-完成时'
						FROM  {0}	sl_cx_zf_wancheng
						WHERE sl_cx_zf_wancheng.`币种` = '{1}'
							AND sl_cx_zf_wancheng.`父级分类` IS NOT NULL
							AND sl_cx_zf_wancheng.`系统订单状态` IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)') 
							AND sl_cx_zf_wancheng.`是否改派` = "直发"
							AND sl_cx_zf_wancheng.`最终状态`IN ('拒收', '理赔', '已签收', '已退货') 
						GROUP BY 币种,年月,物流方式,父级分类,旬
						ORDER BY 币种,年月
					) sl_cx_zf_wc
					 ON sl_cx_zf_wc.`币种` = sl_zong.`币种` 
						AND sl_cx_zf_wc.`年月` = sl_zong.`年月`
						AND sl_cx_zf_wc.`物流方式` = sl_zong.`物流方式`
						AND sl_cx_zf_wc.`父级分类` = sl_zong.`父级分类` 
						AND sl_cx_zf_wc.`旬` = sl_zong.`旬`
				GROUP BY sl_zong.年月,sl_zong.物流方式,sl_zong.父级分类,sl_zong.旬
				with rollup) sl_rb;'''.format('qsb_缓存_month', team)
        listT.append(sqltime4)
        show_name.append(' 月(各月)时效…………')
        listTValue = []                                # 查询sql的结果 存放池
        for i, sql in enumerate(listT):
            print('正在获取 ' + team + show_name[i])
            df = pd.read_sql_query(sql=sql, con=self.engine1)
            print(df)
            columns = list(df.columns)                 # 获取数据的标题名，转为列表
            columns_value = ['采购/销售额', '直发采购/销售额', '运费占比', '手续费占比', '金额签收/完成', '金额签收/总计', '金额完成占比', '数量签收/完成', '数量完成占比',
                             '签收/完成', '签收/总计', '完成占比', '总签收/完成', '总签收/总计', '退款率', '总完成占比', '直发签收/完成', '直发签收/总计', '直发完成占比',
                             '改派签收/完成', '改派签收/总计', '改派完成占比', '总签收/完成(金额)', '总签收/总计(金额)', '退款率(金额)', '总完成占比(金额)', '直发签收/完成(金额)',
                             '直发签收/总计(金额)', '直发完成占比(金额)', '改派签收/完成(金额)', '改派签收/总计(金额)', '改派完成占比(金额)', '订单品类占比', '直发采购额/销售额',
                             '花费占比', '总成本', '利润率', '改派占比', '采购占比', '广告占比', '总成本占比', '签收/完成', '签收/总计', '完成占比']
            for column_val in columns_value:
                if column_val in columns:
                    try:
                        df[column_val] = df[column_val].fillna(value=0)
                        df[column_val] = df[column_val].apply(lambda x: format(x, '.2%'))
                    except Exception as e:
                        print('修改失败：', str(Exception) + str(e) + df[column_val])
            listTValue.append(df)
        print('查询耗时：', datetime.datetime.now() - start)
        today = datetime.date.today().strftime('%Y.%m.%d')
        sheet_name = ['签率(天)_', '签率(月)_', '签率(旬)_', '签率(总)_', '物流(天)_', '物流(月)_', '时效(天)_', '时效(旬)_', '时效(总)_']  # 生成的工作表的表名
        file_Path = []                                # 发送邮箱文件使用
        filePath = ''
        if team == '日本':
            filePath = 'F:\\查询\\日本监控\\{} {}上月数据监控表.xlsx'.format(today, team)
        elif team == '泰国':
            filePath = 'F:\\查询\\泰国监控\\{} {}上月数据监控表.xlsx'.format(today, team)
        elif team == '新加坡' or team == '马来西亚' or team == '菲律宾':
            filePath = 'F:\\查询\\新马监控\\{} {}上月数据监控表.xlsx'.format(today, team)
        elif team == '香港' or team == '台湾':
            filePath = 'F:\\查询\\港台监控\\{} {}上月数据监控表.xlsx'.format(today, team)
        if os.path.exists(filePath):                  # 判断是否有需要的表格
            print("正在使用(上月)文件......")
            filePath = filePath
        else:                                         # 判断是否无需要的表格，进行初始化创建
            print("正在创建文件......")
            df0 = pd.DataFrame([])                    # 创建空的dataframe数据框
            df0.to_excel(filePath, index=False)       # 备用：可以向不同的sheet写入数据（创建新的工作表并进行写入）
            filePath = filePath
        print('正在写入excel…………')
        writer = pd.ExcelWriter(filePath, engine='openpyxl')  # 初始化写入对象
        book = load_workbook(filePath)                # 可以向不同的sheet写入数据（对现有工作表的追加）
        writer.book = book                            # 将数据写入excel中的sheet2表,sheet_name改变后即是新增一个sheet
        for i in range(len(listTValue)):
            listTValue[i].to_excel(excel_writer=writer, sheet_name=sheet_name[i] + team, index=False)
        if 'Sheet1' in book.sheetnames:               # 删除新建文档时的第一个工作表
            del book['Sheet1']
        writer.save()
        writer.close()
        print('正在运行宏…………')
        app = xl.App(visible=False, add_book=False)   # 运行宏调整
        app.display_alerts = False
        wbsht = app.books.open('D:/Users/Administrator/Desktop/新版-格式转换(工具表).xlsm')
        wbsht1 = app.books.open(filePath)
        wbsht.macro('sl_总监控运行2')()
        wbsht1.save()
        wbsht1.close()
        wbsht.close()
        app.quit()
        print('输出(监控)文件成功…………')
        file_Path.append(filePath)
        self.e.send('{} {}上月数据监控表.xlsx'.format(today, team), file_Path,
                    emailAdd[team])
        if team =='泰国':
            self.e.send('{} {}上月数据监控表.xlsx'.format(today, team), file_Path,
                        emailAdd2[team])
        print('处理耗时：', datetime.datetime.now() - start)

    # 获取签收表内容
    def readForm(self, team, startday):
        match2 = {'slgat': '港台',
                 'sltg': '泰国',
                 'slxmt': '新马',
                 'slrb': '日本'}
        match3 = {'新加坡': 'slxmt',
                  '马来西亚': 'slxmt',
                  '菲律宾': 'slxmt',
                  '新马': 'slxmt',
                  '日本': 'slrb',
                  '香港': 'slgat',
                  '台湾': 'slgat',
                  '港台': 'slgat',
                  '泰国': 'sltg'}
        start = datetime.datetime.now()
        # startday = datetime.datetime.now().strftime('%Y.%m.%d')
        path = r'F:\\查询\\订单数据'
        dirs = os.listdir(path=path)
        # ---读取execl文件---
        for dir in dirs:
            filePath = os.path.join(path, dir)
            if dir[:2] != '~$':
                if team in dir and startday in dir:
                    print(filePath)
                    self.wbsheet(filePath, match3[team], startday)
        print('处理耗时：', datetime.datetime.now() - start)
    # 工作表的订单信息
    def wbsheet(self, filePath, team, startday):
        print('---正在获取签收表的详情++++++')
        fileType = os.path.splitext(filePath)[1]
        app = xlwings.App(visible=False, add_book=False)
        app.display_alerts = False
        if 'xls' in fileType:
            wb = app.books.open(filePath, update_links=False, read_only=True)
            for sht in wb.sheets:
                try:
                    # db = sht.used_range.value
                    db = sht.used_range.options(pd.DataFrame, header=1, numbers=int, index=False).value
                    columns = list(db.columns)  # 获取数据的标题名，转为列表
                    columns_value = ['团队', '区域', '电话号码', '运单编号', '物流状态', '物流状态代码', '货物类型', '付款方式', '价格',
                                     '包裹重量', '包裹体积', '邮编', '签收表是否存在', '签收表订单编号', '签收表运单编号',
                                     '原运单号', '签收表物流状态', '添加时间', '成本价', '物流花费', '打包花费', '其它花费',
                                     '添加物流单号时间', '订单删除原因', 'Nan']
                    for column_val in columns_value:
                        if column_val in columns:
                            db.drop(labels=[column_val], axis=1, inplace=True)  # 去掉多余的旬列表
                    # print(db.columns)
                except Exception as e:
                    print('xxxx查看失败：' + sht.name, str(Exception) + str(e))
                if db is not None and len(db) > 0:
                    print('++++正在导入：' + sht.name + ' 共：' + str(len(db)) + '行',
                          'sheet共：' + str(sht.used_range.last_cell.row) + '行')
                    # 将返回的dateFrame导入数据库的临时表
                    self.writeCache(db)
                    print('++++正在更新：' + sht.name + '--->>>到总订单')
                    # 将数据库的临时表替换进指定的总表
                    self.replaceSql(team, list(db.columns), startday)
                    print('++++----->>>' + sht.name + '：订单更新完成++++')
                else:
                    print('----------数据为空导入失败：' + sht.name)
            wb.close()
        app.quit()
    # 写入临时缓存表
    def writeCache(self, dataFrame):
        dataFrame.to_sql('qsb_缓存', con=self.engine1, index=False, if_exists='replace')
    # 写入总表
    def replaceSql(self, team, dfColumns, startday):
        columns = list(dfColumns)
        columns = ', '.join(columns)
        # sql = '''INSERT IGNORE INTO qsb_{0}_copy1({1}, 记录时间) SELECT *, '{2}' 记录时间 FROM qsb_缓存; '''.format(team, columns, startday)
        sql = '''INSERT IGNORE INTO qsb_{0}({1}, 记录时间) SELECT *, '{2}' 记录时间 FROM qsb_缓存; '''.format(team, columns, startday)
        try:
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=2000)
        except Exception as e:
            print('插入失败：', str(Exception) + str(e))

if __name__ == '__main__':
    m = SltemMonitoring()
    start: datetime = datetime.datetime.now()
    match1 = {'slgat': '港台',
              'sltg': '泰国',
              'slxmt': '新马',
              'slzb': '直播团队',
              'slyn': '越南',
              'slrb': '日本'}
    # messagebox.showinfo("提示！！！", "当前查询已完成--->>> 请前往（ 输出文件 ）查看")
    # # 成本查询
    # for team in ['新加坡', '马来西亚', '日本', '香港', '台湾', '泰国']:
    # # for team in ['日本']:
    #     m.costWaybill(team)

    # -----------------------------------------------监控运行的主要程序和步骤-----------------------------------------
    # # # 测试监控运行（三）
    # for team in ['泰国']:
    # for team in ['香港', '台湾', '日本', '菲律宾', '新加坡', '马来西亚', '泰国']:
    #     m.order_Monitoring(team)    # 各月缓存
    #     m.data_Monitoring(team)     # 两月数据
    #     # m.costWaybill(team)       # 成本缓存 与 成本两月数据
    #     m.sl_Monitoring(team)       # 输出数据
    #     m.sl_Monitoring_two(team)  # 输出上月数据

    # 获取签收表内容（二）
    # startday = '2021.02.02'
    # for team in ['香港', '台湾', '日本', '新加坡', '马来西亚', '泰国']:
    #     m.readForm(team, startday)

    # 获取监控表以上传的时间---监控运行（一）
    for team in ['香港', '台湾', '日本', '菲律宾', '新加坡', '马来西亚', '泰国']:
    # for team in ['菲律宾']:
        m.check_time(team)